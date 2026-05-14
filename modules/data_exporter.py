"""
数据导出模块
支持导出为 Excel/CSV 格式
"""

import csv
import json
import logging
import os
from datetime import datetime
from typing import Dict, List, Optional, Any

logger = logging.getLogger(__name__)


class DataExporter:
    """数据导出器"""

    def __init__(self, output_dir: str = "exports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def export_to_csv(self, data: List[Dict], filename: str = None) -> str:
        """导出为 CSV 文件"""
        if not filename:
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = os.path.join(self.output_dir, filename)

        if not data:
            logger.warning("No data to export")
            return ""

        try:
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=data[0].keys())
                writer.writeheader()
                writer.writerows(data)

            logger.info(f"Exported to CSV: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Failed to export CSV: {e}")
            return ""

    def export_to_excel(self, data: List[Dict], filename: str = None, sheet_name: str = "Data") -> str:
        """导出为 Excel 文件"""
        if not filename:
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        if not data:
            logger.warning("No data to export")
            return ""

        try:
            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
            except ImportError:
                logger.error("openpyxl not installed. Run: pip install openpyxl")
                return ""

            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # 写入表头
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # 写入数据
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, key in enumerate(headers, 1):
                    value = row_data.get(key, '')
                    if isinstance(value, (dict, list)):
                        value = json.dumps(value, ensure_ascii=False)
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # 自动调整列宽
            for col_idx in range(1, len(headers) + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for cell in ws[column_letter]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filepath)
            logger.info(f"Exported to Excel: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Failed to export Excel: {e}")
            return ""

    def export_to_json(self, data: Any, filename: str = None, indent: int = 2) -> str:
        """导出为 JSON 文件"""
        if not filename:
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        filepath = os.path.join(self.output_dir, filename)

        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=indent)

            logger.info(f"Exported to JSON: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Failed to export JSON: {e}")
            return ""

    def export_table_to_excel(self, tables: List[List[List[str]]], filename: str = None) -> str:
        """导出多个表格到 Excel（每个表格一个sheet）"""
        if not filename:
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        if not tables:
            logger.warning("No tables to export")
            return ""

        try:
            from openpyxl import Workbook
            wb = Workbook()
            wb.remove(wb.active)  # 删除默认sheet

            for table_idx, table in enumerate(tables):
                if not table:
                    continue

                sheet_name = f"Table_{table_idx + 1}"
                ws = wb.create_sheet(title=sheet_name[:31])

                for row_idx, row in enumerate(table, 1):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(filepath)
            logger.info(f"Exported tables to Excel: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Failed to export tables: {e}")
            return ""


class ExportTask:
    """导出任务"""

    def __init__(self, task_data: Dict, database, exporter: DataExporter = None):
        self.task_data = task_data
        self.database = database
        self.config = task_data.get('config', {})
        self.exporter = exporter or DataExporter()

    def execute(self) -> bool:
        """执行导出任务"""
        task_id = self.task_data['id']
        logger.info(f"Starting export task: {task_id}")

        try:
            # 获取数据源
            data_source = self.config.get('data_source')
            data = self._get_data(data_source)

            if not data:
                self._log_error("No data to export")
                return False

            # 确定导出格式
            export_format = self.config.get('format', 'csv').lower()

            # 生成文件名
            filename = self.config.get('filename')
            if not filename:
                prefix = data_source.replace('scraper_', '').replace('browser_', '')[:20]
                filename = f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

            # 导出
            filepath = ""
            if export_format == 'excel' or export_format == 'xlsx':
                filepath = self.exporter.export_to_excel(data, f"{filename}.xlsx")
            elif export_format == 'json':
                filepath = self.exporter.export_to_json(data, f"{filename}.json")
            else:  # csv
                filepath = self.exporter.export_to_csv(data, f"{filename}.csv")

            if filepath:
                self._log_success(f"Exported to: {filepath}")
                return True
            else:
                self._log_error("Export failed")
                return False

        except Exception as e:
            self._log_error(str(e))
            return False

    def _get_data(self, data_source: str) -> List[Dict]:
        """获取数据"""
        if not data_source:
            return []

        # 从数据库配置中获取数据
        data = self.database.get_config(data_source)
        if data:
            # 提取实际数据
            if isinstance(data, dict):
                return [data] if 'data' not in data else data.get('data', [])

        # 直接从数据库读取
        if 'scraper' in data_source:
            # 网页抓取数据
            config_data = self.database.get_config(data_source)
            if config_data and isinstance(config_data, dict):
                raw = config_data.get('raw', {})
                if isinstance(raw, dict) and 'items' in raw:
                    items = raw.get('items', [])
                    return [{'value': item} for item in items] if isinstance(items[0], str) else items

        elif 'browser' in data_source:
            # 浏览器数据
            config_data = self.database.get_config(data_source)
            if config_data and isinstance(config_data, dict):
                data_list = config_data.get('data', [])
                if data_list and isinstance(data_list[0], list):
                    # 表格数据转换
                    headers = data_list[0] if data_list else []
                    return [
                        {headers[i]: row[i] if i < len(row) else '' for i in range(len(headers))}
                        for row in data_list[1:]
                    ]
                return data_list

        # 任务日志
        elif data_source == 'task_logs':
            cursor = self.database._get_cursor()
            cursor.execute("SELECT * FROM task_logs ORDER BY executed_at DESC LIMIT 1000")
            columns = [desc[0] for desc in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

        return []

    def _log_success(self, message: str):
        """记录成功日志"""
        self.database.add_log(self.task_data['id'], 'success', message)

    def _log_error(self, message: str):
        """记录错误日志"""
        self.database.add_log(self.task_data['id'], 'failed', message)


class ExportManager:
    """导出任务管理器"""

    def __init__(self, database):
        self.database = database
        self.exporter = DataExporter()
        self.tasks: Dict[str, ExportTask] = {}

    def create_task(self, task_data: Dict) -> ExportTask:
        """创建导出任务"""
        task_id = task_data.get('id', 'unknown')
        task = ExportTask(task_data, self.database, self.exporter)
        self.tasks[task_id] = task
        return task

    def get_task(self, task_id: str) -> Optional[ExportTask]:
        """获取任务"""
        return self.tasks.get(task_id)

    def export_now(self, data: List[Dict], format: str = 'csv', filename: str = None) -> str:
        """立即导出数据"""
        if format.lower() in ['xlsx', 'excel']:
            return self.exporter.export_to_excel(data, filename)
        elif format.lower() == 'json':
            return self.exporter.export_to_json(data, filename)
        else:
            return self.exporter.export_to_csv(data, filename)

    def export_table_now(self, tables: List[List[List[str]]], filename: str = None) -> str:
        """立即导出表格"""
        return self.exporter.export_table_to_excel(tables, filename)

    def list_exports(self) -> List[str]:
        """列出已导出的文件"""
        try:
            files = os.listdir(self.output_dir)
            files.sort(reverse=True)
            return files
        except:
            return []


def create_export_handler(export_manager: ExportManager):
    """创建导出任务处理器"""
    def handler(task):
        export_task = export_manager.create_task({
            'id': task.id,
            'name': task.name,
            'config': task.config
        })
        return export_task.execute()

    return handler