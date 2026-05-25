# -*- coding: utf-8 -*-
import openpyxl

file_path = r'C:\Users\admin\Desktop\西安神州数码科技园22F办公室装修项目-四家投标单位报价对比.xlsx'

wb = openpyxl.load_workbook(file_path)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n\n===== Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols) =====\n')

    for row in range(1, ws.max_row + 1):
        row_data = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val is not None:
                row_data.append(f'Col{col}:{val}')
        if row_data:
            print(f'R{row}: ' + ' | '.join(row_data))