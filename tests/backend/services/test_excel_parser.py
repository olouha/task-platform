"""
测试 Excel 解析服务 ExcelParserService
测试 Excel 文件解析、sheet 处理、数据合并等功能
"""

import pytest
import sys
import os
import tempfile
from pathlib import Path
from typing import Dict, Any, List

sys.path.insert(0, 'web/backend')

from services.excel_parser_service import ExcelParserService


class TestExcelParserServiceInit:
    """ExcelParserService 初始化测试"""

    def test_init_with_valid_file(self):
        """有效文件路径应成功初始化"""
        import openpyxl
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            wb.create_sheet('汇总')
            wb.create_sheet('明细')
            wb.save(tmp.name)
            tmp_path = tmp.name

        try:
            parser = ExcelParserService(tmp_path)
            assert parser.file_path.exists()
            assert parser.file_path.name.endswith('.xlsx')
        finally:
            os.unlink(tmp_path)

    def test_init_with_invalid_file_raises(self):
        """无效文件路径应抛出 FileNotFoundError"""
        with pytest.raises(FileNotFoundError):
            ExcelParserService('/nonexistent/path/file.xlsx')


class TestColumnMapping:
    """列映射测试"""

    def test_summary_column_mapping_defined(self):
        """汇总 sheet 列映射应已定义"""
        mapping = ExcelParserService.SUMMARY_COLUMN_MAPPING
        assert "项目名称" in mapping
        assert "业态" in mapping
        assert "项目所在地" in mapping
        assert mapping["项目名称"] == "name"
        assert mapping["业态"] == "category"

    def test_detail_column_mapping_defined(self):
        """明细 sheet 列映射应已定义"""
        mapping = ExcelParserService.DETAIL_COLUMN_MAPPING
        assert "项目名称" in mapping
        assert "单方造价（元/m2）" in mapping
        assert mapping["项目名称"] == "name"
        assert mapping["单方造价（元/m2）"] == "unit_cost"


class TestParseSummarySheet:
    """解析汇总 sheet 测试"""

    def test_parse_summary_with_valid_data(self):
        """解析有效汇总数据"""
        import openpyxl
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.create_sheet('汇总')
            # 表头
            ws.append(["序号", "项目名称", "业态", "项目所在地", "结构形式", "总面积（m2）"])
            # 数据行
            ws.append([1, "测试项目1", "住宅", "山东烟台", "框架结构", 25000])
            ws.append([2, "测试项目2", "商业", "北京", "框架剪力墙", 50000])

            wb.create_sheet('明细')
            wb.save(tmp.name)
            tmp_path = tmp.name

        try:
            parser = ExcelParserService(tmp_path)
            result = parser.parse()
            assert result["success"] is True
            assert len(result["projects"]) == 2
        finally:
            os.unlink(tmp_path)

    def test_parse_summary_empty_sheet(self):
        """解析空 sheet 应返回空列表"""
        import openpyxl
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.create_sheet('汇总')
            ws.append([])  # 只有空行

            wb.create_sheet('明细')
            wb.save(tmp.name)
            tmp_path = tmp.name

        try:
            parser = ExcelParserService(tmp_path)
            result = parser.parse()
            assert result["success"] is True
            assert len(result["projects"]) == 0
        finally:
            os.unlink(tmp_path)


class TestFloorInfoParsing:
    """楼层信息解析测试"""

    def test_parse_floor_info_above_below(self):
        """解析地上地下楼层信息"""
        parser = ExcelParserService.__new__(ExcelParserService)
        above, below = parser._parse_floor_info("地上30/地下3")
        assert above == 30
        assert below == 3

    def test_parse_floor_info_above_only(self):
        """解析只有地上的楼层信息"""
        parser = ExcelParserService.__new__(ExcelParserService)
        above, below = parser._parse_floor_info("地上20层")
        assert above == 20
        assert below is None

    def test_parse_floor_info_below_only(self):
        """解析只有地下的楼层信息"""
        parser = ExcelParserService.__new__(ExcelParserService)
        above, below = parser._parse_floor_info("地下2层")
        assert above is None
        assert below == 2

    def test_parse_floor_info_with_spaces(self):
        """解析带空格的楼层信息"""
        parser = ExcelParserService.__new__(ExcelParserService)
        above, below = parser._parse_floor_info("地上  15 / 地下 2")
        assert above == 15
        assert below == 2


class TestCellValueCleaning:
    """单元格值清理测试"""

    def test_clean_none(self):
        """None 值应返回 None"""
        parser = ExcelParserService.__new__(ExcelParserService)
        result = parser._clean_cell_value(None)
        assert result is None

    def test_clean_empty_string(self):
        """空字符串应返回 None"""
        parser = ExcelParserService.__new__(ExcelParserService)
        result = parser._clean_cell_value("")
        assert result is None

    def test_clean_whitespace_string(self):
        """空白字符串应清理后返回"""
        parser = ExcelParserService.__new__(ExcelParserService)
        result = parser._clean_cell_value("  测试  ")
        assert result == "测试"

    def test_clean_ref_error(self):
        """#REF! 错误应返回 None"""
        parser = ExcelParserService.__new__(ExcelParserService)
        result = parser._clean_cell_value("#REF!")
        assert result is None

    def test_clean_other_hash_error(self):
        """其他 # 开头的错误应返回 None"""
        parser = ExcelParserService.__new__(ExcelParserService)
        result = parser._clean_cell_value("#DIV/0!")
        assert result is None

    def test_clean_normal_string(self):
        """普通字符串应正常返回"""
        parser = ExcelParserService.__new__(ExcelParserService)
        result = parser._clean_cell_value("框架结构")
        assert result == "框架结构"

    def test_clean_number(self):
        """数字应正常返回"""
        parser = ExcelParserService.__new__(ExcelParserService)
        result = parser._clean_cell_value(2500.5)
        assert result == 2500.5


class TestEmptyRowDetection:
    """空行检测测试"""

    def test_is_empty_row_all_none(self):
        """全 None 行应为空"""
        parser = ExcelParserService.__new__(ExcelParserService)
        row = (None, None, None)
        assert parser._is_empty_row(row) is True

    def test_is_empty_row_all_empty_strings(self):
        """全空字符串行应为空"""
        parser = ExcelParserService.__new__(ExcelParserService)
        row = ("", "  ", "")
        assert parser._is_empty_row(row) is True

    def test_is_empty_row_with_data(self):
        """有数据的行应不为空"""
        parser = ExcelParserService.__new__(ExcelParserService)
        row = ("项目名", "业态", "")
        assert parser._is_empty_row(row) is False


class TestColumnLetterHelper:
    """列字母辅助函数测试"""

    def test_column_letter_a(self):
        """A 列"""
        result = ExcelParserService.get_column_letter(0)
        assert result == "A"

    def test_column_letter_z(self):
        """Z 列"""
        result = ExcelParserService.get_column_letter(25)
        assert result == "Z"

    def test_column_letter_aa(self):
        """AA 列"""
        result = ExcelParserService.get_column_letter(26)
        assert result == "AA"

    def test_column_letter_az(self):
        """AZ 列"""
        result = ExcelParserService.get_column_letter(51)
        assert result == "AZ"

    def test_column_letter_ba(self):
        """BA 列"""
        result = ExcelParserService.get_column_letter(52)
        assert result == "BA"


class TestValidateExcelStructure:
    """Excel 结构验证测试"""

    def test_validate_valid_excel(self):
        """验证有效 Excel 结构"""
        import openpyxl
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            wb.create_sheet('汇总')
            wb.create_sheet('明细')
            wb.save(tmp.name)
            tmp_path = tmp.name

        try:
            result = ExcelParserService.validate_excel_structure(tmp_path)
            assert result["valid"] is True
            assert result["has_summary"] is True
            assert result["has_detail"] is True
            assert len(result["errors"]) == 0
        finally:
            os.unlink(tmp_path)

    def test_validate_missing_summary_sheet(self):
        """验证缺少汇总 sheet"""
        import openpyxl
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            wb.create_sheet('明细')  # 只有明细
            wb.create_sheet('其他')
            wb.save(tmp.name)
            tmp_path = tmp.name

        try:
            result = ExcelParserService.validate_excel_structure(tmp_path)
            assert result["valid"] is False
            assert result["has_summary"] is False
            assert "汇总" in result["errors"][0]
        finally:
            os.unlink(tmp_path)

    def test_validate_missing_detail_sheet(self):
        """验证缺少明细 sheet"""
        import openpyxl
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            wb.create_sheet('汇总')  # 只有汇总
            wb.create_sheet('其他')
            wb.save(tmp.name)
            tmp_path = tmp.name

        try:
            result = ExcelParserService.validate_excel_structure(tmp_path)
            assert result["valid"] is False
            assert result["has_detail"] is False
            assert "明细" in result["errors"][0]
        finally:
            os.unlink(tmp_path)


class TestParseResultStructure:
    """解析结果结构测试"""

    def test_parse_result_has_required_fields(self):
        """解析结果应包含必要字段"""
        import openpyxl
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.create_sheet('汇总')
            ws.append(["序号", "项目名称", "业态", "项目所在地", "结构形式", "总面积（m2）"])
            ws.append([1, "测试项目", "住宅", "山东烟台", "框架结构", 25000])

            wb.create_sheet('明细')
            wb.save(tmp.name)
            tmp_path = tmp.name

        try:
            parser = ExcelParserService(tmp_path)
            result = parser.parse()
            assert "success" in result
            assert "projects" in result
            assert "metadata" in result
            assert "source_file" in result["metadata"]
            assert "entry_date" in result["metadata"]
        finally:
            os.unlink(tmp_path)


class TestMergeData:
    """数据合并测试"""

    def test_merge_summary_and_detail(self):
        """汇总和明细数据应正确合并"""
        import openpyxl
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()

            # 汇总 sheet
            ws_summary = wb.create_sheet('汇总')
            ws_summary.append(["序号", "项目名称", "业态", "项目所在地", "结构形式", "总面积（m2）"])
            ws_summary.append([1, "测试项目", "住宅", "山东烟台", "框架结构", 25000])

            # 明细 sheet
            ws_detail = wb.create_sheet('明细')
            ws_detail.append(["序号", "项目名称", "业态", "项目所在地", "结构形式", "总面积（m2）", "单方造价（元/m2）"])
            ws_detail.append([1, "测试项目", "住宅", "山东烟台", "框架结构", 25000, 2350])

            wb.save(tmp.name)
            tmp_path = tmp.name

        try:
            parser = ExcelParserService(tmp_path)
            result = parser.parse()
            assert result["success"] is True
            assert len(result["projects"]) >= 1
            project = result["projects"][0]
            # 应合并汇总和明细的数据
            assert "name" in project
            assert "category" in project
        finally:
            os.unlink(tmp_path)
