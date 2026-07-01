"""
修复后的导入脚本
"""
import sqlite3
import openpyxl
from pathlib import Path
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

EXCEL_FILE = Path("services/data/山东烟台钢筋价格_完整版_数据+截图.xlsx")
DB_FILE = Path("data/yantai_rebar.db")

def import_all_data():
    """导入所有年份数据"""
    logger.info(f"[import] 开始导入数据")
    
    # 清空表
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM rebar_prices")
    conn.commit()
    logger.info("[import] 已清空旧数据")
    
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    
    inserted = 0
    errors = []
    
    for sheet_name in wb.sheetnames:
        # 只处理日期格式的sheet
        if not sheet_name.startswith("2024") and not sheet_name.startswith("2025") and not sheet_name.startswith("2026"):
            continue
        
        try:
            ws = wb[sheet_name]
            
            # 从第3行开始（第1、2行是标题）
            for row_num in range(3, ws.max_row + 1):
                material_name = ws.cell(row=row_num, column=1).value
                spec = ws.cell(row=row_num, column=2).value
                brand = ws.cell(row=row_num, column=3).value
                price = ws.cell(row=row_num, column=4).value
                
                # 跳过空行或无效数据
                if not material_name or not str(material_name).strip():
                    continue
                if not price:
                    continue
                
                # 解析材质类型
                material_type = "未知"
                name_str = str(material_name)
                if "高线" in name_str or "HPB" in name_str:
                    material_type = "HPB300"
                elif "螺纹" in name_str or "HRB" in name_str:
                    material_type = "HRB400E"
                
                try:
                    cursor.execute("""
                        INSERT INTO rebar_prices (date, material_name, spec, material_type, brand, price, region)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (sheet_name, str(material_name), str(spec) if spec else "", material_type, str(brand) if brand else "", int(float(price)), "山东烟台"))
                    inserted += 1
                except Exception as e:
                    errors.append(f"{sheet_name}: {e}")
            
            if inserted % 1000 == 0 and inserted > 0:
                conn.commit()
                logger.info(f"[import] 已导入 {inserted} 条")
                
        except Exception as e:
            logger.error(f"[import] 处理失败 {sheet_name}: {e}")
    
    conn.commit()
    conn.close()
    wb.close()
    
    logger.info(f"[import] 导入完成，共 {inserted} 条")
    if errors:
        logger.info(f"[import] 错误数: {len(errors)}")

if __name__ == "__main__":
    import_all_data()
