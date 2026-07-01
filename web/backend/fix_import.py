"""
从Excel文件导入钢筋价格数据到SQLite数据库
修复列索引问题
"""
import sqlite3
import openpyxl
from pathlib import Path
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# 配置
EXCEL_FILE = Path("services/data/山东烟台钢筋价格_完整版_数据+截图.xlsx")
DB_FILE = Path("data/yantai_rebar.db")

def import_data():
    """从Excel导入数据"""
    logger.info(f"[import] 开始导入数据")
    logger.info(f"Excel: {EXCEL_FILE}")
    logger.info(f"DB: {DB_FILE}")

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    inserted = 0
    skipped = 0
    errors = []
    
    # 获取所有sheet (跳过第一个)
    all_sheets = wb.sheetnames[1:] if len(wb.sheetnames) > 1 else wb.sheetnames
    
    for sheet_name in all_sheets:
        if not sheet_name.startswith(("2024", "2025", "2026")):
            continue
            
        try:
            ws = wb[sheet_name]
            date = sheet_name
            
            # 从第3行开始读取数据（跳过标题行）
            for row_num in range(3, ws.max_row + 1):
                material_name = ws.cell(row=row_num, column=1).value
                spec = ws.cell(row=row_num, column=2).value
                brand = ws.cell(row=row_num, column=3).value
                price = ws.cell(row=row_num, column=4).value  # 上午价格在第4列！
                
                # 跳过空行
                if not material_name or not price:
                    continue
                
                # 解析材质
                material_type = "未知"
                if "高线" in str(material_name):
                    material_type = "HPB300"
                elif "螺纹" in str(material_name):
                    material_type = "HRB400E"
                
                try:
                    cursor.execute("""
                        INSERT OR REPLACE INTO rebar_prices
                        (date, material_name, spec, material_type, brand, price, region)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (date, material_name, spec, material_type, brand, int(price), "山东烟台"))
                    inserted += 1
                except sqlite3.IntegrityError:
                    skipped += 1
                except Exception as e:
                    errors.append(f"{sheet_name} row {row_num}: {e}")
            
            if (all_sheets.index(sheet_name) + 1) % 20 == 0:
                conn.commit()
                logger.info(f"[import] 进度: {all_sheets.index(sheet_name) + 1}/{len(all_sheets)}")
                
        except Exception as e:
            logger.error(f"[import] 处理失败 {sheet_name}: {e}")
            errors.append(f"{sheet_name}: {e}")

    conn.commit()
    conn.close()
    wb.close()

    logger.info(f"[import] 导入完成")
    logger.info(f"  插入: {inserted}条")
    logger.info(f"  跳过: {skipped}条")
    if errors:
        for err in errors[:5]:
            logger.error(f"  错误: {err}")
    
    # 验证
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM rebar_prices WHERE price IS NOT NULL")
    print(f"价格为NULL: {cursor.fetchone()[0]}")
    cursor.execute("SELECT COUNT(*) FROM rebar_prices WHERE price IS NOT NULL")
    print(f"价格有值: {cursor.fetchone()[0]}")
    conn.close()

if __name__ == "__main__":
    import_data()
