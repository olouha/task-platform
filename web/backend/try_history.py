"""尝试通过修改URL获取历史数据"""
import asyncio
import json
from pathlib import Path
from playwright.async_api import async_playwright

DATA_DIR = Path(r'E:\E\任务\task-platform\web\backend\services\data')

async def try_history_urls():
    """尝试历史URL"""
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(viewport={'width': 1920, 'height': 3000}, locale='zh-CN')
        page = await context.new_page()

        # 加载 Cookie
        cookie_file = DATA_DIR / 'mysteel_cookies.json'
        if cookie_file.exists():
            with open(cookie_file, 'r') as f:
                cookies = json.load(f)
            await context.add_cookies(cookies)

        # 尝试 5月13日的不同URL格式
        possible_urls = [
            'https://jiancai.mysteel.com/m/26051310/C7E274318523C3AE.html',  # 之前成功的
            'https://jiancai.mysteel.com/m/26051316/06AC8B0B0D2BB9BF.html',  # 使用相同hash
            'https://jiancai.mysteel.com/m/26051316/C7E274318523C3AE.html',  # 混合
        ]

        for url in possible_urls:
            print(f"\n尝试: {url}")
            try:
                await page.goto(url, wait_until='domcontentloaded', timeout=30000)
                await page.wait_for_timeout(5000)

                # 检查是否有数据
                data = await page.evaluate('''() => {
                    const tables = document.querySelectorAll('table');
                    return {
                        tableCount: tables.length,
                        hasData: document.body.textContent.includes('螺纹钢'),
                        text: document.body.textContent?.slice(0, 200) || ''
                    };
                }''')

                if data['hasData']:
                    print(f"  [OK] 有数据! 表格数: {data['tableCount']}")

                    # 提取数据
                    prices = await page.evaluate('''() => {
                        const tables = document.querySelectorAll('table');
                        const allPrices = [];

                        tables.forEach((table) => {
                            const rows = table.querySelectorAll('tr');
                            rows.forEach((row) => {
                                const cells = row.querySelectorAll('td');
                                if (cells.length >= 4) {
                                    const rowData = [];
                                    cells.forEach(c => rowData.push(c.textContent?.trim() || ''));

                                    if ((rowData[0].includes('螺纹钢') || rowData[0].includes('盘螺') ||
                                         rowData[0].includes('高线') || rowData[0].includes('圆钢'))) {
                                        for (let i = 0; i < rowData.length; i++) {
                                            if (/^\\d{3,4}$/.test(rowData[i])) {
                                                allPrices.push({
                                                    material_name: rowData[0] || '',
                                                    spec: rowData[1] || '',
                                                    material_type: rowData[2] || '',
                                                    brand: rowData[3] || '',
                                                    price: rowData[i],
                                                    fullRow: rowData.slice(0, 8)
                                                });
                                                break;
                                            }
                                        }
                                    }
                                }
                            });
                        });

                        return allPrices;
                    }''')

                    print(f"  提取到 {len(prices)} 条数据")

                    # 保存到 Excel
                    if prices:
                        import openpyxl
                        from datetime import datetime
                        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                        from openpyxl.utils import get_column_letter

                        excel_file = DATA_DIR / '山东烟台钢筋价格.xlsx'
                        wb = openpyxl.load_workbook(excel_file)

                        # 创建5月13日的sheet
                        sheet_name = '2026-05-13_180000'
                        if sheet_name in wb.sheetnames:
                            del wb[sheet_name]

                        ws = wb.create_sheet(title=sheet_name)

                        # 样式
                        header_font = Font(bold=True, size=12, color='FFFFFF')
                        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        thin_border = Border(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )

                        # 标题
                        ws.merge_cells('A1:K1')
                        ws.cell(row=1, column=1, value='山东烟台钢筋价格 - 2026-05-13 (历史恢复)').font = Font(bold=True, size=14)
                        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

                        # 表头
                        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=3, column=col, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = thin_border

                        # 数据
                        row = 4
                        for p in prices:
                            data = [
                                '2026-05-13', '18:00:00',
                                p['material_name'], p['spec'], p['material_type'], p['brand'],
                                p['price'], '', '', '', '山东烟台'
                            ]
                            for col, value in enumerate(data, 1):
                                cell = ws.cell(row=row, column=col, value=value)
                                cell.border = thin_border
                            row += 1

                        # 列宽
                        widths = [12, 12, 10, 10, 12, 14, 12, 10, 25, 10, 10]
                        for i, w in enumerate(widths, 1):
                            ws.column_dimensions[get_column_letter(i)].width = w

                        wb.save(excel_file)
                        wb.close()
                        print(f"  [OK] 数据已保存到 Excel")
                        break  # 成功后退出

                else:
                    print(f"  [X] 无数据: {data['text'][:100]}")

            except Exception as e:
                print(f"  错误: {e}")

        await browser.close()

if __name__ == '__main__':
    asyncio.run(try_history_urls())