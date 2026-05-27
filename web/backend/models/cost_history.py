"""
造价管理 - 历史参考价格数据
从Excel截图OCR识别结果导入
支持按年份和季度查询
"""

from typing import List, Dict, Optional
from collections import defaultdict

# ============================================================
# 钢筋价格历史数据
# 格式: {年份: {季度: [{spec, grade, size, price}]}}
# ============================================================

STEEL_REBAR_HISTORY: Dict[str, Dict[str, List[Dict]]] = {}

# ============================================================
# 混凝土价格历史数据
# 格式: {年份: {季度: [{grade, yantai, rushan}]}}
# ============================================================

CONCRETE_HISTORY: Dict[str, Dict[str, List[Dict]]] = {}


def load_concrete_from_excel():
    """从Excel加载混凝土历史数据"""
    from openpyxl import load_workbook
    import os

    excel_path = r'C:\Users\admin\Desktop\近五年钢筋混凝土造价管理截图(1)\近五年钢筋混凝土造价管理截图\造价参考价数据.xlsx'

    if not os.path.exists(excel_path):
        print(f"Excel文件不存在: {excel_path}")
        return

    wb = load_workbook(excel_path)
    ws = wb['混凝土信息价']

    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
        year, quarter, grade, yantai, rushan = row[:5]
        if not year or not quarter or not grade:
            continue

        year = str(year)
        if year not in CONCRETE_HISTORY:
            CONCRETE_HISTORY[year] = {}
        if quarter not in CONCRETE_HISTORY[year]:
            CONCRETE_HISTORY[year][quarter] = []

        item = {
            'grade': str(grade),
            'yantai': float(yantai) if yantai else None,
            'rushan': float(rushan) if rushan else None
        }
        CONCRETE_HISTORY[year][quarter].append(item)

    wb.close()
    print(f"加载混凝土数据: {len(CONCRETE_HISTORY)} 年")


def load_steel_from_excel():
    """从Excel加载钢筋历史数据"""
    from openpyxl import load_workbook
    import os

    excel_path = r'C:\Users\admin\Desktop\近五年钢筋混凝土造价管理截图(1)\近五年钢筋混凝土造价管理截图\造价参考价数据.xlsx'

    if not os.path.exists(excel_path):
        print(f"Excel文件不存在: {excel_path}")
        return

    wb = load_workbook(excel_path)
    ws = wb['钢筋信息价']

    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
        year, quarter, grade, size, price = row[:5]
        if not year or not quarter or not grade:
            continue

        year = str(year)
        if year not in STEEL_REBAR_HISTORY:
            STEEL_REBAR_HISTORY[year] = {}
        if quarter not in STEEL_REBAR_HISTORY[year]:
            STEEL_REBAR_HISTORY[year][quarter] = []

        item = {
            'grade': str(grade),
            'size': str(size),
            'price': float(price) if price else None,
            'spec': f"{grade}Φ{size}"
        }
        STEEL_REBAR_HISTORY[year][quarter].append(item)

    wb.close()
    print(f"加载钢筋数据: {len(STEEL_REBAR_HISTORY)} 年")


def get_available_periods() -> List[Dict]:
    """获取所有可用时期"""
    periods = []
    for year in sorted(CONCRETE_HISTORY.keys()):
        for quarter in sorted(CONCRETE_HISTORY[year].keys()):
            periods.append({
                'year': year,
                'quarter': quarter,
                'label': f"{year}年{quarter}",
                'concrete_count': len(CONCRETE_HISTORY[year][quarter])
            })

    # 添加钢筋数据但混凝土没有的时期
    for year in sorted(STEEL_REBAR_HISTORY.keys()):
        for quarter in sorted(STEEL_REBAR_HISTORY[year].keys()):
            label = f"{year}年{quarter}"
            if not any(p['label'] == label for p in periods):
                periods.append({
                    'year': year,
                    'quarter': quarter,
                    'label': label,
                    'concrete_count': 0,
                    'rebar_count': len(STEEL_REBAR_HISTORY[year][quarter])
                })

    return sorted(periods, key=lambda x: (x['year'], x['quarter']))


def get_concrete_prices(year: str, quarter: str) -> List[Dict]:
    """获取指定时期的混凝土价格"""
    if year in CONCRETE_HISTORY and quarter in CONCRETE_HISTORY[year]:
        return CONCRETE_HISTORY[year][quarter]
    return []


def get_steel_prices(year: str, quarter: str) -> List[Dict]:
    """获取指定时期的钢筋价格"""
    if year in STEEL_REBAR_HISTORY and quarter in STEEL_REBAR_HISTORY[year]:
        return STEEL_REBAR_HISTORY[year][quarter]
    return []


# 初始化时加载数据
load_concrete_from_excel()
load_steel_from_excel()
