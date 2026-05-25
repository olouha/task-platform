"""
从已知URL列表抓取历史数据
"""
import asyncio
import json
import base64
from pathlib import Path
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

DATA_DIR = Path('services/data')
COOKIE_FILE = DATA_DIR / 'mysteel_cookies.json'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
LINKS_FILE = DATA_DIR / 'yantai_links.json'

USERNAME = 'M6616592358'
PASSWORD = 'mysteel573005'

def get_sheet_name(date_str, period):
    fetch_time = datetime.now().strftime('%H%M%S')
    return f'{date_str}_{period}_{fetch_time}'

async def get_existing_sheets():
    existing_sheets = set()
    if EXCEL_FILE.exists():
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            existing_sheets = set(wb.sheetnames)
            wb.close()
        except Exception:
            pass
    return existing_sheets

async def fetch_prices(page, url):
    await page.goto(url, wait_until='domcontentloaded', timeout=60000)
    await asyncio.sleep(3)

    screenshot = await page.screenshot(full_page=True)
    screenshot_b64 = base64.b64encode(screenshot).decode('utf-8')

    data = await page.evaluate('''() => {
        const tables = document.querySelectorAll('table');
        const results = [];
        tables.forEach(table => {
            const rows = table.querySelectorAll('tr');
            rows.forEach(row => {
                const cells = row.querySelectorAll('td');
                if (cells.length >= 5) {
                    const material_name = cells[0]?.textContent?.trim();
                    const spec = cells[1]?.textContent?.trim();
                    const material_type = cells[2]?.textContent?.trim();
                    const brand = cells[3]?.textContent?.trim();
                    const price = cells[4]?.textContent?.trim();

                    if (['高线', '螺纹钢', '盘螺', '圆钢'].includes(material_name) &&
                        spec && spec.startsWith('Φ') && price && /^\d+$/.test(price)) {
                        results.push({
                            material_name,
                            spec,
                            material_type,
                            brand,
                            price: parseFloat(price),
                            region: '山东烟台'
                        });
                    }
                }
            });
        });
        return results;
    }''')

    return list(data), screenshot_b64

async def save_to_excel(prices, period, date_str, screenshot_b64):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE) if EXCEL_FILE.exists() else openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        sheet_name = get_sheet_name(date_str, period)
        ws = wb.create_sheet(title=sheet_name)

        period_text = '下午(晚)' if period == 'PM' else '上午'
        ws.merge_cells('A1:K1')
        ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        header_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid') if period == 'PM' else \n            PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        fetch_time = datetime.now().strftime('%H:%M:%S')
        for i, price in enumerate(prices):
            row = 4 + i
            for col, val in enumerate([date_str, fetch_time, price['material_name'], price['spec'],
                                       price['material_type'], price['brand'], price['price'],
                                       '', '', '', '山东烟台'], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border

        if screenshot_b64:
            screenshot_path = DATA_DIR / f'screenshot_{date_str.replace("-", "")}_{period}.png'
            with open(screenshot_path, 'wb') as f:
                f.write(base64.b64decode(screenshot_b64))

            row = 4 + len(prices) + 2
            ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)
            img = Image(str(screenshot_path))
            img.width = 900
            img.height = 500
            img.anchor = f'A{row + 1}'
            ws.add_image(img)

        wb.save(EXCEL_FILE)
        wb.close()
        return sheet_name
    except Exception as e:
        print(f'保存失败: {e}')
        import traceback
        traceback.print_exc()
        return None

async def main():
    from playwright.async_api import async_playwright

    print('=' * 60)
    print('从URL列表抓取历史数据')
    print('=' * 60)

    # 加载链接
    with open(LINKS_FILE, 'r', encoding='utf-8') as f:
        all_links = json.load(f)

    print(f'总链接数: {len(all_links)}')

    existing_sheets = await get_existing_sheets()
    print(f'已有sheet: {len(existing_sheets)}个')

    # 解析日期并过滤
    import re
    tasks = []
    for link in all_links:
        href = link['href']
        match = re.search(r'/m/(\d{8})/', href)
        if match:
            full = match.group(1)
            year = 2000 + int(full[0:2])
            if year < 2025:
                continue
            month = int(full[2:4])
            day = int(full[4:6])
            hour = int(full[6:8])

            date_str = f'{year}-{month:02d}-{day:02d}'
            period = 'AM' if hour < 12 else 'PM'

            has_existing = any(s.startswith(f'{date_str}_{period}') for s in existing_sheets)
            if not has_existing:
                tasks.append((date_str, period, href))

    print(f'需要抓取: {len(tasks)} 条\n')

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(viewport={'width': 1920, 'height': 3000}, locale='zh-CN')
        page = await context.new_page()

        # 加载cookies
        if COOKIE_FILE.exists():
            try:
                cookies = json.loads(COOKIE_FILE.read_text(encoding='utf-8'))
                if cookies:
                    await page.context.add_cookies(cookies)
                    print('已加载Cookie')
            except:
                pass

        # 登录
        print('登录中...')
        await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
        await asyncio.sleep(3)

        await page.evaluate(f'''() => {{
            const inputs = document.querySelectorAll('input');
            for (const inp of inputs) {{
                const ph = inp.placeholder || '';
                if (ph.includes(
