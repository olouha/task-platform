"""
批量OCR解析烟台钢筋价格截图
从 screenshot_*.png 文件中识别价格数据并保存到 Excel
"""

import os
import re
import json
from pathlib import Path
from datetime import datetime
from collections import Counter

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PIL import Image

# 尝试导入 ddddocr
try:
    import ddddocr
    HAS_DDDDOCR = True
except ImportError:
    HAS_DDDDOCR = False

# 配置
SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR / 'data'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
PROGRESS_FILE = DATA_DIR / 'ocr_parse_progress.json'

# 设置 pytesseract 路径（如果已安装）
TESSERACT_PATH = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
if os.path.exists(TESSERACT_PATH):
    import pytesseract
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
    HAS_TESSERACT = True
else:
    HAS_TESSERACT = False


def init_ocr():
    """初始化 OCR 引擎"""
    if HAS_DDDDOCR:
        ocr = ddddocr.DdddOcr(beta=True)
        print('[OCR] 使用 ddddocr')
        return ocr
    elif HAS_TESSERACT:
        print('[OCR] 使用 pytesseract')
        return 'pytesseract'
    else:
        print('[WARN] 无可用OCR，将使用表格提取')
        return None


def parse_screenshot_ocr(ocr, screenshot_path):
    """使用 OCR 识别截图中的价格数据"""
    try:
        img = Image.open(screenshot_path)
        img_bytes = open(screenshot_path, 'rb').read()

        if ocr == 'pytesseract':
            # 使用 pytesseract
            text = pytesseract.image_to_string(img, lang='chi_sim+eng')
            return parse_text_to_prices(text)
        else:
            # 使用 ddddocr
            result = ocr.classification(img_bytes)
            return {'raw': result, 'type': 'ddddocr'}
    except Exception as e:
        print(f'[WARN] OCR识别失败: {e}')
        return None


def parse_text_to_prices(text):
    """从OCR文本中解析价格数据"""
    prices = []

    # 匹配品名
    materials = ['高线', '螺纹钢', '盘螺', '圆钢']

    lines = text.split('\n')
    current_material = None

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # 检查是否是品名行
        for mat in materials:
            if mat in line and ('Φ' in line or re.search(r'Φ\d', line)):
                current_material = mat
                break

        # 匹配规格和价格
        # 格式: 规格 品牌 价格
        # 例如: Φ6 永锋 3930

        # 尝试多种模式
        patterns = [
            # Φ规格 品牌 价格
            r'Φ(\d+(?:-\d+)?)\s+(\S+)\s+(\d{3,4})',
            # 品牌 Φ规格 价格
            r'(\S+)\s+Φ(\d+(?:-\d+)?)\s+(\d{3,4})',
        ]

        for pattern in patterns:
            matches = re.findall(pattern, line)
            for match in matches:
                if current_material:
                    if len(match) == 3:
                        # Φ规格 品牌 价格
                        spec = f'Φ{match[0]}'
                        brand = match[1]
                        price = int(match[2])
                    else:
                        brand = match[0]
                        spec = f'Φ{match[1]}'
                        price = int(match[2])

                    prices.append({
                        'material_name': current_material,
                        'spec': spec,
                        'brand': brand,
                        'price': price
                    })

    return prices


def parse_table_from_image(img):
    """尝试从图片中提取表格数据（不依赖OCR）"""
    # 这个方法需要更复杂的图像处理
    # 目前先用 OCR
    return []


def extract_prices_from_screenshot(screenshot_path):
    """从截图提取价格数据"""
    try:
        img = Image.open(screenshot_path)

        # 使用 pytesseract（如果可用）
        if HAS_TESSERACT:
            text = pytesseract.image_to_string(img, lang='chi_sim+eng')
            return parse_text_to_prices(text)
        elif HAS_DDDDOCR:
            # ddddocr 主要用于验证码，不是通用OCR
            # 尝试截取表格区域后使用
            return []
        else:
            return []

    except Exception as e:
        print(f'[ERROR] 解析截图失败: {e}')
        return []


def load_progress():
    """加载解析进度"""
    if PROGRESS_FILE.exists():
        with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'parsed': [], 'failed': []}


def save_progress(progress):
    """保存解析进度"""
    with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def get_all_screenshots():
    """获取所有截图文件"""
    screenshots = []
    for f in os.listdir(DATA_DIR):
        if f.startswith('screenshot_') and f.endswith('.png'):
            screenshots.append(f)
    return sorted(screenshots)


def get_existing_dates():
    """获取Excel中已有的日期"""
    if not EXCEL_FILE.exists():
        return set()

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    dates = set()
    for sheet_name in wb.sheetnames:
        if '_' in sheet_name:
            date_part = sheet_name.split('_')[0]
            dates.add(date_part)
    wb.close()
    return dates


def save_prices_to_excel(prices, date_str, period, fetch_time, screenshot_path=None):
    """保存价格数据到Excel"""
    if not prices:
        return False

    try:
        # 打开或创建 workbook
        if EXCEL_FILE.exists():
            wb = openpyxl.load_workbook(EXCEL_FILE)
        else:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        # 创建 sheet 名称
        time_str = fetch_time.replace(':', '') if fetch_time else '000000'
        sheet_name = f'{date_str}_{period}_{time_str}'

        # 检查是否已存在同名sheet
        if sheet_name in wb.sheetnames:
            print(f'[SKIP] Sheet已存在: {sheet_name}')
            wb.close()
            return False

        ws = wb.create_sheet(title=sheet_name)

        # 设置样式
        header_font = Font(bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid') if period == 'AM' else PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 标题行
        period_text = '上午' if period == 'AM' else '下午'
        ws.merge_cells('A1:K1')
        ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        # 表头
        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 数据
        for i, price in enumerate(prices):
            row = 4 + i
            values = [
                date_str,
                fetch_time or '00:00:00',
                price.get('material_name', ''),
                price.get('spec', ''),
                price.get('material_type', ''),
                price.get('brand', ''),
                price.get('price', 0),
                price.get('price_change', ''),
                price.get('remark', ''),
                price.get('steel_code', ''),
                '山东烟台'
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border

        wb.save(EXCEL_FILE)
        wb.close()

        print(f'[OK] 保存: {sheet_name} ({len(prices)} 条数据)')
        return True

    except Exception as e:
        print(f'[ERROR] 保存失败: {e}')
        return False


def parse_filename(screenshot_name):
    """从截图文件名解析日期和时段"""
    # screenshot_20240105_AM.png
    parts = screenshot_name.replace('screenshot_', '').replace('.png', '').split('_')
    if len(parts) >= 2:
        date_str = parts[0]
        year = date_str[:4]
        month = date_str[4:6]
        day = date_str[6:8]
        period = parts[1]  # AM or PM
        return f'{year}-{month}-{day}', period, '09:00:00' if period == 'AM' else '15:00:00'
    return None, None, None


def run_batch_parse(limit=None, force=False):
    """批量解析截图"""
    print('=' * 60)
    print('批量OCR解析烟台钢筋价格截图')
    print('=' * 60)

    # 获取所有截图
    screenshots = get_all_screenshots()
    print(f'总截图数: {len(screenshots)}')

    # 获取已解析的日期
    existing_dates = get_existing_dates()
    print(f'Excel已有: {len(existing_dates)} 个日期')

    # 加载进度
    progress = load_progress()
    parsed = set(progress.get('parsed', []))
    failed = set(progress.get('failed', []))

    # 初始化OCR
    ocr = init_ocr()

    # 筛选待处理
    to_process = []
    for s in screenshots:
        if s in parsed:
            continue
        if not force and s in failed:
            continue

        date_str, period, _ = parse_filename(s)
        if date_str and date_str in existing_dates:
            # 日期已存在，跳过
            continue

        to_process.append(s)

    print(f'待解析: {len(to_process)} 个')

    if limit:
        to_process = to_process[:limit]

    if not to_process:
        print('没有需要解析的截图')
        return

    # 按日期排序
    to_process.sort()

    # 开始解析
    success_count = 0
    fail_count = 0

    for i, screenshot in enumerate(to_process):
        print(f'\n[{i+1}/{len(to_process)}] 解析: {screenshot}')

        screenshot_path = DATA_DIR / screenshot

        # 解析文件名
        date_str, period, fetch_time = parse_filename(screenshot)
        if not date_str:
            print(f'[SKIP] 无法解析文件名')
            failed.add(screenshot)
            fail_count += 1
            continue

        # 提取价格
        prices = extract_prices_from_screenshot(screenshot_path)

        if prices:
            # 保存到Excel
            if save_prices_to_excel(prices, date_str, period, fetch_time, screenshot_path):
                parsed.add(screenshot)
                success_count += 1
            else:
                fail_count += 1
        else:
            print(f'[WARN] 未提取到价格数据')
            failed.add(screenshot)
            fail_count += 1

        # 每10个保存一次进度
        if (i + 1) % 10 == 0:
            progress['parsed'] = list(parsed)
            progress['failed'] = list(failed)
            save_progress(progress)

    # 最终保存
    progress['parsed'] = list(parsed)
    progress['failed'] = list(failed)
    save_progress(progress)

    print('\n' + '=' * 60)
    print('完成')
    print(f'成功: {success_count}')
    print(f'失败: {fail_count}')
    print('=' * 60)


def main():
    import argparse

    parser = argparse.ArgumentParser(description='批量OCR解析烟台钢筋价格截图')
    parser.add_argument('--limit', '-n', type=int, default=None, help='限制处理数量')
    parser.add_argument('--force', '-f', action='store_true', help='强制重新解析已失败的')
    parser.add_argument('--reset', action='store_true', help='重置进度')

    args = parser.parse_args()

    if args.reset:
        if PROGRESS_FILE.exists():
            os.remove(PROGRESS_FILE)
        print('进度已重置')

    run_batch_parse(limit=args.limit, force=args.force)


if __name__ == '__main__':
    main()