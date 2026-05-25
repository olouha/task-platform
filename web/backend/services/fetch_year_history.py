"""
批量抓取一年历史价格数据
从烟台钢筋市场页面获取所有每日链接
添加反爬虫保护机制
"""

import asyncio
from playwright.async_api import async_playwright
from pathlib import Path
import json
import base64
import re
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
import random

DATA_DIR = Path('services/data')
DATA_DIR.mkdir(exist_ok=True)

COOKIE_FILE = DATA_DIR / 'mysteel_cookies.json'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
MARKET_URL = 'https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html'

# 凭据
USERNAME = 'M6616592358'
PASSWORD = 'mysteel573005'


def random_sleep(min_sec=1, max_sec=3):
    """随机睡眠"""
    return random.uniform(min_sec, max_sec)


async def human_like_mouse(page):
    """模拟人类鼠标操作"""
    try:
        # 随机移动鼠标
        viewport = page.viewport_size
        if viewport:
            x = random.randint(100, viewport['width'] - 100)
            y = random.randint(100, viewport['height'] - 100)
            await page.mouse.move(x, y)
            await asyncio.sleep(random.uniform(0.1, 0.3))

            # 再次移动
            x2 = random.randint(100, viewport['width'] - 100)
            y2 = random.randint(100, viewport['height'] - 100)
            await page.mouse.move(x2, y2)
    except:
        pass  # 忽略鼠标移动错误


def get_sheet_name(date_str, period):
    fetch_time = datetime.now().strftime('%H%M%S')
    return f'{date_str}_{period}_{fetch_time}'


def get_date_range():
    """获取从现在往前一年的日期范围"""
    end_date = datetime.now()
    start_date = end_date - timedelta(days=365)
    return start_date, end_date


def is_date_in_range(date_str, start_date, end_date):
    """检查日期是否在范围内"""
    try:
        date = datetime.strptime(date_str, '%Y-%m-%d')
        return start_date <= date <= end_date
    except:
        return False


async def get_existing_sheets():
    """获取已存在的sheet"""
    existing_sheets = set()
    if EXCEL_FILE.exists():
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            existing_sheets = set(wb.sheetnames)
            wb.close()
        except Exception:
            pass
    return existing_sheets


async def login(page):
    """登录 - 模拟人类操作"""
    print('登录中...')

    cookies = []
    if COOKIE_FILE.exists():
        try:
            cookies = json.loads(COOKIE_FILE.read_text(encoding='utf-8'))
            if cookies:
                await page.context.add_cookies(cookies)
                print('已加载Cookie')
        except:
            pass

    await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
    await asyncio.sleep(random_sleep(2, 4))

    # 模拟鼠标移动
    await human_like_mouse(page)
    await asyncio.sleep(random_sleep(0.5, 1))

    try:
        account_tab = await page.query_selector('.form-tab-account, a[data-tab="account"]')
        if account_tab:
            await account_tab.click()
            await asyncio.sleep(random_sleep(1, 2))
    except:
        pass

    await human_like_mouse(page)

    # 模拟人类输入
    await page.evaluate(f'''() => {{
        const inputs = document.querySelectorAll('input');
        for (const inp of inputs) {{
            const ph = inp.placeholder || '';
            if (ph.includes('用户名')) inp.value = '{USERNAME}';
            if (ph.includes('密码') && inp.type === 'password') inp.value = '{PASSWORD}';
        }}
    }}''')

    await asyncio.sleep(random_sleep(1, 2))
    await human_like_mouse(page)

    try:
        checkbox = await page.query_selector('input[type="checkbox"]')
        if checkbox and not await checkbox.is_checked():
            await checkbox.click()
            await asyncio.sleep(random_sleep(0.5, 1))
    except:
        pass

    await asyncio.sleep(random_sleep(0.5, 1))

    try:
        login_btn = await page.query_selector('.form-button-login, button:has-text("登录")')
        if login_btn:
            await login_btn.click()
    except:
        pass

    await asyncio.sleep(random_sleep(8, 12))

    cookies = await page.context.cookies()
    with open(COOKIE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cookies, f, ensure_ascii=False)
    print('登录完成\n')


async def fetch_all_market_links(page, max_pages=100):
    """从市场页面获取所有每日链接（支持分页）"""
    all_links = []

    # 先访问第一页
    print(f'访问市场页面...')
    await page.goto(MARKET_URL, wait_until='domcontentloaded', timeout=60000)
    await asyncio.sleep(random_sleep(3, 5))

    await human_like_mouse(page)

    # 检查是否被重定向到验证码
    if 'captcha' in page.url.lower():
        print('检测到验证码，等待5秒后重试...')
        await asyncio.sleep(5)

        # 尝试直接访问
        await page.goto(MARKET_URL, wait_until='domcontentloaded', timeout=60000)
        await asyncio.sleep(random_sleep(3, 5))

        if 'captcha' in page.url.lower():
            print('仍然被验证码拦截，尝试从首页进入...')
            await page.goto('https://jiancai.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
            await asyncio.sleep(random_sleep(3, 5))

            # 查找烟台链接
            yantai_links = await page.evaluate('''() => {
                const links = [];
                document.querySelectorAll('a[href]').forEach(a => {
                    const href = a.href;
                    const text = a.textContent.trim();
                    if (text.includes('烟台') && href.includes('market')) {
                        links.push(href);
                    }
                });
                return links;
            }''')

            if yantai_links:
                await page.goto(yantai_links[0], wait_until='domcontentloaded', timeout=60000)
                await asyncio.sleep(random_sleep(3, 5))
            else:
                print('无法找到烟台市场链接')
                return []

    for page_num in range(1, max_pages + 1):
        print(f'获取第{page_num}页链接...', end=' ')

        if page_num > 1:
            page_url = f'https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa{page_num}.html?keyWord='
        else:
            page_url = page.url

        try:
            await page.goto(page_url, wait_until='domcontentloaded', timeout=60000)
        except:
            print('访问失败')
            break

        await asyncio.sleep(random_sleep(2, 4))
        await human_like_mouse(page)

        # 只获取烟台的链接
        links = await page.evaluate('''() => {
            const links = [];
            document.querySelectorAll('a[href]').forEach(a => {
                const href = a.href;
                const text = a.textContent.trim();
                if (href.includes('jiancai.mysteel.com/m/') &&
                    href.includes('.html') &&
                    text.includes('烟台')) {
                    links.push({href: href, text: text});
                }
            });
            return links;
        }''')

        if not links:
            print('无数据，停止')
            break

        all_links.extend(links)
        print(f'{len(links)}条 (累计{len(all_links)})')

        # 检查是否有下一页
        has_next = await page.evaluate('''() => {
            const nextBtn = Array.from(document.querySelectorAll('a'))
                .find(a => a.textContent.trim() === '下一页');
            return nextBtn && nextBtn.href;
        }''')

        if not has_next:
            print('没有更多页面')
            break

    print(f'总共获取到 {len(all_links)} 个烟台链接')
    return all_links


async def parse_links_to_dates(links, start_date, end_date):
    """解析链接并按日期组织"""
    available_data = {}

    for link in links:
        href = link['href']
        # URL格式: /m/26051516/ (YYMMDDHH)
        match = re.search(r'/m/(\d{8})/', href)
        if match:
            full = match.group(1)

            try:
                year = 2000 + int(full[0:2])
                month = int(full[2:4])
                day = int(full[4:6])
                hour = int(full[6:8])

                if year < 2025:
                    continue

                date_str = f'{year}-{month:02d}-{day:02d}'

                if not is_date_in_range(date_str, start_date, end_date):
                    continue

                period = 'AM' if hour < 12 else 'PM'

                if date_str not in available_data:
                    available_data[date_str] = {}

                if period not in available_data[date_str]:
                    available_data[date_str][period] = href

            except:
                continue

    return available_data


async def fetch_prices(page, url):
    """抓取价格数据"""
    await page.goto(url, wait_until='domcontentloaded', timeout=60000)
    await asyncio.sleep(random_sleep(3, 5))

    await human_like_mouse(page)

    screenshot = await page.screenshot(full_page=True)
    screenshot_b64 = base64.b64encode(screenshot).decode('utf-8')

    data = await page.evaluate('''() => {
        const tables = document.querySelectorAll('table');
        const results = [];
        tables.forEach(table => {
            const rows = table.querySelectorAll('tr');
            rows.forEach(row => {
                const cells = row.querySelectorAll('td');
                if (cells.length >= 5) {
                    const material_name = cells[0]?.textContent?.trim();
                    const spec = cells[1]?.textContent?.trim();
                    const material_type = cells[2]?.textContent?.trim();
                    const brand = cells[3]?.textContent?.trim();
                    const price = cells[4]?.textContent?.trim();

                    if (['高线', '螺纹钢', '盘螺', '圆钢'].includes(material_name) &&
                        spec && spec.startsWith('Φ') && price && /^\\d+$/.test(price)) {
                        results.push({
                            material_name,
                            spec,
                            material_type,
                            brand,
                            price: parseFloat(price),
                            region: '山东烟台'
                        });
                    }
                }
            });
        });
        return results;
    }''')

    return list(data), screenshot_b64


async def save_to_excel(prices, period, date_str, screenshot_b64):
    """保存到Excel"""
    try:
        # 尝试加载现有文件
        if EXCEL_FILE.exists():
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE)
            except Exception:
                wb = openpyxl.Workbook()
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']
        else:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        sheet_name = get_sheet_name(date_str, period)
        ws = wb.create_sheet(title=sheet_name)

        # 标题
        period_text = '下午(晚)' if period == 'PM' else '上午'
        ws.merge_cells('A1:K1')
        ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        # 表头
        header_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid') if period == 'PM' else \
            PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 数据
        fetch_time = datetime.now().strftime('%H:%M:%S')
        for i, price in enumerate(prices):
            row = 4 + i
            for col, val in enumerate([date_str, fetch_time, price['material_name'], price['spec'],
                                       price['material_type'], price['brand'], price['price'],
                                       '', '', '', '山东烟台'], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border

        # 截图
        if screenshot_b64:
            screenshot_path = DATA_DIR / f'screenshot_{date_str.replace("-", "")}_{period}.png'
            with open(screenshot_path, 'wb') as f:
                f.write(base64.b64decode(screenshot_b64))

            row = 4 + len(prices) + 2
            ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)
            img = Image(str(screenshot_path))
            img.width = 900
            img.height = 500
            img.anchor = f'A{row + 1}'
            ws.add_image(img)

        wb.save(EXCEL_FILE)
        wb.close()
        return sheet_name
    except Exception as e:
        print(f'保存失败: {e}')
        import traceback
        traceback.print_exc()
        return None


async def main():
    print('=' * 60)
    print('山东烟台钢筋价格 - 一年历史数据抓取')
    print('=' * 60)

    start_date, end_date = get_date_range()
    print(f'日期范围: {start_date.strftime("%Y-%m-%d")} ~ {end_date.strftime("%Y-%m-%d")}')
    print()

    existing_sheets = await get_existing_sheets()
    print(f'已有sheet: {len(existing_sheets)}个\n')

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)  # 使用非无头模式
        context = await browser.new_context(
            viewport={'width': 1920, 'height': 3000},
            locale='zh-CN',
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        )
        page = await context.new_page()

        await login(page)

        # 获取所有市场页面链接 - 限制页数
        links = await fetch_all_market_links(page, max_pages=50)

        # 解析日期
        print('\n解析日期...')
        available_data = await parse_links_to_dates(links, start_date, end_date)
        print(f'找到 {len(available_data)} 个日期的数据\n')

        if not available_data:
            print('没有可抓取的数据')
            await browser.close()
            return

        # 过滤已抓取的
        tasks = []
        for date_str in sorted(available_data.keys(), reverse=True):
            for period in ['AM', 'PM']:
                if period in available_data[date_str]:
                    sheet_name = get_sheet_name(date_str, period)
                    has_existing = any(s.startswith(f'{date_str}_{period}') for s in existing_sheets)
                    if not has_existing:
                        tasks.append((date_str, period, available_data[date_str][period]))

        print(f'需要抓取: {len(tasks)} 条数据\n')

        if not tasks:
            print('所有数据已抓取完毕')
            await browser.close()
            return

        # 开始抓取
        results = {}
        success_count = 0
        fail_count = 0

        for i, (date_str, period, url) in enumerate(tasks):
            print(f'[{i+1}/{len(tasks)}] {date_str} {period}...', end='')

            try:
                prices, screenshot_b64 = await fetch_prices(page, url)
                if prices:
                    sheet_name = await save_to_excel(prices, period, date_str, screenshot_b64)
                    if sheet_name:
                        results[date_str] = results.get(date_str, {})
                        results[date_str][period] = {'sheet': sheet_name, 'count': len(prices)}
                        print(f' {len(prices)}条 ✓')
                        success_count += 1
                    else:
                        print(' 保存失败')
                        fail_count += 1
                else:
                    print(' 无数据')
                    fail_count += 1
            except Exception as e:
                print(f' 失败: {e}')
                fail_count += 1

            if (i + 1) % 10 == 0:
                print(f'  进度: 成功{success_count}, 失败{fail_count}')
                await asyncio.sleep(random_sleep(3, 5))  # 每10个休息一下

        await browser.close()

    print()
    print('=' * 60)
    print('抓取完成')
    print('=' * 60)
    print(f'成功: {success_count} 条')
    print(f'失败: {fail_count} 条')
    print()

    for date_str in sorted(results.keys(), reverse=True):
        am = results[date_str].get('AM', {}).get('count', 0)
        pm = results[date_str].get('PM', {}).get('count', 0)
        if am > 0 or pm > 0:
            print(f'{date_str}: AM={am}, PM={pm}')


if __name__ == '__main__':
    asyncio.run(main())