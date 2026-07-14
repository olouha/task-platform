# -*- coding: utf-8 -*-
"""
Excel解析服务
解析指标库Excel文件，包含汇总sheet和明细sheet
"""

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.cell.cell import TYPE_FORMULA

logger = logging.getLogger(__name__)


class ExcelParserService:
    """Excel解析服务"""

    # 汇总sheet列映射: Excel列名 -> 数据库字段名
    SUMMARY_COLUMN_MAPPING = {
        "序号": "index",
        "项目名称": "name",
        "业态": "category",
        "项目所在地": "location",
        "结构形式": "structure",
        "交付形式": "delivery_form",
        "层数（地上/下）": "floor_info",
        "总面积（m2）": "area_total",
        "檐高（m）": "height",
        "总造价": "total_cost",
    }

    # 明细sheet列映射: Excel列名 -> 数据库字段名
    DETAIL_COLUMN_MAPPING = {
        "序号": "index",
        "项目名称": "name",
        "业态": "category",
        "项目所在地": "location",
        "结构形式": "structure",
        "交付形式": "delivery_form",
        "层数（地上/下）": "floor_info",
        "总面积（m2）": "area_total",
        "檐高（m）": "height",
        "地上建筑面积（m2）": "area_above",
        "地下建筑面积（m2）": "area_below",
        "平米造价（元/m2）": "unit_cost",
        "总造价（元）": "total_cost",
        "地上土建造价": "cost_above_structure",
        "地上安装造价": "cost_above_installation",
        "地下土建造价": "cost_underground_structure",
        "地下安装造价": "cost_underground_installation",
        "措施费（元）": "cost_measures",
        "室外造价（元）": "cost_outdoor",
        "桩基造价（元）": "cost_pile",
        "基坑支护造价（元）": "cost_foundation_support",
        "幕墙造价（元）": "cost_curtain_wall",
        "精装修造价（元）": "cost_decoration",
        "地上砼用量（m3）": "above_concrete",
        "地上砼平米含量": "above_concrete_unit",
        "地上钢筋用量（t）": "above_rebar",
        "地上钢筋平米含量": "above_rebar_unit",
        "地下砼用量（m3）": "underground_concrete",
        "地下砼平米含量": "underground_concrete_unit",
        "地下钢筋用量（t）": "underground_rebar",
        "地下钢筋平米含量": "underground_rebar_unit",
        "开工时间": "start_date",
        "竣工时间": "end_date",
        "备注": "remarks",
    }

    # 新模板列映射（简化版单Sheet）
    SIMPLE_COLUMN_MAPPING = {
        "项目名称": "name",
        "业态": "category",
        "项目所在地": "location",
        "结构形式": "structure",
        "交付形式": "delivery_form",
        "层数（地上/下）": "floor_info",
        "总面积（m2）": "area_total",
        "檐高（m）": "height",
        "地上建筑面积（m2）": "area_above",
        "地下建筑面积（m2）": "area_below",
        "平米造价（元/m2）": "unit_cost",
        "总造价（元）": "total_cost",
        "地上土建造价": "cost_above_structure",
        "地上安装造价": "cost_above_installation",
        "地下土建造价": "cost_underground_structure",
        "地下安装造价": "cost_underground_installation",
        "措施费（元）": "cost_measures",
        "室外造价（元）": "cost_outdoor",
        "桩基造价（元）": "cost_pile",
        "基坑支护造价（元）": "cost_foundation_support",
        "幕墙造价（元）": "cost_curtain_wall",
        "精装修造价（元）": "cost_decoration",
        "地上砼用量（m3）": "above_concrete",
        "地上砼平米含量": "above_concrete_unit",
        "地上钢筋用量（t）": "above_rebar",
        "地上钢筋平米含量": "above_rebar_unit",
        "地下砼用量（m3）": "underground_concrete",
        "地下砼平米含量": "underground_concrete_unit",
        "地下钢筋用量（t）": "underground_rebar",
        "地下钢筋平米含量": "underground_rebar_unit",
        "开工时间": "start_date",
        "竣工时间": "end_date",
        "备注": "remarks",
    }

    def __init__(self, file_path: str):
        """
        初始化Excel解析服务

        Args:
            file_path: Excel文件路径
        """
        self.file_path = Path(file_path)
        logger.info(f"[ExcelParserService] 初始化解析器 | file_path={self.file_path}")

        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel文件不存在: {self.file_path}")

        self.workbook: Optional[openpyxl.Workbook] = None
        self.summary_headers: List[str] = []
        self.detail_headers: List[str] = []
        self.merged_headers: Dict[int, str] = {}

    def parse(self) -> Dict[str, Any]:
        """
        解析Excel文件

        支持两种格式：
        1. 新格式：单Sheet "指标库数据"
        2. 旧格式：双Sheet "汇总" + "明细"

        Returns:
            包含projects列表和metadata的字典
        """
        logger.info(f"[ExcelParserService] 开始解析 | file={self.file_path.name}")

        try:
            # 加载工作簿
            self.workbook = openpyxl.load_workbook(
                self.file_path,
                data_only=True,  # 只读取值，不读取公式
                read_only=True
            )
            logger.info(f"[ExcelParserService] 工作簿加载成功 | sheets={self.workbook.sheetnames}")

            # 检测并解析文件格式
            if "明细" in self.workbook.sheetnames and "汇总" in self.workbook.sheetnames:
                # 新格式：明细表（员工填写）+ 汇总表（自动提取）
                # 优先从明细表读取数据
                logger.info("[ExcelParserService] 使用明细表解析")
                projects = self._parse_detail_sheet()
            elif "指标库数据" in self.workbook.sheetnames:
                # 单Sheet格式（兼容旧模板）
                logger.info("[ExcelParserService] 使用单Sheet解析")
                projects = self._parse_simple_sheet()
            else:
                raise ValueError("Excel文件格式不正确，需要包含'明细'+'汇总'Sheet或'指标库数据'Sheet")

            logger.info(f"[ExcelParserService] 解析完成 | rows={len(projects)}")

            # 构建返回结果
            result = {
                "success": True,
                "projects": projects,
                "metadata": {
                    "source_file": self.file_path.name,
                    "entry_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "total_projects": len(projects),
                }
            }

            logger.info(f"[ExcelParserService] 解析完成 | total_projects={len(projects)}")
            return result

        except Exception as e:
            logger.error(f"[ExcelParserService] 解析失败 | error={e}", exc_info=True)
            return {
                "success": False,
                "error": str(e),
                "projects": [],
                "metadata": {
                    "source_file": self.file_path.name,
                    "entry_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "error": str(e),
                }
            }
        finally:
            if self.workbook:
                self.workbook.close()

    def _parse_simple_sheet(self) -> List[Dict[str, Any]]:
        """
        解析新格式单Sheet

        Returns:
            项目数据列表
        """
        logger.info("[ExcelParserService] 解析单Sheet模板")

        ws = self.workbook["指标库数据"]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            logger.warning("[ExcelParserService] 数据行数不足")
            return []

        # 获取表头（第一行）
        header_row = rows[0]
        headers = [str(cell) if cell else "" for cell in header_row]
        logger.debug(f"[ExcelParserService] 表头 | headers={headers}")

        # 解析数据行
        projects = []
        for row_idx, row in enumerate(rows[1:], start=2):
            if self._is_empty_row(row):
                continue

            row_data = self._parse_simple_row(row, headers)
            if row_data:
                row_data["_row_index"] = row_idx
                projects.append(row_data)

        logger.info(f"[ExcelParserService] 单Sheet解析完成 | count={len(projects)}")
        return projects

    def _parse_simple_row(self, row: Tuple, headers: List[str]) -> Optional[Dict[str, Any]]:
        """
        解析单Sheet的数据行

        Args:
            row: 行数据元组
            headers: 表头列表

        Returns:
            解析后的行数据字典
        """
        row_data = {}
        has_valid_data = False

        for col_idx, cell_value in enumerate(row):
            if col_idx >= len(headers):
                break

            header = headers[col_idx]
            if header not in self.SIMPLE_COLUMN_MAPPING:
                continue

            field_name = self.SIMPLE_COLUMN_MAPPING[header]
            value = self._clean_cell_value(cell_value)
            row_data[field_name] = value

            # 检查是否有有效数据
            if field_name in ("name", "category", "location") and value is not None:
                has_valid_data = True

        # 如果没有有效数据，返回None
        if not has_valid_data:
            return None

        # 解析楼层信息
        floor_info = row_data.get("floor_info")
        if floor_info:
            floor_above, floor_below = self._parse_floor_info(str(floor_info))
            row_data["floor_above"] = floor_above
            row_data["floor_below"] = floor_below
            row_data.pop("floor_info", None)

        return row_data

    def _parse_summary_sheet(self) -> List[Dict[str, Any]]:
        """
        解析汇总sheet

        Returns:
            汇总数据列表
        """
        logger.info("[ExcelParserService] 解析汇总sheet")

        ws = self.workbook["汇总"]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            logger.warning("[ExcelParserService] 汇总sheet数据行数不足")
            return []

        # 获取表头
        header_row = rows[0]
        self.summary_headers = [str(cell) if cell else "" for cell in header_row]
        logger.debug(f"[ExcelParserService] 汇总表头 | headers={self.summary_headers}")

        # 解析数据行
        data_rows = []
        for row_idx, row in enumerate(rows[1:], start=2):
            if self._is_empty_row(row):
                continue

            row_data = self._parse_summary_row(row)
            if row_data:
                row_data["_row_index"] = row_idx
                data_rows.append(row_data)

        return data_rows

    def _parse_summary_row(self, row: Tuple) -> Optional[Dict[str, Any]]:
        """
        解析汇总sheet的单行数据

        Args:
            row: 行数据元组

        Returns:
            解析后的行数据字典，如果无有效数据则返回None
        """
        row_data = {}
        has_valid_data = False

        for col_idx, cell_value in enumerate(row):
            if col_idx >= len(self.summary_headers):
                break

            header = self.summary_headers[col_idx]
            if header not in self.SUMMARY_COLUMN_MAPPING:
                continue

            field_name = self.SUMMARY_COLUMN_MAPPING[header]
            value = self._clean_cell_value(cell_value)
            row_data[field_name] = value

            # 检查是否有有效数据（name, category, location 至少有一个有值）
            if field_name in ("name", "category", "location") and value is not None:
                has_valid_data = True

        # 如果没有有效数据，返回None
        if not has_valid_data:
            return None

        return row_data

    def _parse_detail_sheet(self) -> List[Dict[str, Any]]:
        """
        解析明细sheet

        Returns:
            明细数据列表
        """
        logger.info("[ExcelParserService] 解析明细sheet")

        ws = self.workbook["明细"]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            logger.warning("[ExcelParserService] 明细sheet数据行数不足")
            return []

        # 检查表头行数
        header_row = rows[0]
        self.detail_headers = [str(cell) if cell else "" for cell in header_row]
        logger.debug(f"[ExcelParserService] 明细表头 | headers={self.detail_headers}")

        # 确定数据起始行
        # data_start_row 表示表头所在的行索引
        data_start_row = 0  # 默认第1行是表头

        # 检查是否是旧格式多级表头（检查第2行是否为空）
        if len(rows) > 4 and not any(rows[1]):
            # 第2行为空，说明可能是多级表头格式
            if not any(rows[2]):
                data_start_row = 3  # 第4行是最终表头
                self.detail_headers = [str(cell) if cell else "" for cell in rows[3]]
                logger.debug(f"[ExcelParserService] 检测到多级表头，使用第4行 | headers={self.detail_headers}")

        # 解析数据行（跳过表头行，从下一行开始）
        data_rows = []
        for row_idx, row in enumerate(rows[data_start_row + 1:], start=data_start_row + 2):
            if self._is_empty_row(row):
                continue

            row_data = self._parse_detail_row(row)
            if row_data:
                row_data["_row_index"] = row_idx
                data_rows.append(row_data)

        logger.info(f"[ExcelParserService] 明细解析完成 | count={len(data_rows)}")
        return data_rows

    def _build_merged_headers(self, header_rows: List[Tuple]) -> Dict[int, str]:
        """
        构建合并后的多级表头

        Args:
            header_rows: 前4行表头数据

        Returns:
            列索引到合并后表头名称的映射
        """
        merged = {}

        if len(header_rows) < 4:
            # 如果没有足够的多级表头，直接使用第一行
            for col_idx, cell in enumerate(header_rows[0]):
                if cell:
                    merged[col_idx] = str(cell)
            return merged

        # 第1行: 大类分组（如：基本信息、造价指标、分部工程等）
        # 第2-4行: 具体列名
        group_row = header_rows[0]
        sub_rows = header_rows[1:4]

        for col_idx in range(max(len(row) for row in header_rows)):
            # 获取子类表头（从第2-4行）
            sub_header = ""
            for row in sub_rows:
                if col_idx < len(row) and row[col_idx]:
                    sub_val = str(row[col_idx]).strip()
                    if sub_val:
                        sub_header = sub_val
                        break

            # 如果没有子类表头，尝试使用第一行
            if not sub_header and col_idx < len(group_row):
                group_val = str(group_row[col_idx]).strip() if group_row[col_idx] else ""
                sub_header = group_val

            if sub_header:
                merged[col_idx] = sub_header

        return merged

    def _parse_detail_row(self, row: Tuple) -> Dict[str, Any]:
        """
        解析明细sheet的单行数据

        Args:
            row: 行数据元组

        Returns:
            解析后的行数据字典
        """
        row_data = {}

        for col_idx, cell_value in enumerate(row):
            if col_idx >= len(self.detail_headers):
                break

            header = self.detail_headers[col_idx]
            if header not in self.DETAIL_COLUMN_MAPPING:
                continue

            field_name = self.DETAIL_COLUMN_MAPPING[header]
            value = self._clean_cell_value(cell_value)
            row_data[field_name] = value

        # 检查是否有有效数据（通过序号判断）
        if not row_data.get("index"):
            return None

        # 解析楼层信息
        floor_info = row_data.get("floor_info")
        if floor_info:
            floor_above, floor_below = self._parse_floor_info(str(floor_info))
            row_data["floor_above"] = floor_above
            row_data["floor_below"] = floor_below
            row_data.pop("floor_info", None)

        return row_data

    def _merge_data(
        self,
        summary_data: List[Dict[str, Any]],
        detail_data: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """
        合并汇总和明细数据

        Args:
            summary_data: 汇总数据
            detail_data: 明细数据

        Returns:
            合并后的数据列表
        """
        logger.info(f"[ExcelParserService] 合并数据 | summary={len(summary_data)}, detail={len(detail_data)}")

        merged = []

        # 使用行索引进行匹配
        summary_by_index = {row.get("_row_index", 0): row for row in summary_data}
        detail_by_index = {row.get("_row_index", 0): row for row in detail_data}

        # 获取所有唯一的行索引
        all_indices = set(summary_by_index.keys()) | set(detail_by_index.keys())

        for row_idx in sorted(all_indices):
            summary_row = summary_by_index.get(row_idx, {})
            detail_row = detail_by_index.get(row_idx, {})

            # 合并数据，detail优先（包含更详细的信息）
            merged_row = {**summary_row, **detail_row}

            # 移除内部使用的_row_index
            merged_row.pop("_row_index", None)

            # 解析楼层信息
            floor_info = merged_row.get("floor_info")
            if floor_info:
                floor_above, floor_below = self._parse_floor_info(str(floor_info))
                merged_row["floor_above"] = floor_above
                merged_row["floor_below"] = floor_below
                merged_row.pop("floor_info", None)

            merged.append(merged_row)

        logger.info(f"[ExcelParserService] 数据合并完成 | merged={len(merged)}")
        return merged

    def _parse_floor_info(self, floor_info: str) -> Tuple[Optional[int], Optional[int]]:
        """
        解析楼层信息字符串

        Args:
            floor_info: 楼层信息字符串，格式如 "地上30/地下3" 或 "18/2"

        Returns:
            (地上楼层数, 地下楼层数) 元组
        """
        floor_above = None
        floor_below = None

        try:
            # 先尝试 "地上X/地下Y" 格式
            above_match = re.search(r"地上\s*(\d+)", floor_info)
            if above_match:
                floor_above = int(above_match.group(1))

            below_match = re.search(r"地下\s*(\d+)", floor_info)
            if below_match:
                floor_below = int(below_match.group(1))

            # 如果上面没匹配到，尝试 "X/Y" 格式（斜杠分隔）
            if floor_above is None and floor_below is None:
                parts = floor_info.split('/')
                if len(parts) == 2:
                    try:
                        floor_above = int(parts[0].strip())
                    except ValueError:
                        pass
                    try:
                        floor_below = int(parts[1].strip())
                    except ValueError:
                        pass
            if below_match:
                floor_below = int(below_match.group(1))
        except Exception as e:
            logger.warning(f"[ExcelParserService] 解析楼层信息失败 | floor_info={floor_info}, error={e}")

        return floor_above, floor_below

    def _clean_cell_value(self, value: Any) -> Any:
        """
        清理单元格值

        Args:
            value: 原始单元格值

        Returns:
            清理后的值
        """
        if value is None:
            return None

        # 如果是字符串，处理常见问题
        if isinstance(value, str):
            # 去除首尾空白
            value = value.strip()

            # 处理公式引用错误
            if "#REF!" in value or "#REF" in value:
                logger.debug(f"[ExcelParserService] 跳过公式错误值 | value={value}")
                return None

            # 处理其他Excel错误
            if value.startswith("#"):
                return None

            # 处理空字符串
            if not value:
                return None

            return value

        return value

    def _is_empty_row(self, row: Tuple) -> bool:
        """
        检查行是否为空或仅为说明行

        Args:
            row: 行数据元组

        Returns:
            是否为空行或说明行
        """
        # 检查是否全是空值
        if all(cell is None or str(cell).strip() == "" for cell in row):
            return True

        # 检查第一个单元格是否为说明性文字
        first_cell = row[0] if row else None
        if first_cell is not None:
            first_str = str(first_cell).strip()
            # 跳过表头行（序号列包含"序号"）
            if first_str == "序号":
                return True
            # 跳过填写指导行（以【开头）
            if first_str.startswith("【") and first_str.endswith("】"):
                return True
            # 跳过说明行（包含特定关键词的行）
            skip_keywords = ["填写说明", "说明：", "备注：", "注：", "注意：", "样例："]
            if any(kw in first_str for kw in skip_keywords):
                return True
            # 跳过中文章节序号行（一、二、三... 或 1.、2. 等）
            import re
            if re.match(r'^[一二三四五六七八九十]+[、.：]', first_str):
                return True
            if re.match(r'^[0-9]+[.、:：]', first_str):
                return True
            # 跳过以数字开头+点的序号格式（说明行）
            if re.match(r'^\d+[.、:：]', first_str):
                return True

            # 如果第一列是文字且包含常见说明特征，跳过
            # 指导行特征：第二列开始有长文字
            if first_str and not first_str.isdigit():
                # 检查后续列是否有长文字（超过20字符）
                has_long_text = any(
                    cell and isinstance(cell, str) and len(cell.strip()) > 20
                    for cell in row[1:5]
                )
                if has_long_text:
                    return True

                # 检查第一列是否是字段说明（如"项目名称"、"业态"等）
                guide_field_indicators = ["必须填写", "填写项目", "从下拉列表", "填写地区", "只填数字", "格式："]
                for indicator in guide_field_indicators:
                    if indicator in first_str:
                        return True

        return False

    @staticmethod
    def get_column_letter(col_idx: int) -> str:
        """
        获取列字母标识

        Args:
            col_idx: 列索引（从0开始）

        Returns:
            列字母（如：A, B, C, ...）
        """
        result = ""
        col_idx += 1  # 转换为从1开始
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    @staticmethod
    def validate_excel_structure(file_path: str) -> Dict[str, Any]:
        """
        验证Excel文件结构

        Args:
            file_path: Excel文件路径

        Returns:
            验证结果
        """
        logger.info(f"[ExcelParserService] 验证Excel结构 | file={file_path}")

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames

            result = {
                "valid": True,
                "has_summary": "汇总" in sheets,
                "has_detail": "明细" in sheets,
                "sheets": sheets,
                "errors": []
            }

            if "汇总" not in sheets:
                result["valid"] = False
                result["errors"].append("缺少'汇总'sheet")

            if "明细" not in sheets:
                result["valid"] = False
                result["errors"].append("缺少'明细'sheet")

            wb.close()
            logger.info(f"[ExcelParserService] 验证完成 | valid={result['valid']}")
            return result

        except Exception as e:
            logger.error(f"[ExcelParserService] 验证失败 | error={e}", exc_info=True)
            return {
                "valid": False,
                "error": str(e),
                "errors": [str(e)]
            }