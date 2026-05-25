"""
补抓缺失的历史数据 - 通过市场列表页面
"""
import asyncio
import sys
import json
import base64
import random
from pathlib import Path
from datetime import datetime, timedelta

sys.path.insert(0, '.')

from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

DATA_DIR = Path('services/data')
DATA_DIR.mkdir(exist_ok=True)

COOKIE_FILE = DATA_DIR / 'mysteel_cookies.json'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
MARKET_URL = 'https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html'

# 缺失的时间范围
MISSING_RANGES = [
    (datetime(2024, 1, 1), datetime(2024, 6, 30)),  # 2024年1-6月
    (datetime(2025, 2, 20), datetime(2025, 2, 28)),  # 2025年2月后半
    (datetime(2025, 3, 1), datetime(2025, 5, 31)),   # 2025年3-5月
]

USERNAME = 'M6616592358'
PASSWORD = 'mysteel573005'


def random_sleep(min_sec=14, max_sec=16):
    return random.uniform(min_sec, max_sec)


async def human_like_mouse(page):
    try:
        viewport = page.viewport_size
        if viewport:
            for _ in range(random.randint(2, 4)):
                x = random.randint(100, viewport['width'] - 100)
                y = random.randint(100, viewport['height'] - 100)
                await page.mouse.move(x, y)
                await asyncio.sleep(random.uniform(0.2, 0.5))
    except:
        pass


async def get_existing_sheets():
    existing_sheets = set()
    if EXCEL_FILE.exists():
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            existing_sheets = set(wb.sheetnames)
            wb.close()
        except Exception:
            pass
    return existing_sheets


async def login(page):
    print('登录中...')

    if COOKIE_FILE.exists():
        try:
            cookies = json.loads(COOKIE_FILE.read_text(encoding='utf-8'))
            if cookies:
                await page.context.add_cookies(cookies)
                print('已加载Cookie')
        except:
            pass

    await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
    await asyncio.sleep(random_sleep(2, 4))

    await human_like_mouse(page)

    try:
        account_tab = await page.query_selector('.form-tab-account, a[data-tab="account"]')
        if account_tab:
            await account_tab.click()
            await asyncio.sleep(random_sleep(1, 2))
    except:
        pass

    await human_like_mouse(page)

    await page.evaluate(f'''() => {{
        const inputs = document.querySelectorAll('input');
        for (const inp of inputs) {{
            const ph = inp.placeholder || '';
            if (ph.includes('用户名')) inp.value = '{USERNAME}';
            if (ph.includes('密码') && inp.type === 'password') inp.value = '{PASSWORD}';
        }}
    }}''')

    await asyncio.sleep(random_sleep(1, 2))
    await human_like_mouse(page)

    try:
        login_btn = await page.query_selector('.form-button-login, button:has-text("登录")')
        if login_btn:
            await login_btn.click()
    except:
        pass

    await asyncio.sleep(random_sleep(10, 15))

    cookies = await page.context.cookies()
    with open(COOKIE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cookies, f, ensure_ascii=False)
    print('登录完成\n')


def extract_date_from_url(url):
    import re
    match = re.search(r'/m/(\d{8})/', url)
    if match:
        full = match.group(1)
        try:
            y = 2000 + int(full[0:2])
            m = int(full[2:4])
            d = int(full[4:6])
            h = int(full[6:8])
            date_obj = datetime(y, m, d)
            period = 'AM' if h < 12 else 'PM'
            return date_obj, period
        except:
            return None, None
    return None, None


async def get_page_dates(page):
    links = await page.evaluate('''() => {
        const links = [];
        document.querySelectorAll('a[href]').forEach(a => {
            const href = a.href;
            const text = a.textContent.trim();
            if (href.match(/\/m\/\d{8}\//)) {
                links.push({href, text});
            }
        });
        return links;
    }''')
    return links


async def fetch_prices(page, date_str, period):
    screenshot = await page.screenshot(full_page=True)
    screenshot_b64 = base64.b64encode(screenshot).decode('utf-8')

    data = await page.evaluate('''() => {
        const tables = document.querySelectorAll('table');
        const results = [];
        tables.forEach(table => {
            const rows = table.querySelectorAll('tr');
            rows.forEach(row => {
                const cells = row.querySelectorAll('td');
                if (cells.length >= 5) {
                    const material_name = cells[0]?.textContent?.trim();
                    const spec = cells[1]?.textContent?.trim();
                    const material_type = cells[2]?.textContent?.trim();
                    const brand = cells[3]?.textContent?.trim();
                    const price = cells[4]?.textContent?.trim();

                    if (['高线', '螺纹钢', '盘螺', '圆钢'].includes(material_name) &&
                        spec && spec.startsWith('Φ') && price && /^\\d+$/.test(price)) {
                        results.push({
                            material_name,
                            spec,
                            material_type,
                            brand,
                            price: parseFloat(price),
                            region: '山东烟台'
                        });
                    }
                }
            });
        });
        return results;
    }''')

    return list(data), screenshot_b64


async def save_to_excel(prices, period, date_str, screenshot_b64):
    try:
        if EXCEL_FILE.exists():
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE)
            except Exception:
                wb = openpyxl.Workbook()
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']
        else:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        if period == 'AM':
            sheet_name = date_str
        else:
            sheet_name = f'{date_str}_PM'

        if sheet_name in wb.sheetnames:
            return sheet_name

        ws = wb.create_sheet(title=sheet_name)

        period_text = '下午(晚)' if period == 'PM' else '上午'
        ws.merge_cells('A1:K1')
        ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        header_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid') if period == 'PM' else \
            PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        fetch_time = datetime.now().strftime('%H:%M:%S')
        for i, price in enumerate(prices):
            row = 4 + i
            for col, val in enumerate([date_str, fetch_time, price['material_name'], price['spec'],
                                       price['material_type'], price['brand'], price['price'],
                                       '', '', '', '山东烟台'], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border

        if screenshot_b64:
            screenshot_path = DATA_DIR / f'screenshot_{date_str.replace("-", "")}_{period}.png'
            with open(screenshot_path, 'wb') as f:
                f.write(base64.b64decode(screenshot_b64))

            row = 4 + len(prices) + 2
            ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)
            img = Image(str(screenshot_path))
            img.width = 900
            img.height = 500
            img.anchor = f'A{row + 1}'
            ws.add_image(img)

        wb.save(EXCEL_FILE)
        wb.close()
        return sheet_name
    except Exception as e:
        print(f'保存失败: {e}')
        return None


def is_in_range(date_obj, ranges):
    for start, end in ranges:
        if start <= date_obj <= end:
            return True
    return False


async def main():
    print('=' * 60)
    print('山东烟台钢筋价格 - 补抓缺失数据')
    print('=' * 60)
    print()

    existing_sheets = await get_existing_sheets()
    print(f'已有sheet: {len(existing_sheets)}个')
    print()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            viewport={'width': 1920, 'height': 3000},
            locale='zh-CN',
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        )
        page = await context.new_page()

        await login(page)

        # 打开市场总页面
        print('打开市场页面...')
        await page.goto(MARKET_URL, wait_until='domcontentloaded', timeout=60000)
        await asyncio.sleep(random_sleep(3, 5))

        success_count = 0
        fail_count = 0

        # 收集所有日期链接
        print('\n收集日期链接...')
        all_links = []
        page_num = 1

        while page_num <= 200:  # 增加到200页
            await human_like_mouse(page)
            await asyncio.sleep(random_sleep(1, 2))

            links = await get_page_dates(page)
            print(f'  第{page_num}页: 找到 {len(links)} 条')

            if len(links) == 0:
                break

            for link in links:
                date_obj, period = extract_date_from_url(link['href'])
                if date_obj and is_in_range(date_obj, MISSING_RANGES):
                    all_links.append({
                        'url': link['href'],
                        'date_obj': date_obj,
                        'date_str': date_obj.strftime('%Y-%m-%d'),
                        'period': period,
                        'text': link['text']
                    })

            # 检查是否有下一页
            try:
                next_btn = await page.query_selector('a:has-text("下一页")')
                if next_btn and await next_btn.is_visible():
                    await next_btn.click()
                    await asyncio.sleep(random_sleep(2, 4))
                    page_num += 1
                else:
                    break
            except:
                break

        print(f'在目标范围内找到 {len(all_links)} 条日期链接')

        # 去重
        seen = {}
        unique_links = []
        for link in all_links:
            key = (link['date_str'], link['period'])
            if key not in seen:
                seen[key] = True
                unique_links.append(link)

        unique_links.sort(key=lambda x: x['date_obj'])

        # 过滤已存在的
        tasks = []
        for link in unique_links:
            if link['period'] == 'AM':
                sheet_name = link['date_str']
            else:
                sheet_name = f"{link['date_str']}_PM"

            if sheet_name not in existing_sheets:
                tasks.append(link)
            else:
                print(f'  {link["date_str"]} {link["period"]} 已存在，跳过')

        print(f'需要抓取: {len(tasks)} 条')
        print()

        # 按月份分组抓取
        tasks_by_month = {}
        for task in tasks:
            month_key = task['date_obj'].strftime('%Y-%m')
            if month_key not in tasks_by_month:
                tasks_by_month[month_key] = []
            tasks_by_month[month_key].append(task)

        for month_key in sorted(tasks_by_month.keys()):
            month_tasks = tasks_by_month[month_key]
            print(f'\n{"="*40}')
            print(f'处理月份: {month_key} ({len(month_tasks)}条)')
            print(f'{"="*40}')

            # 每3个月重新登录
            if int(month_key.split('-')[1]) % 3 == 1:
                print('重新登录...')
                await login(page)

            # 重新打开市场页面
            await page.goto(MARKET_URL, wait_until='domcontentloaded', timeout=60000)
            await asyncio.sleep(random_sleep(3, 5))

            for task in month_tasks:
                date_str = task['date_str']
                period = task['period']
                url = task['url']

                print(f'  抓取 {date_str} {period}...', end='', flush=True)

                try:
                    await page.goto(url, wait_until='domcontentloaded', timeout=60000)
                    await asyncio.sleep(random_sleep(4, 6))

                    await human_like_mouse(page)

                    prices, screenshot_b64 = await fetch_prices(page, date_str, period)

                    if prices:
                        sheet_name = await save_to_excel(prices, period, date_str, screenshot_b64)
                        if sheet_name:
                            existing_sheets.add(sheet_name)
                            print(f' {len(prices)}条 [OK]')
                            success_count += 1
                        else:
                            print(' 保存失败')
                            fail_count += 1
                    else:
                        print(' 无数据')
                        fail_count += 1

                    print(f'  停顿15秒...')
                    await asyncio.sleep(random_sleep(14, 16))

                except Exception as e:
                    print(f'失败: {e}')
                    fail_count += 1
                    print(f'  停顿15秒...')
                    await asyncio.sleep(random_sleep(14, 16))

            print(f'  月份完成，停顿30秒...')
            await asyncio.sleep(30)

    await browser.close()

    print()
    print('=' * 60)
    print('补抓完成')
    print('=' * 60)
    print(f'成功: {success_count} 条')
    print(f'失败: {fail_count} 条')


if __name__ == '__main__':
    asyncio.run(main())