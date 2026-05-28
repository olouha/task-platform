# -*- coding: utf-8 -*-
"""
导入OCR识别的造价数据到SQLite数据库
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import sqlite3
import openpyxl
from datetime import datetime
from pathlib import Path
import os

DATA_DIR = Path('services/data')
DB_FILE = DATA_DIR / 'cost_reference.db'

def init_database():
    """初始化数据库"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    # 混凝土价格表
    c.execute('''
        CREATE TABLE IF NOT EXISTS concrete_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year TEXT NOT NULL,
            quarter TEXT NOT NULL,
            grade TEXT NOT NULL,
            yantai_price INTEGER,
            rushan_price INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(year, quarter, grade)
        )
    ''')

    # 钢筋价格表
    c.execute('''
        CREATE TABLE IF NOT EXISTS rebar_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year TEXT NOT NULL,
            quarter TEXT NOT NULL,
            grade TEXT NOT NULL,
            spec TEXT NOT NULL,
            price INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(year, quarter, grade, spec)
        )
    ''')

    # 创建索引
    c.execute('CREATE INDEX IF NOT EXISTS idx_concrete_year ON concrete_prices(year)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_concrete_grade ON concrete_prices(grade)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_rebar_year ON rebar_prices(year)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_rebar_grade ON rebar_prices(grade)')

    conn.commit()
    return conn

def import_excel(conn, excel_path):
    """从Excel导入数据"""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    c = conn.cursor()
    concrete_count = 0
    rebar_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        # 解析标题行
        headers = [str(h).strip() if h else '' for h in rows[0]]
        print(f'\n处理工作表: {sheet_name}')
        print(f'  标题: {headers}')

        # 处理数据行
        for row in rows[1:]:
            try:
                if sheet_name == '混凝土信息价':
                    year = str(row[0]) if row[0] else None
                    quarter = str(row[1]) if row[1] else None
                    grade = str(row[2]) if row[2] else None
                    yantai = int(float(row[3])) if row[3] and str(row[3]).isdigit() else None
                    rushan = int(float(row[4])) if row[4] and str(row[4]).isdigit() else None

                    if year and quarter and grade:
                        c.execute('''
                            INSERT OR REPLACE INTO concrete_prices
                            (year, quarter, grade, yantai_price, rushan_price)
                            VALUES (?, ?, ?, ?, ?)
                        ''', (year, quarter, grade, yantai, rushan))
                        concrete_count += 1

                elif sheet_name == '钢筋信息价':
                    year = str(row[0]) if row[0] else None
                    quarter = str(row[1]) if row[1] else None
                    grade = str(row[2]) if row[2] else None
                    spec = str(row[3]) if row[3] else None
                    price = int(float(row[4])) if row[4] and str(row[4]).replace('.','').replace('-','').isdigit() else None

                    if year and quarter and grade and spec and price:
                        c.execute('''
                            INSERT OR REPLACE INTO rebar_prices
                            (year, quarter, grade, spec, price)
                            VALUES (?, ?, ?, ?, ?)
                        ''', (year, quarter, grade, spec, price))
                        rebar_count += 1

            except Exception as e:
                print(f'  导入错误: {e}, 行: {row}')
                continue

    wb.close()
    return concrete_count, rebar_count

def verify_data(conn):
    """验证数据"""
    c = conn.cursor()

    print('\n' + '=' * 60)
    print('数据库统计')
    print('=' * 60)

    # 混凝土统计
    c.execute('SELECT COUNT(*) FROM concrete_prices')
    concrete_total = c.fetchone()[0]
    c.execute('SELECT COUNT(DISTINCT year) FROM concrete_prices')
    concrete_years = c.fetchone()[0]
    print(f'\n混凝土价格: {concrete_total} 条')
    print(f'  覆盖年份: {concrete_years} 年')

    # 钢筋统计
    c.execute('SELECT COUNT(*) FROM rebar_prices')
    rebar_total = c.fetchone()[0]
    c.execute('SELECT COUNT(DISTINCT year) FROM rebar_prices')
    rebar_years = c.fetchone()[0]
    c.execute('SELECT DISTINCT grade FROM rebar_prices ORDER BY grade')
    grades = [r[0] for r in c.fetchall()]
    print(f'\n钢筋价格: {rebar_total} 条')
    print(f'  覆盖年份: {rebar_years} 年')
    print(f'  等级: {grades}')

    # 按年份统计
    print('\n按年份统计:')
    c.execute('''
        SELECT year, COUNT(*) as cnt
        FROM concrete_prices
        GROUP BY year
        ORDER BY year
    ''')
    print('  混凝土:')
    for row in c.fetchall():
        print(f'    {row[0]}: {row[1]} 条')

    c.execute('''
        SELECT year, COUNT(*) as cnt
        FROM rebar_prices
        GROUP BY year
        ORDER BY year
    ''')
    print('  钢筋:')
    for row in c.fetchall():
        print(f'    {row[0]}: {row[1]} 条')

def main():
    print('=' * 60)
    print('导入造价参考价数据到SQLite数据库')
    print('=' * 60)

    # 初始化数据库
    os.makedirs(DATA_DIR, exist_ok=True)
    conn = init_database()
    print(f'\n数据库: {DB_FILE}')

    # Excel文件
    excel_files = [
        r'C:\Users\admin\Desktop\近五年钢筋混凝土造价管理截图(1)\近五年钢筋混凝土造价管理截图\造价参考价数据_20260528_130227.xlsx',
    ]

    total_concrete = 0
    total_rebar = 0

    for excel_file in excel_files:
        if os.path.exists(excel_file):
            print(f'\n导入: {os.path.basename(excel_file)}')
            concrete, rebar = import_excel(conn, excel_file)
            total_concrete += concrete
            total_rebar += rebar
        else:
            print(f'\n文件不存在: {excel_file}')

    conn.commit()

    print(f'\n总导入: 混凝土 {total_concrete} 条, 钢筋 {total_rebar} 条')

    # 验证
    verify_data(conn)

    conn.close()
    print(f'\n完成! 数据库: {DB_FILE}')

if __name__ == '__main__':
    main()
