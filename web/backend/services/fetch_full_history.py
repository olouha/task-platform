"""
全量历史数据抓取 - 查漏补缺
从2024年1月1日到现在的所有数据
安全模式：只追加不覆盖，已有数据跳过
"""
import asyncio
import sys
import json
import base64
import random
from pathlib import Path
from datetime import datetime, timedelta
from collections import defaultdict

sys.path.insert(0, '.')

from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

DATA_DIR = Path('services/data')
DATA_DIR.mkdir(exist_ok=True)

COOKIE_FILE = DATA_DIR / 'mysteel_cookies.json'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
MARKET_URL = 'https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html'

# 目标时间范围：2024年1月1日 到 今天
START_DATE = datetime(2024, 1, 1)
END_DATE = datetime.now()

USERNAME = 'M6616592358'
PASSWORD = 'mysteel573005'

LOG_FILE = DATA_DIR / 'logs' / f'full_history_fetch_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'


def log(msg):
    """记录日志到文件和控制台"""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f'[{timestamp}] {msg}'
    print(line)
    LOG_FILE.parent.mkdir(exist_ok=True)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(line + '\n')


def random_sleep(min_sec=14, max_sec=16):
    """随机睡眠14-16秒（反爬措施）"""
    return random.uniform(min_sec, max_sec)


async def human_like_mouse(page):
    """模拟人类鼠标操作"""
    try:
        viewport = page.viewport_size
        if viewport:
            for _ in range(random.randint(2, 4)):
                x = random.randint(100, viewport['width'] - 100)
                y = random.randint(100, viewport['height'] - 100)
                await page.mouse.move(x, y)
                await asyncio.sleep(random.uniform(0.2, 0.5))
    except:
        pass


def get_existing_dates():
    """获取Excel中已存在的所有日期"""
    existing = set()
    if EXCEL_FILE.exists():
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
            for sheet_name in wb.sheetnames:
                # 从sheet名提取日期
                date_str = sheet_name[:10]
                try:
                    datetime.strptime(date_str, '%Y-%m-%d')
                    existing.add(date_str)
                except:
                    pass
            wb.close()
        except Exception as e:
            log(f'读取已有数据失败: {e}')
    return existing


def get_target_months():
    """生成目标月份列表"""
    months = []
    current = START_DATE
    while current <= END_DATE:
        months.append((current.year, current.month))
        # 移到下个月
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)
    return months


async def login(page):
    """登录"""
    log('登录中...')

    if COOKIE_FILE.exists():
        try:
            cookies = json.loads(COOKIE_FILE.read_text(encoding='utf-8'))
            if cookies:
                await page.context.add_cookies(cookies)
                log('已加载Cookie')
        except:
            pass

    await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
    await asyncio.sleep(random_sleep(2, 4))

    await human_like_mouse(page)

    try:
        account_tab = await page.query_selector('.form-tab-account, a[data-tab="account"]')
        if account_tab:
            await account_tab.click()
            await asyncio.sleep(random_sleep(1, 2))
    except:
        pass

    await human_like_mouse(page)

    await page.evaluate(f'''() => {{
        const inputs = document.querySelectorAll('input');
        for (const inp of inputs) {{
            const ph = inp.placeholder || '';
            if (ph.includes('用户名')) inp.value = '{USERNAME}';
            if (ph.includes('密码') && inp.type === 'password') inp.value = '{PASSWORD}';
        }}
    }}''')

    await asyncio.sleep(random_sleep(1, 2))

    try:
        login_btn = await page.query_selector('button:has-text("登录"), input[type="submit"]')
        if login_btn:
            await login_btn.click()
            await asyncio.sleep(random_sleep(3, 5))

            # 保存cookie
            cookies = await page.context.cookies()
            COOKIE_FILE.write_text(json.dumps(cookies, ensure_ascii=False), encoding='utf-8')
            log('登录成功，已保存Cookie')
    except Exception as e:
        log(f'登录按钮点击失败: {e}')


async def select_month_filter(page, year, month):
    """选择月份筛选器"""
    try:
        # 查找月份选择器
        month_input = await page.query_selector('input[placeholder*="月"], input[placeholder*="日期"]')
        if month_input:
            await month_input.click()
            await asyncio.sleep(random_sleep(0.5, 1))

            # 输入月份
            month_str = f'{year}-{month:02d}'
            await month_input.fill(month_str)
            await asyncio.sleep(random_sleep(0.5, 1))

            # 按回车
            await month_input.press('Enter')
            await asyncio.sleep(random_sleep(2, 3))
    except Exception as e:
        log(f'月份选择失败: {e}')


async def get_page_dates(page):
    """获取当前页所有日期链接"""
    links = []
    try:
        items = await page.query_selector_all('.list-item, .data-item, tr[data-id], .price-item')
        for item in items:
            try:
                # 提取日期
                text = await item.inner_text()
                date_match = None

                # 多种日期格式匹配
                for pattern in [
                    r'(\d{4})-(\d{1,2})-(\d{1,2})',
                    r'(\d{4})/(\d{1,2})/(\d{1,2})',
                ]:
                    import re
                    m = re.search(pattern, text)
                    if m:
                        date_match = f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'
                        break

                if date_match:
                    # 判断是上午还是下午
                    period = 'AM' if '上午' in text or '早' in text else 'PM'

                    link = await item.query_selector('a[href*="mysteel"], a[href*="jiancai"]')
                    if link:
                        href = await link.get_attribute('href')
                        if href:
                            links.append({
                                'date_str': date_match,
                                'period': period,
                                'href': href if href.startswith('http') else 'https://jiancai.mysteel.com' + href
                            })
            except:
                pass
    except:
        pass
    return links


def extract_date_from_url(url):
    """从URL提取日期"""
    import re
    patterns = [
        r'(\d{4})-(\d{2})-(\d{2})',
        r'(\d{4})(\d{2})(\d{2})',
    ]
    for pattern in patterns:
        m = re.search(pattern, url)
        if m:
            date_str = f'{m.group(1)}-{m.group(2)}-{m.group(3)}'
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                period = 'PM' if 'pm' in url.lower() else 'AM'
                return date_obj, period
            except:
                pass
    return None, 'AM'


async def fetch_prices(page, date_str, period):
    """抓取单日价格数据"""
    prices = []
    screenshot_b64 = None

    try:
        # 等待数据加载
        await asyncio.sleep(random_sleep(2, 4))

        # 尝试多种选择器
        selectors = [
            'table.price-table tr',
            '.price-list tr',
            '.data-table tr',
            'tbody tr',
        ]

        for selector in selectors:
            rows = await page.query_selector_all(selector)
            if rows:
                for row in rows:
                    try:
                        cells = await row.query_selector_all('td, .cell')
                        if len(cells) >= 5:
                            material_name = await cells[0].inner_text()
                            spec = await cells[1].inner_text()
                            material_type = await cells[2].inner_text()
                            brand = await cells[3].inner_text()
                            price_text = await cells[4].inner_text()

                            # 解析价格
                            price = 0
                            import re
                            price_match = re.search(r'([\d,]+)', price_text.replace(',', ''))
                            if price_match:
                                price = int(price_match.group(1))

                            if material_name and price > 0:
                                prices.append({
                                    'material_name': material_name.strip(),
                                    'spec': spec.strip(),
                                    'material_type': material_type.strip(),
                                    'brand': brand.strip(),
                                    'price': price
                                })
                    except:
                        pass
                if prices:
                    break

        # 截图
        try:
            screenshot_b64 = await page.screenshot(full_page=True)
            screenshot_b64 = base64.b64encode(screenshot_b64).decode()
        except:
            pass

    except Exception as e:
        log(f'抓取价格失败: {e}')

    return prices, screenshot_b64


async def save_to_excel(prices, period, date_str, screenshot_b64):
    """保存到Excel - 追加模式，不覆盖已有数据"""
    try:
        if EXCEL_FILE.exists():
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE)
            except Exception:
                log(f'文件损坏，创建新文件')
                wb = openpyxl.Workbook()
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']
        else:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        # 生成sheet名
        if period == 'AM':
            sheet_name = date_str
        else:
            sheet_name = f'{date_str}_PM'

        # 检查是否已存在
        if sheet_name in wb.sheetnames:
            log(f'  Sheet已存在，跳过: {sheet_name}')
            wb.close()
            return sheet_name

        # 创建新sheet
        ws = wb.create_sheet(title=sheet_name)

        # 标题
        period_text = '下午(晚)' if period == 'PM' else '上午'
        ws.merge_cells('A1:K1')
        ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        # 表头
        header_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid') if period == 'PM' else \
            PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 数据
        fetch_time = datetime.now().strftime('%H:%M:%S')
        for i, price in enumerate(prices):
            row = 4 + i
            for col, val in enumerate([date_str, fetch_time, price['material_name'], price['spec'],
                                       price['material_type'], price['brand'], price['price'],
                                       '', '', '', '山东烟台'], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border

        # 截图
        if screenshot_b64:
            screenshot_path = DATA_DIR / f'screenshot_{date_str.replace("-", "")}_{period}.png'
            with open(screenshot_path, 'wb') as f:
                f.write(base64.b64decode(screenshot_b64))

            row = 4 + len(prices) + 2
            ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)
            img = Image(str(screenshot_path))
            img.width = 900
            img.height = 500
            img.anchor = f'A{row + 1}'
            ws.add_image(img)

        wb.save(EXCEL_FILE)
        wb.close()

        log(f'  保存成功: {sheet_name}, {len(prices)}条数据')
        return sheet_name

    except Exception as e:
        log(f'  保存失败: {e}')
        return None


async def main():
    """主函数"""
    print()
    print('=' * 70)
    print('山东烟台钢筋价格 - 全量历史数据抓取')
    print('=' * 70)
    print()

    # 获取已有数据
    existing_dates = get_existing_dates()
    log(f'已有日期数据: {len(existing_dates)} 个')
    if existing_dates:
        log(f'已有数据范围: {min(existing_dates)} 至 {max(existing_dates)}')

    # 计算需要抓取的日期
    target_months = get_target_months()
    log(f'目标月份: {len(target_months)} 个')

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            viewport={'width': 1920, 'height': 3000},
            locale='zh-CN',
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        )
        page = await context.new_page()

        await login(page)

        success_count = 0
        fail_count = 0
        skip_count = 0

        # 按月份抓取
        for year, month in target_months:
            month_str = f'{year}-{month:02d}'
            log(f'\n{"="*40}')
            log(f'处理月份: {month_str}')
            log(f'{"="*40}')

            # 打开市场总页面
            log('打开市场页面...')
            await page.goto(MARKET_URL, wait_until='domcontentloaded', timeout=60000)
            await asyncio.sleep(random_sleep(3, 5))

            # 尝试选择月份
            await select_month_filter(page, year, month)

            await human_like_mouse(page)
            await asyncio.sleep(random_sleep(1, 2))

            # 收集当月所有日期链接
            month_links = []
            page_num = 1

            while page_num <= 50:
                await human_like_mouse(page)
                await asyncio.sleep(random_sleep(1, 2))

                links = await get_page_dates(page)
                log(f'  第{page_num}页: 找到 {len(links)} 条')

                if len(links) == 0:
                    break

                for link in links:
                    date_obj, period = extract_date_from_url(link['href'])
                    if date_obj:
                        link['date_obj'] = date_obj
                        link['period'] = period
                        month_links.append(link)

                # 翻页
                try:
                    next_btn = await page.query_selector('a:has-text("下一页")')
                    if next_btn and await next_btn.is_visible():
                        await next_btn.click()
                        await asyncio.sleep(random_sleep(2, 4))
                        page_num += 1
                    else:
                        break
                except:
                    break

            # 去重
            seen = {}
            unique_links = []
            for link in month_links:
                key = (link['date_str'], link['period'])
                if key not in seen:
                    seen[key] = True
                    unique_links.append(link)
            unique_links.sort(key=lambda x: x.get('date_obj', datetime.min))

            log(f'  该月找到 {len(unique_links)} 条数据')

            # 过滤已存在的
            tasks = []
            for link in unique_links:
                if link['period'] == 'AM':
                    sheet_name = link['date_str']
                else:
                    sheet_name = f"{link['date_str']}_PM"

                # 检查Excel中是否已存在
                if link['date_str'] in existing_dates:
                    # 进一步检查AM/PM
                    am_sheet = link['date_str']
                    pm_sheet = f"{link['date_str']}_PM"

                    if sheet_name in existing_dates:
                        log(f'    {link["date_str"]} {link["period"]} 已存在，跳过')
                        skip_count += 1
                        continue

                tasks.append(link)

            log(f'  需要抓取: {len(tasks)} 条 (已有: {len(unique_links) - len(tasks)} 条)')

            # 抓取每月数据
            month_success = 0
            month_fail = 0

            for task in tasks:
                date_str = task['date_str']
                period = task['period']
                url = task['url']

                log(f'  抓取 {date_str} {period}...', end='', flush=True)

                try:
                    await page.goto(url, wait_until='domcontentloaded', timeout=60000)
                    await asyncio.sleep(random_sleep(4, 6))

                    await human_like_mouse(page)

                    prices, screenshot_b64 = await fetch_prices(page, date_str, period)

                    if prices:
                        sheet_name = await save_to_excel(prices, period, date_str, screenshot_b64)
                        if sheet_name:
                            existing_dates.add(date_str)
                            log(f' {len(prices)}条 [OK]')
                            success_count += 1
                            month_success += 1
                        else:
                            log(' 保存失败')
                            fail_count += 1
                            month_fail += 1
                    else:
                        log(' 无数据')
                        fail_count += 1
                        month_fail += 1
                except Exception as e:
                    log(f'失败: {e}')
                    fail_count += 1
                    month_fail += 1

                # 停顿15秒
                log('  停顿15秒...')
                await asyncio.sleep(random_sleep(14, 16))

            log(f'  月份完成: 成功 {month_success} 条, 失败 {month_fail} 条, 跳过 {len(unique_links) - len(tasks)} 条')

            # 月份之间停顿30秒
            log('  月份完成，停顿30秒...')
            await asyncio.sleep(30)

    await browser.close()

    print()
    print('=' * 70)
    print('抓取完成')
    print('=' * 70)
    log(f'成功: {success_count} 条')
    log(f'失败: {fail_count} 条')
    log(f'跳过: {skip_count} 条')

    # 最终统计
    final_dates = get_existing_dates()
    log(f'最终数据: {len(final_dates)} 个日期')
    if final_dates:
        log(f'数据范围: {min(final_dates)} 至 {max(final_dates)}')


if __name__ == '__main__':
    asyncio.run(main())