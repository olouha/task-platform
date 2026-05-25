"""
山东烟台钢筋价格抓取 - 支持每日多次抓取
规则：
1. 每天抓取两次：上午10:00和下午16:00
2. sheet命名格式：YYYY-MM-DD_AM/PM_HHMM
3. 较晚发布的价格标注为"晚"
"""

import asyncio
import sys
import json
import base64
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass

sys.path.insert(0, '.')

from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

DATA_DIR = Path('services/data')
DATA_DIR.mkdir(exist_ok=True)


@dataclass
class MaterialPrice:
    material_name: str = ''
    spec: str = ''
    material_type: str = ''
    brand: str = ''
    price: float = 0.0
    unit: str = '元/吨'
    region: str = '山东烟台'


def load_credentials():
    """从配置文件加载凭据"""
    config_file = DATA_DIR / 'mysteel_config.json'
    if config_file.exists():
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
                return config.get('username', 'M6616592358'), config.get('password', 'mysteel573005')
        except: pass
    return 'M6616592358', 'mysteel573005'


async def get_yantai_urls_for_date(page, date_str):
    """
    从市场页面获取指定日期的URL

    date_str格式: YYYY-MM-DD
    返回: [(url, period_label), ...]
    period_label: 'AM' 或 'PM'
    """
    year, month, day = date_str.split('-')
    date_display = f'{month}月{day}日'

    # 访问烟台市场列表页
    await page.goto('https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html',
                    wait_until='domcontentloaded', timeout=60000)
    await page.wait_for_timeout(3000)

    # 查找包含日期的链接
    links = await page.evaluate(f'''() => {{
        const links = [];
        document.querySelectorAll('a[href]').forEach(a => {{
            const href = a.href;
            const text = a.textContent.trim();
            if (href.includes('/m/26') && href.includes('.html') &&
                (text.includes('{date_display}') || text.includes('{year}-{month}-{day}'))) {{
                links.push({{text, href}});
            }}
        }});
        return links;
    }}''')

    # 按时间排序，上午在前，下午在后
    results = []
    for link in links:
        href = link['href']
        # 从URL提取时间部分
        import re
        match = re.search(r'/m/26(\d{{8}})/', href)
        if match:
            time_str = match.group(1)  # YYMMDDHH
            hour = time_str[6:8]
            if hour == '10':
                results.append((href, 'AM'))
            elif hour in ['15', '16', '17']:
                results.append((href, 'PM'))

    # 去重并排序（AM在前，PM在后）
    seen = set()
    unique_results = []
    for url, period in results:
        if url not in seen:
            seen.add(url)
            unique_results.append((url, period))

    # 排序确保AM在前
    unique_results.sort(key=lambda x: 0 if x[1] == 'AM' else 1)

    return unique_results


def save_to_excel(prices, period, date_str, screenshot_b64=None, excel_file=None):
    """保存到Excel"""
    if excel_file is None:
        excel_file = DATA_DIR / '山东烟台钢筋价格.xlsx'

    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # PM使用不同的颜色以示区分
    if period == 'PM':
        header_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')

    if Path(excel_file).exists():
        wb = openpyxl.load_workbook(excel_file)
    else:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    fetch_time = datetime.now().strftime('%H:%M:%S')
    sheet_name = f'{date_str}_{period}_{fetch_time.replace(":", "")}'

    # 删除同日期同period的旧sheet
    to_delete = [s for s in wb.sheetnames if s.startswith(f'{date_str}_{period}')]
    for s in to_delete:
        del wb[s]

    ws = wb.create_sheet(title=sheet_name)

    # 标题
    period_text = '下午(晚)' if period == 'PM' else '上午'
    ws.merge_cells('A1:K1')
    ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    # 表头
    headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 数据
    for i, price in enumerate(prices):
        row = 4 + i
        if isinstance(price, dict):
            material_name = price.get('material_name', '')
            spec = price.get('spec', '')
            material_type = price.get('material_type', '')
            brand = price.get('brand', '')
            price_val = price.get('price', 0)
            region = price.get('region', '山东烟台')
        else:
            material_name = price.material_name
            spec = price.spec
            material_type = price.material_type
            brand = price.brand
            price_val = price.price
            region = price.region

        for col, val in enumerate([date_str, fetch_time, material_name, spec,
                                   material_type, brand, price_val, '', '', '', region], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border

    # 嵌入截图
    if screenshot_b64:
        screenshot_path = DATA_DIR / f'screenshot_{date_str}_{period}.png'
        with open(screenshot_path, 'wb') as f:
            f.write(base64.b64decode(screenshot_b64))

        row = 4 + len(prices) + 2
        ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)

        img = Image(str(screenshot_path))
        img.width = 900
        img.height = 500
        img.anchor = f'A{row + 1}'
        ws.add_image(img)

    wb.save(excel_file)
    wb.close()
    print(f'Excel已保存: {excel_file}')
    print(f'数据: {len(prices)}条, Sheet: {sheet_name}')


async def fetch_date_prices(date_str):
    """抓取指定日期的上午和下午价格"""
    cookie_file = DATA_DIR / 'mysteel_cookies.json'
    username, password = load_credentials()
    print(f'抓取日期: {date_str}')
    print(f'使用凭据: {username[:3]}***')

    results = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(viewport={'width': 1920, 'height': 3000}, locale='zh-CN')
        page = await context.new_page()

        # 登录
        print('1. 登录...')
        await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
        await page.wait_for_timeout(5000)

        try:
            account_tab = await page.query_selector('.form-tab-account, a[data-tab="account"]')
            if account_tab:
                await account_tab.click()
                await page.wait_for_timeout(2000)
        except: pass

        await page.evaluate(f'''() => {{
            const inputs = document.querySelectorAll('input');
            for (const inp of inputs) {{
                const ph = inp.placeholder || '';
                if (ph.includes('用户名')) inp.value = '{username}';
                if (ph.includes('密码') && inp.type === 'password') inp.value = '{password}';
            }}
        }}''')

        await page.wait_for_timeout(500)

        try:
            login_btn = await page.query_selector('.form-button-login, button:has-text("登录")')
            if login_btn:
                await login_btn.click()
        except: pass

        await page.wait_for_timeout(8000)

        # 保存Cookie
        cookies = await context.cookies()
        with open(cookie_file, 'w', encoding='utf-8') as f:
            json.dump(cookies, f, ensure_ascii=False)
        print('  登录完成')

        # 获取URL
        print('2. 获取价格URL...')
        urls = await get_yantai_urls_for_date(page, date_str)
        print(f'  找到 {len(urls)} 个URL')

        # 抓取上午和下午价格
        print('3. 抓取价格...')
        for url, period in urls:
            print(f'  {period}...', end='')
            await page.goto(url, wait_until='domcontentloaded', timeout=60000)
            await page.wait_for_timeout(5000)

            # 截图
            screenshot = await page.screenshot(full_page=True)
            screenshot_b64 = base64.b64encode(screenshot).decode('utf-8')

            # 提取价格
            data = await page.evaluate('''() => {
                const tables = document.querySelectorAll('table');
                const results = [];
                tables.forEach(table => {
                    const rows = table.querySelectorAll('tr');
                    rows.forEach(row => {
                        const cells = row.querySelectorAll('td');
                        if (cells.length >= 5) {
                            const material_name = cells[0]?.textContent?.trim();
                            const spec = cells[1]?.textContent?.trim();
                            const material_type = cells[2]?.textContent?.trim();
                            const brand = cells[3]?.textContent?.trim();
                            const price = cells[4]?.textContent?.trim();

                            if (['高线', '螺纹钢', '盘螺', '圆钢'].includes(material_name) &&
                                spec && spec.startsWith('Φ') && price && /^\\d+$/.test(price)) {
                                results.push({
                                    material_name,
                                    spec,
                                    material_type,
                                    brand,
                                    price: parseFloat(price),
                                    region: '山东烟台'
                                });
                            }
                        }
                    });
                });
                return results;
            }''')

            prices = list(data)
            print(f' {len(prices)}条')

            # 保存
            if prices:
                save_to_excel(prices, period, date_str, screenshot_b64)
                results.append({
                    'period': period,
                    'count': len(prices),
                    'prices': prices
                })

        await browser.close()

    return results


async def main():
    """重新抓取5.13-5.16的数据"""
    dates = [
        '2026-05-13',
        '2026-05-14',
        '2026-05-15',
        '2026-05-16',
    ]

    all_results = {}
    for date_str in dates:
        results = await fetch_date_prices(date_str)
        all_results[date_str] = results
        print(f'  完成 {date_str}')

    print()
    print('=' * 60)
    print('抓取汇总')
    print('=' * 60)
    for date_str, results in all_results.items():
        am_count = next((r['count'] for r in results if r['period'] == 'AM'), 0)
        pm_count = next((r['count'] for r in results if r['period'] == 'PM'), 0)
        print(f'{date_str}: 上午{am_count}条, 下午{pm_count}条, 共{am_count + pm_count}条')

    print(f'总计: {sum(len(r) for r in all_results.values())}天数据')


if __name__ == '__main__':
    asyncio.run(main())