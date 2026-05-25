"""
批量抓取一个月的历史价格数据
"""

import asyncio
from playwright.async_api import async_playwright
from pathlib import Path
import json
import base64
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

DATA_DIR = Path('services/data')
DATA_DIR.mkdir(exist_ok=True)

COOKIE_FILE = DATA_DIR / 'mysteel_cookies.json'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
MARKET_URL = 'https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html'


async def login(page):
    """登录我的钢铁网"""
    username = 'M6616592358'
    password = 'mysteel573005'

    print('登录中...')
    await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
    await page.wait_for_timeout(5000)

    try:
        account_tab = await page.query_selector('.form-tab-account, a[data-tab="account"]')
        if account_tab:
            await account_tab.click()
            await page.wait_for_timeout(2000)
    except:
        pass

    await page.evaluate(f'''() => {{
        const inputs = document.querySelectorAll('input');
        for (const inp of inputs) {{
            const ph = inp.placeholder || '';
            if (ph.includes('用户名')) inp.value = '{username}';
            if (ph.includes('密码') && inp.type === 'password') inp.value = '{password}';
        }}
    }}''')

    await page.wait_for_timeout(500)

    try:
        login_btn = await page.query_selector('.form-button-login, button:has-text("登录")')
        if login_btn:
            await login_btn.click()
    except:
        pass

    await page.wait_for_timeout(8000)

    # 保存Cookie
    cookies = await page.context.cookies()
    with open(COOKIE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cookies, f, ensure_ascii=False)
    print('登录完成')


def get_sheet_name(date_str, period):
    """生成sheet名称"""
    fetch_time = datetime.now().strftime('%H:%M:%S')
    return f'{date_str}_{period}_{fetch_time.replace(":", "")}'


def save_to_excel(prices, period, date_str, screenshot_b64):
    """保存到Excel"""
    wb = None
    try:
        if EXCEL_FILE.exists():
            wb = openpyxl.load_workbook(EXCEL_FILE)
        else:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        header_font = Font(bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid') if period == 'PM' else \
            PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        sheet_name = get_sheet_name(date_str, period)
        ws = wb.create_sheet(title=sheet_name)

        # 标题
        period_text = '下午(晚)' if period == 'PM' else '上午'
        ws.merge_cells('A1:K1')
        ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        # 表头
        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 数据
        fetch_time = datetime.now().strftime('%H:%M:%S')
        for i, price in enumerate(prices):
            row = 4 + i
            for col, val in enumerate([date_str, fetch_time, price['material_name'], price['spec'],
                                       price['material_type'], price['brand'], price['price'],
                                       '', '', '', '山东烟台'], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border

        # 截图
        if screenshot_b64:
            screenshot_path = DATA_DIR / f'screenshot_{date_str.replace("-", "")}_{period}.png'
            with open(screenshot_path, 'wb') as f:
                f.write(base64.b64decode(screenshot_b64))

            row = 4 + len(prices) + 2
            ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)

            img = Image(str(screenshot_path))
            img.width = 900
            img.height = 500
            img.anchor = f'A{row + 1}'
            ws.add_image(img)

        wb.save(EXCEL_FILE)
        wb.close()

        return sheet_name, len(prices)
    except Exception as e:
        if wb:
            wb.close()
        raise e


async def fetch_date_urls(page, date_str):
    """获取指定日期的URL列表"""
    # 访问烟台市场页面
    market_url = 'https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html'
    await page.goto(market_url, wait_until='domcontentloaded', timeout=60000)
    await page.wait_for_timeout(5000)

    # 获取所有/m/26链接
    links = await page.evaluate('''() => {
        const links = [];
        document.querySelectorAll('a[href]').forEach(a => {
            const href = a.href;
            if (href.includes('/m/26') && href.includes('.html')) {
                links.push(href);
            }
        });
        return links;
    }''')

    results = []
    for href in links:
        import re
        match = re.search(r'/m/26(\d{8})/', href)
        if match:
            time_str = match.group(1)  # YYMMDDHH
            url_date = f"20{time_str[0:2]}-{time_str[2:4]}-{time_str[4:6]}"
            if url_date == date_str:
                hour = int(time_str[6:8])
                if hour == 10:
                    results.append((href, 'AM'))
                elif hour >= 15:
                    results.append((href, 'PM'))

    # 去重并排序
    seen = set()
    unique_results = []
    for url, period in results:
        if url not in seen:
            seen.add(url)
            unique_results.append((url, period))

    unique_results.sort(key=lambda x: 0 if x[1] == 'AM' else 1)
    return unique_results


async def fetch_prices(page, url):
    """抓取指定URL的价格"""
    await page.goto(url, wait_until='domcontentloaded', timeout=60000)
    await page.wait_for_timeout(5000)

    # 截图
    screenshot = await page.screenshot(full_page=True)
    screenshot_b64 = base64.b64encode(screenshot).decode('utf-8')

    # 提取价格
    data = await page.evaluate('''() => {
        const tables = document.querySelectorAll('table');
        const results = [];
        tables.forEach(table => {
            const rows = table.querySelectorAll('tr');
            rows.forEach(row => {
                const cells = row.querySelectorAll('td');
                if (cells.length >= 5) {
                    const material_name = cells[0]?.textContent?.trim();
                    const spec = cells[1]?.textContent?.trim();
                    const material_type = cells[2]?.textContent?.trim();
                    const brand = cells[3]?.textContent?.trim();
                    const price = cells[4]?.textContent?.trim();

                    if (['高线', '螺纹钢', '盘螺', '圆钢'].includes(material_name) &&
                        spec && spec.startsWith('Φ') && price && /^\\d+$/.test(price)) {
                        results.push({
                            material_name,
                            spec,
                            material_type,
                            brand,
                            price: parseFloat(price),
                            region: '山东烟台'
                        });
                    }
                }
            });
        });
        return results;
    }''')

    return list(data), screenshot_b64


async def main():
    """主程序"""
    # 获取过去一个月的日期
    today = datetime.now().date()
    dates = []
    for i in range(30):
        date = today - timedelta(days=i + 1)
        dates.append(date.isoformat())

    print(f'计划抓取 {len(dates)} 天的数据')
    print(f'日期范围: {dates[-1]} 至 {dates[0]}')
    print()

    # 加载Cookie
    cookies = []
    if COOKIE_FILE.exists():
        cookies = json.loads(COOKIE_FILE.read_text(encoding='utf-8'))

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(viewport={'width': 1920, 'height': 3000}, locale='zh-CN')
        if cookies:
            await context.add_cookies(cookies)

        page = await context.new_page()

        await login(page)

        results = {}

        for i, date_str in enumerate(dates):
            print(f'[{i+1}/{len(dates)}] 抓取 {date_str}...', end='')

            try:
                urls = await fetch_date_urls(page, date_str)

                if not urls:
                    print(' 未找到URL')
                    continue

                date_results = {}
                for url, period in urls:
                    try:
                        prices, screenshot_b64 = await fetch_prices(page, url)
                        if prices:
                            sheet_name = await save_to_excel(prices, period, date_str, screenshot_b64)
                            date_results[period] = {'sheet': sheet_name, 'count': len(prices)}
                    except Exception as e:
                        print(f' {period}抓取失败: {e}', end='')

                if date_results:
                    am = date_results.get('AM', {}).get('count', 0)
                    pm = date_results.get('PM', {}).get('count', 0)
                    results[date_str] = date_results
                    print(f' AM={am}, PM={pm}')
                else:
                    print(' 无数据')

            except Exception as e:
                print(f' 错误: {e}')

        await browser.close()

    print()
    print('=' * 60)
    print('抓取完成')
    print('=' * 60)
    print(f'成功天数: {len(results)}')
    for date_str, data in sorted(results.items()):
        am = data.get('AM', {}).get('count', 0)
        pm = data.get('PM', {}).get('count', 0)
        print(f'  {date_str}: AM={am}, PM={pm}')


if __name__ == '__main__':
    asyncio.run(main())