"""
批量抓取历史价格数据 - 改进版
每次抓取前先检查登录状态
"""

import asyncio
from playwright.async_api import async_playwright
from pathlib import Path
import json
import base64
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

DATA_DIR = Path('services/data')
COOKIE_FILE = DATA_DIR / 'mysteel_cookies.json'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
URLS_FILE = DATA_DIR / 'fetch_urls_list.json'
PROCESSED_FILE = DATA_DIR / 'processed_urls.json'

# 从文件加载可用URL
AVAILABLE_URLS = []
if URLS_FILE.exists():
    with open(URLS_FILE, 'r', encoding='utf-8') as f:
        AVAILABLE_URLS = [(d, p, u) for d, p, u in json.load(f)]


def get_sheet_name(date_str, period):
    fetch_time = datetime.now().strftime('%H%M%S')
    return f'{date_str}_{period}_{fetch_time}'


async def check_logged_in(page):
    """检查是否已登录"""
    await page.goto('https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html', wait_until='domcontentloaded', timeout=60000)
    await page.wait_for_timeout(3000)

    # 检查是否有登录按钮或是否跳转到登录页
    body_text = await page.evaluate('() => document.body.textContent')
    if '登录' in body_text or 'passport' in page.url:
        return False
    return True


async def login(page):
    """登录"""
    print('登录中...')
    await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
    await page.wait_for_timeout(5000)

    username = 'M6616592358'
    password = 'mysteel573005'

    try:
        account_tab = await page.query_selector('.form-tab-account, a[data-tab="account"]')
        if account_tab:
            await account_tab.click()
            await page.wait_for_timeout(2000)
    except:
        pass

    await page.evaluate(f'''() => {{
        const inputs = document.querySelectorAll('input');
        for (const inp of inputs) {{
            const ph = inp.placeholder || '';
            if (ph.includes('用户名')) inp.value = '{username}';
            if (ph.includes('密码') && inp.type === 'password') inp.value = '{password}';
        }}
    }}''')

    await page.wait_for_timeout(500)

    try:
        login_btn = await page.query_selector('.form-button-login, button:has-text("登录")')
        if login_btn:
            await login_btn.click()
    except:
        pass

    await page.wait_for_timeout(8000)

    # 保存Cookie
    cookies = await page.context.cookies()
    with open(COOKIE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cookies, f, ensure_ascii=False)


async def fetch_prices(page, url):
    await page.goto(url, wait_until='domcontentloaded', timeout=60000)
    await page.wait_for_timeout(5000)

    screenshot = await page.screenshot(full_page=True)
    screenshot_b64 = base64.b64encode(screenshot).decode('utf-8')

    data = await page.evaluate('''() => {
        const tables = document.querySelectorAll('table');
        const results = [];
        tables.forEach(table => {
            const rows = table.querySelectorAll('tr');
            rows.forEach(row => {
                const cells = row.querySelectorAll('td');
                if (cells.length >= 5) {
                    const material_name = cells[0]?.textContent?.trim();
                    const spec = cells[1]?.textContent?.trim();
                    const material_type = cells[2]?.textContent?.trim();
                    const brand = cells[3]?.textContent?.trim();
                    const price = cells[4]?.textContent?.trim();

                    if (['高线', '螺纹钢', '盘螺', '圆钢'].includes(material_name) &&
                        spec && spec.startsWith('Φ') && price && /^\\d+$/.test(price)) {
                        results.push({
                            material_name,
                            spec,
                            material_type,
                            brand,
                            price: parseFloat(price),
                            region: '山东烟台'
                        });
                    }
                }
            });
        });
        return results;
    }''')

    return list(data), screenshot_b64


async def save_to_excel(prices, period, date_str, screenshot_b64):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE) if EXCEL_FILE.exists() else openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        sheet_name = get_sheet_name(date_str, period)
        ws = wb.create_sheet(title=sheet_name)

        # 标题
        period_text = '下午(晚)' if period == 'PM' else '上午'
        ws.merge_cells('A1:K1')
        ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        # 表头
        header_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid') if period == 'PM' else \
            PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 数据
        fetch_time = datetime.now().strftime('%H:%M:%S')
        for i, price in enumerate(prices):
            row = 4 + i
            for col, val in enumerate([date_str, fetch_time, price['material_name'], price['spec'],
                                       price['material_type'], price['brand'], price['price'],
                                       '', '', '', '山东烟台'], 1):
                cell = ws.cell(row=row, column=col, value=val)

        # 截图
        if screenshot_b64:
            screenshot_path = DATA_DIR / f'screenshot_{date_str.replace("-", "")}_{period}.png'
            with open(screenshot_path, 'wb') as f:
                f.write(base64.b64decode(screenshot_b64))

            row = 4 + len(prices) + 2
            ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)
            img = Image(str(screenshot_path))
            img.width = 900
            img.height = 500
            img.anchor = f'A{row + 1}'
            ws.add_image(img)

        wb.save(EXCEL_FILE)
        wb.close()
        return sheet_name
    except Exception as e:
        print(f'保存失败: {e}')
        import traceback
        traceback.print_exc()
        return None


async def main():
    # 获取已存在的sheet
    existing_sheets = set()
    if EXCEL_FILE.exists():
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            existing_sheets = set(wb.sheetnames)
            wb.close()
        except:
            pass

    print(f'已有sheet: {len(existing_sheets)}个')
    print(f'待抓取URL: {len(AVAILABLE_URLS)}条')

    # 过滤已抓取的
    tasks = []
    for date_str, period, url in AVAILABLE_URLS:
        # 检查是否已有该日期该时段的sheet
        has_existing = any(s.startswith(f'{date_str}_{period}_') for s in existing_sheets)
        if not has_existing:
            tasks.append((date_str, period, url))

    print(f'需要抓取: {len(tasks)} 条')
    if not tasks:
        print('所有数据已抓取')
        return

    # 限制每次抓取数量（分批抓取）
    BATCH_SIZE = 100
    tasks_to_fetch = tasks[:BATCH_SIZE]
    print(f'本次抓取: {len(tasks_to_fetch)} 条 (剩余: {len(tasks) - BATCH_SIZE})')

    # 登录
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(viewport={'width': 1920, 'height': 3000}, locale='zh-CN')

        page = await context.new_page()
        await login(page)
        print()

        results = {}
        success_count = 0
        fail_count = 0

        for i, (date_str, period, url) in enumerate(tasks_to_fetch):
            print(f'[{i+1}/{len(tasks_to_fetch)}] {date_str} {period}...', end='', flush=True)

            try:
                prices, screenshot_b64 = await fetch_prices(page, url)
                if prices:
                    sheet_name = await save_to_excel(prices, period, date_str, screenshot_b64)
                    if sheet_name:
                        results[date_str] = results.get(date_str, {})
                        results[date_str][period] = {'sheet': sheet_name, 'count': len(prices)}
                        print(f' {len(prices)}条 OK', flush=True)
                        success_count += 1
                    else:
                        print(' Save failed', flush=True)
                        fail_count += 1
                else:
                    print(' No data', flush=True)
                    fail_count += 1
            except Exception as e:
                print(f' Error: {e}', flush=True)
                fail_count += 1

            # 每10个保存进度
            if (i + 1) % 10 == 0:
                print(f'  进度: OK={success_count}, Fail={fail_count}', flush=True)
                # 重新加载Excel以获取最新sheet列表
                if EXCEL_FILE.exists():
                    try:
                        wb = openpyxl.load_workbook(EXCEL_FILE)
                        existing_sheets = set(wb.sheetnames)
                        wb.close()
                    except:
                        pass

        await browser.close()

    print()
    print('=' * 60)
    print(f'抓取完成: OK={success_count}, Fail={fail_count}')
    print('=' * 60)


if __name__ == '__main__':
    asyncio.run(main())