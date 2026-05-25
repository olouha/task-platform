"""
全量数据抓取 - 从yantai_links.json读取链接
每个sheet一页截图，直接追加不覆盖
"""
import asyncio
import sys
import json
import base64
import random
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict

sys.path.insert(0, '.')

from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

DATA_DIR = Path('services/data')
DATA_DIR.mkdir(exist_ok=True)

COOKIE_FILE = DATA_DIR / 'mysteel_cookies.json'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
SCREENSHOT_DIR = DATA_DIR / 'screenshots'
SCREENSHOT_DIR.mkdir(exist_ok=True)

LOG_FILE = DATA_DIR / 'logs' / f'full_fetch_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'

# 登录凭据
USERNAME = 'M6616592358'
PASSWORD = 'mysteel573005'


def log(msg):
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f'[{ts}] {msg}'
    print(line)
    LOG_FILE.parent.mkdir(exist_ok=True)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(line + '\n')


def random_sleep(min_sec=14, max_sec=16):
    return random.uniform(min_sec, max_sec)


def get_existing_sheets():
    """获取Excel中已存在的所有sheet名"""
    existing = set()
    if EXCEL_FILE.exists():
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
            existing = set(wb.sheetnames)
            wb.close()
        except:
            pass
    return existing


def parse_date_from_url(url):
    """从URL解析日期"""
    # 格式: /m/26051516/ -> 2026-05-15 PM 16:00
    m = re.search(r'/m/(\d{8})/', url)
    if m:
        code = m.group(1)
        year = 2000 + int(code[:2])
        month = int(code[2:4])
        day = int(code[4:6])
        time_hour = int(code[6:8]) if len(code) >= 8 else 12
        return f'{year}-{month:02d}-{day:02d}', time_hour
    return None, None


async def human_like_mouse(page):
    try:
        vp = page.viewport_size
        if vp:
            for _ in range(random.randint(2, 4)):
                x = random.randint(100, vp['width'] - 100)
                y = random.randint(100, vp['height'] - 100)
                await page.mouse.move(x, y)
                await asyncio.sleep(random.uniform(0.2, 0.5))
    except:
        pass


async def login(page):
    log('登录中...')

    # 加载Cookie
    if COOKIE_FILE.exists():
        try:
            cookies = json.loads(COOKIE_FILE.read_text(encoding='utf-8'))
            if cookies:
                await page.context.add_cookies(cookies)
                log('已加载Cookie')
        except:
            pass

    await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
    await asyncio.sleep(random_sleep(2, 4))

    await human_like_mouse(page)

    # 填入账号密码
    await page.evaluate(f'''() => {{
        const inputs = document.querySelectorAll('input');
        for (const inp of inputs) {{
            const ph = inp.placeholder || '';
            if (ph.includes('用户名')) inp.value = '{USERNAME}';
            if (ph.includes('密码') && inp.type === 'password') inp.value = '{PASSWORD}';
        }}
    }}''')

    await asyncio.sleep(random_sleep(1, 2))

    # 点击登录
    try:
        btn = await page.query_selector('button:has-text("登录"), input[type="submit"]')
        if btn:
            await btn.click()
            await asyncio.sleep(random_sleep(3, 5))
            cookies = await page.context.cookies()
            COOKIE_FILE.write_text(json.dumps(cookies, ensure_ascii=False), encoding='utf-8')
            log('登录成功，已保存Cookie')
    except Exception as e:
        log(f'登录失败: {e}')


async def fetch_page_prices(page):
    """从当前页面抓取价格数据"""
    prices = []
    try:
        await asyncio.sleep(random_sleep(2, 4))

        # 多种选择器尝试
        selectors = [
            'table tr',
            '.price-list tr',
            '.data-table tr',
            '.market-table tr',
            'tbody tr',
        ]

        for sel in selectors:
            rows = await page.query_selector_all(sel)
            if len(rows) > 3:
                for row in rows:
                    cells = await row.query_selector_all('td, .cell')
                    if len(cells) >= 5:
                        try:
                            material = await cells[0].inner_text()
                            spec = await cells[1].inner_text()
                            mtype = await cells[2].inner_text()
                            brand = await cells[3].inner_text()
                            price_text = await cells[4].inner_text()

                            price_match = re.search(r'([\d,]+)', price_text.replace(',', ''))
                            if price_match and material:
                                prices.append({
                                    'material_name': material.strip(),
                                    'spec': spec.strip(),
                                    'material_type': mtype.strip(),
                                    'brand': brand.strip(),
                                    'price': int(price_match.group(1))
                                })
                        except:
                            pass
                if prices:
                    break
    except Exception as e:
        log(f'解析价格失败: {e}')

    return prices


async def save_sheet(prices, date_str, period, screenshot_b64):
    """保存一个sheet到Excel"""
    try:
        if EXCEL_FILE.exists():
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE)
            except:
                wb = openpyxl.Workbook()
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']
        else:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        # 生成sheet名
        sheet_name = f'{date_str}_PM' if period == 'PM' else date_str
        if sheet_name in wb.sheetnames:
            log(f'  {sheet_name} 已存在，跳过')
            wb.close()
            return None

        ws = wb.create_sheet(title=sheet_name)

        # 标题
        period_text = '下午' if period == 'PM' else '上午'
        ws.merge_cells('A1:K1')
        ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        # 表头
        header_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid') if period == 'PM' else \
                      PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 数据行
        fetch_time = datetime.now().strftime('%H:%M:%S')
        for i, p in enumerate(prices):
            row = 4 + i
            for col, val in enumerate([
                date_str, fetch_time, p['material_name'], p['spec'],
                p['material_type'], p['brand'], p['price'],
                '', '', '', '山东烟台'
            ], 1):
                ws.cell(row=row, column=col, value=val).border = thin_border

        # 保存截图
        if screenshot_b64:
            ss_path = SCREENSHOT_DIR / f'{date_str.replace("-", "")}_{period}.png'
            with open(ss_path, 'wb') as f:
                f.write(base64.b64decode(screenshot_b64))

            row = 4 + len(prices) + 2
            ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)
            img = Image(str(ss_path))
            img.width = 900
            img.height = 500
            img.anchor = f'A{row + 1}'
            ws.add_image(img)

        wb.save(EXCEL_FILE)
        wb.close()

        log(f'  保存成功: {sheet_name}, {len(prices)}条数据, 截图: {ss_path.name}')
        return sheet_name

    except Exception as e:
        log(f'  保存失败: {e}')
        return None


async def main():
    print()
    print('=' * 70)
    print('山东烟台钢筋价格 - 全量数据抓取')
    print('=' * 70)
    print()

    # 读取链接文件
    links_file = DATA_DIR / 'yantai_links.json'
    if not links_file.exists():
        log('错误: yantai_links.json 不存在')
        return

    with open(links_file, 'r', encoding='utf-8') as f:
        all_links = json.load(f)

    log(f'读取到 {len(all_links)} 条链接')

    # 获取已有数据
    existing = get_existing_sheets()
    log(f'已有sheet: {len(existing)} 个')

    # 解析所有链接，生成任务列表
    tasks = []
    for item in all_links:
        url = item.get('href', '')
        date_str, hour = parse_date_from_url(url)
        if date_str:
            period = 'PM' if hour >= 12 else 'AM'
            sheet_name = f'{date_str}_PM' if period == 'PM' else date_str

            # 跳过已存在的
            if sheet_name in existing:
                continue

            tasks.append({
                'url': url,
                'date': date_str,
                'period': period,
                'sheet': sheet_name
            })

    log(f'需要抓取: {len(tasks)} 条')

    # 按日期排序
    tasks.sort(key=lambda x: x['date'])

    # 统计
    by_month = defaultdict(list)
    for t in tasks:
        by_month[t['date'][:7]].append(t)

    log('按月统计:')
    for month in sorted(by_month.keys()):
        log(f'  {month}: {len(by_month[month])} 条')

    # 启动浏览器
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            viewport={'width': 1920, 'height': 3000},
            locale='zh-CN',
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        )
        page = await context.new_page()

        await login(page)

        success = 0
        fail = 0

        # 批量抓取
        for i, task in enumerate(tasks):
            log(f'[{i+1}/{len(tasks)}] 抓取 {task["date"]} {task["period"]}...')

            try:
                await page.goto(task['url'], wait_until='domcontentloaded', timeout=60000)
                await asyncio.sleep(random_sleep(4, 6))

                await human_like_mouse(page)

                # 截图
                screenshot_b64 = await page.screenshot(full_page=True)
                screenshot_b64 = base64.b64encode(screenshot_b64).decode()

                # 抓取数据
                prices = await fetch_page_prices(page)

                if prices:
                    sheet = await save_sheet(prices, task['date'], task['period'], screenshot_b64)
                    if sheet:
                        success += 1
                        existing.add(sheet)
                    else:
                        fail += 1
                else:
                    log(f'  无数据')
                    fail += 1

            except Exception as e:
                log(f'  失败: {e}')
                fail += 1

            # 停顿反爬
            log(f'  停顿15秒...')
            await asyncio.sleep(random_sleep(14, 16))

            # 每50条报告一次
            if (i + 1) % 50 == 0:
                log(f'\n进度报告: {i+1}/{len(tasks)} 完成, 成功: {success}, 失败: {fail}\n')

        await browser.close()

    print()
    print('=' * 70)
    print('抓取完成')
    print('=' * 70)
    log(f'成功: {success} 条')
    log(f'失败: {fail} 条')

    # 最终统计
    final = get_existing_sheets()
    log(f'Excel现有: {len(final)} 个sheet')


if __name__ == '__main__':
    asyncio.run(main())