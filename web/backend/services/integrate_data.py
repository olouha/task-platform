"""
整合所有本地数据到一个Excel文件
- 从备份Excel中提取所有价格数据
- 嵌入对应的截图
"""
import os
import json
import base64
from pathlib import Path
from datetime import datetime
from PIL import Image

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage

DATA_DIR = Path('services/data')
BACKUP_DIR = DATA_DIR

def get_all_data_from_backup():
    """从备份文件中提取所有数据"""
    print('=' * 70)
    print('整合所有本地数据到一个Excel')
    print('=' * 70)
    print()

    # 找到最大的备份文件
    files = [f for f in os.listdir(BACKUP_DIR) if f.startswith('山东烟台钢筋价格_backup') and f.endswith('.xlsx')]
    files.sort(key=lambda x: os.path.getsize(os.path.join(BACKUP_DIR, x)), reverse=True)

    if not files:
        print('错误: 未找到备份文件')
        return

    backup_file = os.path.join(BACKUP_DIR, files[0])
    print(f'使用备份文件: {files[0]}')
    print(f'文件大小: {os.path.getsize(backup_file) / 1024 / 1024:.2f}MB')
    print()

    # 打开备份文件
    wb = openpyxl.load_workbook(backup_file, read_only=True)
    all_sheets = wb.sheetnames
    print(f'总Sheet数: {len(all_sheets)}')

    # 创建新Excel
    new_wb = openpyxl.Workbook()
    if 'Sheet' in new_wb.sheetnames:
        del new_wb['Sheet']

    # 统计
    total_prices = 0
    skipped = 0

    # 定义样式
    header_font = Font(bold=True, size=12, color='FFFFFF')
    am_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    pm_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 处理每个sheet
    for i, sheet_name in enumerate(all_sheets):
        print(f'处理 [{i+1}/{len(all_sheets)}]: {sheet_name}', end=' ')

        try:
            ws = wb[sheet_name]

            # 提取数据
            headers = []
            data_rows = []

            for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
                if row[0] is None and len([x for x in row if x is not None]) == 0:
                    continue

                if len(headers) == 0:
                    # 检查是否是表头行
                    if row[0] and '品名' in str(row[0]):
                        headers = list(row)
                        continue

                # 检查是否是标题行
                if row[0] and isinstance(row[0], str) and '价格' in row[0]:
                    continue

                # 检查是否是截图说明行
                if row[0] and isinstance(row[0], str) and '截图' in row[0]:
                    continue

                # 数据行
                if row[0] and isinstance(row[0], str) and row[0].startswith('20'):
                    data_rows.append(list(row))
                elif row[1] and isinstance(row[1], str) and row[1].startswith('20'):
                    data_rows.append(list(row))

            if not data_rows:
                # 可能是一个特殊格式的sheet，尝试其他方式
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
                    if row and row[0]:
                        text = str(row[0])
                        if text.startswith('20'):
                            date_str = text[:10]
                            # 这可能是一个数据行
                            break

                print('(无数据行) ', end='')
                skipped += 1
                continue

            # 确定是AM还是PM
            period = 'AM' if '_AM' in sheet_name or '上午' in sheet_name else 'PM'
            if '_PM' in sheet_name:
                period = 'PM'

            # 获取截图
            screenshot_path = None
            date_str = sheet_name[:10].replace('-', '')
            for suffix in ['_AM.png', '_PM.png', '.png']:
                potential = DATA_DIR / f'screenshot_{date_str}{suffix}'
                if potential.exists():
                    screenshot_path = potential
                    break

            # 创建新sheet
            new_ws = new_wb.create_sheet(title=sheet_name[:31])  # Excel sheet名限制31字符

            # 标题
            period_text = '下午(晚)' if period == 'PM' else '上午'
            new_ws.merge_cells('A1:K1')
            new_ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {sheet_name[:10]} {period_text}').font = Font(bold=True, size=14)
            new_ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

            # 表头
            header_list = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
            for col, header in enumerate(header_list, 1):
                cell = new_ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = pm_fill if period == 'PM' else am_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # 数据
            for row_idx, row in enumerate(data_rows):
                row_num = 4 + row_idx
                for col, val in enumerate(row[:11], 1):
                    cell = new_ws.cell(row=row_num, column=col, value=val)
                    cell.border = thin_border

                total_prices += 1

            # 嵌入截图
            if screenshot_path and screenshot_path.exists():
                try:
                    # 压缩截图并保存
                    img = Image.open(screenshot_path)
                    img = img.resize((900, 600), Image.LANCZOS)
                    compressed_path = DATA_DIR / f'temp_compressed_{i}.png'
                    img.save(str(compressed_path), 'PNG', quality=85)

                    row = 4 + len(data_rows) + 2
                    new_ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)

                    xl_img = XLImage(str(compressed_path))
                    xl_img.width = 900
                    xl_img.height = 600
                    xl_img.anchor = f'A{row + 1}'
                    new_ws.add_image(xl_img)

                    # 在保存前删除临时文件
                    # 延迟删除，等Excel保存完
                except Exception as e:
                    print(f'截图失败: {e}')

            print(f'{len(data_rows)}条数据')

        except Exception as e:
            print(f'错误: {e}')
            skipped += 1

    wb.close()

    # 保存
    output_file = DATA_DIR / '山东烟台钢筋价格_整合版.xlsx'
    new_wb.save(str(output_file))
    new_wb.close()

    # 清理临时文件
    for f in DATA_DIR.glob('temp_compressed_*.png'):
        try:
            os.remove(f)
        except:
            pass

    print()
    print('=' * 70)
    print('整合完成')
    print('=' * 70)
    print(f'输出文件: {output_file}')
    print(f'总价格记录: {total_prices}')
    print(f'跳过: {skipped}')
    print(f'文件大小: {os.path.getsize(output_file) / 1024 / 1024:.2f}MB')


if __name__ == '__main__':
    get_all_data_from_backup()