"""
将Excel中的价格数据导入到SQLite数据库
"""
import sqlite3
import openpyxl
import logging
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)

# 配置
DATA_DIR = Path(__file__).parent / 'data'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格_完整版.xlsx'
DB_FILE = DATA_DIR / 'yantai_rebar.db'

# 回退选项
BACKUP_FILES = [
    '山东烟台钢筋价格.xlsx',
    '山东烟台钢筋价格_current.xlsx',
]


def get_excel_file():
    """获取可用的Excel文件"""
    for candidate in [EXCEL_FILE] + [DATA_DIR / f for f in BACKUP_FILES]:
        if candidate.exists():
            return candidate
    return None


def init_database():
    """初始化数据库表"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    # 创建价格表
    c.execute('''
        CREATE TABLE IF NOT EXISTS rebar_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            fetch_time TEXT,
            material_name TEXT,
            spec TEXT,
            material_type TEXT,
            brand TEXT,
            price INTEGER,
            price_change TEXT,
            remark TEXT,
            region TEXT DEFAULT '山东烟台',
            UNIQUE(date, material_name, spec, brand, price)
        )
    ''')

    # 创建索引
    c.execute('CREATE INDEX IF NOT EXISTS idx_date ON rebar_prices(date)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_material ON rebar_prices(material_name)')

    conn.commit()
    conn.close()
    logger.info("[init_database] 数据库初始化完成")


def import_excel_to_db(excel_file: str, clear_existing: bool = False):
    """
    从Excel文件导入价格数据到数据库

    Args:
        excel_file: Excel文件路径
        clear_existing: 是否清除现有数据
    """
    logger.info(f"[import_excel_to_db] 开始导入 | file={excel_file}")

    # 初始化数据库
    init_database()

    # 如果需要清除现有数据
    if clear_existing:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute('DELETE FROM rebar_prices')
        conn.commit()
        conn.close()
        logger.info("[import_excel_to_db] 已清除现有数据")

    # 打开Excel
    wb = openpyxl.load_workbook(excel_file, read_only=True)
    sheet_names = wb.sheetnames

    logger.info(f"[import_excel_to_db] 共 {len(sheet_names)} 个sheet")

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    total_inserted = 0
    total_skipped = 0
    errors = []

    for sheet_name in sheet_names:
        try:
            ws = wb[sheet_name]

            # 解析日期
            date_str = sheet_name[:10] if len(sheet_name) >= 10 else sheet_name

            # 查找数据起始行（通常从第4行开始）
            data_start_row = None
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                if row and len(row) >= 7 and str(row[0]).startswith('20'):
                    # 可能是数据行，但先找表头
                    continue
                if row and len(row) >= 2 and row[0] and '品名' in str(row[0]):
                    data_start_row = row_idx + 1
                    break

            if data_start_row is None:
                data_start_row = 4  # 默认从第4行开始

            # 读取数据
            sheet_inserted = 0
            for row in ws.iter_rows(min_row=data_start_row, values_only=True):
                if not row or len(row) < 7:
                    continue

                # 解析数据
                date = row[0] or date_str
                fetch_time = row[1] if row[1] else None
                material_name = row[2]
                spec = row[3]
                material_type = row[4]
                brand = row[5]
                price = row[6]
                price_change = row[7] if len(row) > 7 else None
                remark = row[8] if len(row) > 8 else None
                region = row[10] if len(row) > 10 else '山东烟台'

                # 跳过无效行
                if not material_name or not spec or price is None:
                    continue

                # 跳过品名不是钢筋的行
                valid_names = ['高线', '螺纹钢', '盘螺', '圆钢']
                if material_name not in valid_names:
                    continue

                # 确保价格是数字
                try:
                    price = int(float(price)) if price is not None else 0
                    if price <= 0:
                        continue
                except (ValueError, TypeError):
                    continue

                # 插入数据库
                try:
                    c.execute('''
                        INSERT INTO rebar_prices
                        (date, fetch_time, material_name, spec, material_type, brand, price, price_change, remark, region)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (date, fetch_time, material_name, spec, material_type, brand, price, price_change, remark, region))
                    sheet_inserted += 1
                except sqlite3.IntegrityError:
                    total_skipped += 1
                except Exception as e:
                    errors.append((sheet_name, str(e), row))

            total_inserted += sheet_inserted
            logger.info(f"[import_excel_to_db] {sheet_name}: 插入 {sheet_inserted} 条")

        except Exception as e:
            logger.error(f"[import_excel_to_db] 处理sheet失败 | sheet={sheet_name} | error={e}")
            errors.append((sheet_name, str(e), None))

    conn.commit()
    conn.close()
    wb.close()

    logger.info(f"[import_excel_to_db] 导入完成 | 插入={total_inserted} | 跳过={total_skipped}")

    if errors:
        logger.warning(f"[import_excel_to_db] 错误数: {len(errors)}")
        for sheet, err, row in errors[:5]:
            logger.warning(f"  {sheet}: {err}")

    return {
        'success': True,
        'inserted': total_inserted,
        'skipped': total_skipped,
        'errors': len(errors)
    }


def verify_import():
    """验证导入结果"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    # 总记录数
    c.execute('SELECT COUNT(*) FROM rebar_prices')
    total = c.fetchone()[0]

    # 有价格的记录数
    c.execute('SELECT COUNT(*) FROM rebar_prices WHERE price IS NOT NULL AND price > 0')
    with_price = c.fetchone()[0]

    # 日期数
    c.execute('SELECT COUNT(DISTINCT date) FROM rebar_prices')
    dates = c.fetchone()[0]

    # 日期范围
    c.execute('SELECT MIN(date), MAX(date) FROM rebar_prices')
    date_range = c.fetchone()

    # 示例数据
    c.execute('SELECT * FROM rebar_prices WHERE price > 0 LIMIT 3')
    sample = c.fetchall()

    conn.close()

    return {
        'total_records': total,
        'records_with_price': with_price,
        'unique_dates': dates,
        'date_range': date_range,
        'sample': sample
    }


def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(description='将Excel价格数据导入SQLite数据库')
    parser.add_argument('--file', '-f', help='Excel文件路径')
    parser.add_argument('--clear', '-c', action='store_true', help='清除现有数据')
    parser.add_argument('--verify', '-v', action='store_true', help='只验证不导入')

    args = parser.parse_args()

    excel_file = args.file or get_excel_file()

    if not excel_file:
        print('错误: 找不到Excel文件')
        return

    print(f'Excel文件: {excel_file}')

    if args.verify:
        # 只验证
        result = verify_import()
        print(f'\n验证结果:')
        print(f'  总记录数: {result["total_records"]}')
        print(f'  有价格记录: {result["records_with_price"]}')
        print(f'  日期数: {result["unique_dates"]}')
        print(f'  日期范围: {result["date_range"][0]} 至 {result["date_range"][1]}')
        if result['sample']:
            print(f'\n示例数据:')
            for row in result['sample']:
                print(f'  {row}')
        return

    # 导入
    result = import_excel_to_db(excel_file, clear_existing=args.clear)

    print(f'\n导入结果:')
    print(f'  插入: {result["inserted"]} 条')
    print(f'  跳过: {result["skipped"]} 条')
    print(f'  错误: {result["errors"]} 个')

    # 验证
    print(f'\n验证导入结果:')
    verify_result = verify_import()
    print(f'  总记录数: {verify_result["total_records"]}')
    print(f'  有价格记录: {verify_result["records_with_price"]}')
    print(f'  日期数: {verify_result["unique_dates"]}')


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    main()
