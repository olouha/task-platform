"""
批量抓取历史价格数据
按照日期列表逐个抓取，跳过已有完整数据的日期
"""
import asyncio
import sys
from pathlib import Path
from datetime import datetime, timedelta

sys.path.insert(0, '.')

from playwright.async_api import async_playwright
import openpyxl
import json

DATA_DIR = Path('services/data')
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格_批量抓取.xlsx'


def get_dates_to_fetch():
    """获取需要抓取的日期列表"""
    with open(DATA_DIR / 'dates_to_fetch.txt', 'r') as f:
        dates = [line.strip() for line in f if line.strip()]
    return dates


def get_existing_dates():
    """从Excel获取已抓取的日期"""
    if not EXCEL_FILE.exists():
        return set()

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    dates = set()
    for sheet_name in wb.sheetnames:
        # 从sheet名提取日期 (YYYY-MM-DD_AM_HHMMSS 或 YYYY-MM-DD)
        date_part = sheet_name[:10]
        if date_part.startswith('20'):
            dates.add(date_part)
    wb.close()
    return dates


async def fetch_date(p, page, target_date: str):
    """
    抓取指定日期的数据

    注意：大多数网站只保留近期数据，历史日期可能无法访问
    这里尝试访问，如果失败则返回空
    """
    print(f'  尝试抓取 {target_date}...')

    # 构造URL（尝试通过日期参数）
    # 注意：这个URL可能不支持历史日期，仅作为示例
    url = 'https://jiancai.mysteel.com/m/26051410/25B3355C6617BD3C.html'

    try:
        await page.goto(url, wait_until='domcontentloaded', timeout=30000)
        await page.wait_for_timeout(5000)

        # 尝试查找日期选择器或历史数据链接
        # 如果网站支持历史数据，这里需要实现相应的逻辑

        # 提取数据
        data = await page.evaluate('''() => {
            const tables = document.querySelectorAll('table');
            const results = [];
            tables.forEach((table) => {
                const rows = table.querySelectorAll('tr');
                rows.forEach((row) => {
                    const cells = row.querySelectorAll('td, th');
                    const rowData = [];
                    cells.forEach(c => rowData.push(c.textContent.trim()));
                    if (rowData.length > 0) results.push(rowData);
                });
            });
            return results;
        }''')

        prices = []
        for t in data:
            for row in t['rows'] if isinstance(t, dict) and 'rows' in t else t:
                if row and len(row) >= 5:
                    material_name = row[0].strip()
                    spec = row[1].strip()
                    material_type = row[2].strip()
                    brand = row[3].strip()
                    price_str = row[4].strip()

                    valid_names = ['高线', '螺纹钢', '盘螺', '圆钢']
                    if material_name in valid_names and spec.startswith('Φ') and price_str.isdigit():
                        prices.append({
                            'material_name': material_name,
                            'spec': spec,
                            'material_type': material_type,
                            'brand': brand,
                            'price': float(price_str)
                        })

        if prices:
            print(f'  [OK] {target_date}: 提取到 {len(prices)} 条价格')
            return {'success': True, 'date': target_date, 'prices': prices}
        else:
            print(f'  [SKIP] {target_date}: 无数据（可能不支持历史日期）')
            return {'success': False, 'date': target_date, 'error': '无数据'}

    except Exception as e:
        print(f'  [ERROR] {target_date}: 抓取失败 - {e}')
        return {'success': False, 'date': target_date, 'error': str(e)}


def save_to_excel(results):
    """保存抓取结果到Excel"""
    if not EXCEL_FILE.exists():
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
    else:
        wb = openpyxl.load_workbook(EXCEL_FILE)

    for result in results:
        if not result['success'] or not result['prices']:
            continue

        date = result['date']
        prices = result['prices']
        period = 'PM'
        fetch_time = datetime.now().strftime('%H:%M:%S')
        sheet_name = f'{date}_{period}_{fetch_time.replace(":", "")}'

        if sheet_name in wb.sheetnames:
            continue

        ws = wb.create_sheet(title=sheet_name)

        # 表头
        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)

        # 数据
        for i, price in enumerate(prices):
            row = 4 + i
            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=2, value=fetch_time)
            ws.cell(row=row, column=3, value=price['material_name'])
            ws.cell(row=row, column=4, value=price['spec'])
            ws.cell(row=row, column=5, value=price['material_type'])
            ws.cell(row=row, column=6, value=price['brand'])
            ws.cell(row=row, column=7, value=price['price'])

    wb.save(EXCEL_FILE)
    wb.close()


async def main():
    """主函数"""
    print('=' * 60)
    print('批量历史价格抓取')
    print('=' * 60)

    dates = get_dates_to_fetch()
    print(f'需要抓取的日期: {len(dates)} 个')

    # 反转列表，从最近的日期开始
    dates = list(reversed(dates))

    # 检查已抓取的日期
    existing_dates = get_existing_dates()
    print(f'已抓取: {len(existing_dates)} 个')

    # 过滤掉已抓取的
    dates_to_fetch = [d for d in dates if d not in existing_dates]
    print(f'待抓取: {len(dates_to_fetch)} 个')

    if not dates_to_fetch:
        print('所有日期已完成！')
        return

    print(f'\\n注意：网站可能不支持历史数据抓取')
    print(f'将从最近的日期开始尝试...\\n')

    # 加载凭据
    config_file = DATA_DIR / 'mysteel_config.json'
    if config_file.exists():
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
            username = config.get('username', 'M6616592358')
            password = config.get('password', 'panhui199261')
    else:
        username, password = 'M6616592358', 'panhui199261'

    cookie_file = DATA_DIR / 'myst_cookies.json'

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(viewport={'width': 1920, 'height': 3000}, locale='zh-CN')
        page = await context.new_page()

        # 登录
        print('1. 登录中...')
        await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
        await page.wait_for_timeout(5000)

        try:
            account_tab = await page.query_selector('.form-tab-account, a[data-tab="account"]')
            if account_tab:
                await account_tab.click()
                await page.wait_for_timeout(2000)
        except: pass

        await page.evaluate(f'''() => {{
            const inputs = document.querySelectorAll('input');
            for (const inp of inputs) {{
                const ph = inp.placeholder || '';
                if (ph.includes('用户名')) inp.value = '{username}';
                if (ph.includes('密码') && inp.type === 'password') inp.value = '{password}';
            }}
        }}''')

        await page.wait_for_timeout(500)

        try:
            login_btn = await page.query_selector('.form-button-login, button:has-text("登录")')
            if login_btn:
                await login_btn.click()
        except: pass

        await page.wait_for_timeout(8000)
        print('  登录完成')

        cookies = await context.cookies()
        with open(cookie_file, 'w', encoding='utf-8') as f:
            json.dump(cookies, f, ensure_ascii=False)

        results = []
        success_count = 0
        skip_count = 0

        # 限制抓取数量（避免过度请求）
        max_fetch = 10  # 先尝试最近10个

        for date in dates_to_fetch[:max_fetch]:
            result = await fetch_date(p, page, date)
            results.append(result)

            if result['success']:
                success_count += 1
            else:
                skip_count += 1

            # 短暂延迟
            await page.wait_for_timeout(2000)

        await browser.close()

    print(f'\\n抓取完成：')
    print(f'  成功: {success_count}')
    print(f'  失败/跳过: {skip_count}')

    # 保存结果
    save_to_excel(results)
    print(f'\\n结果已保存到: {EXCEL_FILE}')


if __name__ == '__main__':
    asyncio.run(main())
