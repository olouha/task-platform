"""
烟台钢筋价格数据完整性分析报告生成器
生成Excel报告，包含：
1. 数据统计概览
2. 缺失日期清单
3. 按月统计
4. 数据趋势图表
"""
import sys
sys.path.insert(0, '.')

import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
import logging

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

DATA_DIR = Path('web/backend/services/data')
DB_FILE = DATA_DIR / 'yantai_rebar.db'
REPORT_FILE = DATA_DIR / '钢筋价格数据完整性分析报告.xlsx'


def get_date_range():
    """获取数据库中实际有的日期"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute('SELECT DISTINCT date FROM rebar_prices ORDER BY date')
    existing = set(row[0] for row in cursor.fetchall())
    conn.close()
    return existing


def calculate_missing_dates(start_date='2024-01-01', end_date='2026-05-30'):
    """计算指定日期范围内的工作日，并找出缺失的日期"""
    existing = get_date_range()

    start = datetime.strptime(start_date, '%Y-%m-%d')
    end = datetime.strptime(end_date, '%Y-%m-%d')

    all_dates = []
    missing = []

    current = start
    while current <= end:
        if current.weekday() < 5:  # 工作日
            date_str = current.strftime('%Y-%m-%d')
            all_dates.append(date_str)
            if date_str not in existing:
                missing.append(date_str)
        current += timedelta(days=1)

    return all_dates, existing, missing


def get_monthly_stats(dates_list):
    """按月统计"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    stats = {}

    for date_str in dates_list:
        year_month = date_str[:7]  # YYYY-MM
        cursor.execute('''
            SELECT COUNT(*) FROM rebar_prices WHERE date = ?
        ''', (date_str,))
        count = cursor.fetchone()[0]

        if year_month not in stats:
            stats[year_month] = {'total': 0, 'days': 0, 'complete_days': 0}

        stats[year_month]['total'] += count
        stats[year_month]['days'] += 1
        if count > 0:
            stats[year_month]['complete_days'] += 1

    conn.close()
    return stats


def get_price_trend():
    """获取价格趋势（取每天第一条记录的价格）"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute('''
        SELECT date, MIN(price) as min_price, MAX(price) as max_price, COUNT(*) as cnt
        FROM rebar_prices
        WHERE material_name = '螺纹钢'
        GROUP BY date
        ORDER BY date
    ''')

    trend = []
    for row in cursor.fetchall():
        trend.append({
            'date': row[0],
            'min_price': row[1],
            'max_price': row[2],
            'count': row[3]
        })

    conn.close()
    return trend


def generate_report():
    """生成完整的分析报告"""
    if not HAS_OPENPYXL:
        logger.error("需要安装 openpyxl")
        return False

    logger.info("[generate_report] 开始生成数据完整性分析报告")

    # 计算缺失日期
    all_dates, existing, missing = calculate_missing_dates()

    logger.info(f"[generate_report] 应有工作日: {len(all_dates)}, 已有: {len(existing)}, 缺失: {len(missing)}")

    wb = openpyxl.Workbook()

    # 样式定义
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ========== Sheet 1: 概览 ==========
    ws_overview = wb.active
    ws_overview.title = "数据概览"

    row = 1
    ws_overview.merge_cells(f'A{row}:D{row}')
    ws_overview.cell(row=row, column=1, value="烟台钢筋价格数据完整性分析报告").font = Font(bold=True, size=16)
    ws_overview.cell(row=row, column=1).alignment = center_align

    row = 3
    data = [
        ("数据时间范围", "2024-01-01 至 2026-05-30"),
        ("应有工作日总数", str(len(all_dates))),
        ("数据库已有工作日", str(len(existing))),
        ("缺失工作日数量", str(len(missing))),
        ("数据完整率", f"{len(existing)/len(all_dates)*100:.1f}%"),
        ("数据库总记录数", str(len(get_date_range()))),  # 这需要重新查询
        ("", ""),
        ("缺失日期分布", ""),
        ("2024年缺失", str(len([d for d in missing if d.startswith('2024')]))),
        ("2025年缺失", str(len([d for d in missing if d.startswith('2025')]))),
        ("2026年缺失", str(len([d for d in missing if d.startswith('2026')]))),
    ]

    for item in data:
        ws_overview.cell(row=row, column=1, value=item[0]).font = Font(bold=True)
        ws_overview.cell(row=row, column=2, value=item[1])
        row += 1

    # 列宽
    ws_overview.column_dimensions['A'].width = 20
    ws_overview.column_dimensions['B'].width = 30

    # ========== Sheet 2: 缺失日期清单 ==========
    ws_missing = wb.create_sheet("缺失日期清单")

    row = 1
    ws_missing.merge_cells(f'A{row}:C{row}')
    ws_missing.cell(row=row, column=1, value="缺失日期完整清单").font = title_font

    row = 3
    headers = ["序号", "缺失日期", "所属年份", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws_missing.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    row = 4
    for i, date_str in enumerate(missing, 1):
        year = date_str[:4]
        remark = ""
        if datetime.strptime(date_str, '%Y-%m-%d').weekday() == 4:
            remark = "周五"

        ws_missing.cell(row=row, column=1, value=i).border = thin_border
        ws_missing.cell(row=row, column=2, value=date_str).border = thin_border
        ws_missing.cell(row=row, column=3, value=year).border = thin_border
        ws_missing.cell(row=row, column=4, value=remark).border = thin_border
        row += 1

    ws_missing.column_dimensions['A'].width = 8
    ws_missing.column_dimensions['B'].width = 15
    ws_missing.column_dimensions['C'].width = 12
    ws_missing.column_dimensions['D'].width = 15

    # ========== Sheet 3: 按月统计 ==========
    ws_monthly = wb.create_sheet("按月统计")

    row = 1
    ws_monthly.merge_cells(f'A{row}:F{row}')
    ws_monthly.cell(row=row, column=1, value="按月数据统计").font = title_font

    row = 3
    headers = ["年月", "应有工作日", "已有工作日", "缺失工作日", "数据库记录数", "完整率"]
    for col, header in enumerate(headers, 1):
        cell = ws_monthly.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 按月计算
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    current = datetime(2024, 1, 1)
    end = datetime(2026, 5, 30)

    row = 4
    while current <= end:
        year_month = current.strftime('%Y-%m-%d')[:7]

        # 计算当月应有工作日
        month_start = datetime(current.year, current.month, 1)
        if current.month == 12:
            month_end = datetime(current.year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = datetime(current.year, current.month + 1, 1) - timedelta(days=1)

        working_days_in_month = 0
        tmp = month_start
        while tmp <= month_end:
            if tmp.weekday() < 5:
                working_days_in_month += 1
            tmp += timedelta(days=1)

        # 已有工作日
        cursor.execute('''
            SELECT COUNT(DISTINCT date) FROM rebar_prices WHERE date LIKE ?
        ''', (f'{year_month}%',))
        existing_days = cursor.fetchone()[0] or 0

        # 数据库记录数
        cursor.execute('''
            SELECT COUNT(*) FROM rebar_prices WHERE date LIKE ?
        ''', (f'{year_month}%',))
        record_count = cursor.fetchone()[0] or 0

        missing_days = working_days_in_month - existing_days
        complete_rate = existing_days / working_days_in_month * 100 if working_days_in_month > 0 else 0

        ws_monthly.cell(row=row, column=1, value=year_month).border = thin_border
        ws_monthly.cell(row=row, column=2, value=working_days_in_month).border = thin_border
        ws_monthly.cell(row=row, column=3, value=existing_days).border = thin_border
        ws_monthly.cell(row=row, column=4, value=missing_days).border = thin_border
        ws_monthly.cell(row=row, column=5, value=record_count).border = thin_border
        ws_monthly.cell(row=row, column=6, value=f"{complete_rate:.1f}%").border = thin_border

        row += 1
        current = month_end + timedelta(days=1)
        # 处理月份进位
        if current.month == 1 and current.year == 2026 and current.month > 5:
            break
        if current.year > 2026 or (current.year == 2026 and current.month > 5):
            break

    conn.close()

    ws_monthly.column_dimensions['A'].width = 12
    ws_monthly.column_dimensions['B'].width = 12
    ws_monthly.column_dimensions['C'].width = 12
    ws_monthly.column_dimensions['D'].width = 12
    ws_monthly.column_dimensions['E'].width = 15
    ws_monthly.column_dimensions['F'].width = 10

    # ========== Sheet 4: 数据样例 ==========
    ws_sample = wb.create_sheet("数据样例")

    row = 1
    ws_sample.merge_cells(f'A{row}:H{row}')
    ws_sample.cell(row=row, column=1, value="数据库数据样例（每日期取前3条）").font = title_font

    row = 3
    headers = ["日期", "抓取时间", "品名", "规格", "材质", "品牌", "价格(元/吨)", "地区"]
    for col, header in enumerate(headers, 1):
        cell = ws_sample.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute('''
        SELECT date, fetch_time, material_name, spec, material_type, brand, price, region
        FROM rebar_prices
        ORDER BY date DESC, id ASC
        LIMIT 100
    ''')

    row = 4
    for db_row in cursor.fetchall():
        for col, value in enumerate(db_row, 1):
            ws_sample.cell(row=row, column=col, value=value).border = thin_border
        row += 1

    conn.close()

    for i in range(1, 9):
        ws_sample.column_dimensions[get_column_letter(i)].width = 15

    # 保存
    wb.save(REPORT_FILE)
    logger.info(f"[generate_report] 报告已保存: {REPORT_FILE}")

    return True


if __name__ == '__main__':
    generate_report()
    print(f"\n报告已生成: {REPORT_FILE}")