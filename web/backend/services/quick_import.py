"""
快速导入Excel数据到SQLite数据库
"""
import sqlite3
import openpyxl
from datetime import datetime
from pathlib import Path

DATA_DIR = Path('services/data')
DB_FILE = DATA_DIR / 'yantai_rebar.db'

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute('DROP TABLE IF EXISTS rebar_prices')
    c.execute('''
        CREATE TABLE rebar_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            fetch_time TEXT,
            material_name TEXT,
            spec TEXT,
            material_type TEXT,
            brand TEXT,
            price INTEGER,
            price_change TEXT,
            remark TEXT,
            steel_code TEXT,
            region TEXT DEFAULT '山东烟台'
        )
    ''')
    c.execute('CREATE INDEX idx_date ON rebar_prices(date)')
    c.execute('CREATE INDEX idx_material ON rebar_prices(material_name)')

    conn.commit()
    return conn

def import_data(conn, excel_file):
    c = conn.cursor()
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    count = 0

    data = []
    for sheet_name in wb.sheetnames:
        if '-' not in sheet_name:
            continue
        date = sheet_name.split('_')[0]
        try:
            datetime.strptime(date, '%Y-%m-%d')
        except:
            continue

        ws = wb[sheet_name]
        if ws.max_row < 5:
            continue

        for row_idx in range(2, ws.max_row + 1):
            row_data = [ws.cell(row=row_idx, column=col).value for col in range(1, min(12, ws.max_column + 1))]
            material = row_data[2] if len(row_data) > 2 else None
            if not material or '品名' in str(material):
                continue

            spec = row_data[3] if len(row_data) > 3 else ''
            mat_type = row_data[4] if len(row_data) > 4 else ''
            brand = row_data[5] if len(row_data) > 5 else ''
            price_text = row_data[6] if len(row_data) > 6 else ''

            import re
            price = 0
            if price_text:
                match = re.search(r'(\d{3,5})', str(price_text))
                if match:
                    price = int(match.group(1))

            if price > 0:
                data.append((date, material, spec, mat_type, brand, price))

            count += 1
            if count % 10000 == 0:
                print(f'  读取: {count} 行...')

    wb.close()

    # 批量插入
    print(f'  插入 {len(data)} 条数据...')
    c.executemany('INSERT INTO rebar_prices (date, material_name, spec, material_type, brand, price) VALUES (?, ?, ?, ?, ?, ?)', data)
    conn.commit()

    return len(data)

def main():
    print('=' * 60)
    print('导入数据到SQLite')
    print('=' * 60)

    conn = init_db()
    print(f'数据库: {DB_FILE}')

    # 只使用完整版数据
    excel_file = DATA_DIR / '山东烟台钢筋价格_完整版_数据+截图.xlsx'

    if excel_file.exists():
        print(f'\n导入: {excel_file.name}')
        total = import_data(conn, excel_file)
        print(f'导入完成: {total} 条')

        # 统计
        c = conn.cursor()
        c.execute('SELECT COUNT(*) FROM rebar_prices')
        print(f'总记录: {c.execute("SELECT COUNT(*) FROM rebar_prices").fetchone()[0]}')
        c.execute('SELECT COUNT(DISTINCT date) FROM rebar_prices')
        print(f'日期数: {c.execute("SELECT COUNT(DISTINCT date) FROM rebar_prices").fetchone()[0]}')
        c.execute('SELECT MIN(date), MAX(date) FROM rebar_prices')
        r = c.execute('SELECT MIN(date), MAX(date) FROM rebar_prices').fetchone()
        print(f'范围: {r[0]} 到 {r[1]}')

    conn.close()
    print(f'\n完成! 数据库: {DB_FILE}')

if __name__ == '__main__':
    main()