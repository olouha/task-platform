"""
步骤2: 从保存的链接中抓取数据
"""

import asyncio
from playwright.async_api import async_playwright
from pathlib import Path
import json
import base64
import re
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
import random

DATA_DIR = Path('services/data')
COOKIE_FILE = DATA_DIR / 'mysteel_cookies.json'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
LINKS_FILE = DATA_DIR / 'yantai_links.json'

USERNAME = 'M6616592358'
PASSWORD = 'mysteel573005'


def random_sleep(min_sec=2, max_sec=5):
    return random.uniform(min_sec, max_sec)


def get_sheet_name(date_str, period):
    fetch_time = datetime.now().strftime('%H%M%S')
    return f'{date_str}_{period}_{fetch_time}'


def get_date_range():
    end_date = datetime.now()
    start_date = end_date - timedelta(days=365)
    return start_date, end_date


def is_date_in_range(date_str, start_date, end_date):
    try:
        date = datetime.strptime(date_str, '%Y-%m-%d')
        return start_date <= date <= end_date
    except:
        return False


def parse_date_from_url(href):
    """从URL解析日期"""
    match = re.search(r'/m/(\d{8})/', href)
    if match:
        full = match.group(1)
        year = 2000 + int(full[0:2])
        month = int(full[2:4])
        day = int(full[4:6])
        hour = int(full[6:8])
        return f'{year}-{month:02d}-{day:02d}', 'AM' if hour < 12 else 'PM'
    return None, None


async def get_existing_sheets():
    existing_sheets = set()
    if EXCEL_FILE.exists():
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            existing_sheets = set(wb.sheetnames)
            wb.close()
        except Exception:
            pass
    return existing_sheets


async def login(page):
    """登录"""
    print('登录中...')
    cookies = []
    if COOKIE_FILE.exists():
        try:
            cookies = json.loads(COOKIE_FILE.read_text(encoding='utf-8'))
            if cookies:
                await page.context.add_cookies(cookies)
        except:
            pass

    await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
    await asyncio.sleep(random_sleep(2, 4))

    try:
        account_tab = await page.query_selector('.form-tab-account')
        if account_tab:
            await account_tab.click()
            await asyncio.sleep(random_sleep(1, 2))
    except:
        pass

    await page.evaluate(f'''() => {{
        const inputs = document.querySelectorAll('input');
        for (const inp of inputs) {{
            const ph = inp.placeholder || '';
            if (ph.includes('用户名')) inp.value = '{USERNAME}';
            if (ph.includes('密码') && inp.type === 'password') inp.value = '{PASSWORD}';
        }}
    }}''')

    await asyncio.sleep(random_sleep(1, 2))

    try:
        checkbox = await page.query_selector('input[type="checkbox"]')
        if checkbox and not await checkbox.is_checked():
            await checkbox.click()
    except:
        pass

    await asyncio.sleep(random_sleep(0.5, 1))

    try:
        login_btn = await page.query_selector('.form-button-login')
        if login_btn:
            await login_btn.click()
    except:
        pass

    await asyncio.sleep(random_sleep(8, 12))

    cookies = await page.context.cookies()
    with open(COOKIE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cookies, f, ensure_ascii=False)
    print('登录完成\n')


async def fetch_prices(page, url):
    """抓取价格数据"""
    await page.goto(url, wait_until='domcontentloaded', timeout=60000)
    await asyncio.sleep(random_sleep(3, 5))

    screenshot = await page.screenshot(full_page=True)
    screenshot_b64 = base64.b64encode(screenshot).decode('utf-8')

    data = await page.evaluate('''() => {
        const tables = document.querySelectorAll('table');
        const results = [];
        tables.forEach(table => {
            const rows = table.querySelectorAll('tr');
            rows.forEach(row => {
                const cells = row.querySelectorAll('td');
                if (cells.length >= 5) {
                    const material_name = cells[0]?.textContent?.trim();
                    const spec = cells[1]?.textContent?.trim();
                    const material_type = cells[2]?.textContent?.trim();
                    const brand = cells[3]?.textContent?.trim();
                    const price = cells[4]?.textContent?.trim();

                    if (['高线', '螺纹钢', '盘螺', '圆钢'].includes(material_name) &&
                        spec && spec.startsWith('Φ') && price && /^\\d+$/.test(price)) {
                        results.push({
                            material_name,
                            spec,
                            material_type,
                            brand,
                            price: parseFloat(price),
                            region: '山东烟台'
                        });
                    }
                }
            });
        });
        return results;
    }''')

    return list(data), screenshot_b64


async def save_to_excel(prices, period, date_str, screenshot_b64):
    """保存到Excel"""
    try:
        if EXCEL_FILE.exists():
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE)
            except Exception:
                wb = openpyxl.Workbook()
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']
        else:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        sheet_name = get_sheet_name(date_str, period)
        ws = wb.create_sheet(title=sheet_name)

        period_text = '下午(晚)' if period == 'PM' else '上午'
        ws.merge_cells('A1:K1')
        ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        header_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid') if period == 'PM' else \
            PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        fetch_time = datetime.now().strftime('%H:%M:%S')
        for i, price in enumerate(prices):
            row = 4 + i
            for col, val in enumerate([date_str, fetch_time, price['material_name'], price['spec'],
                                       price['material_type'], price['brand'], price['price'],
                                       '', '', '', '山东烟台'], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border

        if screenshot_b64:
            screenshot_path = DATA_DIR / f'screenshot_{date_str.replace("-", "")}_{period}.png'
            with open(screenshot_path, 'wb') as f:
                f.write(base64.b64decode(screenshot_b64))

            row = 4 + len(prices) + 2
            ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)
            img = Image(str(screenshot_path))
            img.width = 900
            img.height = 500
            img.anchor = f'A{row + 1}'
            ws.add_image(img)

        wb.save(EXCEL_FILE)
        wb.close()
        return sheet_name
    except Exception as e:
        print(f'保存失败: {e}')
        return None


async def main():
    print('=' * 60)
    print('步骤2: 从保存的链接中抓取数据')
    print('=' * 60)

    # 加载链接
    if not LINKS_FILE.exists():
        print(f'错误: 链接文件不存在 {LINKS_FILE}')
        print('请先运行 step1_get_links.py')
        return

    with open(LINKS_FILE, 'r', encoding='utf-8') as f:
        all_links = json.load(f)

    print(f'加载了 {len(all_links)} 个链接')

    start_date, end_date = get_date_range()
    print(f'日期范围: {start_date.strftime("%Y-%m-%d")} ~ {end_date.strftime("%Y-%m-%d")}')

    existing_sheets = await get_existing_sheets()
    print(f'已有sheet: {len(existing_sheets)}个')

    # 准备抓取任务
    tasks = []
    for link in all_links:
        date_str, period = parse_date_from_url(link['href'])
        if date_str and is_date_in_range(date_str, start_date, end_date):
            sheet_name = get_sheet_name(date_str, period)
            has_existing = any(s.startswith(f'{date_str}_{period}') for s in existing_sheets)
            if not has_existing:
                tasks.append((date_str, period, link['href']))

    print(f'需要抓取: {len(tasks)} 条数据\n')

    if not tasks:
        print('所有数据已抓取完毕')
        return

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            viewport={'width': 1920, 'height': 3000},
            locale='zh-CN',
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        )
        page = await context.new_page()

        await login(page)

        results = {}
        success_count = 0
        fail_count = 0

        for i, (date_str, period, url) in enumerate(tasks):
            print(f'[{i+1}/{len(tasks)}] {date_str} {period}...', end='')

            try:
                prices, screenshot_b64 = await fetch_prices(page, url)
                if prices:
                    sheet_name = await save_to_excel(prices, period, date_str, screenshot_b64)
                    if sheet_name:
                        results[date_str] = results.get(date_str, {})
                        results[date_str][period] = {'sheet': sheet_name, 'count': len(prices)}
                        print(f' {len(prices)}条 ✓')
                        success_count += 1
                    else:
                        print(' 保存失败')
                        fail_count += 1
                else:
                    print(' 无数据')
                    fail_count += 1
            except Exception as e:
                print(f' 失败: {e}')
                fail_count += 1

            # 每5个休息一下
            if (i + 1) % 5 == 0:
                await asyncio.sleep(random_sleep(5, 10))
                print(f'  进度: 成功{success_count}, 失败{fail_count}')

        await browser.close()

    print()
    print('=' * 60)
    print('抓取完成')
    print('=' * 60)
    print(f'成功: {success_count} 条')
    print(f'失败: {fail_count} 条')

    # 保存剩余未抓取的链接
    remaining = []
    for date_str, period, url in tasks[success_count + fail_count:]:
        remaining.append({'date': date_str, 'period': period, 'url': url})

    if remaining:
        with open(LINKS_FILE, 'w', encoding='utf-8') as f:
            json.dump(remaining, f, ensure_ascii=False, indent=2)
        print(f'剩余 {len(remaining)} 条链接已保存')


if __name__ == '__main__':
    asyncio.run(main())