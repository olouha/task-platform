"""
从 '完整版_数据+截图.xlsx' 导入历史价格数据
格式: 品名 | 规格 | 品牌 | 最低价格 | 最高价格 | 涨跌 | 备注
"""
import sqlite3
import openpyxl
import logging
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)

# 配置
DATA_DIR = Path(__file__).parent / 'data'
SOURCE_FILE = DATA_DIR / '山东烟台钢筋价格_完整版_数据+截图.xlsx'
DB_FILE = DATA_DIR / 'yantai_rebar.db'


def init_database():
    """初始化数据库表"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS rebar_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            fetch_time TEXT,
            material_name TEXT,
            spec TEXT,
            material_type TEXT,
            brand TEXT,
            price INTEGER,
            price_change TEXT,
            remark TEXT,
            region TEXT DEFAULT '山东烟台',
            UNIQUE(date, material_name, spec, brand, price)
        )
    ''')

    c.execute('CREATE INDEX IF NOT EXISTS idx_date ON rebar_prices(date)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_material ON rebar_prices(material_name)')

    conn.commit()
    conn.close()
    logger.info("[init_database] 数据库初始化完成")


def parse_sheet(ws, sheet_name):
    """
    解析单个sheet的数据

    格式:
    Row 1: 标题（山东烟台钢筋价格 - 日期）
    Row 2: 表头（品名|规格|品牌|最低价格|最高价格|涨跌|备注）
    Row 3+: 数据行
    """
    # 从sheet名称提取日期
    date_str = sheet_name[:10] if len(sheet_name) >= 10 else sheet_name

    prices = []
    fetch_time = '09:00'  # 默认上午

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        # 跳过空行
        if not row or all(cell is None for cell in row):
            continue

        # Row 1: 标题行
        if i == 1 and row[0] and '山东烟台钢筋价格' in str(row[0]):
            continue

        # Row 2: 表头行
        if i == 2 and row[0] and '品名' in str(row[0]):
            continue

        # 截图标记行
        if row[0] and '截图' in str(row[0]):
            break

        # 数据行
        if len(row) >= 4 and row[3]:  # 至少有品名、规格、品牌、价格
            material_name = row[0]
            spec = row[1]
            brand = row[2]
            price_low = row[3]  # 最低价格
            price_high = row[4]  # 最高价格
            price_change = row[5]
            remark = row[6] if len(row) > 6 else None

            # 验证数据
            if not material_name or not spec or not brand:
                continue

            # 转换价格
            try:
                # 优先使用最高价，如果没有则使用最低价
                price = int(float(price_high if price_high else price_low))
                if price <= 0:
                    continue
            except (ValueError, TypeError):
                continue

            # 品名映射（处理可能的编码问题）
            valid_names = ['高线', '螺纹钢', '盘螺', '圆钢', '螺纹']
            material_name_str = str(material_name).strip()
            if not any(vn in material_name_str or material_name_str in vn for vn in valid_names):
                # 如果不是标准品名，可能是编码问题，尝试处理
                if '螺纹' not in material_name_str and '高线' not in material_name_str:
                    continue

            prices.append({
                'date': date_str,
                'fetch_time': fetch_time,
                'material_name': material_name_str,
                'spec': str(spec).strip(),
                'material_type': '',  # 这种格式没有材质信息
                'brand': str(brand).strip(),
                'price': price,
                'price_change': str(price_change) if price_change else None,
                'remark': str(remark) if remark else None,
                'region': '山东烟台'
            })

    return prices


def import_from_large_file(clear_existing: bool = False):
    """
    从完整版_数据+截图.xlsx导入数据
    """
    logger.info(f"[import_from_large_file] 开始导入 | file={SOURCE_FILE}")

    if not SOURCE_FILE.exists():
        logger.error(f"[import_from_large_file] 文件不存在: {SOURCE_FILE}")
        return {'success': False, 'error': '文件不存在'}

    # 初始化数据库
    init_database()

    # 如果需要清除现有数据
    if clear_existing:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute('DELETE FROM rebar_prices')
        conn.commit()
        conn.close()
        logger.info("[import_from_large_file] 已清除现有数据")

    # 打开Excel
    wb = openpyxl.load_workbook(SOURCE_FILE, read_only=True)
    sheet_names = wb.sheetnames

    logger.info(f"[import_from_large_file] 共 {len(sheet_names)} 个sheet")

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    total_inserted = 0
    total_skipped = 0
    errors = []
    dates_processed = []

    for sheet_name in sheet_names:
        try:
            ws = wb[sheet_name]
            prices = parse_sheet(ws, sheet_name)

            if not prices:
                logger.debug(f"[import_from_large_file] {sheet_name}: 无数据")
                continue

            # 插入数据库
            sheet_inserted = 0
            for price in prices:
                try:
                    c.execute('''
                        INSERT INTO rebar_prices
                        (date, fetch_time, material_name, spec, material_type, brand, price, price_change, remark, region)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        price['date'], price['fetch_time'], price['material_name'],
                        price['spec'], price['material_type'], price['brand'],
                        price['price'], price['price_change'], price['remark'], price['region']
                    ))
                    sheet_inserted += 1
                except sqlite3.IntegrityError:
                    total_skipped += 1
                except Exception as e:
                    errors.append((sheet_name, str(e), price))

            total_inserted += sheet_inserted
            dates_processed.append(sheet_name[:10])
            logger.info(f"[import_from_large_file] {sheet_name}: 插入 {sheet_inserted} 条")

        except Exception as e:
            logger.error(f"[import_from_large_file] 处理sheet失败 | sheet={sheet_name} | error={e}")
            errors.append((sheet_name, str(e), None))

    conn.commit()
    conn.close()
    wb.close()

    logger.info(f"[import_from_large_file] 导入完成 | 插入={total_inserted} | 跳过={total_skipped}")

    if errors:
        logger.warning(f"[import_from_large_file] 错误数: {len(errors)}")

    return {
        'success': True,
        'inserted': total_inserted,
        'skipped': total_skipped,
        'errors': len(errors),
        'dates_processed': len(set(dates_processed))
    }


def verify_import():
    """验证导入结果"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute('SELECT COUNT(*) FROM rebar_prices')
    total = c.fetchone()[0]

    c.execute('SELECT COUNT(*) FROM rebar_prices WHERE price IS NOT NULL AND price > 0')
    with_price = c.fetchone()[0]

    c.execute('SELECT COUNT(DISTINCT date) FROM rebar_prices')
    dates = c.fetchone()[0]

    c.execute('SELECT MIN(date), MAX(date) FROM rebar_prices')
    date_range = c.fetchone()

    conn.close()

    return {
        'total_records': total,
        'records_with_price': with_price,
        'unique_dates': dates,
        'date_range': date_range
    }


def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(description='从完整版_数据+截图.xlsx导入历史价格数据')
    parser.add_argument('--clear', '-c', action='store_true', help='清除现有数据')
    parser.add_argument('--verify', '-v', action='store_true', help='只验证不导入')

    args = parser.parse_args()

    if args.verify:
        result = verify_import()
        print(f'\n验证结果:')
        print(f'  总记录数: {result["total_records"]}')
        print(f'  有价格记录: {result["records_with_price"]}')
        print(f'  日期数: {result["unique_dates"]}')
        print(f'  日期范围: {result["date_range"][0]} 至 {result["date_range"][1]}')
        return

    # 导入
    print('开始导入历史价格数据...')
    result = import_from_large_file(clear_existing=args.clear)

    print(f'\n导入结果:')
    print(f'  插入: {result["inserted"]} 条')
    print(f'  跳过: {result["skipped"]} 条')
    print(f'  错误: {result.get("errors", 0)} 个')
    print(f'  处理日期数: {result.get("dates_processed", 0)}')

    # 验证
    print(f'\n验证导入结果:')
    verify_result = verify_import()
    print(f'  总记录数: {verify_result["total_records"]}')
    print(f'  有价格记录: {verify_result["records_with_price"]}')
    print(f'  日期数: {verify_result["unique_dates"]}')
    print(f'  日期范围: {verify_result["date_range"][0]} 至 {verify_result["date_range"][1]}')


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    main()
