"""
每日烟台钢筋价格抓取脚本（增强反检测版）
- 随机User-Agent
- 模拟人类行为（滚动、延迟）
- 反自动化检测脚本注入
"""

import asyncio
import json
import hashlib
import random
from pathlib import Path
from datetime import datetime, date
from playwright.async_api import async_playwright
import openpyxl

# 配置
DATA_DIR = Path(__file__).parent / 'data'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
COOKIE_FILE = DATA_DIR / 'mysteel_cookies.json'
TODAY = date.today().isoformat()
TIME_NOW = datetime.now().strftime('%H%M%S')

# 时段判断
CURRENT_HOUR = datetime.now().hour
IS_AM = CURRENT_HOUR < 14
PERIOD = 'AM' if IS_AM else 'PM'
PERIOD_LABEL = '上午' if IS_AM else '下午(较晚)'

# 真实User-Agent池
USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
]

# 反检测脚本
ANTI_DETECTION_JS = '''
// 隐藏webdriver属性
Object.defineProperty(navigator, 'webdriver', {
    get: () => undefined
});

// 伪造plugins
Object.defineProperty(navigator, 'plugins', {
    get: () => [1, 2, 3, 4, 5]
});

// 伪造languages
Object.defineProperty(navigator, 'languages', {
    get: () => ['zh-CN', 'zh', 'en']
});

// 添加chrome对象
window.chrome = {
    runtime: {}
};

// 伪造permissions
Object.defineProperty(navigator, 'permissions', {
    get: () => ({
        query: () => Promise.resolve({ state: 'granted' })
    })
});
'''


class YantaiPriceScraper:
    """烟台钢筋价格抓取器（增强反检测版）"""

    def __init__(self):
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None

    async def __aenter__(self):
        self.playwright = await async_playwright().start()
        await self.start()
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.stop()
        await self.playwright.stop()

    async def start(self):
        """启动浏览器"""
        random_ua = random.choice(USER_AGENTS)

        self.browser = await self.playwright.chromium.launch(
            headless=False,  # 非headless更容易通过验证
            args=[
                '--disable-blink-features=AutomationControlled',
                '--no-sandbox',
                '--disable-setuid-sandbox',
                '--disable-dev-shm-usage',
                '--exclude-switches=enable-automation',
                '--disable-infobars',
                '--disable-extensions'
            ]
        )

        self.context = await self.browser.new_context(
            viewport={'width': 1920, 'height': 1080},
            locale='zh-CN',
            timezone_id='Asia/Shanghai',
            user_agent=random_ua
        )

        # 注入反检测脚本
        await self.context.add_init_script(ANTI_DETECTION_JS)

        self.page = await self.context.new_page()

        # 加载Cookie
        if COOKIE_FILE.exists():
            try:
                with open(COOKIE_FILE) as f:
                    cookies = json.load(f)
                await self.context.add_cookies(cookies)
                print(f'[OK] 已加载 {len(cookies)} 个Cookie')
            except Exception as e:
                print(f'[WARN] Cookie加载失败: {e}')

    async def stop(self):
        """停止浏览器"""
        if self.browser:
            await self.browser.close()

    async def random_sleep(self, min_sec=2, max_sec=5):
        """随机休眠"""
        sleep_time = random.uniform(min_sec, max_sec)
        await asyncio.sleep(sleep_time)

    async def human_like_scroll(self):
        """模拟人类滚动"""
        await self.page.evaluate('''
            const scroll = () => {
                window.scrollBy(0, 100);
                if (window.scrollY < document.body.scrollHeight) {
                    setTimeout(scroll, 100 + Math.random() * 200);
                }
            };
            scroll();
        ''')
        await self.random_sleep(1, 2)

    async def check_login_status(self) -> bool:
        """检查登录状态"""
        print('[CHECK] 检查登录状态...')

        test_url = 'https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html'
        await self.page.goto(test_url, wait_until='domcontentloaded', timeout=60000)
        await self.random_sleep(3, 5)

        current_url = self.page.url

        if 'captcha' in current_url:
            print('[WARN] 检测到验证码页面')
            return False
        if 'passport' in current_url or 'login' in current_url:
            print('[WARN] 检测到登录页面')
            return False

        body_text = await self.page.evaluate('document.body.textContent')
        if len(body_text) < 100:
            print('[WARN] 页面内容异常')
            return False

        print('[OK] 登录状态正常')
        return True

    async def login(self):
        """自动登录"""
        from config.mysteel import MYSTEEL_USERNAME, MYSTEEL_PASSWORD

        print('[LOGIN] 开始登录...')

        # 先访问首页
        await self.page.goto('https://www.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
        await self.random_sleep(2, 4)

        # 跳转到登录页
        await self.page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
        await self.random_sleep(2, 3)

        # 填写表单
        await self.page.evaluate(f'''
            const inputs = document.querySelectorAll('input');
            for (const inp of inputs) {{
                const ph = inp.placeholder || "";
                const type = inp.type || "";
                if (ph.includes("用户名") || ph.includes("账号")) {{
                    inp.value = "{MYSTEEL_USERNAME}";
                    inp.dispatchEvent(new Event('input', {{ bubbles: true }}));
                }}
                if (ph.includes("密码") && type === "password") {{
                    inp.value = "{MYSTEEL_PASSWORD}";
                    inp.dispatchEvent(new Event('input', {{ bubbles: true }}));
                }}
            }}
        ''')

        await self.random_sleep(1, 2)

        # 查找并点击登录按钮
        try:
            selectors = [
                'button:has-text("登录")',
                '.form-button-login',
                'button[type="submit"]'
            ]
            for selector in selectors:
                try:
                    btn = await self.page.query_selector(selector)
                    if btn:
                        await btn.click()
                        break
                except:
                    continue
        except Exception as e:
            print(f'[WARN] 点击登录按钮: {e}')

        await self.random_sleep(10, 15)

        # 保存Cookie
        cookies = await self.context.cookies()
        with open(COOKIE_FILE, 'w') as f:
            json.dump(cookies, f, ensure_ascii=False)
        print(f'[OK] 已保存 {len(cookies)} 个Cookie')

        return True

    async def fetch_prices(self) -> list:
        """获取今日价格数据"""
        url = 'https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html'
        print(f'[DATA] 获取价格数据: {url}')

        await self.page.goto(url, wait_until='domcontentloaded', timeout=60000)
        await self.random_sleep(3, 5)
        await self.human_like_scroll()

        data = await self.page.evaluate('''() => {
            const results = [];
            const tables = document.querySelectorAll('table');
            tables.forEach(table => {
                const rows = table.querySelectorAll('tr');
                rows.forEach(row => {
                    const cells = row.querySelectorAll('td');
                    if (cells.length >= 5) {
                        const material_name = cells[0]?.textContent?.trim();
                        const spec = cells[1]?.textContent?.trim();
                        const material_type = cells[2]?.textContent?.trim();
                        const brand = cells[3]?.textContent?.trim();
                        const price_str = cells[4]?.textContent?.trim();
                        if (['高线', '螺纹钢', '盘螺', '圆钢'].includes(material_name) &&
                            spec && spec.startsWith('Φ') &&
                            price_str && /^\\d+$/.test(price_str)) {
                            results.push({ material_name, spec, material_type, brand, price: parseInt(price_str), region: '山东烟台' });
                        }
                    }
                });
            });
            return results;
        }''')

        print(f'[OK] 获取到 {len(data)} 条价格数据')
        return data

    def calculate_data_hash(self, data: list) -> str:
        """计算数据哈希"""
        sorted_data = sorted([(d['material_name'], d['spec'], d['brand'], d['price']) for d in data])
        return hashlib.md5(json.dumps(sorted_data, ensure_ascii=False).encode()).hexdigest()

    def check_today_data_exists(self, data_hash: str) -> tuple:
        """检查今日数据是否已存在"""
        if not EXCEL_FILE.exists():
            return False, None

        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sheet_names = wb.sheetnames
            wb.close()

            am_exists = any(f'{TODAY}_AM_' in s for s in sheet_names)
            pm_exists = any(f'{TODAY}_PM_' in s for s in sheet_names)

            if PERIOD == 'AM' and am_exists:
                return True, 'AM'
            if PERIOD == 'PM' and pm_exists:
                return True, 'PM'

            hash_file = DATA_DIR / 'data_hashes.json'
            if hash_file.exists():
                with open(hash_file) as f:
                    hashes = json.load(f)
                    key = f'{TODAY}_{PERIOD}'
                    if key in hashes and hashes[key] == data_hash:
                        return True, PERIOD

            return False, None
        except Exception as e:
            return False, None

    def save_to_excel(self, data: list, data_hash: str) -> bool:
        """保存数据到Excel"""
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE) if EXCEL_FILE.exists() else openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

            sheet_name = f'{TODAY}_{PERIOD}_{TIME_NOW}'
            ws = wb.create_sheet(title=sheet_name)

            period_text = '下午(较晚)' if PERIOD == 'PM' else '上午'
            ws.merge_cells('A1:K1')
            ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {TODAY} {period_text}').font = openpyxl.styles.Font(bold=True, size=14)
            ws.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')

            from openpyxl.styles import PatternFill
            header_fill = PatternFill(start_color='FF6B6B' if PERIOD == 'PM' else '4472C4', end_color='FF6B6B' if PERIOD == 'PM' else '4472C4', fill_type='solid')
            headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True, size=12, color='FFFFFF')
                cell.fill = header_fill
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            fetch_time = datetime.now().strftime('%H:%M:%S')
            for i, item in enumerate(data):
                row = 4 + i
                for col, val in enumerate([TODAY, fetch_time, item['material_name'], item['spec'],
                    item['material_type'], item['brand'], item['price'], '', '', '', '山东烟台'], 1):
                    ws.cell(row=row, column=col, value=val)

            wb.save(EXCEL_FILE)
            wb.close()

            hash_file = DATA_DIR / 'data_hashes.json'
            hashes = {}
            if hash_file.exists():
                with open(hash_file) as f:
                    hashes = json.load(f)
            hashes[f'{TODAY}_{PERIOD}'] = data_hash
            with open(hash_file, 'w') as f:
                json.dump(hashes, f, ensure_ascii=False)

            print(f'[OK] 数据已保存: {sheet_name}')
            return True
        except Exception as e:
            print(f'[FAIL] 保存失败: {e}')
            return False


async def main():
    """主函数"""
    print('=' * 60)
    print(f'[START] 烟台钢筋价格每日抓取 - {TODAY} {PERIOD_LABEL}')
    print('=' * 60)

    async with YantaiPriceScraper() as scraper:
        is_logged_in = await scraper.check_login_status()

            if not is_logged_in:
                print('[INFO] 需要登录，尝试自动登录...')
                await scraper.login()
                await scraper.random_sleep(2, 3)
                if not await scraper.check_login_status():
                    print('[FAIL] 登录失败')
                    return

            data = await scraper.fetch_prices()
            if not data:
                print('[WARN] 未获取到价格数据')
                return

            data_hash = scraper.calculate_data_hash(data)
            exists, _ = scraper.check_today_data_exists(data_hash)

            if exists:
                print(f'[SKIP] 今日数据已存在，跳过保存')
                return

            success = await asyncio.to_thread(scraper.save_to_excel, data, data_hash)

            if success:
                print('=' * 60)
                print('[OK] 抓取完成')
                print(f'[COUNT] 数据条数: {len(data)}')
                print(f'[TIME] 时段: {PERIOD_LABEL}')
                print('=' * 60)


if __name__ == '__main__':
    asyncio.run(main())