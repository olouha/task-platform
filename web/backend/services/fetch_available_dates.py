"""
批量抓取历史价格数据 - 直接构造URL
"""

import asyncio
from playwright.async_api import async_playwright
from pathlib import Path
import json
import base64
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

DATA_DIR = Path('services/data')
DATA_DIR.mkdir(exist_ok=True)

COOKIE_FILE = DATA_DIR / 'mysteel_cookies.json'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'


def get_sheet_name(date_str, period):
    fetch_time = datetime.now().strftime('%H:%M:%S')
    return f'{date_str}_{period}_{fetch_time.replace(":", "")}'


async def save_to_excel(prices, period, date_str, screenshot_b64):
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid') if period == 'PM' else \
        PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    sheet_name = get_sheet_name(date_str, period)
    ws = wb.create_sheet(title=sheet_name)

    # 标题
    period_text = '下午(晚)' if period == 'PM' else '上午'
    ws.merge_cells('A1:K1')
    ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    # 表头
    headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 数据
    fetch_time = datetime.now().strftime('%H:%M:%S')
    for i, price in enumerate(prices):
        row = 4 + i
        for col, val in enumerate([date_str, fetch_time, price['material_name'], price['spec'],
                                   price['material_type'], price['brand'], price['price'],
                                   '', '', '', '山东烟台'], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border

    # 截图
    if screenshot_b64:
        screenshot_path = DATA_DIR / f'screenshot_{date_str.replace("-", "")}_{period}.png'
        with open(screenshot_path, 'wb') as f:
            f.write(base64.b64decode(screenshot_b64))

        row = 4 + len(prices) + 2
        ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)

        img = Image(str(screenshot_path))
        img.width = 900
        img.height = 500
        img.anchor = f'A{row + 1}'
        ws.add_image(img)

    wb.save(EXCEL_FILE)
    wb.close()
    return sheet_name


async def fetch_prices(page, url):
    await page.goto(url, wait_until='domcontentloaded', timeout=60000)
    await page.wait_for_timeout(5000)

    screenshot = await page.screenshot(full_page=True)
    screenshot_b64 = base64.b64encode(screenshot).decode('utf-8')

    data = await page.evaluate('''() => {
        const tables = document.querySelectorAll('table');
        const results = [];
        tables.forEach(table => {
            const rows = table.querySelectorAll('tr');
            rows.forEach(row => {
                const cells = row.querySelectorAll('td');
                if (cells.length >= 5) {
                    const material_name = cells[0]?.textContent?.trim();
                    const spec = cells[1]?.textContent?.trim();
                    const material_type = cells[2]?.textContent?.trim();
                    const brand = cells[3]?.textContent?.trim();
                    const price = cells[4]?.textContent?.trim();

                    if (['高线', '螺纹钢', '盘螺', '圆钢'].includes(material_name) &&
                        spec && spec.startsWith('Φ') && price && /^\\d+$/.test(price)) {
                        results.append({
                            material_name,
                            spec,
                            material_type,
                            brand,
                            price: parseFloat(price),
                            region: '山东烟台'
                        });
                    }
                }
            });
        });
        return results;
    }''')

    return list(data), screenshot_b64


async def main():
    # 从市场页面获取可用日期
    cookies = json.loads(COOKIE_FILE.read_text(encoding='utf-8')) if COOKIE_FILE.exists() else []

    available_dates = {}

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(viewport={'width': 1920, 'height': 3000}, locale='zh-CN')
        if cookies:
            await context.add_cookies(cookies)
        page = await context.new_page()

        market_url = 'https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html'
        await page.goto(market_url, wait_until='domcontentloaded', timeout=60000)
        await page.wait_for_timeout(5000)

        links = await page.evaluate('''() => {
            const links = [];
            document.querySelectorAll('a[href]').forEach(a => {
                const href = a.href;
                if (href.includes('/m/26') && href.includes('.html')) {
                    links.push(href);
                }
            });
            return links;
        }''')

        # 解析日期
        import re
        for href in links:
            match = re.search(r'/m/26(\d{8})/', href)
            if match:
                time_str = match.group(1)
                url_date = f"20{time_str[0:2]}-{time_str[2:4]}-{time_str[4:6]}"
                hour = int(time_str[6:8])
                period = 'AM' if hour == 10 else 'PM'

                if url_date not in available_dates:
                    available_dates[url_date] = {}
                if period not in available_dates[url_date]:
                    available_dates[url_date][period] = href

        await browser.close()

    # 获取已存在的日期
    existing_sheets = set()
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        existing_sheets = set(wb.sheetnames)
        wb.close()

    print(f'从市场页面获取到 {len(available_dates)} 个可用日期')

    # 过滤掉已抓取的日期
    dates_to_fetch = []
    for date_str, periods in sorted(available_dates.items()):
        for period in ['AM', 'PM']:
            sheet_name = get_sheet_name(date_str, period)
            # 检查是否已有该日期该时段的sheet
            has_existing = any(s.startswith(f'{date_str}_{period}') for s in existing_sheets)
            if not has_existing and period in periods:
                dates_to_fetch.append((date_str, period, periods[period]))

    if not dates_to_fetch:
        print('所有可用日期已抓取完毕')
        return

    print(f'需要抓取 {len(dates_to_fetch)} 条数据')
    print()

    # 重新登录
    cookies = []
    if COOKIE_FILE.exists():
        cookies = json.loads(COOKIE_FILE.read_text(encoding='utf-8'))

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(viewport={'width': 1920, 'height': 3000}, locale='zh-CN')
        if cookies:
            await context.add_cookies(cookies)
        page = await context.new_page()

        # 登录
        print('登录中...')
        await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
        await page.wait_for_timeout(5000)

        username = 'M6616592358'
        password = 'mysteel573005'

        try:
            account_tab = await page.query_selector('.form-tab-account, a[data-tab="account"]')
            if account_tab:
                await account_tab.click()
                await page.wait_for_timeout(2000)
        except:
            pass

        await page.evaluate(f'''() => {{
            const inputs = document.querySelectorAll('input');
            for (const inp of inputs) {{
                const ph = inp.placeholder || '';
                if (ph.includes('用户名')) inp.value = '{username}';
                if (ph.includes('密码') && inp.type === 'password') inp.value = '{password}';
            }}
        }}''')

        await page.wait_for_timeout(500)

        try:
            login_btn = await page.query_selector('.form-button-login, button:has-text("登录")')
            if login_btn:
                await login_btn.click()
        except:
            pass

        await page.wait_for_timeout(8000)

        # 保存Cookie
        cookies = await context.cookies()
        with open(COOKIE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cookies, f, ensure_ascii=False)
        print('登录完成')
        print()

        results = {}
        for i, (date_str, period, url) in enumerate(dates_to_fetch):
            print(f'[{i+1}/{len(dates_to_fetch)}] {date_str} {period}...', end='')

            try:
                prices, screenshot_b64 = await fetch_prices(page, url)
                if prices:
                    sheet_name = await save_to_excel(prices, period, date_str, screenshot_b64)
                    results[date_str] = results.get(date_str, {})
                    results[date_str][period] = {'sheet': sheet_name, 'count': len(prices)}
                    print(f' {len(prices)}条')
                else:
                    print(' 无数据')
            except Exception as e:
                print(f' 失败: {e}')

        await browser.close()

    print()
    print('=' * 60)
    print('抓取完成')
    print('=' * 60)
    for date_str, data in sorted(results.items()):
        am = data.get('AM', {}).get('count', 0)
        pm = data.get('PM', {}).get('count', 0)
        print(f'{date_str}: AM={am}, PM={pm}')


if __name__ == '__main__':
    asyncio.run(main())