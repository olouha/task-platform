"""
山东烟台钢筋价格抓取 - 补充5.13-5.16的下午价格
"""

import asyncio
import sys
import json
import base64
from pathlib import Path
from datetime import datetime

sys.path.insert(0, '.')

from playwright.async_api import async_playwright

DATA_DIR = Path('services/data')
DATA_DIR.mkdir(exist_ok=True)


# 已知正确的URL（从之前调试获取）
DATES_URLS = {
    '2026-05-13': {
        'AM': 'https://jiancai.mysteel.com/m/26051310/C7E274318523C3AE.html',
        'PM': 'https://jiancai.mysteel.com/m/26051316/B7EAA4BE8AB3DA35.html',
    },
    '2026-05-14': {
        'AM': 'https://jiancai.mysteel.com/m/26051410/25B3355C6617BD3C.html',
        'PM': 'https://jiancai.mysteel.com/m/26051416/C9998D9EC5FE17F0.html',
    },
    '2026-05-15': {
        'AM': 'https://jiancai.mysteel.com/m/26051510/19B77109BDE6183C.html',
        'PM': 'https://jiancai.mysteel.com/m/26051516/06AC8B0B0D2BB9BF.html',
    },
    '2026-05-16': {
        'AM': 'https://jiancai.mysteel.com/m/26051610/25B3355C6617BD3C.html',  # 需要获取PM URL
    },
}


async def fetch_url_prices(page, url, period, date_str):
    """抓取指定URL的价格"""
    print(f'  {period}...', end='')
    await page.goto(url, wait_until='domcontentloaded', timeout=60000)
    await page.wait_for_timeout(5000)

    # 截图
    screenshot = await page.screenshot(full_page=True)
    screenshot_b64 = base64.b64encode(screenshot).decode('utf-8')

    # 提取价格
    data = await page.evaluate('''() => {
        const tables = document.querySelectorAll('table');
        const results = [];
        tables.forEach(table => {
            const rows = table.querySelectorAll('tr');
            rows.forEach(row => {
                const cells = row.querySelectorAll('td');
                if (cells.length >= 5) {
                    const material_name = cells[0]?.textContent?.trim();
                    const spec = cells[1]?.textContent?.trim();
                    const material_type = cells[2]?.textContent?.trim();
                    const brand = cells[3]?.textContent?.trim();
                    const price = cells[4]?.textContent?.trim();

                    if (['高线', '螺纹钢', '盘螺', '圆钢'].includes(material_name) &&
                        spec && spec.startsWith('Φ') && price && /^\\d+$/.test(price)) {
                        results.push({
                            material_name,
                            spec,
                            material_type,
                            brand,
                            price: parseFloat(price),
                            region: '山东烟台'
                        });
                    }
                }
            });
        });
        return results;
    }''')

    prices = list(data)
    print(f' {len(prices)}条')
    return prices, screenshot_b64


async def main():
    cookie_file = DATA_DIR / 'mysteel_cookies.json'
    username = 'M6616592358'
    password = 'mysteel573005'

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(viewport={'width': 1920, 'height': 3000}, locale='zh-CN')
        page = await context.new_page()

        # 登录
        print('1. 登录...')
        await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
        await page.wait_for_timeout(5000)

        try:
            account_tab = await page.query_selector('.form-tab-account, a[data-tab="account"]')
            if account_tab:
                await account_tab.click()
                await page.wait_for_timeout(2000)
        except: pass

        await page.evaluate(f'''() => {{
            const inputs = document.querySelectorAll('input');
            for (const inp of inputs) {{
                const ph = inp.placeholder || '';
                if (ph.includes('用户名')) inp.value = '{username}';
                if (ph.includes('密码') && inp.type === 'password') inp.value = '{password}';
            }}
        }}''')

        await page.wait_for_timeout(500)

        try:
            login_btn = await page.query_selector('.form-button-login, button:has-text("登录")')
            if login_btn:
                await login_btn.click()
        except: pass

        await page.wait_for_timeout(8000)

        # 保存Cookie
        cookies = await context.cookies()
        with open(cookie_file, 'w', encoding='utf-8') as f:
            json.dump(cookies, f, ensure_ascii=False)
        print('  登录完成')

        # 抓取各日期数据
        print('2. 抓取价格...')
        all_results = {}

        for date_str, urls in DATES_URLS.items():
            print(f'{date_str}:')

            # 先获取5.16的PM URL
            if date_str == '2026-05-16' and 'PM' not in urls:
                # 从首页查找今日下午的链接
                await page.goto('https://jiancai.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
                await page.wait_for_timeout(5000)

                links = await page.evaluate('''() => {
                    const links = [];
                    document.querySelectorAll('a[href]').forEach(a => {
                        const href = a.href;
                        const text = a.textContent.trim();
                        if (href.includes('jiancai.mysteel.com') && text.includes('烟台') &&
                            href.includes('/m/26') && (text.includes('16') || text.includes('15'))) {
                            links.push({text, href});
                        }
                    });
                    return links;
                }''')

                if links:
                    # 找包含16（下午）的链接
                    for link in links:
                        if '16' in link['text']:
                            urls['PM'] = link['href']
                            break
                pm_url = urls.get('PM', '未找到')
                print(f'  PM URL: {pm_url}')

            # 抓取AM和PM
            for period in ['AM', 'PM']:
                if period in urls and urls[period]:
                    prices, screenshot_b64 = await fetch_url_prices(page, urls[period], period, date_str)

                    if prices:
                        # 保存到Excel
                        import openpyxl
                        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                        from openpyxl.drawing.image import Image
                        from openpyxl.utils import get_column_letter

                        excel_file = DATA_DIR / '山东烟台钢筋价格.xlsx'

                        header_font = Font(bold=True, size=12, color='FFFFFF')
                        if period == 'PM':
                            header_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                        else:
                            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        thin_border = Border(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )

                        wb = openpyxl.load_workbook(excel_file) if excel_file.exists() else openpyxl.Workbook()
                        if 'Sheet' in wb.sheetnames:
                            del wb['Sheet']

                        fetch_time = datetime.now().strftime('%H:%M:%S')
                        sheet_name = f'{date_str}_{period}_{fetch_time.replace(":", "")}'

                        # 删除同日期同period的旧sheet
                        to_delete = [s for s in wb.sheetnames if s.startswith(f'{date_str}_{period}_')]
                        for s in to_delete:
                            del wb[s]

                        ws = wb.create_sheet(title=sheet_name)

                        # 标题
                        period_text = '下午(晚)' if period == 'PM' else '上午'
                        ws.merge_cells('A1:K1')
                        ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
                        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

                        # 表头
                        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=3, column=col, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = thin_border

                        # 数据
                        for i, price in enumerate(prices):
                            row = 4 + i
                            for col, val in enumerate([date_str, fetch_time, price['material_name'], price['spec'],
                                                       price['material_type'], price['brand'], price['price'], '', '', '', '山东烟台'], 1):
                                cell = ws.cell(row=row, column=col, value=val)
                                cell.border = thin_border

                        # 截图
                        if screenshot_b64:
                            screenshot_path = DATA_DIR / f'screenshot_{date_str}_{period}.png'
                            with open(screenshot_path, 'wb') as f:
                                f.write(base64.b64decode(screenshot_b64))

                            row = 4 + len(prices) + 2
                            ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)

                            img = Image(str(screenshot_path))
                            img.width = 900
                            img.height = 500
                            img.anchor = f'A{row + 1}'
                            ws.add_image(img)

                        wb.save(excel_file)
                        wb.close()
                        print(f'    已保存: Sheet={sheet_name}')

                        all_results[date_str] = all_results.get(date_str, {})
                        all_results[date_str][period] = len(prices)

        await browser.close()

    print()
    print('=' * 60)
    print('抓取汇总')
    print('=' * 60)
    for date_str, counts in sorted(all_results.items()):
        am = counts.get('AM', 0)
        pm = counts.get('PM', 0)
        print(f'{date_str}: 上午{am}条, 下午{pm}条')


if __name__ == '__main__':
    asyncio.run(main())