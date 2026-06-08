"""
烟台钢筋价格历史数据批量抓取
URL格式: https://jiancai.mysteel.com/m/YYMMDDHH/XXXXXXXX.html
每天AM和PM各抓取一次
"""
import asyncio
import sys
import json
import base64
from pathlib import Path
from datetime import datetime, timedelta

sys.path.insert(0, '.')

from playwright.async_api import async_playwright
import openpyxl
import sqlite3

DATA_DIR = Path('services/data')
DB_FILE = DATA_DIR / 'yantai_rebar.db'
EXCEL_FILE = DATA_DIR / '烟台钢筋价格_完整历史.xlsx'


def init_database():
    """初始化数据库"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS rebar_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            fetch_time TEXT,
            material_name TEXT,
            spec TEXT,
            material_type TEXT,
            brand TEXT,
            price INTEGER,
            region TEXT DEFAULT '山东烟台',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(date, fetch_time, material_name, spec, brand, price)
        )
    ''')

    c.execute('CREATE INDEX IF NOT EXISTS idx_date ON rebar_prices(date)')
    conn.commit()
    conn.close()


def get_existing_data():
    """获取已抓取的数据键"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT date, fetch_time, material_name, spec, brand FROM rebar_prices')
    existing = set(f"{r[0]}_{r[1]}_{r[2]}_{r[3]}_{r[4]}" for r in c.fetchall())
    conn.close()
    return existing


def save_to_excel(date: str, period: str, prices: list):
    """保存数据到Excel"""
    if not prices:
        return False

    if not EXCEL_FILE.exists():
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        wb.save(EXCEL_FILE)

    wb = openpyxl.load_workbook(EXCEL_FILE)

    time_str = datetime.now().strftime('%H%M%S')
    sheet_name = f'{date}_{period}_{time_str}'

    if sheet_name in wb.sheetnames:
        wb.close()
        return False

    ws = wb.create_sheet(title=sheet_name)

    # 标题
    ws.merge_cells('A1:K1')
    period_text = '上午' if period == 'AM' else '下午'
    ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date} {period_text}')
    ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=14)

    # 表头
    headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
        ws.cell(row=3, column=col).font = openpyxl.styles.Font(bold=True)

    # 数据
    fetch_time = datetime.now().strftime('%H:%M:%S')
    for i, price in enumerate(prices):
        row = 4 + i
        ws.cell(row=row, column=1, value=date)
        ws.cell(row=row, column=2, value=fetch_time)
        ws.cell(row=row, column=3, value=price.get('material_name', ''))
        ws.cell(row=row, column=4, value=price.get('spec', ''))
        ws.cell(row=row, column=5, value=price.get('material_type', ''))
        ws.cell(row=row, column=6, value=price.get('brand', ''))
        ws.cell(row=row, column=7, value=price.get('price', 0))
        ws.cell(row=row, column=11, value='山东烟台')

    wb.save(EXCEL_FILE)
    wb.close()
    return True


async def fetch_url_data(page, url: str, date: str, period: str, existing: set):
    """抓取单个URL的数据"""
    try:
        print(f'    访问: {url}')

        # 保存当前cookie
        cookies = await page.context.cookies()
        with open(DATA_DIR / 'myst_cookies.json', 'w', encoding='utf-8') as f:
            json.dump(cookies, f, ensure_ascii=False)

        await page.goto(url, wait_until='domcontentloaded', timeout=30000)
        await page.wait_for_timeout(5000)

        # 截图
        date_clean = date.replace('-', '')
        screenshot_path = DATA_DIR / f'screenshot_{date_clean}_{period}.png'
        await page.screenshot(path=str(screenshot_path), full_page=True)

        # 提取数据
        data = await page.evaluate('''() => {
            const tables = document.querySelectorAll('table');
            const results = [];
            tables.forEach((table) => {
                const rows = table.querySelectorAll('tr');
                rows.forEach((row) => {
                    const cells = row.querySelectorAll('td, th');
                    const rowData = [];
                    cells.forEach(c => rowData.push(c.textContent.trim()));
                    if (rowData.length > 0) results.push(rowData);
                });
            });
            return results;
        }''')

        prices = []
        for row in data:
            if row and len(row) >= 5:
                material_name = str(row[0]).strip()
                spec = str(row[1]).strip()
                material_type = str(row[2]).strip()
                brand = str(row[3]).strip()
                price_str = str(row[4]).strip()

                valid_names = ['高线', '螺纹钢', '盘螺', '圆钢']
                if material_name in valid_names and spec.startswith('Φ') and price_str.isdigit():
                    try:
                        price = int(price_str)
                        if price > 0:
                            prices.append({
                                'material_name': material_name,
                                'spec': spec,
                                'material_type': material_type,
                                'brand': brand,
                                'price': price
                            })
                    except:
                        pass

        if prices:
            # 保存到数据库
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()

            inserted = 0
            fetch_time_str = period

            for p_data in prices:
                key = f"{date}_{fetch_time_str}_{p_data['material_name']}_{p_data['spec']}_{p_data['brand']}"
                if key not in existing:
                    try:
                        c.execute('''
                            INSERT INTO rebar_prices
                            (date, fetch_time, material_name, spec, material_type, brand, price, region)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            date, fetch_time_str, p_data['material_name'],
                            p_data['spec'], p_data['material_type'], p_data['brand'],
                            p_data['price'], '山东烟台'
                        ))
                        inserted += 1
                        existing.add(key)
                    except:
                        pass

            conn.commit()
            conn.close()

            # 保存到Excel
            save_to_excel(date, period, prices)

            print(f'    [OK] 提取 {len(prices)} 条，新增 {inserted} 条')
            return inserted
        else:
            print(f'    [FAIL] 未提取到钢筋数据')
            return 0

    except Exception as e:
        print(f'    [ERROR] {e}')
        return 0


async def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(description='烟台钢筋价格历史批量抓取')
    parser.add_argument('--start', '-s', default='2024-01-01', help='开始日期 (YYYY-MM-DD)')
    parser.add_argument('--end', '-e', default=None, help='结束日期 (YYYY-MM-DD)')
    parser.add_argument('--interval', '-i', type=int, default=30, help='抓取间隔秒数')

    args = parser.parse_args()

    end_date = args.end or datetime.now().strftime('%Y-%m-%d')

    print('=' * 60)
    print('烟台钢筋价格历史批量抓取')
    print(f'日期范围: {args.start} 至 {end_date}')
    print(f'抓取间隔: {args.interval} 秒')
    print('=' * 60)

    init_database()
    existing = get_existing_data()
    print(f'数据库已有 {len(existing)} 条记录')

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(viewport={'width': 1920, 'height': 3000}, locale='zh-CN')
        page = await context.new_page()

        # 登录
        print('\n1. 登录中...')
        await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
        await page.wait_for_timeout(5000)

        try:
            await page.click('.form-tab-account', timeout=5000)
        except: pass

        await page.evaluate('''() => {
            const inputs = document.querySelectorAll('input');
            for (const inp of inputs) {
                const ph = inp.placeholder || '';
                if (ph.includes('用户名')) inp.value = 'M6616592358';
                if (ph.includes('密码') && inp.type === 'password') inp.value = 'panhui199261';
            }
        }''')
        await page.wait_for_timeout(500)

        try:
            await page.click('button:has-text("登录")', timeout=5000)
        except: pass

        await page.wait_for_timeout(8000)
        print('  登录完成')

        # 生成日期列表
        current = datetime.strptime(args.start, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        dates = []

        while current <= end:
            if current.weekday() < 5:  # 工作日
                dates.append(current.strftime('%Y-%m-%d'))
            current += timedelta(days=1)

        print(f'\n2. 需要抓取 {len(dates)} 个工作日')
        print(f'   每天抓取上午(AM)和下午(PM)两次\n')

        total_inserted = 0
        success_count = 0

        for i, date in enumerate(dates):
            print(f'\n[{i+1}/{len(dates)}] {date}')

            # AM
            url_am = f'https://jiancai.mysteel.com/m/{date[2:4]}{date[5:7]}{date[8:10]}10/placeholder.html'
            print(f'  AM...')
            inserted_am = await fetch_url_data(page, url_am, date, 'AM', existing)

            await page.wait_for_timeout(args.interval * 1000)

            # PM
            url_pm = f'https://jiancai.mysteel.com/m/{date[2:4]}{date[5:7]}{date[8:10]}16/placeholder.html'
            print(f'  PM...')
            inserted_pm = await fetch_url_data(page, url_pm, date, 'PM', existing)

            if inserted_am > 0 or inserted_pm > 0:
                success_count += 1
                total_inserted += inserted_am + inserted_pm

            if i < len(dates) - 1:
                print(f'  等待 {args.interval} 秒...')
                await page.wait_for_timeout(args.interval * 1000)

        await browser.close()

        print('\n' + '=' * 60)
        print('抓取完成')
        print(f'  成功日期: {success_count}/{len(dates)}')
        print(f'  新增记录: {total_inserted} 条')
        print('=' * 60)


if __name__ == '__main__':
    asyncio.run(main())