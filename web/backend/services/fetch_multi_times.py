"""
烟台钢筋价格多时段抓取脚本
每隔指定秒数抓取一次，保存到Excel和数据库
用于获取每天多个时段的价格数据（上午场、下午场等）
"""
import asyncio
import sys
import json
import base64
import time
from pathlib import Path
from datetime import datetime

sys.path.insert(0, '.')

from playwright.async_api import async_playwright
import openpyxl
import sqlite3

DATA_DIR = Path('services/data')
DB_FILE = DATA_DIR / 'yantai_rebar.db'
EXCEL_FILE = DATA_DIR / '烟台钢筋价格_多时段数据.xlsx'
COOKIE_FILE = DATA_DIR / 'myst_cookies.json'


def init_database():
    """初始化数据库"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS rebar_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            fetch_time TEXT,
            material_name TEXT,
            spec TEXT,
            material_type TEXT,
            brand TEXT,
            price INTEGER,
            region TEXT DEFAULT '山东烟台',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(date, fetch_time, material_name, spec, brand, price)
        )
    ''')

    c.execute('CREATE INDEX IF NOT EXISTS idx_date ON rebar_prices(date)')
    conn.commit()
    conn.close()


def get_existing_data():
    """获取已抓取的数据键"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT date, fetch_time, material_name, spec, brand FROM rebar_prices')
    existing = set(f"{r[0]}_{r[1]}_{r[2]}_{r[3]}_{r[4]}" for r in c.fetchall())
    conn.close()
    return existing


def save_to_excel(date: str, period: str, prices: list, screenshot_b64: str = None):
    """保存数据到Excel（含截图）"""
    if not prices:
        return False

    if not EXCEL_FILE.exists():
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        wb.save(EXCEL_FILE)

    wb = openpyxl.load_workbook(EXCEL_FILE)

    time_str = datetime.now().strftime('%H%M%S')
    sheet_name = f'{date}_{period}_{time_str}'

    if sheet_name in wb.sheetnames:
        wb.close()
        return False

    ws = wb.create_sheet(title=sheet_name)

    # 标题
    ws.merge_cells('A1:K1')
    period_text = '上午' if period == 'AM' else '下午'
    ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date} {period_text}')
    ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=14)

    # 表头
    headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
        ws.cell(row=3, column=col).font = openpyxl.styles.Font(bold=True)

    # 数据
    fetch_time = datetime.now().strftime('%H:%M:%S')
    for i, price in enumerate(prices):
        row = 4 + i
        ws.cell(row=row, column=1, value=date)
        ws.cell(row=row, column=2, value=fetch_time)
        ws.cell(row=row, column=3, value=price.get('material_name', ''))
        ws.cell(row=row, column=4, value=price.get('spec', ''))
        ws.cell(row=row, column=5, value=price.get('material_type', ''))
        ws.cell(row=row, column=6, value=price.get('brand', ''))
        ws.cell(row=row, column=7, value=price.get('price', 0))
        ws.cell(row=row, column=11, value='山东烟台')

    # 嵌入截图
    if screenshot_b64:
        try:
            from openpyxl.drawing.image import Image as XLImage
            date_for_file = date.replace("-", "")
            screenshot_path = DATA_DIR / f'screenshot_{date_for_file}_{period}_{time_str}.png'
            with open(screenshot_path, 'wb') as f:
                f.write(base64.b64decode(screenshot_b64))

            img = XLImage(str(screenshot_path))
            img.width = 800
            img.height = 450
            img.anchor = f'A{4 + len(prices) + 2}'
            ws.add_image(img)
        except Exception as e:
            print(f'    [WARN] 截图保存失败: {e}')

    wb.save(EXCEL_FILE)
    wb.close()
    return True


async def fetch_current_prices(page, existing: set):
    """抓取当前页面数据"""
    try:
        # 获取当前时间
        now = datetime.now()
        date_str = now.strftime('%Y-%m-%d')
        period = 'AM' if now.hour < 12 else 'PM'

        # 截图
        screenshot = await page.screenshot(full_page=True)
        screenshot_b64 = base64.b64encode(screenshot).decode('utf-8')

        # 提取数据
        data = await page.evaluate('''() => {
            const tables = document.querySelectorAll('table');
            const results = [];
            tables.forEach((table) => {
                const rows = table.querySelectorAll('tr');
                rows.forEach((row) => {
                    const cells = row.querySelectorAll('td, th');
                    const rowData = [];
                    cells.forEach(c => rowData.push(c.textContent.trim()));
                    if (rowData.length > 0) results.push(rowData);
                });
            });
            return results;
        }''')

        prices = []
        for row in data:
            if row and len(row) >= 5:
                material_name = str(row[0]).strip()
                spec = str(row[1]).strip()
                material_type = str(row[2]).strip()
                brand = str(row[3]).strip()
                price_str = str(row[4]).strip()

                valid_names = ['高线', '螺纹钢', '盘螺', '圆钢']
                if material_name in valid_names and spec.startswith('Φ') and price_str.isdigit():
                    try:
                        price = int(price_str)
                        if price > 0:
                            prices.append({
                                'material_name': material_name,
                                'spec': spec,
                                'material_type': material_type,
                                'brand': brand,
                                'price': price
                            })
                    except:
                        pass

        if prices:
            # 保存到数据库
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()

            inserted = 0
            fetch_time_str = now.strftime('%H:%M:%S')

            for p_data in prices:
                key = f"{date_str}_{fetch_time_str}_{p_data['material_name']}_{p_data['spec']}_{p_data['brand']}"
                if key not in existing:
                    try:
                        c.execute('''
                            INSERT INTO rebar_prices
                            (date, fetch_time, material_name, spec, material_type, brand, price, region)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            date_str, fetch_time_str, p_data['material_name'],
                            p_data['spec'], p_data['material_type'], p_data['brand'],
                            p_data['price'], '山东烟台'
                        ))
                        inserted += 1
                        existing.add(key)
                    except:
                        pass

            conn.commit()
            conn.close()

            # 保存到Excel
            save_to_excel(date_str, period, prices, screenshot_b64)

            return inserted, len(prices)

        return 0, 0

    except Exception as e:
        print(f'    [ERROR] {e}')
        return 0, 0


async def main():
    """主函数 - 持续抓取模式"""
    import argparse

    parser = argparse.ArgumentParser(description='烟台钢筋价格多时段抓取')
    parser.add_argument('--interval', '-i', type=int, default=60, help='抓取间隔秒数')
    parser.add_argument('--count', '-c', type=int, default=10, help='抓取次数')

    args = parser.parse_args()

    print('=' * 60)
    print('烟台钢筋价格多时段抓取')
    print(f'抓取间隔: {args.interval} 秒')
    print(f'抓取次数: {args.count}')
    print('=' * 60)

    # 加载凭据
    with open(DATA_DIR / 'mysteel_config.json', 'r', encoding='utf-8') as f:
        config = json.load(f)
        username = config.get('username')
        password = config.get('password')

    init_database()
    existing = get_existing_data()
    print(f'数据库已有 {len(existing)} 条记录')

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(viewport={'width': 1920, 'height': 3000}, locale='zh-CN')
        page = await context.new_page()

        # 登录
        print('\n1. 登录中...')
        await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
        await page.wait_for_timeout(5000)

        try:
            account_tab = await page.query_selector('.form-tab-account, a[data-tab="account"]')
            if account_tab:
                await account_tab.click()
                await page.wait_for_timeout(2000)
        except: pass

        await page.evaluate(f'''() => {{
            const inputs = document.querySelectorAll('input');
            for (const inp of inputs) {{
                const ph = inp.placeholder || '';
                if (ph.includes('用户名')) inp.value = '{username}';
                if (ph.includes('密码') && inp.type === 'password') inp.value = '{password}';
            }}
        }}''')
        await page.wait_for_timeout(500)

        try:
            login_btn = await page.query_selector('.form-button-login, button:has-text("登录")')
            if login_btn:
                await login_btn.click()
        except: pass

        await page.wait_for_timeout(8000)
        print('  登录完成')

        # 访问价格页
        print('\n2. 访问价格页...')
        url = 'https://jiancai.mysteel.com/m/26051410/25B3355C6617BD3C.html'
        await page.goto(url, wait_until='domcontentloaded', timeout=60000)
        await page.wait_for_timeout(10000)
        print(f'  页面已加载: {page.url}')

        # 开始抓取
        print(f'\n3. 开始抓取 ({args.count} 次，间隔 {args.interval} 秒)...')

        total_inserted = 0
        total_extracted = 0

        for i in range(args.count):
            now = datetime.now()
            print(f'\n[{i+1}/{args.count}] {now.strftime(\"%Y-%m-%d %H:%M:%S\")}')

            inserted, extracted = await fetch_current_prices(page, existing)

            if extracted > 0:
                print(f'    提取 {extracted} 条，新增 {inserted} 条')
                total_inserted += inserted
                total_extracted += extracted
            else:
                print(f'    [WARN] 未提取到数据')

            if i < args.count - 1:
                print(f'    等待 {args.interval} 秒...')
                await page.wait_for_timeout(args.interval * 1000)

        await browser.close()

        print('\n' + '=' * 60)
        print('抓取完成')
        print(f'  总提取: {total_extracted} 条')
        print(f'  新增记录: {total_inserted} 条')
        print(f'  Excel文件: {EXCEL_FILE}')
        print('=' * 60)


if __name__ == '__main__':
    asyncio.run(main())