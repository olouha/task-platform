"""
山东烟台钢筋价格抓取服务
从我的钢铁网抓取实时价格数据
"""
import asyncio
import sys
import json
import base64
import logging
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple

sys.path.insert(0, '.')

from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

logger = logging.getLogger(__name__)

DATA_DIR = Path('services/data')
DATA_DIR.mkdir(exist_ok=True)


@dataclass
class MaterialPrice:
    """钢材价格数据"""
    material_name: str = ''
    spec: str = ''
    material_type: str = ''
    brand: str = ''
    price: float = 0.0
    unit: str = '元/吨'
    region: str = '山东烟台'


def save_to_excel_with_screenshot(prices: List, screenshot_b64: str, excel_file: Optional[Path] = None) -> None:
    """
    保存到Excel（包含截图）
    - 累积sheet页，不删除旧数据
    - Sheet命名: YYYY-MM-DD_{AM/PM}_HHMMSS
    - AM: 0-12点, PM: 12-24点
    """
    logger.info(f"[save_to_excel] 开始保存Excel | 数据条数={len(prices)}")

    if excel_file is None:
        excel_file = DATA_DIR / '山东烟台钢筋价格.xlsx'

    header_font = Font(bold=True, size=12, color='FFFFFF')
    # PM使用不同颜色
    now_hour = datetime.now().hour
    if now_hour >= 12:
        header_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        period = 'PM'
        period_text = '下午(晚)'
    else:
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        period = 'AM'
        period_text = '上午'

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 累积sheet页，不删除旧数据
    if Path(excel_file).exists():
        wb = openpyxl.load_workbook(excel_file)
    else:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    today_str = datetime.now().strftime('%Y-%m-%d')
    fetch_time_str = datetime.now().strftime('%H:%M:%S')
    sheet_name = f'{today_str}_{period}_{fetch_time_str.replace(":", "")}'

    # 不删除旧sheet，创建新的
    ws = wb.create_sheet(title=sheet_name)
    logger.info(f"[save_to_excel] 创建Sheet | name={sheet_name}")

    # 标题
    ws.merge_cells('A1:K1')
    ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {today_str} {period_text}').font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    # 表头
    headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 数据
    for i, price in enumerate(prices):
        row = 4 + i
        # 支持字典和MaterialPrice两种格式
        if isinstance(price, dict):
            material_name = price.get('material_name', '')
            spec = price.get('spec', '')
            material_type = price.get('material_type', '')
            brand = price.get('brand', '')
            price_val = price.get('price', 0)
            region = price.get('region', '山东烟台')
        else:
            material_name = price.material_name
            spec = price.spec
            material_type = price.material_type
            brand = price.brand
            price_val = price.price
            region = price.region

        for col, val in enumerate([today_str, fetch_time_str, material_name, spec,
                                   material_type, brand, price_val, '', '', '', region], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border

    # 嵌入截图
    if screenshot_b64:
        screenshot_path = DATA_DIR / f'screenshot_{today_str.replace("-", "")}_{period}.png'
        try:
            with open(screenshot_path, 'wb') as f:
                f.write(base64.b64decode(screenshot_b64))
            logger.info(f"[save_to_excel] 截图已保存 | path={screenshot_path}")

            row = 4 + len(prices) + 2
            ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)

            img = Image(str(screenshot_path))
            img.width = 900
            img.height = 500
            img.anchor = f'A{row + 1}'
            ws.add_image(img)
        except Exception as e:
            logger.error(f"[save_to_excel] 保存截图失败 | error={e}", exc_info=True)

    wb.save(excel_file)
    wb.close()
    logger.info(f"[save_to_excel] Excel已保存 | file={excel_file} | sheet={sheet_name} | 数据={len(prices)}条")


class YantaiScraper:
    """烟台钢筋价格抓取器"""

    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.cookie_file = DATA_DIR / 'mysteel_cookies.json'
        self.excel_file = DATA_DIR / '山东烟台钢筋价格.xlsx'
        self.username, self.password = self._load_credentials()
        self.logger.info(f"[YantaiScraper] 初始化完成 | username={self.username[:3]}***")

    def _load_credentials(self) -> Tuple[str, str]:
        """从配置文件加载凭据"""
        config_file = DATA_DIR / 'mysteel_config.json'
        if config_file.exists():
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    return (
                        config.get('username', 'M6616592358'),
                        config.get('password', 'mysteel573005')
                    )
            except Exception as e:
                self.logger.warning(f"[YantaiScraper] 加载配置失败 | error={e}")
        return 'M6616592358', 'mysteel573005'

    def save_to_excel_with_screenshot(self, prices: List[Dict], screenshot_b64: str) -> None:
        """保存到Excel（含截图）"""
        save_to_excel_with_screenshot(prices, screenshot_b64, self.excel_file)

    async def _login(self, page) -> bool:
        """执行登录操作"""
        self.logger.info("[YantaiScraper._login] 开始登录")

        try:
            await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
            await page.wait_for_timeout(5000)

            # 切换到账号登录
            try:
                account_tab = await page.query_selector('.form-tab-account, a[data-tab="account"]')
                if account_tab:
                    await account_tab.click()
                    await page.wait_for_timeout(2000)
                    self.logger.info("[YantaiScraper._login] 已切换到账号登录")
            except Exception as e:
                self.logger.warning(f"[YantaiScraper._login] 切换账号标签失败 | {e}")

            # 填写表单
            await page.evaluate(f'''() => {{
                const inputs = document.querySelectorAll('input');
                for (const inp of inputs) {{
                    const ph = inp.placeholder || '';
                    if (ph.includes('用户名')) inp.value = '{self.username}';
                    if (ph.includes('密码') && inp.type === 'password') inp.value = '{self.password}';
                }}
            }}''')
            self.logger.info("[YantaiScraper._login] 表单已填写")

            await page.wait_for_timeout(500)

            # 登录
            try:
                login_btn = await page.query_selector('.form-button-login, button:has-text("登录")')
                if login_btn:
                    await login_btn.click()
                    self.logger.info("[YantaiScraper._login] 点击登录按钮")
            except Exception as e:
                self.logger.warning(f"[YantaiScraper._login] 点击登录按钮失败 | {e}")

            await page.wait_for_timeout(8000)

            # 保存Cookie
            cookies = await page.context.cookies()
            with open(self.cookie_file, 'w', encoding='utf-8') as f:
                json.dump(cookies, f, ensure_ascii=False)
            self.logger.info(f"[YantaiScraper._login] Cookie已保存 | count={len(cookies)}")

            return True

        except Exception as e:
            self.logger.error(f"[YantaiScraper._login] 登录失败 | error={e}", exc_info=True)
            return False

    async def _extract_prices(self, page) -> List[Dict]:
        """提取价格数据（支持分页）"""
        self.logger.info("[YantaiScraper._extract_prices] 开始提取价格")
        prices = []
        page_num = 1
        max_pages = 10  # 最多抓取10页

        while page_num <= max_pages:
            # 提取当前页数据
            data = await page.evaluate('''() => {
                const tables = document.querySelectorAll('table');
                const results = [];
                tables.forEach((table, idx) => {
                    const tableData = [];
                    const rows = table.querySelectorAll('tr');
                    rows.forEach((row) => {
                        const cells = row.querySelectorAll('td, th');
                        const rowData = [];
                        cells.forEach(c => rowData.push(c.textContent.trim()));
                        if (rowData.length > 0) tableData.push(rowData);
                    });
                    if (tableData.length > 0) results.push({idx, rows: tableData});
                });
                return results;
            }''')

            page_prices = []
            for t in data:
                for row in t['rows']:
                    if row and len(row) >= 5:
                        material_name = row[0].strip()
                        spec = row[1].strip()
                        material_type = row[2].strip()
                        brand = row[3].strip()
                        price_str = row[4].strip()

                        valid_names = ['高线', '螺纹钢', '盘螺', '圆钢']
                        if material_name in valid_names and spec.startswith('Φ') and price_str.isdigit():
                            try:
                                page_prices.append({
                                    'material_name': material_name,
                                    'spec': spec,
                                    'material_type': material_type,
                                    'brand': brand,
                                    'price': float(price_str)
                                })
                            except Exception as e:
                                self.logger.warning(f"[YantaiScraper._extract_prices] 解析价格失败 | row={row} | error={e}")

            self.logger.info(f"[YantaiScraper._extract_prices] 第{page_num}页: 提取到 {len(page_prices)} 条")
            prices.extend(page_prices)

            # 检查是否有"下一页"按钮
            has_next = await page.evaluate('''() => {
                const buttons = Array.from(document.querySelectorAll('a'))
                    .filter(a => /下一页|>/i.test(a.textContent) && a.href && a.href !== window.location.href);
                return buttons.length > 0;
            }''')

            if not has_next:
                self.logger.info("[YantaiScraper._extract_prices] 没有更多页面，停止抓取")
                break

            # 点击下一页
            try:
                next_btn = await page.query_selector('a:has-text("下一页"), a:has-text(">")')
                if next_btn:
                    await next_btn.click()
                    await page.wait_for_timeout(5000)
                    page_num += 1
                else:
                    self.logger.info("[YantaiScraper._extract_prices] 未找到下一页按钮")
                    break
            except Exception as e:
                self.logger.error(f"[YantaiScraper._extract_prices] 点击下一页失败 | error={e}")
                break

        self.logger.info(f"[YantaiScraper._extract_prices] 总共提取到 {len(prices)} 条")

        # 去重
        seen = set()
        unique_prices = []
        for p in prices:
            key = (p['material_name'], p['spec'], p['brand'], p['price'])
            if key not in seen:
                seen.add(key)
                unique_prices.append(p)
        self.logger.info(f"[YantaiScraper._extract_prices] 去重后: {len(unique_prices)} 条")
        return unique_prices

    async def fetch(self) -> Dict:
        """执行抓取流程"""
        self.logger.info("[YantaiScraper.fetch] 开始抓取流程")

        # 导入WebSocket管理器
        try:
            from services.websocket_manager import ws_manager
        except ImportError as e:
            self.logger.error(f"[YantaiScraper.fetch] 导入ws_manager失败 | error={e}")
            ws_manager = None

        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            context = await browser.new_context(viewport={'width': 1920, 'height': 3000}, locale='zh-CN')
            page = await context.new_page()

            try:
                # 1. 登录
                self.logger.info("[YantaiScraper.fetch] 步骤1: 登录")
                if ws_manager:
                    await ws_manager.notify_fetch_started()

                login_success = await self._login(page)
                if not login_success:
                    self.logger.error("[YantaiScraper.fetch] 登录失败")
                    if ws_manager:
                        await ws_manager.notify_fetch_failed('登录失败')
                    return {
                        'success': False,
                        'error': '登录失败',
                        'prices': []
                    }

                # 2. 访问价格页
                self.logger.info("[YantaiScraper.fetch] 步骤2: 访问价格页")
                url = 'https://jiancai.mysteel.com/m/26051410/25B3355C6617BD3C.html'
                await page.goto(url, wait_until='domcontentloaded', timeout=60000)
                await page.wait_for_timeout(10000)

                self.logger.info(f"[YantaiScraper.fetch] 页面URL: {page.url}")

                # 截图
                screenshot = await page.screenshot(full_page=True)
                screenshot_b64 = base64.b64encode(screenshot).decode('utf-8')
                self.logger.info("[YantaiScraper.fetch] 截图已完成")

                # 3. 提取价格
                self.logger.info("[YantaiScraper.fetch] 步骤3: 提取价格")
                prices = await self._extract_prices(page)

                # 4. 保存到Excel
                today_str = datetime.now().strftime('%Y-%m-%d')
                if prices:
                    self.logger.info("[YantaiScraper.fetch] 步骤4: 保存到Excel")
                    save_to_excel_with_screenshot(prices, screenshot_b64, self.excel_file)
                    if ws_manager:
                        await ws_manager.notify_fetch_success(len(prices), today_str)
                else:
                    self.logger.warning("[YantaiScraper.fetch] 未提取到价格数据")
                    if ws_manager:
                        await ws_manager.notify_fetch_failed('未提取到价格数据')

                await browser.close()

                return {
                    'success': len(prices) > 0,
                    'prices': prices,
                    'source_name': '我的钢铁网-山东烟台',
                    'fetched_at': datetime.now().isoformat(),
                    'screenshot': screenshot_b64
                }

            except Exception as e:
                self.logger.error(f"[YantaiScraper.fetch] 抓取失败 | error={e}", exc_info=True)
                if ws_manager:
                    await ws_manager.notify_fetch_failed(str(e))
                return {
                    'success': False,
                    'error': str(e),
                    'prices': []
                }


# 兼容旧接口
async def run_fetch() -> Dict:
    """
    兼容旧接口 - 从 services/fetch_yantai.py 迁移
    保持原有函数签名以确保向后兼容
    """
    logger.info("[run_fetch] 兼容接口调用")
    scraper = YantaiScraper()
    return await scraper.fetch()


async def main():
    """测试入口"""
    result = await run_fetch()
    logger.info(f"[main] 抓取完成 | success={result['success']} | 数据={len(result.get('prices', []))}")
    print(f"抓取完成: 成功={result['success']}, 数据={len(result.get('prices', []))}")


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='[%(asctime)s] %(levelname)s %(message)s')
    asyncio.run(main())