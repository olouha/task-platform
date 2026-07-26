# -*- coding: utf-8 -*-
"""
Excel 解析服务 - 解析员工用 WPS 转图 / 复制网页得到的钢筋价格 Excel。

通用解析：自动识别表头（按关键字匹配），列顺序任意都能解析；
复用 screenshot_recognizer 的枚举纠错（材质/规格）+ 合理性校验（valid/issues）。
"""
import re
import logging
from typing import Optional, Dict, Any, List, Tuple

logger = logging.getLogger(__name__)

REGION = '山东烟台'

# 表头关键字 → 字段（覆盖 mysteel 价格表常见列名 + WPS 转表/复制粘贴的列名）
HEADER_KEYWORDS: Dict[str, List[str]] = {
    'material_name': ['品名', '材料名称', '产品名称', '名称'],
    'spec': ['规格', '规格型号', '规格/牌号'],
    'material_type': ['材质', '牌号'],
    'brand': ['品牌', '钢厂', '厂家', '生产企业', '供应商', '产地'],
    'price': ['价格', '单价', '报价', '现价'],
}


def _extract_price(val: Any) -> Optional[int]:
    """从单元格提取价格（3-5 位数字，落在 1500-6500）"""
    if val is None:
        return None
    s = str(val).replace(',', '').replace('，', '').strip()
    m = re.search(r'\d{3,5}(?:\.\d+)?', s)
    if not m:
        return None
    try:
        p = int(float(m.group()))
    except ValueError:
        return None
    return p if 1500 <= p <= 6500 else None


def _find_header(rows: List[tuple]) -> Tuple[int, Dict[str, int]]:
    """在前 20 行找表头行，返回 (行索引, {字段: 列索引})。要求至少匹配到价格列 + 1 个其他列。"""
    for i, row in enumerate(rows[:20]):
        col_map: Dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            text = str(cell) if cell is not None else ''
            if not text:
                continue
            for field, keywords in HEADER_KEYWORDS.items():
                if field not in col_map and any(kw in text for kw in keywords):
                    col_map[field] = col_idx
        if 'price' in col_map and len(col_map) >= 2:
            logger.info(f"[excel_parser] 找到表头 | 行={i} | 列映射={col_map}")
            return i, col_map
    logger.warning("[excel_parser] 未找到表头行（需含 价格/单价 + 品名/规格/材质/品牌 之一）")
    return -1, {}


def _cell_str(val: Any) -> str:
    if val is None:
        return ''
    return str(val).strip()


def parse_excel_to_prices(file_path: str) -> Dict[str, Any]:
    """
    解析 Excel 为价格记录列表（仅解析，不入库）。

    Returns:
        {ok: bool, prices: List[Dict], warnings: List[str]}
        prices 每项含 material_name/spec/material_type/brand/price/region/valid/issues
    """
    from openpyxl import load_workbook
    from services.price.screenshot_recognizer import (
        _guess_material_name, _validate_record,
    )

    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        logger.error(f"[excel_parser] 打开 Excel 失败 | {type(e).__name__}: {e}", exc_info=True)
        return {'ok': False, 'prices': [], 'warnings': [f'打开 Excel 失败: {type(e).__name__}: {e}']}

    prices: List[Dict[str, Any]] = []
    parsed_sheet = False
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header_idx, col_map = _find_header(rows)
            if header_idx >= 0 and 'price' in col_map:
                data_start = header_idx + 1
            else:
                # 无表头模式：按 mysteel 标准列序（品名/规格/材质/品牌/价格）
                col_map = {'material_name': 0, 'spec': 1, 'material_type': 2, 'brand': 3, 'price': 4}
                data_start = 0
                logger.info(f"[excel_parser] 未找到表头，按固定列顺序解析 | {col_map}")

            for row in rows[data_start:]:
                try:
                    price = _extract_price(row[col_map['price']])
                    if not price:
                        continue
                    mt_raw = _cell_str(row[col_map['material_type']]) if 'material_type' in col_map else ''
                    spec_raw = _cell_str(row[col_map['spec']]) if 'spec' in col_map else ''
                    brand_raw = _cell_str(row[col_map['brand']]) if 'brand' in col_map else ''
                    name_raw = _cell_str(row[col_map['material_name']]) if 'material_name' in col_map else ''

                    # Excel 是结构化数据（复制粘贴所得），原值即准确，不做 OCR 纠错
                    material_type = mt_raw
                    spec = spec_raw
                    rec = {
                        'material_name': name_raw or _guess_material_name(material_type),
                        'spec': spec,
                        'material_type': material_type,
                        'brand': brand_raw[:20],
                        'price': price,
                        'region': REGION,
                    }
                    rec['valid'], rec['issues'] = _validate_record(rec)
                    prices.append(rec)
                except Exception:
                    continue
            parsed_sheet = True
            break  # 只解析第一个有效 sheet
    finally:
        wb.close()

    logger.info(f"[excel_parser] 解析完成 | records={len(prices)}")
    if not prices:
        hint = 'Excel 未解析到价格行' if parsed_sheet else '未找到有效表头'
        return {'ok': False, 'prices': [],
                'warnings': [f'{hint}（请确保表头含 价格/单价，以及 品名/规格/材质/品牌/钢厂 之一）']}
    return {'ok': True, 'prices': prices, 'warnings': []}
