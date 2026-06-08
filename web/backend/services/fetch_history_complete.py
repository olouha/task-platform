"""
烟台钢筋价格历史数据完整抓取脚本
从 https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html 抓取
支持日期选择、自动翻页、数据保存（Excel+截图+数据库）

用法：
    python fetch_history_complete.py --start 2024-01-01 --end 2026-05-30 --interval 10
"""
import asyncio
import sys
import json
import base64
import time
from pathlib import Path
from datetime import datetime, timedelta

sys.path.insert(0, '.')

from playwright.async_api import async_playwright
import openpyxl
import sqlite3

DATA_DIR = Path('services/data')
DB_FILE = DATA_DIR / 'yantai_rebar.db'
EXCEL_FILE = DATA_DIR / '烟台钢筋价格_完整历史数据.xlsx'
COOKIE_FILE = DATA_DIR / 'myst_cookies.json'
HISTORY_URL = 'https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html'


def init_database():
    """初始化数据库"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS rebar_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            fetch_time TEXT,
            material_name TEXT,
            spec TEXT,
            material_type TEXT,
            brand TEXT,
            price INTEGER,
            region TEXT DEFAULT '山东烟台',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(date, fetch_time, material_name, spec, brand, price)
        )
    ''')

    c.execute('CREATE INDEX IF NOT EXISTS idx_date ON rebar_prices(date)')
    conn.commit()
    conn.close()


def get_existing_data():
    """获取已抓取的数据键"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT date, fetch_time, material_name, spec, brand FROM rebar_prices')
    existing = set(f"{r[0]}_{r[1]}_{r[2]}_{r[3]}_{r[4]}" for r in c.fetchall())
    conn.close()
    return existing


def save_to_excel(date: str, period: str, prices: list, screenshot_b64: str = None):
    """保存数据到Excel（含截图）"""
    if not prices:
        return False

    if not EXCEL_FILE.exists():
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        wb.save(EXCEL_FILE)

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # 创建sheet名
    time_str = datetime.now().strftime('%H%M%S')
    sheet_name = f'{date}_{period}_{time_str}'

    # 如果已存在则跳过
    if sheet_name in wb.sheetnames:
        wb.close()
        return False

    ws = wb.create_sheet(title=sheet_name)

    # 标题
    ws.merge_cells('A1:K1')
    period_text = '上午' if period == 'AM' else '下午'
    ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date} {period_text}')
    ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=14)

    # 表头
    headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
        ws.cell(row=3, column=col).font = openpyxl.styles.Font(bold=True)

    # 数据
    fetch_time = datetime.now().strftime('%H:%M:%S')
    for i, price in enumerate(prices):
        row = 4 + i
        ws.cell(row=row, column=1, value=date)
        ws.cell(row=row, column=2, value=fetch_time)
        ws.cell(row=row, column=3, value=price.get('material_name', ''))
        ws.cell(row=row, column=4, value=price.get('spec', ''))
        ws.cell(row=row, column=5, value=price.get('material_type', ''))
        ws.cell(row=row, column=6, value=price.get('brand', ''))
        ws.cell(row=row, column=7, value=price.get('price', 0))
        ws.cell(row=row, column=11, value='山东烟台')

    # 嵌入截图
    if screenshot_b64:
        try:
            from openpyxl.drawing.image import Image as XLImage
            date_for_file = date.replace("-", "")
            screenshot_path = DATA_DIR / f'screenshot_{date_for_file}_{period}.png'
            with open(screenshot_path, 'wb') as f:
                f.write(base64.b64decode(screenshot_b64))

            img = XLImage(str(screenshot_path))
            img.width = 800
            img.height = 450
            img.anchor = f'A{4 + len(prices) + 2}'
            ws.add_image(img)
        except Exception as e:
            print(f'    [WARN] 截图保存失败: {e}')

    wb.save(EXCEL_FILE)
    wb.close()
    return True


async def click_date_and_fetch(page, date_str: str, existing: set):
    """选择日期并抓取数据"""
    try:
        # 使用JavaScript设置日期值
        await page.evaluate(f'''
            const startInput = document.querySelector('input.startTime');
            if (startInput) {{
                startInput.value = '{date_str}';
                startInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
                startInput.dispatchEvent(new Event('change', {{ bubbles: true }}));
            }}
            const endInput = document.querySelector('input.endTime');
            if (endInput) {{
                endInput.value = '{date_str}';
                endInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
                endInput.dispatchEvent(new Event('change', {{ bubbles: true }}));
            }}
        ''')

        await page.wait_for_timeout(1000)

        # 点击搜索
        search_btn = await page.query_selector('button:has-text("搜索")')
        if not search_btn:
            search_btn = await page.query_selector('.search-btn')

        if search_btn:
            await search_btn.click()
            print(f'    已点击搜索')

        # 等待数据加载
        await page.wait_for_timeout(4000)

        # 截图
        screenshot = await page.screenshot(full_page=True)
        screenshot_b64 = base64.b64encode(screenshot).decode('utf-8')

        # 提取数据
        data = await page.evaluate('''() => {
            const results = [];
            const tables = document.querySelectorAll('table');
            tables.forEach(table => {
                const tbody = table.querySelector('tbody');
                if (tbody) {
                    const rows = tbody.querySelectorAll('tr');
                    rows.forEach(row => {
                        const cells = row.querySelectorAll('td');
                        const rowData = [];
                        cells.forEach(cell => rowData.push(cell.textContent.trim()));
                        if (rowData.length >= 5) results.push(rowData);
                    });
                }
            });
            return results;
        }''')

        prices = []
        for row in data:
            if len(row) >= 6:
                material_name = str(row[0]).strip()
                spec = str(row[1]).strip()
                material_type = str(row[2]).strip()
                brand = str(row[3]).strip()
                price_str = str(row[4]).strip()

                valid_names = ['高线', '螺纹钢', '盘螺', '圆钢']
                if material_name in valid_names and spec.startswith('Φ'):
                    try:
                        price = int(''.join(filter(str.isdigit, price_str)))
                        if price > 0:
                            prices.append({
                                'date': date_str,
                                'fetch_time': datetime.now().strftime('%H:%M:%S'),
                                'material_name': material_name,
                                'spec': spec,
                                'material_type': material_type,
                                'brand': brand,
                                'price': price
                            })
                    except:
                        pass

        if prices:
            # 保存到数据库
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()

            inserted = 0
            for p_data in prices:
                key = f"{p_data['date']}_{p_data['fetch_time']}_{p_data['material_name']}_{p_data['spec']}_{p_data['brand']}"
                if key not in existing:
                    try:
                        c.execute('''
                            INSERT INTO rebar_prices
                            (date, fetch_time, material_name, spec, material_type, brand, price, region)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            p_data['date'], p_data['fetch_time'], p_data['material_name'],
                            p_data['spec'], p_data['material_type'], p_data['brand'],
                            p_data['price'], '山东烟台'
                        ))
                        inserted += 1
                        existing.add(key)
                    except:
                        pass

            conn.commit()
            conn.close()

            # 保存到Excel
            period = 'AM' if '09' in datetime.now().strftime('%H') else 'PM'
            save_to_excel(date_str, period, prices, screenshot_b64)

            print(f'    [OK] 提取 {len(prices)} 条，新增 {inserted} 条')
            return inserted
        else:
            print(f'    [FAIL] 未提取到钢筋数据')
            return 0

    except Exception as e:
        print(f'    [ERROR] {e}')
        return 0


async def fetch_by_date_range(start_date: str, end_date: str, interval_seconds: int = 10):
    """按日期范围抓取"""
    print('=' * 60)
    print(f'开始历史数据抓取')
    print(f'日期范围: {start_date} 至 {end_date}')
    print(f'抓取间隔: {interval_seconds} 秒')
    print('=' * 60)

    # 加载凭据
    with open(DATA_DIR / 'mysteel_config.json', 'r', encoding='utf-8') as f:
        config = json.load(f)
        username = config.get('username')
        password = config.get('password')

    init_database()
    existing = get_existing_data()
    print(f'数据库已有 {len(existing)} 条记录')

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(viewport={'width': 1920, 'height': 3000}, locale='zh-CN')
        page = await context.new_page()

        # 登录
        print('\n1. 登录中...')
        await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
        await page.wait_for_timeout(5000)

        try:
            account_tab = await page.query_selector('.form-tab-account, a[data-tab="account"]')
            if account_tab:
                await account_tab.click()
                await page.wait_for_timeout(2000)
        except: pass

        await page.evaluate(f'''() => {{
            const inputs = document.querySelectorAll('input');
            for (const inp of inputs) {{
                const ph = inp.placeholder || '';
                if (ph.includes('用户名')) inp.value = '{username}';
                if (ph.includes('密码') && inp.type === 'password') inp.value = '{password}';
            }}
        }}''')
        await page.wait_for_timeout(500)

        try:
            login_btn = await page.query_selector('.form-button-login, button:has-text("登录")')
            if login_btn:
                await login_btn.click()
        except: pass

        await page.wait_for_timeout(8000)
        print('  登录完成')

        # 访问历史页面
        print('\n2. 访问历史页面...')
        await page.goto(HISTORY_URL, wait_until='networkidle', timeout=60000)
        await page.wait_for_timeout(5000)

        # 保存初始截图
        await page.screenshot(path=str(DATA_DIR / 'history_initial.png'), full_page=True)

        # 生成日期列表（仅工作日）
        current = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        dates = []

        while current <= end:
            if current.weekday() < 5:  # 周一到周五
                dates.append(current.strftime('%Y-%m-%d'))
            current += timedelta(days=1)

        print(f'\n3. 需要抓取 {len(dates)} 个工作日')
        print('   请在浏览器中操作：')
        print('   - 找到日期选择器')
        print('   - 输入开始日期和结束日期')
        print('   - 点击搜索按钮')
        print('   程序会自动保存数据\n')

        # 等待用户手动操作
        print('   等待 30 秒让您操作...')
        await page.wait_for_timeout(30000)

        # 尝试自动抓取每个日期
        print('\n4. 开始自动逐日抓取...')

        total_inserted = 0
        success_count = 0
        fail_count = 0

        for i, date in enumerate(dates):
            print(f'\n[{i+1}/{len(dates)}] {date}')

            inserted = await click_date_and_fetch(page, date, existing)

            if inserted > 0:
                success_count += 1
                total_inserted += inserted
            else:
                fail_count += 1

            # 等待间隔（防止被封）
            if i < len(dates) - 1:
                print(f'    等待 {interval_seconds} 秒...')
                await page.wait_for_timeout(interval_seconds * 1000)

        await browser.close()

        print('\n' + '=' * 60)
        print('抓取完成')
        print(f'  成功: {success_count} 天')
        print(f'  失败: {fail_count} 天')
        print(f'  新增记录: {total_inserted} 条')
        print(f'  Excel文件: {EXCEL_FILE}')
        print('=' * 60)


async def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(description='烟台钢筋价格历史数据完整抓取')
    parser.add_argument('--start', '-s', default='2024-01-01', help='开始日期 (YYYY-MM-DD)')
    parser.add_argument('--end', '-e', default=None, help='结束日期 (YYYY-MM-DD)')
    parser.add_argument('--interval', '-i', type=int, default=10, help='抓取间隔秒数')

    args = parser.parse_args()

    end_date = args.end or datetime.now().strftime('%Y-%m-%d')

    await fetch_by_date_range(args.start, end_date, args.interval)


if __name__ == '__main__':
    asyncio.run(main())