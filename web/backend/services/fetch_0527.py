"""
抓取5月27日数据
"""
import asyncio
import json
import openpyxl
from pathlib import Path
from playwright.async_api import async_playwright
from openpyxl.styles import Font
from datetime import datetime
import re

DATA_DIR = Path('services/data')
COOKIE_FILE = DATA_DIR / 'mysteel_cookies.json'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格_完整版_数据+截图.xlsx'

async def main():
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(viewport={'width': 1920, 'height': 3000}, locale='zh-CN')
        page = await context.new_page()

        if COOKIE_FILE.exists():
            with open(COOKIE_FILE, 'r', encoding='utf-8') as f:
                cookies = json.load(f)
            await context.add_cookies(cookies)

        url = 'https://jiancai.mysteel.com/m/26052716/F397E9E69DB11079.html'
        print(f'抓取: {url}')

        try:
            await page.goto(url, wait_until='load', timeout=60000)
            await page.wait_for_timeout(5000)

            # 滚动加载
            for i in range(5):
                await page.evaluate('window.scrollBy(0, 500)')
                await page.wait_for_timeout(500)

            # 等待表格出现
            await page.wait_for_selector('table', timeout=10000)

            # 提取数据
            data = await page.evaluate('''() => {
                const tables = document.querySelectorAll('table');
                const results = [];
                tables.forEach((table) => {
                    const rows = table.querySelectorAll('tr');
                    rows.forEach((row) => {
                        const cells = row.querySelectorAll('td');
                        if (cells.length >= 4) {
                            const cellTexts = [];
                            for (let j = 0; j < Math.min(cells.length, 10); j++) {
                                cellTexts.push(cells[j].textContent.trim());
                            }
                            results.push(cellTexts);
                        }
                    });
                });
                return results;
            }''')

            print(f'找到 {len(data)} 行')

            # 过滤有效数据
            prices = []
            for row in data:
                if len(row) >= 4:
                    material = str(row[0]) if row[0] else ''
                    spec = str(row[1]) if row[1] else ''
                    mat_type = str(row[2]) if row[2] else ''
                    price_text = str(row[3]) if row[3] else ''

                    valid_materials = ['高线', '螺纹', '盘螺', '圆钢']
                    if any(m in material for m in valid_materials):
                        match = re.search(r'(\d{3,5})', price_text)
                        if match:
                            price = int(match.group(1))
                            prices.append({
                                'material': material,
                                'spec': spec,
                                'mat_type': mat_type,
                                'price': price
                            })

            print(f'有效价格: {len(prices)} 条')

            if prices:
                print('\n示例:')
                for p in prices[:5]:
                    print(f"  {p['material']} | {p['spec']} | {p['mat_type']} | {p['price']}")

                # 保存到Excel
                wb = openpyxl.load_workbook(EXCEL_FILE)
                sheet_name = '2026-05-27'

                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(title=sheet_name)

                    ws.cell(row=1, column=1, value='山东烟台钢筋价格 - 2026-05-27')
                    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
                    ws.merge_cells('A1:K1')

                    headers = ['品名', '规格', '材质', '价格(元/吨)', '', '品牌', '时间']
                    for col, h in enumerate(headers, 1):
                        ws.cell(row=2, column=col, value=h).font = Font(bold=True)

                    for i, p in enumerate(prices):
                        row = i + 3
                        ws.cell(row=row, column=1, value=p['material'])
                        ws.cell(row=row, column=2, value=p['spec'])
                        ws.cell(row=row, column=3, value=p['mat_type'])
                        ws.cell(row=row, column=4, value=p['price'])
                        ws.cell(row=row, column=7, value='16:15')

                    wb.save(EXCEL_FILE)
                    wb.close()
                    print(f'\n已保存到Excel')

                    # 更新数据库
                    import sqlite3
                    conn = sqlite3.connect(DATA_DIR / 'yantai_rebar.db')
                    c = conn.cursor()

                    for p in prices:
                        c.execute('''INSERT INTO rebar_prices
                            (date, fetch_time, material_name, spec, material_type, brand, price, region)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                            ('2026-05-27', '16:15', p['material'], p['spec'], p['mat_type'], '', p['price'], '山东烟台'))

                    conn.commit()
                    c.execute('SELECT MAX(date) FROM rebar_prices')
                    print(f'数据库最新日期: {c.execute("SELECT MAX(date) FROM rebar_prices").fetchone()[0]}')
                    conn.close()
                    print('已更新数据库')
                else:
                    print(f'Sheet已存在')

        except Exception as e:
            print(f'错误: {e}')

        await browser.close()

if __name__ == '__main__':
    asyncio.run(main())
