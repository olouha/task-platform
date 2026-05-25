"""
每日烟台钢筋价格抓取脚本（OCR验证码版）
使用 ddddocr 识别验证码
"""

import asyncio
import json
import hashlib
import random
import base64
from pathlib import Path
from datetime import datetime, date
from playwright.async_api import async_playwright
import openpyxl
import ddddocr

# 配置
DATA_DIR = Path(__file__).parent / 'data'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
COOKIE_FILE = DATA_DIR / 'mysteel_cookies.json'
TODAY = date.today().isoformat()
TIME_NOW = datetime.now().strftime('%H%M%S')

# 时段判断
CURRENT_HOUR = datetime.now().hour
IS_AM = CURRENT_HOUR < 14
PERIOD = 'AM' if IS_AM else 'PM'
PERIOD_LABEL = '上午' if IS_AM else '下午(较晚)'

# User-Agent池
USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
]

# 反检测脚本
ANTI_DETECTION_JS = '''
Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
Object.defineProperty(navigator, 'languages', {get: () => ['zh-CN', 'zh', 'en']});
window.chrome = {runtime: {}};
'''


class YantaiPriceScraper:
    """烟台钢筋价格抓取器（OCR版）"""

    def __init__(self):
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None
        self.ocr = ddddocr.DdddOcr()

    async def __aenter__(self):
        self.playwright = await async_playwright().start()
        await self.start()
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.stop()
        await self.playwright.stop()

    async def start(self):
        """启动浏览器"""
        random_ua = random.choice(USER_AGENTS)
        self.browser = await self.playwright.chromium.launch(
            headless=False,  # OCR需要非headless
            args=[
                '--disable-blink-features=AutomationControlled',
                '--no-sandbox',
                '--disable-setuid-sandbox'
            ]
        )
        self.context = await self.browser.new_context(
            viewport={'width': 1920, 'height': 1080},
            locale='zh-CN',
            user_agent=random_ua
        )
        await self.context.add_init_script(ANTI_DETECTION_JS)
        self.page = await self.context.new_page()
        if COOKIE_FILE.exists():
            with open(COOKIE_FILE) as f:
                cookies = json.load(f)
            await self.context.add_cookies(cookies)
            print(f'[OK] 已加载 {len(cookies)} 个Cookie')

    async def stop(self):
        """停止浏览器"""
        if self.browser:
            await self.browser.close()

    async def random_sleep(self, min_sec=2, max_sec=5):
        await asyncio.sleep(random.uniform(min_sec, max_sec))

    async def recognize_captcha(self, img_bytes) -> str:
        """使用ddddocr识别验证码"""
        try:
            # 保存图片
            captcha_file = DATA_DIR / 'captcha.png'
            with open(captcha_file, 'wb') as f:
                f.write(img_bytes)

            # OCR识别
            result = self.ocr.classification(captcha_file)
            if isinstance(result, str):
                return result.strip()
            elif isinstance(result, list):
                return result[0].strip() if result else ''
            return ''
        except Exception as e:
            print(f'[WARN] OCR识别失败: {e}')
            return ''

    async def login_with_captcha(self):
        """带验证码的登录"""
        from config.mysteel import MYSTEEL_USERNAME, MYSTEEL_PASSWORD

        print('[LOGIN] 访问登录页...')
        await self.page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
        await self.random_sleep(2, 3)

        # 填写表单
        await self.page.evaluate(f'''
            const inputs = document.querySelectorAll('input');
            for (const inp of inputs) {{
                const ph = inp.placeholder || "";
                const type = inp.type || "";
                if (ph.includes("用户名") || ph.includes("账号")) {{
                    inp.value = "{MYSTEEL_USERNAME}";
                    inp.dispatchEvent(new Event('input', {{ bubbles: true }}));
                }}
                if (ph.includes("密码") && type === "password") {{
                    inp.value = "{MYSTEEL_PASSWORD}";
                    inp.dispatchEvent(new Event('input', {{ bubbles: true }}));
                }}
            }}
        ''')

        await self.random_sleep(1, 2)

        # 查找验证码图片
        captcha_url = None
        captcha_img = None

        # 尝试多种方式获取验证码
        try:
            # 方式1: 查找img标签
            captcha_imgs = await self.page.query_selector_all('img[src*="captcha"], img[alt*="验证"], img[class*="captcha"]')
            if captcha_imgs:
                for img in captcha_imgs:
                    src = await img.get_attribute('src')
                    if src:
                        captcha_url = src
                        break

            # 方式2: 通过API获取
            if not captcha_url:
                await self.page.screenshot(path=str(DATA_DIR / 'page.png'), full_page=False)
                # 从截图OCR验证码位置
                page_img = await self.page.screenshot(type='png')
                # 这部分需要根据实际页面结构调整

        except Exception as e:
            print(f'[WARN] 查找验证码失败: {e}')

        # 如果找到验证码URL，识别并填写
        if captcha_url:
            print('[OCR] 识别验证码...')
            # 尝试直接访问验证码图片
            try:
                captcha_response = await self.page.goto(captcha_url)
                captcha_img = await captcha_response.body()
                captcha_code = await self.recognize_captcha(captcha_img)

                if captcha_code:
                    print(f'[OK] 验证码识别: {captcha_code}')
                    # 填写验证码
                    captcha_input = await self.page.query_selector('input[name*="captcha"], input[placeholder*="验证码"], input[id*="captcha"]')
                    if captcha_input:
                        await captcha_input.fill(captcha_code)
                        await self.random_sleep(1, 2)
                else:
                    print('[WARN] 验证码识别为空')
            except Exception as e:
                print(f'[WARN] 获取验证码图片失败: {e}')

        # 点击登录按钮
        for selector in ['button:has-text("登录")', '.form-button-login', 'button[type="submit"]']:
            try:
                btn = await self.page.query_selector(selector)
                if btn:
                    await btn.click()
                    break
            except:
                continue

        await self.random_sleep(10, 15)

        # 保存Cookie
        cookies = await self.context.cookies()
        with open(COOKIE_FILE, 'w') as f:
            json.dump(cookies, f, ensure_ascii=False)
        print(f'[OK] 已保存 {len(cookies)} 个Cookie')

        return True

    async def check_login_status(self) -> bool:
        """检查登录状态"""
        test_url = 'https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html'
        await self.page.goto(test_url, wait_until='domcontentloaded', timeout=60000)
        await self.random_sleep(3, 5)

        current_url = self.page.url
        if 'captcha' in current_url or 'passport' in current_url:
            return False

        body_text = await self.page.evaluate('document.body.textContent')
        return len(body_text) > 100

    async def fetch_prices(self) -> list:
        """获取价格数据"""
        url = 'https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html'
        print(f'[DATA] 获取价格数据')

        await self.page.goto(url, wait_until='domcontentloaded', timeout=60000)
        await self.random_sleep(3, 5)

        data = await self.page.evaluate('''() => {
            const results = [];
            const tables = document.querySelectorAll('table');
            tables.forEach(table => {
                const rows = table.querySelectorAll('tr');
                rows.forEach(row => {
                    const cells = row.querySelectorAll('td');
                    if (cells.length >= 5) {
                        const material_name = cells[0]?.textContent?.trim();
                        const spec = cells[1]?.textContent?.trim();
                        const brand = cells[3]?.textContent?.trim();
                        const price_str = cells[4]?.textContent?.trim();
                        if (['高线', '螺纹钢', '盘螺', '圆钢'].includes(material_name) &&
                            spec && spec.startsWith('Φ') && price_str && /^\\d+$/.test(price_str)) {
                            results.push({material_name, spec, brand, price: parseInt(price_str)});
                        }
                    }
                });
            });
            return results;
        }''')

        print(f'[OK] 获取到 {len(data)} 条数据')
        return data

    def calculate_data_hash(self, data: list) -> str:
        sorted_data = sorted([(d['material_name'], d['spec'], d['brand'], d['price']) for d in data])
        return hashlib.md5(json.dumps(sorted_data, ensure_ascii=False).encode()).hexdigest()

    def check_today_data_exists(self, data_hash: str) -> bool:
        if not EXCEL_FILE.exists():
            return False
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sheet_names = wb.sheetnames
            wb.close()
            exists = any(f'{TODAY}_{PERIOD}_' in s for s in sheet_names)
            if not exists:
                hash_file = DATA_DIR / 'data_hashes.json'
                if hash_file.exists():
                    hashes = json.load(open(hash_file))
                    key = f'{TODAY}_{PERIOD}'
                    exists = key in hashes and hashes[key] == data_hash
            return exists
        except:
            return False

    def save_to_excel(self, data: list, data_hash: str) -> bool:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE) if EXCEL_FILE.exists() else openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

            sheet_name = f'{TODAY}_{PERIOD}_{TIME_NOW}'
            ws = wb.create_sheet(title=sheet_name)

            period_text = '下午(较晚)' if PERIOD == 'PM' else '上午'
            ws.merge_cells('A1:K1')
            ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {TODAY} {period_text}').font = openpyxl.styles.Font(bold=True, size=14)
            ws.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')

            from openpyxl.styles import PatternFill
            header_fill = PatternFill(start_color='FF6B6B' if PERIOD == 'PM' else '4472C4', end_color='FF6B6B' if PERIOD == 'PM' else '4472C4', fill_type='solid')
            headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True, size=12, color='FFFFFF')
                cell.fill = header_fill

            fetch_time = datetime.now().strftime('%H:%M:%S')
            for i, item in enumerate(data):
                row = 4 + i
                for col, val in enumerate([TODAY, fetch_time, item['material_name'], item['spec'],
                    item.get('material_type', ''), item['brand'], item['price'], '', '', '', '山东烟台'], 1):
                    ws.cell(row=row, column=col, value=val)

            wb.save(EXCEL_FILE)
            wb.close()

            hash_file = DATA_DIR / 'data_hashes.json'
            hashes = json.load(open(hash_file)) if hash_file.exists() else {}
            hashes[f'{TODAY}_{PERIOD}'] = data_hash
            with open(hash_file, 'w') as f:
                json.dump(hashes, f, ensure_ascii=False)

            print(f'[OK] 数据已保存: {sheet_name}')
            return True
        except Exception as e:
            print(f'[FAIL] 保存失败: {e}')
            return False


async def main():
    """主函数"""
    print(f'[START] 抓取 - {TODAY} {PERIOD_LABEL}')

    async with YantaiPriceScraper() as scraper:
        is_logged_in = await scraper.check_login_status()

        if not is_logged_in:
            print('[INFO] 需要登录，使用OCR验证码...')
            await scraper.login_with_captcha()
            await scraper.random_sleep(2, 3)
            if not await scraper.check_login_status():
                print('[FAIL] 登录失败')
                return

        data = await scraper.fetch_prices()
        if not data:
            print('[WARN] 未获取到数据')
            return

        data_hash = scraper.calculate_data_hash(data)
        if scraper.check_today_data_exists(data_hash):
            print('[SKIP] 今日数据已存在')
            return

        success = await asyncio.to_thread(scraper.save_to_excel, data, data_hash)
        if success:
            print(f'[OK] 完成 - {len(data)}条数据, {PERIOD_LABEL}')


if __name__ == '__main__':
    asyncio.run(main())