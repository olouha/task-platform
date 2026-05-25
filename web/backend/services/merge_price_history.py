"""
合并烟台钢筋价格历史数据
将所有sheet合并为一个，按日期汇总
"""

import os
import json
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR / 'data'
SOURCE_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
OUTPUT_FILE = DATA_DIR / '烟台钢筋价格汇总.xlsx'
PROGRESS_FILE = DATA_DIR / 'merge_progress.json'


def load_progress():
    if PROGRESS_FILE.exists():
        with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'last_sheet': None, 'merged': 0}


def save_progress(progress):
    with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def merge_all_sheets():
    """合并所有sheet到汇总文件"""

    print('=' * 60)
    print('合并烟台钢筋价格历史数据')
    print('=' * 60)

    if not SOURCE_FILE.exists():
        print('[ERROR] 源文件不存在')
        return

    # 读取源文件
    print('[1/3] 读取源文件...')
    source_wb = openpyxl.load_workbook(SOURCE_FILE, read_only=True)

    # 按日期分组数据
    all_data = defaultdict(list)

    for i, sheet_name in enumerate(source_wb.sheetnames):
        if '_' not in sheet_name:
            continue

        parts = sheet_name.split('_')
        if len(parts) < 2:
            continue

        date_str = parts[0]
        period = parts[1]  # AM or PM

        try:
            ws = source_wb[sheet_name]
            rows = list(ws.iter_rows(min_row=4, values_only=True))  # 从第4行开始是数据

            for row in rows:
                if not row or not row[0]:
                    continue
                # row: (日期, 时间, 品名, 规格, 材质, 品牌, 单价, 涨跌, 备注, 钢号, 地区)
                if row[2] and row[6]:  # 有品名和价格
                    all_data[date_str].append({
                        'period': period,
                        'date': row[0] or date_str,
                        'time': row[1] or ('09:00:00' if period == 'AM' else '15:00:00'),
                        'material_name': row[2],
                        'spec': row[3] or '',
                        'material_type': row[4] or '',
                        'brand': row[5] or '',
                        'price': row[6],
                        'price_change': row[7] or '',
                        'remark': row[8] or '',
                        'steel_code': row[9] or '',
                        'region': row[10] or '山东烟台'
                    })
        except Exception as e:
            print(f'[WARN] 读取 {sheet_name} 失败: {e}')
            continue

        if (i + 1) % 50 == 0:
            print(f'  已处理 {i + 1}/{len(source_wb.sheetnames)} sheets')

    source_wb.close()
    print(f'  共 {len(all_data)} 个日期')

    # 创建汇总文件
    print('[2/3] 创建汇总文件...')

    if os.path.exists(OUTPUT_FILE):
        os.remove(OUTPUT_FILE)

    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 创建汇总sheet
    ws_summary = wb.create_sheet(title='价格汇总')
    ws_summary.cell(row=1, column=1, value='烟台钢筋价格历史汇总').font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:L1')

    # 表头
    headers = ['日期', '品名', '规格', '品牌', '上午价格', '下午价格', '涨跌', '备注', '时间', '钢号', '地区', '数据类型']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 按日期排序写入
    print('[3/3] 写入汇总数据...')
    row_num = 4
    sorted_dates = sorted(all_data.keys())

    # 去重：将同一日期同一规格同一品牌的价格合并
    for date_str in sorted_dates:
        items = all_data[date_str]

        # 按 (品名, 规格, 品牌) 分组
        grouped = defaultdict(list)
        for item in items:
            key = (item['material_name'], item['spec'], item['brand'])
            grouped[key].append(item)

        for (material, spec, brand), price_list in grouped.items():
            am_price = None
            pm_price = None
            am_time = ''
            pm_time = ''

            for p in price_list:
                if p['period'] == 'AM':
                    am_price = p['price']
                    am_time = p['time']
                else:
                    pm_price = p['price']
                    pm_time = p['time']

            # 计算涨跌（下午-上午）
            price_change = ''
            if am_price and pm_price:
                diff = pm_price - am_price
                price_change = f'+{diff}' if diff > 0 else str(diff)

            # 写一行
            values = [
                date_str,
                material,
                spec,
                brand,
                am_price or '',
                pm_price or '',
                price_change,
                '',
                f'{am_time or ""} / {pm_time or ""}',
                '',
                '山东烟台',
                'AM/PM合并' if (am_price and pm_price) else ('AM' if am_price else 'PM')
            ]
            for col, val in enumerate(values, 1):
                cell = ws_summary.cell(row=row_num, column=col, value=val)
                cell.border = thin_border
            row_num += 1

        if len(sorted_dates) % 20 == 0:
            print(f'  已处理 {sorted_dates.index(date_str) + 1}/{len(sorted_dates)} 日期')

    # 设置列宽
    ws_summary.column_dimensions['A'].width = 12
    ws_summary.column_dimensions['B'].width = 10
    ws_summary.column_dimensions['C'].width = 8
    ws_summary.column_dimensions['D'].width = 10
    ws_summary.column_dimensions['E'].width = 10
    ws_summary.column_dimensions['F'].width = 10

    # 保存
    wb.save(OUTPUT_FILE)
    wb.close()

    print(f'\n完成！')
    print(f'汇总文件: {OUTPUT_FILE}')
    print(f'总日期: {len(sorted_dates)}')
    print(f'总数据行: {row_num - 4}')

    return len(sorted_dates), row_num - 4


if __name__ == '__main__':
    dates, rows = merge_all_sheets()