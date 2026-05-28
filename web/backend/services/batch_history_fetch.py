"""
批量抓取历史价格数据 - 使用URL列表文件
账号: M6616592358 / panhui199261
每15秒抓取一个，每个月后停顿30秒
"""
import asyncio
import json
import base64
import random
from pathlib import Path
from datetime import datetime
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

DATA_DIR = Path('services/data')
COOKIE_FILE = DATA_DIR / 'mysteel_cookies.json'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
# URL列表文件
URLS_FILE = DATA_DIR / 'all_urls_2024_2026.json'

# 账号
USERNAME = 'M6616592358'
PASSWORD = 'panhui199261'

LOG_FILE = DATA_DIR / 'logs' / f'history_fetch_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'


def log(msg):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f'[{timestamp}] {msg}'
    print(line)
    LOG_FILE.parent.mkdir(exist_ok=True)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(line + '\n')


def random_sleep():
    return random.uniform(14, 16)


async def login(page):
    log('登录中...')
    if COOKIE_FILE.exists():
        try:
            cookies = json.loads(COOKIE_FILE.read_text(encoding='utf-8'))
            if cookies:
                await page.context.add_cookies(cookies)
                log('已加载Cookie')
        except:
            pass

    await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
    await asyncio.sleep(random_sleep() / 3)

    try:
        account_tab = await page.query_selector('.form-tab-account, a[data-tab="account"]')
        if account_tab:
            await account_tab.click()
            await asyncio.sleep(random_sleep() / 5)
    except:
        pass

    await page.evaluate(f'''() => {{
        const inputs = document.querySelectorAll('input');
        for (const inp of inputs) {{
            const ph = inp.placeholder || '';
            if (ph.includes('用户名')) inp.value = '{USERNAME}';
            if (ph.includes('密码') && inp.type === 'password') inp.value = '{PASSWORD}';
        }}
    }}''')

    await asyncio.sleep(random_sleep() / 5)

    try:
        login_btn = await page.query_selector('button:has-text("登录"), input[type="submit"]')
        if login_btn:
            await login_btn.click()
            await asyncio.sleep(random_sleep() * 0.3)
            cookies = await page.context.cookies()
            COOKIE_FILE.write_text(json.dumps(cookies, ensure_ascii=False), encoding='utf-8')
            log('登录成功')
    except Exception as e:
        log(f'登录失败: {e}')


def get_existing_sheets():
    """获取Excel中已存在的所有sheet"""
    existing = set()
    if EXCEL_FILE.exists():
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
            for name in wb.sheetnames:
                # 格式: YYYY-MM-DD_AM_HHMMSS
                if name.startswith('20') and '_' in name:
                    parts = name.split('_')
                    if len(parts) >= 2:
                        existing.add(parts[0])  # 日期部分
                        existing.add(f"{parts[0]}_PM")  # 下午
            wb.close()
        except:
            pass
    return existing


def save_to_excel(prices, period, date_str, screenshot_b64):
    """保存到Excel"""
    try:
        if EXCEL_FILE.exists():
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE)
            except:
                log('  文件损坏')
                wb = openpyxl.Workbook()
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']
        else:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        sheet_name = date_str if period == 'AM' else f'{date_str}_PM'

        if sheet_name in wb.sheetnames:
            log(f'  Sheet已存在，跳过: {sheet_name}')
            wb.close()
            return False

        ws = wb.create_sheet(title=sheet_name)

        # 标题
        period_text = '下午(晚)' if period == 'PM' else '上午'
        ws.merge_cells('A1:K1')
        ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        # 表头
        header_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid') if period == 'PM' else \
            PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 数据
        fetch_time = datetime.now().strftime('%H:%M:%S')
        for i, price in enumerate(prices):
            row = 4 + i
            for col, val in enumerate([date_str, fetch_time, price['material_name'], price['spec'],
                                       price['material_type'], price['brand'], price['price'],
                                       '', '', '', '山东烟台'], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border

        # 截图
        if screenshot_b64:
            screenshot_path = DATA_DIR / f'screenshot_{date_str.replace("-", "")}_{period}.png'
            with open(screenshot_path, 'wb') as f:
                f.write(base64.b64decode(screenshot_b64))
            row = 4 + len(prices) + 2
            ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)
            img = Image(str(screenshot_path))
            img.width = 900
            img.height = 500
            img.anchor = f'A{row + 1}'
            ws.add_image(img)

        wb.save(EXCEL_FILE)
        wb.close()
        return True

    except Exception as e:
        log(f'  保存失败: {e}')
        return False


async def fetch_and_save(page, date_str, period, url):
    """抓取单个日期的数据"""
    try:
        await page.goto(url, wait_until='domcontentloaded', timeout=60000)
        await asyncio.sleep(random_sleep() * 0.3)

        # 滚动页面触发懒加载
        for i in range(5):
            await page.evaluate('window.scrollBy(0, 500)')
            await asyncio.sleep(1)

        await asyncio.sleep(2)

        # 截图
        screenshot_b64 = None
        try:
            screenshot = await page.screenshot(full_page=True)
            screenshot_b64 = base64.b64encode(screenshot).decode()
        except:
            pass

        # 提取数据 - 完整的提取逻辑
        data = await page.evaluate('''() => {
            const tables = document.querySelectorAll('table');
            const results = [];
            tables.forEach(table => {
                const rows = table.querySelectorAll('tr');
                rows.forEach(row => {
                    const cells = row.querySelectorAll('td, th');
                    if (cells.length >= 5) {
                        const material_name = cells[0]?.textContent?.trim().replace(/\\s+/g, ' ');
                        const spec = cells[1]?.textContent?.trim().replace(/\\s+/g, ' ');
                        const material_type = cells[2]?.textContent?.trim().replace(/\\s+/g, ' ');
                        const brand = cells[3]?.textContent?.trim().replace(/\\s+/g, ' ');
                        const price_text = cells[4]?.textContent?.trim().replace(/\\s+/g, ' ');

                        // 解析价格 - 可能是纯数字或带单位的
                        let price = 0;
                        const priceMatch = price_text.match(/(\\d{3,5})/);
                        if (priceMatch) {
                            price = parseInt(priceMatch[1]);
                        }

                        const valid_names = ['高线', '螺纹钢', '盘螺', '圆钢', '拉丝材'];
                        if (valid_names.some(n => material_name.includes(n)) && price > 0) {
                            results.push({
                                material_name,
                                spec,
                                material_type,
                                brand,
                                price
                            });
                        }
                    }
                });
            });
            return results;
        }''')

        return list(data), screenshot_b64

    except Exception as e:
        log(f'  抓取失败: {e}')
        return [], None


async def main():
    print()
    print('=' * 70)
    print('烟台钢筋价格 - 历史数据批量抓取')
    print('=' * 70)
    print()

    # 加载URL列表
    urls = []
    if URLS_FILE.exists():
        with open(URLS_FILE, 'r', encoding='utf-8') as f:
            urls = json.load(f)
    else:
        log('错误: URL列表文件不存在')
        return

    log(f'URL总数: {len(urls)}')

    # 获取已存在的数据
    existing = get_existing_sheets()
    log(f'已有数据: {len(existing)} 个')

    # 过滤需要抓取的
    tasks = []
    for date_str, period, url in urls:
        target = date_str if period == 'AM' else f'{date_str}_PM'
        if target not in existing:
            tasks.append((date_str, period, url))

    log(f'需要抓取: {len(tasks)} 条')
    print()

    if not tasks:
        log('所有数据已抓取完成')
        return

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            viewport={'width': 1920, 'height': 3000},
            locale='zh-CN'
        )
        page = await context.new_page()

        await login(page)

        success_count = 0
        fail_count = 0
        last_month = None

        for i, (date_str, period, url) in enumerate(tasks):
            # 检查月份变化
            current_month = date_str[:7]
            if last_month and current_month != last_month:
                log(f'\n月份切换: {last_month} -> {current_month}，停顿30秒')
                await asyncio.sleep(30)

            last_month = current_month

            log(f'[{i+1}/{len(tasks)}] {date_str} {period}...')
            print(f'  -> 抓取中...', end='', flush=True)

            prices, screenshot = await fetch_and_save(page, date_str, period, url)

            if prices:
                if save_to_excel(prices, period, date_str, screenshot):
                    existing.add(date_str)
                    existing.add(f'{date_str}_PM')
                    log(f' {len(prices)}条 [OK]')
                    success_count += 1
                else:
                    log(' 保存失败')
                    fail_count += 1
            else:
                log(' 无数据')
                fail_count += 1

            # 停顿15秒
            await asyncio.sleep(random_sleep())

        await browser.close()

    print()
    print('=' * 70)
    print('抓取完成')
    print('=' * 70)
    log(f'成功: {success_count} 条')
    log(f'失败: {fail_count} 条')

    final_existing = get_existing_sheets()
    log(f'最终数据: {len(final_existing)} 个日期')


if __name__ == '__main__':
    asyncio.run(main())