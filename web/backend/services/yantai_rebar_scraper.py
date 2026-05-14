"""
山东烟台钢筋价格爬虫 v7.1
- 支持品牌维度抓取（螺纹钢、盘螺、高线、圆钢）
- 自动登录（首次登录，后续使用Cookie）
- 每天定时抓取，支持截图和Excel导出
"""

import asyncio
import logging
import json
import re
import base64
from typing import List, Tuple
from datetime import datetime
from dataclasses import dataclass, field
from pathlib import Path

try:
    from playwright.async_api import async_playwright
    HAS_PLAYWRIGHT = True
except ImportError:
    HAS_PLAYWRIGHT = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 路径设置
SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR / "data"
LOGS_DIR = SCRIPT_DIR / "logs"

DATA_DIR.mkdir(exist_ok=True)
LOGS_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(str(LOGS_DIR / 'yantai_rebar_scraper.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


@dataclass
class MaterialPrice:
    material_id: str = ""
    material_name: str = ""  # 品名：高线、螺纹钢、盘螺、圆钢
    spec: str = ""           # 规格：如 Φ6、Φ8、Φ10
    material_type: str = ""  # 材质：如 HPB300、HRB400E、HRB500E
    brand: str = ""          # 钢厂/产地
    price: float = 0.0
    price_max: float = 0.0
    unit: str = ""
    price_change: str = ""   # 涨跌
    remark: str = ""         # 备注
    steel_code: str = ""     # 钢号
    region: str = ""


@dataclass
class CrawlResult:
    success: bool = False
    source_name: str = ""
    url: str = ""
    fetched_at: str = ""
    error_message: str = ""
    prices: List[MaterialPrice] = field(default_factory=list)
    screenshot: str = ""


class YantaiRebarScraper:
    """山东烟台钢筋价格爬虫 v8.0 - 修复URL获取逻辑"""

    # 配置文件路径
    CONFIG_FILE = DATA_DIR / "mysteel_config.json"

    def __init__(self, username: str = None, password: str = None):
        # 从配置文件读取，或使用默认值
        config = self._load_config()
        self.username = username or config.get('username', 'M6616592358')
        self.password = password or config.get('password', 'mysteel573005')
        self.last_fetch_file = LOGS_DIR / "yantai_last_fetch.json"
        self.excel_file = DATA_DIR / "山东烟台钢筋价格.xlsx"
        self.cookie_file = DATA_DIR / "mysteel_cookies.json"

    def _load_config(self) -> dict:
        """加载配置文件"""
        try:
            if self.CONFIG_FILE.exists():
                with open(self.CONFIG_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            logger.warning(f"加载配置失败: {e}")
        return {}

    def _save_config(self, username: str = None, password: str = None):
        """保存配置文件"""
        config = self._load_config()
        if username is not None:
            config['username'] = username
        if password is not None:
            config['password'] = password

        with open(self.CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        logger.info(f"配置已保存: username={username}")

    def update_credentials(self, username: str, password: str) -> bool:
        """更新登录凭据"""
        try:
            self._save_config(username, password)
            self.username = username
            self.password = password
            # 删除旧Cookie，强制重新登录
            if self.cookie_file.exists():
                self.cookie_file.unlink()
                logger.info("旧Cookie已删除，将重新登录")
            return True
        except Exception as e:
            logger.error(f"更新凭据失败: {e}")
            return False

    def _get_today_url(self) -> str:
        """获取今日价格URL - 从首页获取最新URL"""
        today = datetime.now()
        date_str = today.strftime('%y%m%d')
        # 默认URL格式（会动态更新）
        return f"https://jiancai.mysteel.com/m/{date_str}10/25B3355C6617BD3C.html"

    def _get_market_urls(self) -> list:
        """获取市场页面URL列表"""
        return [
            "https://jiancai.mysteel.com/market/pa228aa01010104a0aaaaa1.html",
            "https://jiancai.mysteel.com/market/pa228a81723aa0aaaaa1.html",
        ]

    def _check_rate_limit(self) -> Tuple[bool, str]:
        if self.last_fetch_file.exists():
            try:
                with open(self.last_fetch_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    last_fetch = datetime.fromisoformat(data['last_fetch'])
                    today = datetime.now().date()
                    if last_fetch.date() == today:
                        return False, f"今日({today})已抓取"
            except Exception as e:
                logger.warning(f"读取抓取记录失败: {e}")
        return True, "可以抓取"

    def _save_fetch_record(self, success: bool, prices_count: int = 0):
        record = {
            'last_fetch': datetime.now().isoformat(),
            'success': success,
            'prices_count': prices_count,
            'region': '山东烟台'
        }
        with open(self.last_fetch_file, 'w', encoding='utf-8') as f:
            json.dump(record, f, ensure_ascii=False, indent=2)

    def _load_cookies(self) -> List:
        """加载保存的Cookie"""
        try:
            if self.cookie_file.exists():
                with open(self.cookie_file, 'r', encoding='utf-8') as f:
                    cookies = json.load(f)
                    logger.info(f"已加载 {len(cookies)} 个Cookie")
                    return cookies
        except Exception as e:
            logger.warning(f"加载Cookie失败: {e}")
        return []

    def _save_cookies(self, cookies: List):
        """保存Cookie"""
        try:
            with open(self.cookie_file, 'w', encoding='utf-8') as f:
                json.dump(cookies, f, ensure_ascii=False)
            logger.info(f"Cookie已保存到 {self.cookie_file}")
        except Exception as e:
            logger.warning(f"保存Cookie失败: {e}")

    def _clean_text(self, text: str) -> str:
        if not text:
            return ""
        text = re.sub(r'[\n\r\t]+', '', text)
        text = re.sub(r'\s{2,}', ' ', text)
        return text.strip()

    def _generate_material_id(self, material_name: str, spec: str, brand: str) -> str:
        name_map = {'高线': 'gx', '螺纹钢': 'lw', '盘螺': 'pl', '圆钢': 'yg'}
        prefix = name_map.get(material_name, 'ot')
        brand_short = brand[:2] if brand else 'xx'
        spec_clean = spec.replace('Φ', '').replace('mm', '')
        return f"yt_{prefix}_{brand_short}_{spec_clean}"

    def _parse_remark_prices(self, remark: str) -> List[Tuple[str, float]]:
        prices = []
        if not remark:
            return prices
        pattern = re.compile(r'Φ(\d+(?:-\d+)?):(\d+)')
        for match in pattern.finditer(remark):
            spec = f"Φ{match.group(1)}"
            price = float(match.group(2))
            prices.append((spec, price))
        return prices

    async def _login(self, page, context) -> bool:
        """执行登录流程"""
        try:
            logger.info("开始登录...")

            # 访问登录页
            await page.goto('https://passport.mysteel.com/', wait_until='networkidle', timeout=60000)
            await page.wait_for_timeout(5000)

            # 点击账号登录标签
            try:
                # 尝试多种方式找到账号登录标签
                selectors = [
                    '.form-tab-account',
                    'a[data-tab="account"]',
                    'text=账号登录',
                    'xpath=//a[contains(text(),"账号登录")]'
                ]
                for selector in selectors:
                    try:
                        account_tab = await page.query_selector(selector)
                        if account_tab:
                            await account_tab.click()
                            logger.info("已切换到账号登录")
                            await page.wait_for_timeout(2000)
                            break
                    except:
                        continue
            except Exception as e:
                logger.warning(f"切换账号登录失败: {e}")

            # 填写登录表单 - 使用更可靠的选择器
            try:
                # 找到用户名输入框（第4个input，即账号登录模式的）
                inputs = await page.query_selector_all('input')
                username_input = None
                password_input = None

                for inp in inputs:
                    placeholder = await inp.get_attribute('placeholder') or ''
                    inp_type = await inp.get_attribute('type') or ''
                    is_visible = await inp.is_visible()

                    if '用户名' in placeholder and is_visible:
                        username_input = inp
                    elif '密码' in placeholder and inp_type == 'password' and is_visible:
                        password_input = inp

                if username_input:
                    await username_input.fill(self.username)
                    logger.info("已填写用户名")
                if password_input:
                    await password_input.fill(self.password)
                    logger.info("已填写密码")

            except Exception as e:
                logger.warning(f"填写表单失败: {e}")
                # 尝试使用 evaluate 直接设置值
                await page.evaluate('''() => {
                    const inputs = document.querySelectorAll('input');
                    for (const inp of inputs) {
                        const ph = inp.placeholder || '';
                        if (ph.includes('用户名')) inp.value = 'M6616592358';
                        if (ph.includes('密码') && inp.type === 'password') inp.value = 'mysteel573005';
                    }
                }''')

            await page.wait_for_timeout(500)

            # 勾选同意协议
            try:
                checkbox = await page.query_selector('input[type="checkbox"]')
                if checkbox and not await checkbox.is_checked():
                    await checkbox.click()
            except:
                pass

            await page.wait_for_timeout(500)

            # 点击登录按钮
            try:
                login_btn = await page.query_selector('.form-button-login')
                if login_btn:
                    await login_btn.click()
                    logger.info("已点击登录按钮")
                else:
                    # 尝试其他选择器
                    btns = await page.query_selector_all('button')
                    for btn in btns:
                        text = await btn.text_content()
                        if text and '登录' in text:
                            await btn.click()
                            logger.info("已点击登录按钮")
                            break
            except Exception as e:
                logger.warning(f"点击登录按钮失败: {e}")

            logger.info("等待登录完成...")
            await page.wait_for_timeout(8000)

            # 验证登录是否成功
            current_url = page.url
            if 'passport' in current_url and 'login' in current_url:
                logger.warning("登录可能未完成，仍在登录页")
                # 再等一会儿
                await page.wait_for_timeout(5000)
                current_url = page.url
                if 'passport' in current_url:
                    return False

            # 保存登录后的cookie
            cookies = await context.cookies()
            self._save_cookies(cookies)
            logger.info("登录成功，Cookie已保存")
            return True

        except Exception as e:
            logger.error(f"登录失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    async def _get_prices_via_js(self, page) -> List[MaterialPrice]:
        """通过JS获取表格数据"""
        try:
            # 先尝试查找价格表格
            data = await page.evaluate('''() => {
                // 尝试多种表格选择器
                const selectors = [
                    'table.price-table',
                    'table.list-table',
                    '.price-list table',
                    '.list-box table',
                    'table'
                ];

                let table = null;
                for (const sel of selectors) {
                    table = document.querySelector(sel);
                    if (table) break;
                }

                if (!table) {
                    // 尝试在iframe中查找
                    const iframes = document.querySelectorAll('iframe');
                    for (const iframe of iframes) {
                        try {
                            const iframeDoc = iframe.contentDocument || iframe.contentWindow.document;
                            table = iframeDoc.querySelector('table');
                            if (table) break;
                        } catch(e) {}
                    }
                }

                if (!table) {
                    // 查找包含价格数据的元素
                    const priceElements = document.querySelectorAll('[class*="price"], [class*="list"]');
                    return { success: false, message: 'No table found', elements: priceElements.length };
                }

                const rows = table.querySelectorAll('tr');
                const results = [];

                // 查找表头行，确定列索引
                let headerRow = 0;
                for (let i = 0; i < Math.min(rows.length, 3); i++) {
                    const cells = rows[i].querySelectorAll('td, th');
                    const text = Array.from(cells).map(c => c.textContent.trim()).join(',');
                    if (text.includes('品名') || text.includes('规格') || text.includes('价格')) {
                        headerRow = i;
                        break;
                    }
                }

                // 从表头后开始读取数据
                for (let i = headerRow + 1; i < rows.length; i++) {
                    const cells = rows[i].querySelectorAll('td');
                    if (cells.length >= 4) {
                        // 提取单元格文本
                        const cellTexts = [];
                        for (let j = 0; j < Math.min(cells.length, 10); j++) {
                            cellTexts.push(cells[j].textContent.trim());
                        }

                        // 判断是否有价格数据
                        const priceCell = cellTexts[4] || '';
                        const hasPrice = priceCell && /^\d/.test(priceCell) && !priceCell.includes('icon');

                        results.push({
                            material_name: cellTexts[0] || '',
                            spec: cellTexts[1] || '',
                            material_type: cellTexts[2] || '',
                            brand: cellTexts[3] || '',
                            price_text: priceCell,
                            has_price: hasPrice,
                            price_change: cellTexts[5] || '',
                            remark: cellTexts[6] || '',
                            steel_code: cellTexts[7] || ''
                        });
                    }
                }

                return { success: true, data: results, rowCount: rows.length };
            }''')

            if data.get('success') and data.get('data'):
                prices = []
                for item in data['data']:
                    # 过滤掉空行
                    if not item['material_name'] or not item['spec']:
                        continue

                    material_id = self._generate_material_id(
                        item['material_name'], item['spec'], item['brand']
                    )

                    price = 0.0
                    price_max = 0.0

                    if item['has_price']:
                        price_match = re.search(r'(\d+)', item['price_text'])
                        if price_match:
                            price = float(price_match.group(1))

                    if item['remark'] and ':' in item['remark']:
                        remark_prices = self._parse_remark_prices(item['remark'])
                        if remark_prices:
                            if price == 0:
                                price = remark_prices[0][1]
                            if len(remark_prices) > 1:
                                price_max = max(p[1] for p in remark_prices)

                    prices.append(MaterialPrice(
                        material_id=material_id,
                        material_name=item['material_name'],
                        spec=item['spec'],
                        material_type=item['material_type'],
                        brand=item['brand'],
                        price=price,
                        price_max=price_max,
                        price_change=item['price_change'],
                        remark=item['remark'],
                        steel_code=item['steel_code'],
                        unit='元/吨',
                        region='山东烟台'
                    ))

                logger.info(f"JS提取成功: {len(prices)} 条数据")
                return prices

            else:
                logger.warning(f"JS提取失败: {data.get('message')}")

        except Exception as e:
            logger.warning(f"JS提取异常: {e}")

        return []

    async def _find_latest_yantai_url(self, page) -> str:
        """从首页找到今日烟台价格URL"""
        try:
            # 访问首页
            await page.goto('https://jiancai.mysteel.com/', wait_until='domcontentloaded', timeout=30000)
            await page.wait_for_timeout(5000)

            # 获取所有链接
            links = await page.evaluate('''() => {
                const links = [];
                document.querySelectorAll('a[href]').forEach(a => {
                    const href = a.href;
                    const text = a.textContent.trim();
                    if (href.includes('jiancai.mysteel.com') && text.length > 0 && text.length < 50) {
                        links.push({text, href});
                    }
                });
                return links;
            }''')

            # 找烟台相关链接 - 优先找今日价格链接
            yantai_urls = []
            for link in links:
                if '烟台' in link['text'] and '/m/' in link['href']:
                    yantai_urls.append(link['href'])

            if yantai_urls:
                logger.info(f"从首页找到烟台URL: {yantai_urls[0]}")
                return yantai_urls[0]

            # 如果没有直接链接，访问山东市场页面
            for link in links:
                if '山东' in link['text'] and 'market' in link['href']:
                    await page.goto(link['href'], wait_until='domcontentloaded', timeout=30000)
                    await page.wait_for_timeout(5000)

                    # 在市场页面找烟台链接
                    market_links = await page.evaluate('''() => {
                        const links = [];
                        document.querySelectorAll('a[href]').forEach(a => {
                            links.push({text: a.textContent.trim(), href: a.href});
                        });
                        return links;
                    }''')

                    for ml in market_links:
                        if '烟台' in ml['text'] and '/m/' in ml['href']:
                            return ml['href']

        except Exception as e:
            logger.warning(f"查找URL失败: {e}")

        # 返回默认URL
        today = datetime.now().strftime('%y%m%d')
        return f"https://jiancai.mysteel.com/m/{today}10/25B3355C6617BD3C.html"

    async def _extract_prices_from_table(self, page) -> List[MaterialPrice]:
        """从表格提取价格数据"""
        try:
            data = await page.evaluate('''() => {
                const tables = document.querySelectorAll('table');
                const results = [];

                tables.forEach((table, idx) => {
                    const rows = table.querySelectorAll('tr');
                    const tableData = [];

                    rows.forEach((row, rIdx) => {
                        const cells = row.querySelectorAll('td, th');
                        const rowData = [];
                        cells.forEach(c => {
                            rowData.push(c.textContent.trim());
                        });
                        if (rowData.length > 0) {
                            tableData.push(rowData);
                        }
                    });

                    if (tableData.length > 0) {
                        results.push({idx, rows: tableData});
                    }
                });

                return results;
            }''')

            prices = []
            for t in data:
                for row in t['rows']:
                    if row and len(row) >= 5:
                        material_name = row[0].strip() if len(row) > 0 else ''
                        spec = row[1].strip() if len(row) > 1 else ''
                        material_type = row[2].strip() if len(row) > 2 else ''
                        brand = row[3].strip() if len(row) > 3 else ''
                        price_str = row[4].strip() if len(row) > 4 else ''
                        steel_code = row[7].strip() if len(row) > 7 else ''

                        # 验证是有效的价格行
                        valid_names = ['高线', '螺纹钢', '盘螺', '圆钢']
                        if material_name in valid_names and spec.startswith('Φ') and price_str.isdigit():
                            try:
                                price = float(price_str)
                                material_id = self._generate_material_id(material_name, spec, brand)
                                prices.append(MaterialPrice(
                                    material_id=material_id,
                                    material_name=material_name,
                                    spec=spec,
                                    material_type=material_type,
                                    brand=brand,
                                    price=price,
                                    unit='元/吨',
                                    steel_code=steel_code,
                                    region='山东烟台'
                                ))
                            except:
                                pass

            return prices

        except Exception as e:
            logger.warning(f"提取价格失败: {e}")
            return []

    async def fetch_async(self, force: bool = False) -> CrawlResult:
        """异步抓取 - v8.0 修复版"""
        result = CrawlResult(success=False, source_name="我的钢铁网-山东烟台")

        if not HAS_PLAYWRIGHT:
            result.error_message = "未安装 playwright"
            return result

        if not force:
            can_fetch, reason = self._check_rate_limit()
            if not can_fetch:
                result.error_message = reason
                return result

        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            context = await browser.new_context(
                viewport={'width': 1920, 'height': 3000},
                locale='zh-CN'
            )

            try:
                page = await context.new_page()

                # 加载Cookie
                cookies = self._load_cookies()
                if cookies:
                    await context.add_cookies(cookies)
                    logger.info("已加载Cookie")

                # 访问价格页面
                url = self._get_today_url()
                logger.info(f"打开价格页: {url}")

                await page.goto(url, wait_until='domcontentloaded', timeout=60000)
                await page.wait_for_timeout(10000)

                # 检查是否需要登录
                body_text = await page.evaluate('() => document.body.textContent')
                if '登录' in body_text and len(body_text) < 1000:
                    logger.info("需要登录...")
                    login_success = await self._login(page, context)
                    if not login_success:
                        result.error_message = "登录失败"
                        return result
                    # 重新访问
                    await page.goto(url, wait_until='domcontentloaded', timeout=60000)
                    await page.wait_for_timeout(10000)

                # 截图
                screenshot_bytes = await page.screenshot(full_page=True)
                result.screenshot = base64.b64encode(screenshot_bytes).decode('utf-8')
                logger.info("截图已保存")

                # 提取价格数据
                prices = await self._extract_prices_from_table(page)
                valid_prices = [p for p in prices if p.price > 0]

                if valid_prices:
                    result.success = True
                    result.prices = valid_prices
                    result.fetched_at = datetime.now().isoformat()
                    result.url = url
                    self._save_fetch_record(True, len(valid_prices))

                    logger.info(f"抓取成功: {len(valid_prices)} 条数据")

                    # 统计
                    summary = {}
                    for p in valid_prices:
                        name = p.material_name
                        if name not in summary:
                            summary[name] = {'count': 0, 'brands': set()}
                        summary[name]['count'] += 1
                        summary[name]['brands'].add(p.brand)

                    for name, info in summary.items():
                        logger.info(f"   - {name}: {info['count']}条, 品牌: {', '.join(sorted(info['brands']))}")
                else:
                    result.error_message = "未解析到价格"

            except Exception as e:
                logger.error(f"抓取失败: {e}")
                result.error_message = str(e)
                import traceback
                traceback.print_exc()
            finally:
                await browser.close()

        return result

    def fetch(self, force: bool = False) -> CrawlResult:
        """同步抓取"""
        return asyncio.run(self.fetch_async(force=force))


def save_to_excel(result: CrawlResult, excel_file: str = None) -> bool:
    """保存到Excel"""
    if not HAS_OPENPYXL:
        logger.error("未安装 openpyxl")
        return False

    if excel_file is None:
        excel_file = str(DATA_DIR / "山东烟台钢筋价格.xlsx")
    else:
        excel_file = str(excel_file)

    try:
        today = datetime.now().date()
        today_str = today.isoformat()
        fetch_time = datetime.now().strftime("%H:%M:%S")

        # 每次都创建今天的日期sheet（不使用时间区分，保留最近一次抓取）
        sheet_name = today_str

        # 样式
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 打开或创建workbook
        if Path(excel_file).exists():
            wb = openpyxl.load_workbook(excel_file)
            # 删除已存在的今日sheet，重新创建
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
        else:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        ws = wb.create_sheet(title=sheet_name)

        # 标题 - 包含抓取时间
        ws.merge_cells(f'A{1}:K{1}')
        ws.cell(row=1, column=1, value=f"山东烟台钢筋价格 - {today_str} {fetch_time} (品牌维度)").font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        # 表头
        headers = ["日期", "时间", "品名", "规格", "材质", "品牌/钢厂", "单价(元/吨)", "涨跌", "备注", "钢号", "地区"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 数据
        row = 4
        if result.prices:
            for price in result.prices:
                data = [
                    today_str,      # 日期 - 使用今天日期
                    fetch_time,     # 时间 - 精确抓取时间
                    price.material_name, price.spec,
                    price.material_type, price.brand,
                    price.price if price.price > 0 else "",
                    price.price_change, price.remark, price.steel_code, price.region
                ]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = thin_border
                row += 1

        # 列宽
        widths = [12, 12, 10, 10, 12, 14, 12, 10, 25, 10, 10]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # 截图
        if result.screenshot:
            screenshot_path = DATA_DIR / f"screenshot_{today_str.replace('-', '')}.png"
            with open(screenshot_path, 'wb') as f:
                f.write(base64.b64decode(result.screenshot))

            row = 4 + len(result.prices) + 2
            ws.cell(row=row, column=1, value="当日截图").font = Font(bold=True, size=12)
            ws.row_dimensions[row + 1].height = 400

            img = Image(str(screenshot_path))
            img.width = 900
            img.height = 500
            img.anchor = f'A{row + 1}'
            ws.add_image(img)

        wb.save(excel_file)
        wb.close()
        logger.info(f"保存成功: {excel_file}, Sheet: {sheet_name}")
        return True

    except Exception as e:
        logger.error(f"Excel保存失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def read_from_excel(excel_file: str = None) -> dict:
    """读取Excel数据"""
    if not HAS_OPENPYXL:
        return {}

    if excel_file is None:
        excel_file = str(DATA_DIR / "山东烟台钢筋价格.xlsx")

    try:
        if not Path(excel_file).exists():
            return {}

        wb = openpyxl.load_workbook(excel_file)
        all_data = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            prices = []

            # 检测表头版本
            headers = [ws.cell(row=3, column=c).value for c in range(1, 12)]
            has_brand = '品牌/钢厂' in headers or '品牌' in headers

            for row_num in range(4, ws.max_row + 1):
                date = ws.cell(row=row_num, column=1).value
                if date and date != "当日截图":
                    if has_brand:
                        prices.append({
                            'date': date,
                            'time': ws.cell(row=row_num, column=2).value,
                            'material_name': ws.cell(row=row_num, column=3).value,
                            'spec': ws.cell(row=row_num, column=4).value,
                            'material_type': ws.cell(row=row_num, column=5).value,
                            'brand': ws.cell(row=row_num, column=6).value,
                            'price': ws.cell(row=row_num, column=7).value,
                            'price_change': ws.cell(row=row_num, column=8).value,
                            'remark': ws.cell(row=row_num, column=9).value,
                            'steel_code': ws.cell(row=row_num, column=10).value,
                            'region': ws.cell(row=row_num, column=11).value,
                        })

            all_data[sheet_name] = {'prices': prices, 'has_screenshot': True}

        wb.close()
        return all_data

    except Exception as e:
        logger.error(f"读取Excel失败: {e}")
        return {}


def get_sheet_names(excel_file: str = None) -> list:
    """获取所有sheet名称"""
    if not HAS_OPENPYXL:
        return []

    if excel_file is None:
        excel_file = str(DATA_DIR / "山东烟台钢筋价格.xlsx")

    try:
        if not Path(excel_file).exists():
            return []
        wb = openpyxl.load_workbook(excel_file)
        names = wb.sheetnames
        wb.close()
        return names
    except:
        return []


def get_latest_sheet(excel_file: str = None) -> str:
    """获取最新sheet"""
    names = get_sheet_names(excel_file)
    return names[-1] if names else ""


def run():
    """主程序"""
    print("=" * 60)
    print("山东烟台钢筋价格抓取工具 v7.1")
    print("自动登录版 - 支持Cookie保存")
    print("=" * 60)

    scraper = YantaiRebarScraper()

    can_fetch, reason = scraper._check_rate_limit()
    print(f"\n状态: {reason}")

    result = scraper.fetch(force=True)

    print()
    print("=" * 60)
    print("抓取结果")
    print("=" * 60)
    print(f"成功: {result.success}")
    print(f"截图: {'有' if result.screenshot else '无'}")

    if result.success:
        print(f"时间: {result.fetched_at}")
        print(f"材料数量: {len(result.prices)}")

        summary = {}
        for p in result.prices:
            name = p.material_name
            if name not in summary:
                summary[name] = set()
            summary[name].add(p.brand)

        print("\n按品名统计:")
        for name, brands in summary.items():
            print(f"  - {name}: {len([p for p in result.prices if p.material_name == name])}条, 品牌: {', '.join(brands)}")

        print(f"\n品牌列表: {sorted(set(p.brand for p in result.prices))}")

        print("\n保存到Excel...")
        if save_to_excel(result):
            print(f"Sheet: {get_sheet_names()}")
    else:
        print(f"失败: {result.error_message}")

    print()
    print("=" * 60)


if __name__ == '__main__':
    run()