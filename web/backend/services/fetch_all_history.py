"""
使用历史URL抓取所有缺失的价格数据
URL格式: https://jiancai.mysteel.com/m/YYMMDDHH/XXXXXXXXX.html
"""
import asyncio
import sys
import json
import base64
from pathlib import Path
from datetime import datetime

sys.path.insert(0, '.')

from playwright.async_api import async_playwright
import openpyxl
import sqlite3

DATA_DIR = Path('services/data')
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格_历史抓取.xlsx'
DB_FILE = DATA_DIR / 'yantai_rebar.db'
URL_FILE = DATA_DIR / 'all_urls_2024_2026.json'
COOKIE_FILE = DATA_DIR / 'myst_cookies.json'
PROGRESS_FILE = DATA_DIR / 'fetch_progress.json'


def get_existing_records():
    """获取数据库中已有的日期和记录数"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT date, COUNT(*) FROM rebar_prices GROUP BY date')
    existing = {row[0]: row[1] for row in c.fetchall()}
    conn.close()
    return existing


def save_progress(processed, failed):
    """保存进度"""
    with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
        json.dump({
            'processed': processed,
            'failed': failed,
            'last_update': datetime.now().isoformat()
        }, f, ensure_ascii=False, indent=2)


def load_progress():
    """加载进度"""
    if PROGRESS_FILE.exists():
        with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'processed': [], 'failed': []}


async def fetch_url(page, url, date, period):
    """抓取单个URL的数据"""
    try:
        await page.goto(url, wait_until='domcontentloaded', timeout=45000)
        await page.wait_for_timeout(5000)

        # 截图
        screenshot = await page.screenshot(full_page=True)

        # 提取数据
        data = await page.evaluate('''() => {
            const tables = document.querySelectorAll('table');
            const results = [];
            tables.forEach((table) => {
                const rows = table.querySelectorAll('tr');
                rows.forEach((row) => {
                    const cells = row.querySelectorAll('td, th');
                    const rowData = [];
                    cells.forEach(c => rowData.push(c.textContent.trim()));
                    if (rowData.length > 0) results.push(rowData);
                });
            });
            return results;
        }''')

        prices = []
        for t in data:
            for row in (t.get('rows', []) if isinstance(t, dict) else t):
                if row and len(row) >= 5:
                    material_name = row[0].strip()
                    spec = row[1].strip()
                    material_type = row[2].strip()
                    brand = row[3].strip()
                    price_str = row[4].strip()

                    valid_names = ['高线', '螺纹钢', '盘螺', '圆钢']
                    if material_name in valid_names and spec.startswith('Φ') and price_str.isdigit():
                        prices.append({
                            'material_name': material_name,
                            'spec': spec,
                            'material_type': material_type,
                            'brand': brand,
                            'price': float(price_str)
                        })

        return {
            'success': len(prices) > 0,
            'date': date,
            'period': period,
            'url': url,
            'count': len(prices),
            'prices': prices,
            'screenshot': base64.b64encode(screenshot).decode('utf-8') if prices else None
        }

    except Exception as e:
        return {
            'success': False,
            'date': date,
            'period': period,
            'url': url,
            'error': str(e)
        }


def save_to_excel(result):
    """保存抓取结果到Excel"""
    if not result['success']:
        return

    if not EXCEL_FILE.exists():
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        wb.save(EXCEL_FILE)

    wb = openpyxl.load_workbook(EXCEL_FILE)

    date = result['date']
    period = result['period']
    fetch_time = datetime.now().strftime('%H%M%S')
    sheet_name = f'{date}_{period}_{fetch_time}'

    # 跳过已存在的sheet
    if sheet_name in wb.sheetnames:
        wb.close()
        return

    ws = wb.create_sheet(title=sheet_name)

    # 标题
    ws.merge_cells('A1:K1')
    period_text = '上午' if period == 'AM' else '下午'
    ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date} {period_text}')

    # 表头
    headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)

    # 数据
    fetch_time_str = datetime.now().strftime('%H:%M:%S')
    for i, price in enumerate(result['prices']):
        row = 4 + i
        ws.cell(row=row, column=1, value=date)
        ws.cell(row=row, column=2, value=fetch_time_str)
        ws.cell(row=row, column=3, value=price['material_name'])
        ws.cell(row=row, column=4, value=price['spec'])
        ws.cell(row=row, column=5, value=price['material_type'])
        ws.cell(row=row, column=6, value=price['brand'])
        ws.cell(row=row, column=7, value=price['price'])

    # 嵌入截图
    if result.get('screenshot'):
        date_str = date.replace("-", "")
        screenshot_path = DATA_DIR / f'screenshot_{date_str}_{period}.png'
        with open(screenshot_path, 'wb') as f:
            f.write(base64.b64decode(result['screenshot']))
        ws.cell(row=4 + len(result['prices']) + 2, column=1, value='截图已保存')

    wb.save(EXCEL_FILE)
    wb.close()


def save_to_database(result):
    """保存到SQLite数据库"""
    if not result['success'] or not result['prices']:
        return 0

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    inserted = 0
    for price in result['prices']:
        try:
            c.execute('''
                INSERT INTO rebar_prices
                (date, fetch_time, material_name, spec, material_type, brand, price, region)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                result['date'],
                result['period'],
                price['material_name'],
                price['spec'],
                price['material_type'],
                price['brand'],
                int(price['price']),
                '山东烟台'
            ))
            inserted += 1
        except sqlite3.IntegrityError:
            pass
        except Exception as e:
            print(f'  插入错误: {e}')

    conn.commit()
    conn.close()
    return inserted


async def main():
    """主函数"""
    print('=' * 60)
    print('历史价格抓取脚本')
    print('=' * 60)

    # 加载URL列表
    with open(URL_FILE, 'r', encoding='utf-8') as f:
        url_list = json.load(f)

    print(f'总共 {len(url_list)} 个URL')

    # 获取已处理的记录
    progress = load_progress()
    processed_urls = set(progress.get('processed', []))
    failed_urls = set(progress.get('failed', []))

    # 过滤掉已处理的
    urls_to_fetch = [u for u in url_list if u[2] not in processed_urls and u[2] not in failed_urls]
    print(f'待抓取: {len(urls_to_fetch)} 个')

    if not urls_to_fetch:
        print('所有URL已处理完成！')
        return

    # 加载凭据
    with open(DATA_DIR / 'mysteel_config.json', 'r', encoding='utf-8') as f:
        config = json.load(f)
        username = config.get('username')
        password = config.get('password')

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(viewport={'width': 1920, 'height': 3000}, locale='zh-CN')
        page = await context.new_page()

        # 登录
        print('\n1. 登录中...')
        await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
        await page.wait_for_timeout(5000)

        try:
            account_tab = await page.query_selector('.form-tab-account, a[data-tab="account"]')
            if account_tab:
                await account_tab.click()
                await page.wait_for_timeout(2000)
        except: pass

        await page.evaluate(f'''() => {{
            const inputs = document.querySelectorAll('input');
            for (const inp of inputs) {{
                const ph = inp.placeholder || '';
                if (ph.includes('用户名')) inp.value = '{username}';
                if (ph.includes('密码') && inp.type === 'password') inp.value = '{password}';
            }}
        }}''')

        await page.wait_for_timeout(500)

        try:
            login_btn = await page.query_selector('.form-button-login, button:has-text("登录")')
            if login_btn:
                await login_btn.click()
        except: pass

        await page.wait_for_timeout(8000)
        print('  登录完成')

        # 保存cookie
        cookies = await context.cookies()
        with open(COOKIE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cookies, f, ensure_ascii=False)

        # 批量抓取
        total_success = 0
        total_inserted = 0

        for i, item in enumerate(urls_to_fetch):
            date, period, url = item[0], item[1], item[2]

            print(f'\n[{i+1}/{len(urls_to_fetch)}] {date} {period}...')

            result = await fetch_url(page, url, date, period)

            if result['success']:
                print(f'  [OK] 提取 {result["count"]} 条数据')

                # 保存到Excel
                save_to_excel(result)

                # 保存到数据库
                inserted = save_to_database(result)
                total_inserted += inserted
                print(f'  [DB] 插入 {inserted} 条')

                total_success += 1
                processed_urls.add(url)
            else:
                print(f'  [FAIL] {result.get("error", "未知错误")}')
                failed_urls.add(url)

            # 保存进度
            if (i + 1) % 10 == 0:
                save_progress(list(processed_urls), list(failed_urls))
                print(f'  进度已保存: {i+1}/{len(urls_to_fetch)}')

            # 短暂延迟避免被封
            await page.wait_for_timeout(1000)

        await browser.close()

    # 最终保存进度
    save_progress(list(processed_urls), list(failed_urls))

    print('\n' + '=' * 60)
    print('抓取完成')
    print(f'  成功: {total_success} 个URL')
    print(f'  新增记录: {total_inserted} 条')
    print(f'  进度已保存')
    print('=' * 60)


if __name__ == '__main__':
    asyncio.run(main())