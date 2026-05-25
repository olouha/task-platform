"""
每日烟台钢筋价格抓取脚本（智能版）
- 每次抓取前检查登录状态
- 自动登录
- 智能去重
- AM/PM时段标记
"""

import asyncio
import json
import hashlib
from pathlib import Path
from datetime import datetime, time, date
from playwright.async_api import async_playwright
import openpyxl

# 配置
DATA_DIR = Path(__file__).parent / 'data'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
COOKIE_FILE = DATA_DIR / 'mysteel_cookies.json'
TODAY = date.today().isoformat()
TIME_NOW = datetime.now().strftime('%H%M%S')

# 时段判断
CURRENT_HOUR = datetime.now().hour
IS_AM = CURRENT_HOUR < 14
PERIOD = 'AM' if IS_AM else 'PM'
PERIOD_LABEL = '上午' if IS_AM else '下午(较晚)'


class YantaiPriceScraper:
    """烟台钢筋价格抓取器"""

    def __init__(self):
        self.cookie = None
        self.browser = None
        self.context = None
        self.page = None
        self.playwright = None

    async def __aenter__(self):
        self.playwright = await async_playwright().start()
        await self.start()
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.stop()
        await self.playwright.stop()

    async def start(self):
        """启动浏览器"""
        self.browser = await self.playwright.chromium.launch(
            headless=True,
            args=['--no-sandbox', '--disable-setuid-sandbox']
        )
        self.context = await self.browser.new_context(
            viewport={'width': 1920, 'height': 3000},
            locale='zh-CN',
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        )
        self.page = await self.context.new_page()

        # 尝试加载Cookie
        if COOKIE_FILE.exists():
            with open(COOKIE_FILE) as f:
                cookies = json.load(f)
            await self.context.add_cookies(cookies)
            print(f'[OK] 已加载 {len(cookies)} 个Cookie')

    async def stop(self):
        """停止浏览器"""
        if self.browser:
            await self.browser.close()

    async def check_login_status(self) -> bool:
        """检查登录状态"""
        test_url = 'https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html'
        await self.page.goto(test_url, wait_until='domcontentloaded', timeout=30000)
        await self.page.wait_for_timeout(3000)

        # 检查是否跳转到验证码或登录页
        current_url = self.page.url
        if 'captcha' in current_url:
            print('[WARN] 检测到验证码页面')
            return False
        if 'passport' in current_url or 'login' in current_url:
            print('[WARN] 检测到登录页面')
            return False

        # 检查页面内容
        body_text = await self.page.evaluate('document.body.textContent')
        if len(body_text) < 100:
            print('[WARN] 页面内容异常')
            return False

        print('[OK] 登录状态正常')
        return True

    async def login(self):
        """自动登录"""
        from config.mysteel import MYSTEEL_USERNAME, MYSTEEL_PASSWORD

        print('[LOGIN] 开始登录...')
        await self.page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
        await self.page.wait_for_timeout(3000)

        # 填写用户名和密码
        await self.page.evaluate(f'''
            const fillForm = () => {{
                const inputs = document.querySelectorAll('input');
                for (const inp of inputs) {{
                    const ph = inp.placeholder || "";
                    if (ph.includes("用户名") || ph.includes("账号") || ph.includes("手机号")) {{
                        inp.value = "{MYSTEEL_USERNAME}";
                    }}
                    if (ph.includes("密码") && inp.type === "password") {{
                        inp.value = "{MYSTEEL_PASSWORD}";
                    }}
                }}
            }};
            fillForm();
        ''')

        await self.page.wait_for_timeout(1000)

        # 尝试多种方式点击登录
        try:
            # 方式1: 按文本查找
            login_btn = await self.page.query_selector('button:has-text("登录")')
            if login_btn:
                await login_btn.click()
            else:
                # 方式2: 按类名查找
                login_btn = await self.page.query_selector('.form-button-login, .login-btn, [type="submit"]')
                if login_btn:
                    await login_btn.click()
        except Exception as e:
            print(f'点击登录按钮异常: {e}')

        await self.page.wait_for_timeout(10000)

        # 保存Cookie
        cookies = await self.context.cookies()
        with open(COOKIE_FILE, 'w') as f:
            json.dump(cookies, f, ensure_ascii=False)
        print(f'[OK] 已保存 {len(cookies)} 个Cookie')

        # 验证登录
        await self.page.goto('https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html', wait_until='domcontentloaded', timeout=60000)
        await self.page.wait_for_timeout(5000)

        current_url = self.page.url
        if 'captcha' in current_url:
            print('[FAIL] 登录失败：需要验证码')
            return False

        print('[OK] 登录成功')
        return True

    async def fetch_prices(self) -> list:
        """获取今日价格数据"""
        url = 'https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html'
        print(f'[DATA] 获取价格数据: {url}')

        await self.page.goto(url, wait_until='domcontentloaded', timeout=60000)
        await self.page.wait_for_timeout(5000)

        # 获取数据
        data = await self.page.evaluate('''() => {
            const results = [];
            const tables = document.querySelectorAll('table');
            tables.forEach(table => {
                const rows = table.querySelectorAll('tr');
                rows.forEach(row => {
                    const cells = row.querySelectorAll('td');
                    if (cells.length >= 5) {
                        const material_name = cells[0]?.textContent?.trim();
                        const spec = cells[1]?.textContent?.trim();
                        const material_type = cells[2]?.textContent?.trim();
                        const brand = cells[3]?.textContent?.trim();
                        const price_str = cells[4]?.textContent?.trim();

                        if (['高线', '螺纹钢', '盘螺', '圆钢'].includes(material_name) &&
                            spec && spec.startsWith('Φ') &&
                            price_str && /^\\d+$/.test(price_str)) {
                            results.push({
                                material_name,
                                spec,
                                material_type,
                                brand,
                                price: parseInt(price_str),
                                region: '山东烟台'
                            });
                        }
                    }
                });
            });
            return results;
        }''')

        print(f'[OK] 获取到 {len(data)} 条价格数据')
        return data

    def calculate_data_hash(self, data: list) -> str:
        """计算数据哈希（用于去重）"""
        # 按品名+规格+品牌+价格排序后计算哈希
        sorted_data = sorted([
            (d['material_name'], d['spec'], d['brand'], d['price'])
            for d in data
        ])
        data_str = json.dumps(sorted_data, ensure_ascii=False)
        return hashlib.md5(data_str.encode()).hexdigest()

    def check_today_data_exists(self, data_hash: str) -> tuple:
        """检查今日数据是否已存在
        返回: (是否存在, 时段)
        """
        if not EXCEL_FILE.exists():
            return False, None

        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sheet_names = wb.sheetnames
            wb.close()

            # 检查今日的AM和PM数据
            am_exists = any(f'{TODAY}_AM_' in s for s in sheet_names)
            pm_exists = any(f'{TODAY}_PM_' in s for s in sheet_names)

            if PERIOD == 'AM' and am_exists:
                return True, 'AM'
            if PERIOD == 'PM' and pm_exists:
                return True, 'PM'

            # 检查之前保存的数据哈希
            hash_file = DATA_DIR / 'data_hashes.json'
            if hash_file.exists():
                with open(hash_file) as f:
                    hashes = json.load(f)
                    key = f'{TODAY}_{PERIOD}'
                    if key in hashes and hashes[key] == data_hash:
                        return True, PERIOD

            return False, None
        except Exception as e:
            print(f'检查数据失败: {e}')
            return False, None

    def save_to_excel(self, data: list, data_hash: str) -> bool:
        """保存数据到Excel"""
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE) if EXCEL_FILE.exists() else openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

            sheet_name = f'{TODAY}_{PERIOD}_{TIME_NOW}'
            ws = wb.create_sheet(title=sheet_name)

            # 标题
            period_text = '下午(较晚)' if PERIOD == 'PM' else '上午'
            ws.merge_cells('A1:K1')
            ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {TODAY} {period_text}').font = openpyxl.styles.Font(bold=True, size=14)
            ws.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')

            # 表头
            from openpyxl.styles import PatternFill
            header_fill = PatternFill(start_color='FF6B6B' if PERIOD == 'PM' else '4472C4', end_color='FF6B6B' if PERIOD == 'PM' else '4472C4', fill_type='solid')
            headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True, size=12, color='FFFFFF')
                cell.fill = header_fill
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # 数据
            fetch_time = datetime.now().strftime('%H:%M:%S')
            for i, item in enumerate(data):
                row = 4 + i
                for col, val in enumerate([
                    TODAY,
                    fetch_time,
                    item['material_name'],
                    item['spec'],
                    item['material_type'],
                    item['brand'],
                    item['price'],
                    '',
                    '',
                    '',
                    '山东烟台'
                ], 1):
                    ws.cell(row=row, column=col, value=val)

            wb.save(EXCEL_FILE)
            wb.close()

            # 保存哈希
            hash_file = DATA_DIR / 'data_hashes.json'
            hashes = {}
            if hash_file.exists():
                with open(hash_file) as f:
                    hashes = json.load(f)
            hashes[f'{TODAY}_{PERIOD}'] = data_hash
            with open(hash_file, 'w') as f:
                json.dump(hashes, f, ensure_ascii=False)

            print(f'[OK] 数据已保存: {sheet_name}')
            return True
        except Exception as e:
            print(f'[FAIL] 保存失败: {e}')
            import traceback
            traceback.print_exc()
            return False

    def get_summary(self) -> dict:
        """获取数据摘要"""
        if not EXCEL_FILE.exists():
            return {'total_sheets': 0, 'dates': []}

        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sheet_names = wb.sheetnames
            wb.close()

            # 提取日期
            dates = []
            for name in sheet_names:
                if name.startswith('20'):
                    date_part = name.split('_')[0]
                    if date_part not in dates:
                        dates.append(date_part)

            return {
                'total_sheets': len(sheet_names),
                'dates': sorted(dates, reverse=True),
                'latest_date': dates[0] if dates else None
            }
        except Exception as e:
            print(f'获取摘要失败: {e}')
            return {'total_sheets': 0, 'dates': []}


async def main():
    """主函数"""
    print('=' * 60)
    print(f'[START] 烟台钢筋价格每日抓取 - {TODAY} {PERIOD_LABEL}')
    print('=' * 60)

    async with YantaiPriceScraper() as scraper:
        # Step 1: 检查登录状态
        is_logged_in = await scraper.check_login_status()

        # Step 2: 如果未登录，执行登录
        if not is_logged_in:
            login_success = await scraper.login()
            if not login_success:
                print('[FAIL] 登录失败，请手动登录后重试')
                return

        # Step 3: 获取价格数据
        data = await scraper.fetch_prices()

        if not data:
            print('[WARN] 未获取到价格数据')
            return

        # Step 4: 计算数据哈希
        data_hash = scraper.calculate_data_hash(data)

        # Step 5: 检查今日数据是否已存在
        exists, existing_period = scraper.check_today_data_exists(data_hash)

        if exists:
            print(f'[SKIP] 今日数据已存在 ({TODAY} {existing_period})，跳过保存')
            # 检查是否需要更新PM数据
            if PERIOD == 'PM' and existing_period == 'AM':
                # 下午可能有新数据，即使哈希相同也需要保存
                # 这里可以选择覆盖或跳过
                print('[INFO] PM时段数据与AM相同，不覆盖')
            return

        # Step 6: 保存数据
        success = scraper.save_to_excel(data, data_hash)

        if success:
            summary = scraper.get_summary()
            print('=' * 60)
            print('[OK] 抓取完成')
            print(f'[DATA] 总Sheet数: {summary["total_sheets"]}')
            print(f'[DATE] 最新日期: {summary["latest_date"]}')
            print(f'[TIME] 本次时段: {PERIOD_LABEL}')
            print(f'[COUNT] 数据条数: {len(data)}')
            print('=' * 60)

            # 保存抓取记录
            record_file = DATA_DIR / 'fetch_records.json'
            records = []
            if record_file.exists():
                with open(record_file) as f:
                    records = json.load(f)
            records.append({
                'date': TODAY,
                'period': PERIOD,
                'time': datetime.now().isoformat(),
                'count': len(data),
                'hash': data_hash
            })
            with open(record_file, 'w') as f:
                json.dump(records, f, ensure_ascii=False)


if __name__ == '__main__':
    asyncio.run(main())