"""
严格整合本地数据 - 只包含有数据有截图的日期
不允许编造，截图与数据一一对应
"""
import os
import json
from pathlib import Path
from datetime import datetime, timedelta
from PIL import Image

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage

DATA_DIR = Path('services/data')
BACKUP_DIR = DATA_DIR

def strict_integrate():
    """严格整合：只包含有数据+截图的日期"""
    print('=' * 70)
    print('严格整合本地数据 - 有数据有截图')
    print('=' * 70)
    print()
    print('原则：')
    print('  1. 只包含有Excel数据的日期')
    print('  2. 只包含有对应截图的日期')
    print('  3. 不编造任何数据')
    print('  4. 截图与数据一一对应')
    print()

    # ========== 步骤1: 扫描所有截图 ==========
    print('步骤1: 扫描截图文件...')
    screenshots = {}  # {(date_str, period): file_path}
    for f in os.listdir(DATA_DIR):
        if f.startswith('screenshot_') and f.endswith('.png'):
            # 格式: screenshot_YYYYMMDD_AM.png 或 screenshot_YYYYMMDD_PM.png
            parts = f.replace('.png', '').split('_')
            if len(parts) >= 2:
                date_part = parts[1]
                if len(date_part) == 8 and date_part.isdigit():
                    year = date_part[:4]
                    month = date_part[4:6]
                    day = date_part[6:8]
                    date_str = f'{year}-{month}-{day}'

                    period = 'AM'
                    if len(parts) >= 3 and parts[2] == 'PM':
                        period = 'PM'

                    key = (date_str, period)
                    screenshots[key] = DATA_DIR / f

    print(f'  找到 {len(screenshots)} 个截图文件')

    # ========== 步骤2: 扫描Excel数据 ==========
    print('\n步骤2: 扫描Excel数据...')

    # 找到最大的备份文件
    backup_files = [f for f in os.listdir(BACKUP_DIR) if f.startswith('山东烟台钢筋价格_backup') and f.endswith('.xlsx')]
    backup_files.sort(key=lambda x: os.path.getsize(os.path.join(BACKUP_DIR, x)), reverse=True)

    if not backup_files:
        print('  错误: 未找到备份文件')
        return

    backup_file = os.path.join(BACKUP_DIR, backup_files[0])
    print(f'  使用备份: {backup_files[0]}')

    wb = openpyxl.load_workbook(backup_file, read_only=True)

    excel_data = {}  # {(date_str, period): [row_data]}
    for sheet_name in wb.sheetnames:
        # 从sheet名提取日期和时段
        if not sheet_name.startswith('20'):
            continue

        # 解析sheet名
        date_str = sheet_name[:10]
        period = 'AM'
        if '_PM' in sheet_name:
            period = 'PM'

        key = (date_str, period)

        # 提取数据行
        ws = wb[sheet_name]
        data_rows = []
        for row in ws.iter_rows(values_only=True):
            if row and row[0] and isinstance(row[0], str):
                if row[0].startswith('20') and len(row) >= 6:
                    # 数据行
                    data_rows.append(list(row))

        if data_rows:
            excel_data[key] = data_rows

    wb.close()
    print(f'  找到 {len(excel_data)} 个数据条目')

    # ========== 步骤3: 找出有数据+截图的日期 ==========
    print('\n步骤3: 匹配数据与截图...')

    matched = []  # [(date_str, period, data, screenshot_path), ...]
    missing_screenshot = []
    missing_data = []

    for key in sorted(excel_data.keys()):
        date_str, period = key
        if key in screenshots:
            matched.append((date_str, period, excel_data[key], screenshots[key]))
        else:
            missing_screenshot.append((date_str, period))

    print(f'  匹配成功: {len(matched)} 个')
    print(f'  有数据无截图: {len(missing_screenshot)} 个')

    # ========== 步骤4: 按日期排序 ==========
    matched.sort(key=lambda x: (x[0], x[1]))

    # 统计日期范围
    if matched:
        first_date = matched[0][0]
        last_date = matched[-1][0]
        print(f'  数据范围: {first_date} 至 {last_date}')

        # 统计每个日期的数据条数
        date_counts = {}
        for date_str, period, data, _ in matched:
            date_counts[date_str] = date_counts.get(date_str, 0) + len(data)

        total_prices = sum(date_counts.values())
        total_dates = len(date_counts)
        print(f'  总日期数: {total_dates}')
        print(f'  总价格记录: {total_prices}')

    # ========== 步骤5: 检查缺失的工作日 ==========
    print('\n步骤4: 检查工作日连续性...')

    start_date = datetime(2024, 1, 1)
    end_date = datetime(2026, 5, 27)

    # 收集所有有数据的日期
    available_dates = set()
    for date_str, period, _, _ in matched:
        available_dates.add(date_str)

    # 统计缺失
    missing_dates = []
    current = start_date
    while current <= end_date:
        if current.weekday() < 5:  # 工作日
            date_str = current.strftime('%Y-%m-%d')
            if date_str not in available_dates:
                missing_dates.append(date_str)
        current += timedelta(days=1)

    print(f'  目标工作日总数: {sum(1 for d in range((end_date - start_date).days + 1) if (start_date + timedelta(days=d)).weekday() < 5)}')
    print(f'  有数据的工作日: {len(available_dates)}')
    print(f'  缺失的工作日: {len(missing_dates)}')
    print(f'  覆盖率: {len(available_dates) / sum(1 for d in range((end_date - start_date).days + 1) if (start_date + timedelta(days=d)).weekday() < 5) * 100:.1f}%')

    # ========== 步骤6: 生成Excel ==========
    print('\n步骤5: 生成Excel文件...')

    new_wb = openpyxl.Workbook()
    if 'Sheet' in new_wb.sheetnames:
        del new_wb['Sheet']

    # 样式
    header_font = Font(bold=True, size=12, color='FFFFFF')
    am_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    pm_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    temp_files = []  # 跟踪临时文件

    for date_str, period, data_rows, screenshot_path in matched:
        # 生成sheet名（限制31字符）
        sheet_name = f'{date_str[:10]}_{period[:2]}'

        try:
            new_ws = new_wb.create_sheet(title=sheet_name)
        except:
            # 名称冲突，添加后缀
            sheet_name = f'{date_str[:8]}_{period[:2]}'
            new_ws = new_wb.create_sheet(title=sheet_name[:31])

        # 标题
        period_text = '下午' if period == 'PM' else '上午'
        new_ws.merge_cells('A1:K1')
        new_ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
        new_ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        # 表头
        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
        for col, header in enumerate(headers, 1):
            cell = new_ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = pm_fill if period == 'PM' else am_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 数据
        fetch_time = datetime.now().strftime('%H:%M:%S')
        for row_idx, row in enumerate(data_rows):
            row_num = 4 + row_idx
            for col, val in enumerate(row[:11], 1):
                if col == 1:
                    val = date_str  # 确保日期正确
                elif col == 2:
                    val = fetch_time  # 添加时间
                cell = new_ws.cell(row=row_num, column=col, value=val)
                cell.border = thin_border

        # 嵌入截图
        if screenshot_path and screenshot_path.exists():
            try:
                # 压缩并保存
                img = Image.open(screenshot_path)
                img = img.resize((900, 500), Image.LANCZOS)

                temp_path = DATA_DIR / f'temp_{date_str.replace("-", "")}_{period}.png'
                img.save(str(temp_path), 'PNG', quality=85)
                temp_files.append(temp_path)

                row = 4 + len(data_rows) + 2
                new_ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)

                xl_img = XLImage(str(temp_path))
                xl_img.width = 900
                xl_img.height = 500
                xl_img.anchor = f'A{row + 1}'
                new_ws.add_image(xl_img)
            except Exception as e:
                print(f'  截图处理失败: {date_str} {period} - {e}')

    # 保存
    output_file = DATA_DIR / '山东烟台钢筋价格_严格版.xlsx'
    new_wb.save(str(output_file))
    new_wb.close()

    # 清理临时文件
    for f in temp_files:
        try:
            os.remove(f)
        except:
            pass

    # ========== 步骤7: 生成报告 ==========
    print('\n' + '=' * 70)
    print('整合完成')
    print('=' * 70)
    print(f'\n输出文件: {output_file}')
    print(f'文件大小: {os.path.getsize(output_file) / 1024 / 1024:.2f} MB')
    print()
    print('数据统计:')
    print(f'  - 总日期数: {len(available_dates)}')
    print(f'  - 总价格记录: {total_prices}')
    print(f'  - 数据范围: {first_date} 至 {last_date}')
    print()
    print('覆盖率分析:')
    target_weekdays = sum(1 for d in range((end_date - start_date).days + 1) if (start_date + timedelta(days=d)).weekday() < 5)
    print(f'  - 目标工作日: {target_weekdays}')
    print(f'  - 实际覆盖: {len(available_dates)}')
    print(f'  - 缺失: {target_weekdays - len(available_dates)}')
    print(f'  - 覆盖率: {len(available_dates) / target_weekdays * 100:.1f}%')
    print()
    print('注意事项:')
    print('  - 不允许编造数据，只包含有原始数据的日期')
    print('  - 截图与数据一一对应')
    print('  - 2024年1月至2026年5月间缺失大量数据')

    # 保存缺失报告
    report_file = DATA_DIR / '数据覆盖报告.txt'
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write('山东烟台钢筋价格数据覆盖报告\n')
        f.write('=' * 50 + '\n\n')
        f.write(f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n\n')

        f.write('数据来源:\n')
        f.write(f'  - Excel备份: {backup_files[0]}\n')
        f.write(f'  - 截图文件: {len(screenshots)} 个\n\n')

        f.write('数据统计:\n')
        f.write(f'  - 有数据有截图的日期: {len(available_dates)} 个\n')
        f.write(f'  - 总价格记录: {total_prices} 条\n')
        f.write(f'  - 数据范围: {first_date} 至 {last_date}\n\n')

        f.write('覆盖率:\n')
        f.write(f'  - 目标工作日: {target_weekdays}\n')
        f.write(f'  - 实际覆盖: {len(available_dates)}\n')
        f.write(f'  - 覆盖率: {len(available_dates) / target_weekdays * 100:.1f}%\n\n')

        f.write('缺失日期列表 (工作日):\n')
        for date_str in missing_dates:
            f.write(f'  - {date_str}\n')

    print(f'\n覆盖报告: {report_file}')


if __name__ == '__main__':
    strict_integrate()