"""
将Excel数据导入SQLite数据库
"""
import sqlite3
import openpyxl
from datetime import datetime
from pathlib import Path

DATA_DIR = Path('services/data')
DB_FILE = DATA_DIR / 'yantai_rebar.db'

def init_database():
    """初始化数据库"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    # 创建价格表
    c.execute('''
        CREATE TABLE IF NOT EXISTS rebar_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            fetch_time TEXT,
            material_name TEXT,
            spec TEXT,
            material_type TEXT,
            brand TEXT,
            price INTEGER,
            price_change TEXT,
            remark TEXT,
            steel_code TEXT,
            region TEXT DEFAULT '山东烟台',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 创建索引
    c.execute('CREATE INDEX IF NOT EXISTS idx_date ON rebar_prices(date)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_material ON rebar_prices(material_name)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_spec ON rebar_prices(spec)')

    # 创建元数据表
    c.execute('''
        CREATE TABLE IF NOT EXISTS metadata (
            key TEXT PRIMARY KEY,
            value TEXT,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    conn.commit()
    return conn

def import_from_excel(conn, excel_file):
    """从Excel导入数据"""
    c = conn.cursor()

    if not excel_file.exists():
        print(f'文件不存在: {excel_file}')
        return 0

    print(f'导入: {excel_file.name}')

    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    imported = 0

    for sheet_name in wb.sheetnames:
        # 验证日期格式
        if '-' not in sheet_name or len(sheet_name) < 10:
            continue

        date = sheet_name.split('_')[0]
        try:
            datetime.strptime(date, '%Y-%m-%d')
        except ValueError:
            continue

        ws = wb[sheet_name]

        # 检查是否有数据
        if ws.max_row < 5 or ws.max_column < 8:
            continue

        # 读取数据
        for row_idx in range(2, ws.max_row + 1):  # 跳过标题行
            row_data = []
            for col_idx in range(1, min(12, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data.append(cell.value)

            # 跳过空行
            if not row_data[0] and not row_data[2]:
                continue

            # 提取数据
            material = row_data[2] if len(row_data) > 2 else None
            if not material or '品名' in str(material):
                continue

            spec = row_data[3] if len(row_data) > 3 else ''
            material_type = row_data[4] if len(row_data) > 4 else ''
            brand = row_data[5] if len(row_data) > 5 else ''
            price_text = row_data[6] if len(row_data) > 6 else ''

            # 解析价格
            price = 0
            if price_text:
                import re
                match = re.search(r'(\d+)', str(price_text))
                if match:
                    price = int(match.group(1))

            if price > 0:  # 只保存有效价格
                c.execute('''
                    INSERT INTO rebar_prices (date, material_name, spec, material_type, brand, price, region)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (date, material, spec, material_type, brand, price, '山东烟台'))
                imported += 1

        if imported % 1000 == 0 and imported > 0:
            conn.commit()
            print(f'  已导入 {imported} 条...')

    wb.close()
    conn.commit()
    return imported

def verify_data(conn):
    """验证数据"""
    c = conn.cursor()

    # 统计
    c.execute('SELECT COUNT(*) FROM rebar_prices')
    total = c.execute('SELECT COUNT(*) FROM rebar_prices').fetchone()[0]

    c.execute('SELECT COUNT(DISTINCT date) FROM rebar_prices')
    dates = c.execute('SELECT COUNT(DISTINCT date) FROM rebar_prices').fetchone()[0]

    c.execute('SELECT MIN(date), MAX(date) FROM rebar_prices')
    date_range = c.execute('SELECT MIN(date), MAX(date) FROM rebar_prices').fetchone()

    c.execute('SELECT material_name, COUNT(*) as cnt FROM rebar_prices GROUP BY material_name ORDER BY cnt DESC')
    materials = c.execute('SELECT material_name, COUNT(*) as cnt FROM rebar_prices GROUP BY material_name ORDER BY cnt DESC').fetchall()

    print('\n' + '=' * 60)
    print('数据库统计')
    print('=' * 60)
    print(f'总记录数: {total}')
    print(f'日期数: {dates}')
    print(f'日期范围: {date_range[0]} 到 {date_range[1]}')
    print(f'\n材料类型统计:')
    for mat, cnt in materials[:10]:
        print(f'  {mat}: {cnt} 条')

    return total, dates

def main():
    print('=' * 60)
    print('导入Excel数据到SQLite数据库')
    print('=' * 60)

    # 初始化数据库
    conn = init_database()
    print(f'数据库: {DB_FILE}')

    # 源文件
    source_files = [
        '山东烟台钢筋价格_完整版_数据+截图.xlsx',
        '山东烟台钢筋价格_最终版.xlsx',
    ]

    total_imported = 0

    for file_name in source_files:
        excel_file = DATA_DIR / file_name
        if excel_file.exists():
            count = import_from_excel(conn, excel_file)
            total_imported += count
            print(f'  {file_name}: +{count} 条')
        else:
            print(f'  {file_name}: 文件不存在')

    print(f'\n总导入: {total_imported} 条')

    # 验证
    verify_data(conn)

    conn.close()

    print(f'\n完成! 数据库位置: {DB_FILE}')

if __name__ == '__main__':
    main()