"""
合并我的钢铁网钢筋价格数据到单一Excel文件
每个日期只保留最后一张表（包含数据和截图）
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from datetime import datetime, timedelta
from pathlib import Path
import os

DATA_DIR = Path('services/data')
OUTPUT_FILE = DATA_DIR / '烟台钢筋价格_合并版.xlsx'

def log(msg):
    print(f'[{datetime.now().strftime("%H:%M:%S")}] {msg}')

def create_merged_excel():
    log('开始合并数据...')

    # 源文件 - 只使用完整版数据（格式一致）
    source_files = [
        ('山东烟台钢筋价格_完整版_数据+截图.xlsx', '完整版数据'),
    ]

    # 收集所有数据
    all_data = {}  # {date: {sheet_data, screenshot, rows, cols}}

    for file_name, file_type in source_files:
        file_path = DATA_DIR / file_name
        if not file_path.exists():
            log(f'文件不存在: {file_name}')
            continue

        log(f'处理: {file_name}')

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                # 提取日期
                if '-' in sheet_name:
                    parts = sheet_name.split('_')
                    date = parts[0]

                    # 验证日期格式
                    try:
                        datetime.strptime(date, '%Y-%m-%d')
                    except ValueError:
                        continue  # 跳过无效日期

                    # 打开sheet
                    ws = wb[sheet_name]
                    max_row = ws.max_row if ws.max_row else 0

                    # 检查是否已有数据（取数据最多的）
                    if date not in all_data or len(all_data[date].get('rows', [])) < max_row - 1:
                        if ws.max_row > 5 and ws.max_column >= 8:
                            # 收集所有行
                            data_rows = []
                            for row_idx in range(1, ws.max_row + 1):
                                row_data = []
                                for col_idx in range(1, min(12, ws.max_column + 1)):
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    row_data.append(cell.value)
                                data_rows.append(row_data)

                            all_data[date] = {
                                'sheet_name': date,
                                'rows': data_rows,
                                'max_row': ws.max_row,
                                'max_col': ws.max_column,
                                'source': file_type
                            }
                            log(f'  {date}: {ws.max_row}行 (来自 {file_type})')

            wb.close()
        except Exception as e:
            log(f'处理 {file_name} 失败: {e}')

    log(f'\n共收集 {len(all_data)} 个日期的数据')

    # 按日期排序
    sorted_dates = sorted(all_data.keys())

    # 检查连贯性
    log('\n检查日期连贯性...')
    missing_dates = []
    for i, date in enumerate(sorted_dates):
        if i > 0:
            try:
                prev_date = datetime.strptime(sorted_dates[i-1], '%Y-%m-%d')
                curr_date = datetime.strptime(date, '%Y-%m-%d')
                days_diff = (curr_date - prev_date).days
                if days_diff > 1:
                    # 有缺失日期（跳过节假日）
                    for d in range(1, days_diff):
                        missing_date = (prev_date + timedelta(days=d)).strftime('%Y-%m-%d')
                        missing_dates.append(missing_date)
            except ValueError:
                continue

    log(f'缺失日期数: {len(missing_dates)}')
    if missing_dates:
        log(f'缺失日期示例: {missing_dates[:10]}...')

    # 创建新的Excel文件
    log('\n创建合并后的Excel文件...')

    # 删除旧文件
    if OUTPUT_FILE.exists():
        os.remove(OUTPUT_FILE)

    wb_new = openpyxl.Workbook()
    # 删除默认sheet
    if 'Sheet' in wb_new.sheetnames:
        del wb_new['Sheet']

    # 表头样式
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    success_count = 0

    for date in sorted_dates:
        data = all_data[date]
        sheet_name = date[:31]  # Excel sheet名称限制31字符

        try:
            ws = wb_new.create_sheet(title=sheet_name)

            # 写入标题行
            ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date}')
            ws.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)

            # 写入表头
            headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 写入数据
            if data['rows']:
                for row_idx, row_data in enumerate(data['rows'][:100]):  # 最多100行数据
                    for col_idx, value in enumerate(row_data[:11]):
                        cell = ws.cell(row=row_idx + 4, column=col_idx + 1)
                        cell.value = value
                        cell.border = border

                        # 第一列日期格式化
                        if col_idx == 0 and value:
                            cell.alignment = Alignment(horizontal='center')

            wb_new.save(OUTPUT_FILE)
            success_count += 1

            if success_count % 100 == 0:
                log(f'  已处理 {success_count} 个日期')

        except Exception as e:
            log(f'  创建sheet失败 {date}: {e}')

    wb_new.close()

    log(f'\n合并完成!')
    log(f'文件: {OUTPUT_FILE}')
    log(f'日期数: {success_count}')
    log(f'文件大小: {os.path.getsize(OUTPUT_FILE)/1024/1024:.1f} MB')

    return success_count, sorted_dates

if __name__ == '__main__':
    count, dates = create_merged_excel()
    print(f'\n最终结果: {count} 个日期')
    print(f'日期范围: {dates[0]} 到 {dates[-1]}')