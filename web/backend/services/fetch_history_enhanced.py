"""
烟台钢筋价格历史数据增强抓取脚本 v2.0
功能特性：
1. 人机模拟（随机鼠标移动、滚动、延迟）
2. 早晚价格分别抓取（AM/PM）
3. 完整性验证（每天不少于11条）
4. 手动登录支持
5. 账号更新功能
6. 从2024年到2026年6月10日连续抓取

用法：
    python fetch_history_enhanced.py --start 2024-01-01 --end 2026-06-10 --interval 5
    python fetch_history_enhanced.py --login-only  # 仅登录并保存Cookie
"""
import asyncio
import sys
import json
import base64
import random
import time
import argparse
from pathlib import Path
from datetime import datetime, timedelta
from typing import List, Dict, Set, Optional, Tuple

sys.path.insert(0, '.')

try:
    from playwright.async_api import async_playwright, Page, Browser
    HAS_PLAYWRIGHT = True
except ImportError:
    HAS_PLAYWRIGHT = False
    print("错误: 未安装 playwright，请运行: pip install playwright")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("警告: 未安装 openpyxl，无法保存Excel")

# 路径配置
DATA_DIR = Path('services/data')
DATA_DIR.mkdir(exist_ok=True)

DB_FILE = DATA_DIR / 'yantai_rebar.db'
EXCEL_FILE = DATA_DIR / '烟台钢筋价格_完整历史_2024_2026.xlsx'
COOKIE_FILE = DATA_DIR / 'mysteel_cookies_enhanced.json'
CONFIG_FILE = DATA_DIR / 'mysteel_config.json'

# 最小数据条数要求
MIN_PRICES_PER_DAY = 11

# 日志配置
import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(str(DATA_DIR / 'fetch_history_enhanced.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# ==================== 人机模拟模块 ====================

class HumanSimulation:
    """人机模拟 - 模拟真实用户行为"""

    @staticmethod
    async def random_move_mouse(page: Page):
        """随机鼠标移动 - 模拟真实用户浏览"""
        try:
            # 获取页面尺寸
            viewport = page.viewport_size
            if viewport:
                # 随机移动鼠标到几个位置
                for _ in range(random.randint(1, 3)):
                    x = random.randint(100, viewport['width'] - 100)
                    y = random.randint(100, viewport['height'] - 100)
                    await page.mouse.move(x, y)
                    await page.wait_for_timeout(random.randint(50, 200))
        except Exception as e:
            logger.debug(f"鼠标移动模拟失败: {e}")

    @staticmethod
    async def random_scroll(page: Page):
        """随机滚动 - 模拟真实用户浏览"""
        try:
            # 随机滚动几次
            for _ in range(random.randint(1, 2)):
                scroll_amount = random.randint(-200, 200)
                await page.evaluate(f'window.scrollBy(0, {scroll_amount})')
                await page.wait_for_timeout(random.randint(200, 500))
        except Exception as e:
            logger.debug(f"滚动模拟失败: {e}")

    @staticmethod
    async def random_delay(min_ms: int = 2000, max_ms: int = 5000):
        """随机延迟 - 模拟人类思考时间"""
        delay = random.randint(min_ms, max_ms)
        await asyncio.sleep(delay / 1000)
        return delay

    @staticmethod
    async def simulate_reading(page: Page):
        """模拟阅读页面 - 综合模拟行为"""
        await HumanSimulation.random_delay(1000, 2000)
        await HumanSimulation.random_move_mouse(page)
        await HumanSimulation.random_scroll(page)
        await HumanSimulation.random_delay(500, 1500)


# ==================== 数据库模块 ====================

class DatabaseManager:
    """数据库管理"""

    @staticmethod
    def init_db():
        """初始化数据库"""
        import sqlite3
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()

        # 检查 fetch_time 字段
        c.execute("PRAGMA table_info(rebar_prices)")
        columns = [col[1] for col in c.fetchall()]

        c.execute('''
            CREATE TABLE IF NOT EXISTS rebar_prices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                fetch_time TEXT,
                period TEXT,
                material_name TEXT,
                spec TEXT,
                material_type TEXT,
                brand TEXT,
                price INTEGER,
                region TEXT DEFAULT '山东烟台',
                screenshot_path TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(date, period, material_name, spec, brand, price)
            )
        ''')

        # 如果没有 period 字段，添加它
        if 'period' not in columns:
            c.execute('ALTER TABLE rebar_prices ADD COLUMN period TEXT')
            logger.info("数据库已添加 period 字段")

        # 添加 screenshot_path 字段
        if 'screenshot_path' not in columns:
            c.execute('ALTER TABLE rebar_prices ADD COLUMN screenshot_path TEXT')
            logger.info("数据库已添加 screenshot_path 字段")

        c.execute('CREATE INDEX IF NOT EXISTS idx_date ON rebar_prices(date)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_period ON rebar_prices(period)')
        conn.commit()
        conn.close()
        logger.info("数据库初始化完成")

    @staticmethod
    def get_existing_keys() -> Set[str]:
        """获取已存在的数据键"""
        import sqlite3
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute('SELECT date, period, material_name, spec, brand FROM rebar_prices')
        existing = {f"{r[0]}_{r[1] or ''}_{r[2]}_{r[3]}_{r[4]}" for r in c.fetchall()}
        conn.close()
        return existing

    @staticmethod
    def insert_prices(date: str, period: str, prices: List[Dict], existing_keys: Set[str]) -> Tuple[int, int]:
        """插入价格数据，返回 (插入数, 跳过数)"""
        import sqlite3
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()

        inserted = 0
        skipped = 0

        for price in prices:
            key = f"{date}_{period}_{price.get('material_name', '')}_{price.get('spec', '')}_{price.get('brand', '')}"
            if key not in existing_keys:
                try:
                    c.execute('''
                        INSERT INTO rebar_prices
                        (date, period, fetch_time, material_name, spec, material_type, brand, price, region, screenshot_path)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        date, period, price.get('fetch_time', ''), price.get('material_name', ''),
                        price.get('spec', ''), price.get('material_type', ''), price.get('brand', ''),
                        price.get('price', 0), '山东烟台', price.get('screenshot_path', '')
                    ))
                    inserted += 1
                    existing_keys.add(key)
                except sqlite3.IntegrityError:
                    skipped += 1
            else:
                skipped += 1

        conn.commit()
        conn.close()
        return inserted, skipped

    @staticmethod
    def get_date_count(date: str, period: str = None) -> int:
        """获取指定日期的数据条数"""
        import sqlite3
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        if period:
            c.execute('SELECT COUNT(*) FROM rebar_prices WHERE date = ? AND period = ?', (date, period))
        else:
            c.execute('SELECT COUNT(*) FROM rebar_prices WHERE date = ?', (date,))
        count = c.fetchone()[0]
        conn.close()
        return count

    @staticmethod
    def get_missing_dates(start_date: str, end_date: str) -> List[str]:
        """获取缺失数据的日期列表"""
        import sqlite3
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()

        # 获取数据库中的所有日期
        c.execute('SELECT DISTINCT date FROM rebar_prices ORDER BY date')
        existing_dates = {row[0] for row in c.fetchall()}
        conn.close()

        # 生成日期范围
        current = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        all_dates = []

        while current <= end:
            # 只检查工作日
            if current.weekday() < 5:
                date_str = current.strftime('%Y-%m-%d')
                # 检查该日期是否有足够的数据（AM+PM都应该有数据）
                count = DatabaseManager.get_date_count(date_str)
                if count < MIN_PRICES_PER_DAY * 2:  # AM和PM各至少11条
                    all_dates.append(date_str)
            current += timedelta(days=1)

        return all_dates


# ==================== Excel保存模块 ====================

class ExcelManager:
    """Excel管理"""

    @staticmethod
    def save_data(date: str, period: str, prices: List[Dict], screenshot_path: str = None):
        """保存数据到Excel"""
        if not HAS_OPENPYXL or not prices:
            return

        try:
            # 打开或创建工作簿
            if EXCEL_FILE.exists():
                wb = openpyxl.load_workbook(EXCEL_FILE)
            else:
                wb = openpyxl.Workbook()
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']

            # 创建sheet名
            time_str = datetime.now().strftime('%H%M%S')
            sheet_name = f'{date}_{period}_{time_str}'

            # 如果sheet已存在则跳过
            if sheet_name in wb.sheetnames:
                wb.close()
                return

            ws = wb.create_sheet(title=sheet_name)

            # 样式定义
            header_font = Font(bold=True, size=12, color="FFFFFF")
            if period == 'PM':
                header_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
            else:
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # 标题
            ws.merge_cells('A1:L1')
            period_text = '上午' if period == 'AM' else '下午'
            ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date} {period_text}').font = Font(bold=True, size=14)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

            # 表头
            headers = ['日期', '时段', '抓取时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # 数据行
            fetch_time = datetime.now().strftime('%H:%M:%S')
            for i, price in enumerate(prices):
                row = 4 + i
                data = [
                    date, period, fetch_time,
                    price.get('material_name', ''), price.get('spec', ''),
                    price.get('material_type', ''), price.get('brand', ''),
                    price.get('price', 0), '', '', '', '山东烟台'
                ]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = thin_border

            # 设置列宽
            widths = [12, 8, 12, 10, 10, 12, 14, 12, 10, 25, 10, 10]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            # 嵌入截图
            if screenshot_path and Path(screenshot_path).exists():
                try:
                    row = 4 + len(prices) + 2
                    ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)
                    ws.row_dimensions[row + 1].height = 300

                    img = Image(screenshot_path)
                    img.width = 900
                    img.height = 450
                    img.anchor = f'A{row + 1}'
                    ws.add_image(img)
                except Exception as e:
                    logger.warning(f"嵌入截图失败: {e}")

            wb.save(EXCEL_FILE)
            wb.close()
            logger.info(f"Excel保存成功: {sheet_name}")

        except Exception as e:
            logger.error(f"Excel保存失败: {e}")


# ==================== 配置管理模块 ====================

class ConfigManager:
    """配置管理"""

    @staticmethod
    def load_credentials() -> Tuple[str, str]:
        """加载登录凭据"""
        if CONFIG_FILE.exists():
            try:
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    return config.get('username', ''), config.get('password', '')
            except Exception as e:
                logger.warning(f"加载配置失败: {e}")
        return '', ''

    @staticmethod
    def save_credentials(username: str, password: str):
        """保存登录凭据"""
        try:
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump({'username': username, 'password': password}, f, ensure_ascii=False, indent=2)
            logger.info(f"凭据已保存: {username[:3]}***")
        except Exception as e:
            logger.error(f"保存凭据失败: {e}")

    @staticmethod
    def load_cookies() -> List[Dict]:
        """加载Cookie"""
        if COOKIE_FILE.exists():
            try:
                with open(COOKIE_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                logger.warning(f"加载Cookie失败: {e}")
        return []

    @staticmethod
    def save_cookies(cookies: List[Dict]):
        """保存Cookie"""
        try:
            with open(COOKIE_FILE, 'w', encoding='utf-8') as f:
                json.dump(cookies, f, ensure_ascii=False)
            logger.info(f"Cookie已保存: {len(cookies)}条")
        except Exception as e:
            logger.warning(f"保存Cookie失败: {e}")


# ==================== 主抓取模块 ====================

class HistoryFetcher:
    """历史数据抓取器"""

    def __init__(self, start_date: str, end_date: str, interval: int = 5, headless: bool = True):
        self.start_date = start_date
        self.end_date = end_date
        self.interval = interval  # 抓取间隔（秒）
        self.headless = headless
        self.existing_keys = set()
        self.browser = None
        self.context = None
        self.page = None

    async def init_browser(self):
        """初始化浏览器"""
        if not HAS_PLAYWRIGHT:
            raise Exception("未安装 playwright")

        playwright = await async_playwright().start()
        self.browser = await playwright.chromium.launch(headless=self.headless)
        self.context = await self.browser.new_context(
            viewport={'width': 1920, 'height': 3000},
            locale='zh-CN',
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        )
        self.page = await self.context.new_page()

        # 加载Cookie
        cookies = ConfigManager.load_cookies()
        if cookies:
            await self.context.add_cookies(cookies)
            logger.info(f"已加载 {len(cookies)} 个Cookie")

    async def close_browser(self):
        """关闭浏览器"""
        if self.context:
            cookies = await self.context.cookies()
            ConfigManager.save_cookies(cookies)
        if self.browser:
            await self.browser.close()

    async def manual_login(self) -> bool:
        """手动登录 - 打开登录页面等待用户操作"""
        logger.info("打开登录页面，请手动登录...")
        await self.page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)

        # 等待用户手动登录（最多等待2分钟）
        logger.info("等待手动登录（最多2分钟）...")

        for i in range(24):  # 24 * 5秒 = 2分钟
            await asyncio.sleep(5)
            # 检查是否登录成功
            current_url = self.page.url
            if 'passport' not in current_url or 'login' not in current_url:
                logger.info("检测到登录成功！")
                # 保存Cookie
                cookies = await self.context.cookies()
                ConfigManager.save_cookies(cookies)
                return True
            logger.info(f"等待登录... ({i+1}/24)")

        logger.warning("登录超时，请检查网络或账号密码")
        return False

    async def auto_login(self, username: str, password: str) -> bool:
        """自动登录"""
        logger.info(f"尝试自动登录: {username[:3]}***")

        try:
            await self.page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
            await HumanSimulation.random_delay(2000, 4000)

            # 切换到账号登录
            try:
                account_tab = await self.page.query_selector('.form-tab-account, a[data-tab="account"]')
                if account_tab:
                    await account_tab.click()
                    await HumanSimulation.random_delay(1000, 2000)
            except:
                pass

            # 填写表单
            await self.page.evaluate(f'''() => {{
                const inputs = document.querySelectorAll('input');
                for (const inp of inputs) {{
                    const ph = inp.placeholder || '';
                    if (ph.includes('用户名')) inp.value = '{username}';
                    if (ph.includes('密码') && inp.type === 'password') inp.value = '{password}';
                }}
            }}''')

            await HumanSimulation.random_delay(500, 1500)

            # 勾选协议
            try:
                checkbox = await self.page.query_selector('input[type="checkbox"]')
                if checkbox and not await checkbox.is_checked():
                    await checkbox.click()
            except:
                pass

            await HumanSimulation.random_delay(500, 1000)

            # 点击登录
            try:
                login_btn = await self.page.query_selector('.form-button-login, button:has-text("登录")')
                if login_btn:
                    await HumanSimulation.random_move_mouse(self.page)
                    await login_btn.click()
                    logger.info("已点击登录按钮")
            except:
                pass

            # 等待登录完成
            await HumanSimulation.random_delay(5000, 8000)

            # 验证登录状态
            current_url = self.page.url
            if 'passport' in current_url and 'login' in current_url:
                logger.warning("自动登录可能失败，尝试手动登录...")
                return await self.manual_login()

            # 保存Cookie
            cookies = await self.context.cookies()
            ConfigManager.save_cookies(cookies)
            logger.info("自动登录成功，Cookie已保存")
            return True

        except Exception as e:
            logger.error(f"自动登录失败: {e}")
            return await self.manual_login()

    async def fetch_date_data(self, date: str, period: str) -> Tuple[int, List[Dict]]:
        """抓取指定日期的数据"""
        logger.info(f"抓取 {date} {period}...")

        # 生成URL - 根据时段设置不同时间
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        if period == 'AM':
            # 上午 - 使用10点
            time_str = date_obj.strftime('%y%m%d') + '10'
        else:
            # 下午 - 使用16点
            time_str = date_obj.strftime('%y%m%d') + '16'

        # 尝试多个可能的URL
        urls = [
            f'https://jiancai.mysteel.com/m/{time_str}/25B3355C6617BD3C.html',
            f'https://jiancai.mysteel.com/m/{time_str}10/25B3355C6617BD3C.html',
            f'https://jiancai.mysteel.com/m/{time_str}16/25B3355C6617BD3C.html',
        ]

        prices = []
        screenshot_path = None

        for url in urls:
            try:
                logger.debug(f"尝试URL: {url}")

                await self.page.goto(url, wait_until='domcontentloaded', timeout=30000)

                # 人机模拟
                await HumanSimulation.simulate_reading(self.page)

                # 检查是否需要登录
                body_text = await self.page.evaluate('() => document.body.textContent')
                if '登录' in body_text and len(body_text) < 1000:
                    logger.warning("需要重新登录")
                    return 0, []

                # 截图
                date_clean = date.replace('-', '')
                screenshot_path = str(DATA_DIR / f'screenshot_{date_clean}_{period}.png')
                await self.page.screenshot(path=screenshot_path, full_page=True)

                # 提取数据
                data = await self.page.evaluate('''() => {
                    const tables = document.querySelectorAll('table');
                    const results = [];
                    tables.forEach(table => {
                        const rows = table.querySelectorAll('tr');
                        rows.forEach(row => {
                            const cells = row.querySelectorAll('td, th');
                            const rowData = [];
                            cells.forEach(c => rowData.push(c.textContent.trim()));
                            if (rowData.length > 0) results.push(rowData);
                        });
                    });
                    return results;
                }''')

                # 解析价格
                for row in data:
                    if row and len(row) >= 5:
                        material_name = str(row[0]).strip()
                        spec = str(row[1]).strip()
                        material_type = str(row[2]).strip()
                        brand = str(row[3]).strip()
                        price_str = str(row[4]).strip()

                        valid_names = ['高线', '螺纹钢', '盘螺', '圆钢']
                        if material_name in valid_names and spec.startswith('Φ'):
                            try:
                                price = int(''.join(filter(str.isdigit, price_str)))
                                if price > 0:
                                    prices.append({
                                        'material_name': material_name,
                                        'spec': spec,
                                        'material_type': material_type,
                                        'brand': brand,
                                        'price': price,
                                        'fetch_time': datetime.now().strftime('%H:%M:%S'),
                                        'screenshot_path': screenshot_path
                                    })
                            except:
                                pass

                if prices:
                    logger.info(f"成功提取 {len(prices)} 条数据")
                    break
                else:
                    await HumanSimulation.random_delay(2000, 4000)

            except Exception as e:
                logger.warning(f"URL抓取失败: {e}")
                continue

        return len(prices), prices

    async def verify_completeness(self, date: str) -> bool:
        """验证指定日期的数据完整性"""
        # 检查AM和PM的数据量
        am_count = DatabaseManager.get_date_count(date, 'AM')
        pm_count = DatabaseManager.get_date_count(date, 'PM')

        total = am_count + pm_count
        logger.info(f"日期 {date} 完整性检查: AM={am_count}, PM={pm_count}, 总计={total}")

        if total < MIN_PRICES_PER_DAY:
            logger.warning(f"日期 {date} 数据不足: {total} < {MIN_PRICES_PER_DAY}")
            return False

        return True

    async def run(self) -> Dict:
        """运行抓取任务"""
        logger.info("=" * 60)
        logger.info("烟台钢筋价格历史数据增强抓取 v2.0")
        logger.info(f"日期范围: {self.start_date} 至 {self.end_date}")
        logger.info(f"最小数据量: 每天{MIN_PRICES_PER_DAY}条")
        logger.info(f"抓取间隔: {self.interval}秒")
        logger.info("=" * 60)

        # 初始化
        DatabaseManager.init_db()
        self.existing_keys = DatabaseManager.get_existing_keys()
        logger.info(f"数据库已有 {len(self.existing_keys)} 条记录")

        # 初始化浏览器
        await self.init_browser()

        # 登录
        username, password = ConfigManager.load_credentials()
        if username and password:
            login_success = await self.auto_login(username, password)
        else:
            logger.info("未找到凭据，请手动登录")
            login_success = await self.manual_login()

        if not login_success:
            await self.close_browser()
            return {'success': False, 'error': '登录失败'}

        # 生成日期列表
        current = datetime.strptime(self.start_date, '%Y-%m-%d')
        end = datetime.strptime(self.end_date, '%Y-%m-%d')
        dates = []

        while current <= end:
            # 只抓取工作日
            if current.weekday() < 5:
                dates.append(current.strftime('%Y-%m-%d'))
            current += timedelta(days=1)

        logger.info(f"需要抓取 {len(dates)} 个工作日")
        logger.info(f"每天抓取 AM 和 PM 两个时段\n")

        # 开始抓取
        total_inserted = 0
        total_skipped = 0
        success_dates = 0
        incomplete_dates = []

        for i, date in enumerate(dates):
            logger.info(f"\n[{i+1}/{len(dates)}] {date}")

            # 检查是否已有数据
            existing_count = DatabaseManager.get_date_count(date)
            if existing_count >= MIN_PRICES_PER_DAY * 2:
                logger.info(f"日期 {date} 已有足够数据，跳过")
                continue

            # 抓取 AM
            am_count, am_prices = await self.fetch_date_data(date, 'AM')
            if am_prices:
                inserted, skipped = DatabaseManager.insert_prices(date, 'AM', am_prices, self.existing_keys)
                total_inserted += inserted
                total_skipped += skipped
                ExcelManager.save_data(date, 'AM', am_prices)

            await HumanSimulation.random_delay(self.interval * 1000, (self.interval + 3) * 1000)

            # 抓取 PM
            pm_count, pm_prices = await self.fetch_date_data(date, 'PM')
            if pm_prices:
                inserted, skipped = DatabaseManager.insert_prices(date, 'PM', pm_prices, self.existing_keys)
                total_inserted += inserted
                total_skipped += skipped
                ExcelManager.save_data(date, 'PM', pm_prices)

            # 验证完整性
            is_complete = await self.verify_completeness(date)
            if is_complete:
                success_dates += 1
            else:
                incomplete_dates.append(date)

            # 间隔延迟
            if i < len(dates) - 1:
                delay = self.interval + random.randint(0, 3)
                logger.info(f"等待 {delay} 秒后继续...")
                await asyncio.sleep(delay)

        await self.close_browser()

        # 输出结果
        logger.info("\n" + "=" * 60)
        logger.info("抓取完成")
        logger.info(f"  成功日期: {success_dates}/{len(dates)}")
        logger.info(f"  新增记录: {total_inserted} 条")
        logger.info(f"  跳过记录: {total_skipped} 条")
        logger.info(f"  Excel文件: {EXCEL_FILE}")
        logger.info(f"  数据库: {DB_FILE}")

        if incomplete_dates:
            logger.warning(f"  数据不足的日期: {len(incomplete_dates)} 天")
            logger.warning(f"  {incomplete_dates[:10]}...")

        logger.info("=" * 60)

        return {
            'success': True,
            'total_dates': len(dates),
            'success_dates': success_dates,
            'inserted': total_inserted,
            'skipped': total_skipped,
            'incomplete_dates': incomplete_dates
        }


# ==================== 主函数 ====================

async def login_only():
    """仅登录并保存Cookie"""
    print("=" * 60)
    print("手动登录模式")
    print("=" * 60)

    fetcher = HistoryFetcher('2024-01-01', '2024-01-02')
    await fetcher.init_browser()
    success = await fetcher.manual_login()
    await fetcher.close_browser()

    if success:
        print("\n登录成功！Cookie已保存，可以开始抓取数据")
    else:
        print("\n登录失败或超时")


async def main():
    """主函数"""
    parser = argparse.ArgumentParser(description='烟台钢筋价格历史数据增强抓取')
    parser.add_argument('--start', '-s', default='2024-01-01', help='开始日期 (YYYY-MM-DD)')
    parser.add_argument('--end', '-e', default='2026-06-10', help='结束日期 (YYYY-MM-DD)')
    parser.add_argument('--interval', '-i', type=int, default=5, help='抓取间隔秒数')
    parser.add_argument('--headless', action='store_true', help='无头模式')
    parser.add_argument('--login-only', action='store_true', help='仅登录并保存Cookie')
    parser.add_argument('--update-credentials', action='store_true', help='更新登录凭据')

    args = parser.parse_args()

    # 更新凭据
    if args.update_credentials:
        print("请输入登录凭据:")
        username = input("用户名: ").strip()
        password = input("密码: ").strip()
        if username and password:
            ConfigManager.save_credentials(username, password)
            print("凭据已保存！")
        return

    # 仅登录
    if args.login_only:
        await login_only()
        return

    # 执行抓取
    fetcher = HistoryFetcher(args.start, args.end, args.interval, not args.headless)
    result = await fetcher.run()

    if result.get('success'):
        print(f"\n抓取成功完成！")
        print(f"Excel: {EXCEL_FILE}")
        print(f"数据库: {DB_FILE}")
    else:
        print(f"\n抓取失败: {result.get('error')}")


if __name__ == '__main__':
    if HAS_PLAYWRIGHT:
        asyncio.run(main())
    else:
        print("错误: 需要安装 playwright")
        print("请运行: pip install playwright")
        print("然后运行: playwright install chromium")
