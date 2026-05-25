"""
使用 PaddleOCR 批量解析烟台钢筋价格截图
"""

import os
import json
import re
from pathlib import Path
from datetime import datetime
from collections import Counter

from PIL import Image
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 配置
SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR / 'data'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
PROGRESS_FILE = DATA_DIR / 'paddle_ocr_progress.json'

# 全局 OCR 实例
_ocr = None


def init_ocr():
    """初始化 PaddleOCR"""
    global _ocr
    if _ocr is None:
        from paddleocr import PaddleOCR
        print('[OCR] Initializing PaddleOCR...')
        _ocr = PaddleOCR(lang='ch', use_textline_orientation=False, use_angle_cls=False)
        print('[OCR] PaddleOCR ready')
    return _ocr


def parse_filename(screenshot_name):
    """从截图文件名解析日期和时段"""
    # screenshot_20240105_AM.png
    parts = screenshot_name.replace('screenshot_', '').replace('.png', '').split('_')
    if len(parts) >= 2:
        date_str = parts[0]
        year = date_str[:4]
        month = date_str[4:6]
        day = date_str[6:8]
        period = parts[1]
        return f'{year}-{month}-{day}', period
    return None, None


def get_existing_dates():
    """获取 Excel 中已有的日期"""
    if not EXCEL_FILE.exists():
        return set()

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
        dates = set()
        for sheet_name in wb.sheetnames:
            if '_' in sheet_name:
                date_part = sheet_name.split('_')[0]
                dates.add(date_part)
        wb.close()
        return dates
    except Exception as e:
        print(f'[ERROR] 读取Excel失败: {e}')
        return set()


def parse_ocr_result(ocr_result):
    """解析 OCR 结果，提取价格数据"""
    prices = []

    if not ocr_result or not ocr_result[0]:
        return prices

    # 提取所有文本及其位置
    texts = []
    for line in ocr_result[0]:
        text = line[1][0].strip()
        bbox = line[0]
        if text:
            y_center = (bbox[0][1] + bbox[2][1]) / 2
            texts.append((y_center, text))

    # 按 Y 坐标排序
    texts.sort(key=lambda x: x[0])

    # 关键词映射
    material_keywords = ['高线', '螺纹钢', '盘螺', '圆钢']
    price_keywords = ['Φ6', 'Φ8', 'Φ10', 'Φ12', 'Φ14', 'Φ16', 'Φ18', 'Φ20', 'Φ22', 'Φ25', 'Φ28', 'Φ32', 'Φ36']

    # 分析每行数据
    current_material = None
    current_spec = None

    for y, text in texts:
        # 检测品名
        for mat in material_keywords:
            if mat in text:
                current_material = mat
                break

        # 检测规格
        for spec in price_keywords:
            if spec in text:
                current_spec = spec
                break

        # 检测价格（4位数字，通常是 3xxx 或 4xxx）
        import re
        price_match = re.search(r'\b(\d{4})\b', text)
        if price_match and current_material and current_spec:
            price = int(price_match.group(1))
            # 价格合理性检查
            if 3000 <= price <= 6000:
                # 尝试提取品牌（在价格附近的文本）
                brand = ''
                idx = texts.index((y, text))
                # 查找附近的品牌关键词
                for j in range(max(0, idx - 5), min(len(texts), idx + 5)):
                    t = texts[j][1]
                    if any(b in t for b in ['永锋', '石横', '镔鑫', '莱钢', '日照', '敬业']):
                        brand = t
                        break

                if not brand:
                    brand = '未知'

                prices.append({
                    'material_name': current_material,
                    'spec': current_spec,
                    'brand': brand,
                    'price': price
                })

                # 重置规格（每行价格后继续找下一个规格）
                current_spec = None

    # 去重（同一规格只保留一个价格）
    seen = set()
    unique_prices = []
    for p in prices:
        key = (p['material_name'], p['spec'])
        if key not in seen:
            seen.add(key)
            unique_prices.append(p)

    return unique_prices


def save_prices_to_excel(prices, date_str, period, fetch_time='09:00:00'):
    """保存价格数据到 Excel"""
    if not prices:
        return False

    try:
        # 打开或创建 workbook
        if EXCEL_FILE.exists():
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE)
            except Exception:
                wb = openpyxl.Workbook()
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']
        else:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        # 生成 sheet 名称
        time_str = fetch_time.replace(':', '')
        sheet_name = f'{date_str}_{period}_{time_str}'

        # 避免重复
        base_name = sheet_name
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f'{base_name}_{counter}'
            counter += 1

        ws = wb.create_sheet(title=sheet_name)

        # 样式
        period_text = '上午' if period == 'AM' else '下午'
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid') if period == 'AM' else PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 标题行
        ws.merge_cells('A1:K1')
        ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        # 表头
        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 数据行
        for i, price in enumerate(prices):
            row = 4 + i
            values = [
                date_str,
                fetch_time,
                price.get('material_name', ''),
                price.get('spec', ''),
                price.get('material_type', ''),
                price.get('brand', ''),
                price.get('price', 0),
                '',
                '',
                '',
                '山东烟台'
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border

        wb.save(EXCEL_FILE)
        wb.close()

        return True

    except Exception as e:
        print(f'[ERROR] 保存失败: {e}')
        return False


def load_progress():
    """加载解析进度"""
    if PROGRESS_FILE.exists():
        with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'parsed': [], 'failed': []}


def save_progress(progress):
    """保存解析进度"""
    with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def get_all_screenshots():
    """获取所有截图文件"""
    screenshots = []
    for f in os.listdir(DATA_DIR):
        if f.startswith('screenshot_') and f.endswith('.png'):
            screenshots.append(f)
    return sorted(screenshots)


def process_screenshot(ocr, screenshot_path, date_str, period):
    """处理单个截图"""
    try:
        img = Image.open(screenshot_path)

        # 截取表格区域（顶部约 55% 的高度，表格通常在这里）
        table_region = img.crop((0, 0, img.width, int(img.height * 0.55)))

        # 转换为 numpy 数组
        img_array = np.array(table_region)

        # OCR
        result = ocr.ocr(img_array)

        # 解析结果
        prices = parse_ocr_result(result)

        return prices

    except Exception as e:
        print(f'[ERROR] 处理截图失败: {e}')
        return []


def run_batch_parse(limit=None, force=False):
    """批量解析截图"""
    print('=' * 60)
    print('PaddleOCR 批量解析烟台钢筋价格截图')
    print('=' * 60)

    # 初始化 OCR
    ocr = init_ocr()

    # 获取所有截图
    screenshots = get_all_screenshots()
    print(f'总截图数: {len(screenshots)}')

    # 获取已有日期
    existing_dates = get_existing_dates()
    print(f'Excel 已有: {len(existing_dates)} 个日期')

    # 加载进度
    progress = load_progress()
    parsed = set(progress.get('parsed', []))
    failed = set(progress.get('failed', []))

    # 筛选待处理
    to_process = []
    for s in screenshots:
        if s in parsed:
            continue
        if not force and s in failed:
            continue

        date_str, period = parse_filename(s)
        if date_str and date_str in existing_dates:
            continue

        to_process.append(s)

    print(f'待解析: {len(to_process)} 个')

    if limit:
        to_process = to_process[:limit]

    if not to_process:
        print('没有需要解析的截图')
        return

    # 按日期排序
    to_process.sort()

    # 开始解析
    success_count = 0
    fail_count = 0

    for i, screenshot in enumerate(to_process):
        print(f'\n[{i+1}/{len(to_process)}] {screenshot}')

        screenshot_path = DATA_DIR / screenshot

        # 解析文件名
        date_str, period = parse_filename(screenshot)
        if not date_str:
            print(f'[SKIP] 无法解析文件名')
            failed.add(screenshot)
            fail_count += 1
            continue

        # 提取价格
        prices = process_screenshot(ocr, screenshot_path, date_str, period)

        if prices:
            # 保存到 Excel
            if save_prices_to_excel(prices, date_str, period):
                print(f'[OK] {date_str} {period} - {len(prices)} 条数据')
                parsed.add(screenshot)
                success_count += 1
            else:
                fail_count += 1
        else:
            print(f'[WARN] 未提取到价格数据')
            failed.add(screenshot)
            fail_count += 1

        # 每 20 个保存一次进度
        if (i + 1) % 20 == 0:
            progress['parsed'] = list(parsed)
            progress['failed'] = list(failed)
            save_progress(progress)

    # 最终保存
    progress['parsed'] = list(parsed)
    progress['failed'] = list(failed)
    save_progress(progress)

    print('\n' + '=' * 60)
    print('完成')
    print(f'成功: {success_count}')
    print(f'失败: {fail_count}')
    print('=' * 60)


def main():
    import argparse

    parser = argparse.ArgumentParser(description='PaddleOCR 批量解析烟台钢筋价格截图')
    parser.add_argument('--limit', '-n', type=int, default=None, help='限制处理数量')
    parser.add_argument('--force', '-f', action='store_true', help='强制重新解析已失败的')
    parser.add_argument('--reset', action='store_true', help='重置进度')

    args = parser.parse_args()

    if args.reset:
        if PROGRESS_FILE.exists():
            os.remove(PROGRESS_FILE)
        print('进度已重置')

    run_batch_parse(limit=args.limit, force=args.force)


if __name__ == '__main__':
    main()