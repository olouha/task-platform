"""
导入Excel数据到SQLite数据库
"""
import sqlite3
import openpyxl
import re
from datetime import datetime
from pathlib import Path

DATA_DIR = Path('services/data')
DB_FILE = DATA_DIR / 'yantai_rebar.db'

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute('DROP TABLE IF EXISTS rebar_prices')
    c.execute('''
        CREATE TABLE rebar_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            fetch_time TEXT,
            material_name TEXT,
            spec TEXT,
            material_type TEXT,
            brand TEXT,
            price INTEGER,
            price_change TEXT,
            remark TEXT,
            steel_code TEXT,
            region TEXT DEFAULT '山东烟台'
        )
    ''')
    c.execute('CREATE INDEX idx_date ON rebar_prices(date)')
    c.execute('CREATE INDEX idx_material ON rebar_prices(material_name)')
    c.execute('CREATE INDEX idx_spec ON rebar_prices(spec)')

    conn.commit()
    return conn

def import_data(conn, excel_file):
    c = conn.cursor()
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)

    data = []
    for sheet_name in wb.sheetnames:
        if '-' not in sheet_name:
            continue
        date = sheet_name.split('_')[0]
        try:
            datetime.strptime(date, '%Y-%m-%d')
        except:
            continue

        ws = wb[sheet_name]

        # 从第3行开始读取数据（跳过标题和表头）
        for row_idx in range(3, ws.max_row + 1):
            row_data = [ws.cell(row=row_idx, column=col).value for col in range(1, 7)]

            material = row_data[0] if len(row_data) > 0 else None
            spec = row_data[1] if len(row_data) > 1 else ''
            mat_type = row_data[2] if len(row_data) > 2 else ''
            price_val = row_data[3] if len(row_data) > 3 else None

            # 跳过空行和非数据行
            if not material or not isinstance(material, str):
                continue
            if '截图' in material or '状态' in material:
                continue

            # 解析价格
            price = 0
            if price_val:
                if isinstance(price_val, (int, float)):
                    price = int(price_val)
                else:
                    match = re.search(r'(\d{3,5})', str(price_val))
                    if match:
                        price = int(match.group(1))

            if price > 0:
                data.append((date, '', material, spec, mat_type, '', price, '', '', '', '山东烟台'))

    wb.close()

    print(f'读取到 {len(data)} 条数据，准备插入...')
    c.executemany('''INSERT INTO rebar_prices
        (date, fetch_time, material_name, spec, material_type, brand, price, price_change, remark, steel_code, region)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', data)
    conn.commit()

    return len(data)

def main():
    print('=' * 60)
    print('导入数据到SQLite')
    print('=' * 60)

    conn = init_db()
    print(f'数据库: {DB_FILE}')

    excel_file = DATA_DIR / '山东烟台钢筋价格_完整版_数据+截图.xlsx'

    if excel_file.exists():
        print(f'\n导入: {excel_file.name}')
        total = import_data(conn, excel_file)
        print(f'导入完成: {total} 条')

        # 统计
        c = conn.cursor()
        print(f'总记录: {c.execute("SELECT COUNT(*) FROM rebar_prices").fetchone()[0]}')
        print(f'日期数: {c.execute("SELECT COUNT(DISTINCT date) FROM rebar_prices").fetchone()[0]}')
        r = c.execute('SELECT MIN(date), MAX(date) FROM rebar_prices').fetchone()
        print(f'范围: {r[0]} 到 {r[1]}')

        # 材料统计
        print('\n材料统计:')
        rows = c.execute('SELECT material_name, COUNT(*) as cnt FROM rebar_prices GROUP BY material_name ORDER BY cnt DESC LIMIT 10').fetchall()
        for mat, cnt in rows:
            print(f'  {mat}: {cnt}')

        # 规格统计
        print('\n规格统计:')
        rows = c.execute('SELECT spec, COUNT(*) as cnt FROM rebar_prices GROUP BY spec ORDER BY cnt DESC LIMIT 10').fetchall()
        for spec, cnt in rows:
            print(f'  {spec}: {cnt}')

    conn.close()
    print(f'\n完成! 数据库: {DB_FILE}')

if __name__ == '__main__':
    main()