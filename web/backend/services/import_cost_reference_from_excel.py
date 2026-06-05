# -*- coding: utf-8 -*-
"""
导入造价参考价数据到数据库
从 Excel 文件读取数据，存入 SQLite 数据库
"""
import sqlite3
from pathlib import Path
import openpyxl

# 数据库路径 - 使用 services/data 目录
DB_DIR = Path(__file__).parent / "data"
DB_DIR.mkdir(exist_ok=True)
DB_PATH = DB_DIR / "cost_reference.db"

print(f"[INFO] Database path: {DB_PATH}")

# Excel 文件路径 - 使用绝对路径
EXCEL_PATH = Path(r"C:\Users\admin\Desktop\近五年钢筋混凝土造价管理截图(1)\近五年钢筋混凝土造价管理截图\造价参考价数据.xlsx")


def get_db_connection():
    """获取数据库连接"""
    conn = sqlite3.connect(DB_PATH)
    return conn


def init_db():
    """初始化数据库表"""
    conn = get_db_connection()
    c = conn.cursor()

    # 删除旧表并重新创建
    c.execute('DROP TABLE IF EXISTS concrete_prices')
    c.execute('DROP TABLE IF EXISTS rebar_prices')

    # 混凝土价格表
    c.execute('''
        CREATE TABLE concrete_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year TEXT NOT NULL,
            quarter TEXT NOT NULL,
            grade TEXT NOT NULL,
            yantai REAL,
            rushan REAL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 钢筋价格表
    c.execute('''
        CREATE TABLE rebar_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year TEXT NOT NULL,
            quarter TEXT NOT NULL,
            grade TEXT,
            spec TEXT,
            price REAL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    conn.commit()
    conn.close()
    print("[OK] Database initialized")


def import_from_excel():
    """从 Excel 导入数据"""
    if not EXCEL_PATH.exists():
        print(f"[ERROR] Excel file not found: {EXCEL_PATH}")
        return

    print(f"Loading Excel: {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    print(f"Sheets: {wb.sheetnames}")

    conn = get_db_connection()

    # 清空旧数据
    c = conn.cursor()
    c.execute("DELETE FROM concrete_prices")
    c.execute("DELETE FROM rebar_prices")
    conn.commit()
    print("[OK] Old data cleared")

    total_concrete = 0
    total_rebar = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            continue

        # 跳过标题行
        data_rows = rows[1:]

        if '混凝土' in sheet_name:
            # 混凝土数据: 年份, 季度, 强度等级, 烟台含税, 蓬莱含税
            for row in data_rows:
                if not row or not row[0]:
                    continue
                year = str(row[0]).strip() if row[0] else ''
                quarter = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                grade = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                yantai = float(row[3]) if len(row) > 3 and row[3] else None
                rushan = float(row[4]) if len(row) > 4 and row[4] else None

                if year and quarter and grade:
                    c.execute('''
                        INSERT INTO concrete_prices (year, quarter, grade, yantai, rushan)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (year, quarter, grade, yantai, rushan))
                    total_concrete += 1

            conn.commit()
            print(f"[OK] Sheet '{sheet_name}': {total_concrete} concrete records")

        elif '钢筋' in sheet_name:
            # 钢筋数据: 年份, 季度, 等级, 规格(mm), 价格(含税元/吨)
            for row in data_rows:
                if not row or not row[0]:
                    continue
                year = str(row[0]).strip() if row[0] else ''
                quarter = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                grade = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                spec = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                price = float(row[4]) if len(row) > 4 and row[4] else None

                if year and quarter and grade:
                    c.execute('''
                        INSERT INTO rebar_prices (year, quarter, grade, spec, price)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (year, quarter, grade, spec, price))
                    total_rebar += 1

            conn.commit()
            print(f"[OK] Sheet '{sheet_name}': {total_rebar} rebar records")

    conn.close()

    print(f"\n=== Import Complete ===")
    print(f"Concrete records: {total_concrete}")
    print(f"Rebar records: {total_rebar}")


def verify_data():
    """验证导入的数据"""
    conn = get_db_connection()
    c = conn.cursor()

    c.execute("SELECT COUNT(*) FROM concrete_prices")
    concrete_count = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM rebar_prices")
    rebar_count = c.fetchone()[0]

    c.execute("SELECT DISTINCT year FROM concrete_prices ORDER BY year")
    years = [r[0] for r in c.fetchall()]

    print(f"\n=== Data Verification ===")
    print(f"Concrete records: {concrete_count}")
    print(f"Rebar records: {rebar_count}")
    print(f"Years: {years}")

    conn.close()


if __name__ == "__main__":
    import os
    os.chdir(Path(__file__).parent)

    init_db()
    import_from_excel()
    verify_data()
