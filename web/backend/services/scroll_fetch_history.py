"""
烟台钢筋价格历史数据抓取 - 滚动页面版本
从2020年1月开始，逐月滚动抓取
"""
import asyncio
import json
import base64
import random
from pathlib import Path
from datetime import datetime, timedelta
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

DATA_DIR = Path('services/data')
COOKIE_FILE = DATA_DIR / 'mysteel_cookies.json'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
MARKET_URL = 'https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html'

# 新账号
USERNAME = 'M6616672758'
PASSWORD = 'Panhui199261*'

# 目标范围：2020年1月到2026年5月
START_YEAR, START_MONTH = 2020, 1
END_YEAR, END_MONTH = 2026, 5

LOG_FILE = DATA_DIR / 'logs' / f'scroll_fetch_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'


def log(msg):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f'[{timestamp}] {msg}'
    print(line)
    LOG_FILE.parent.mkdir(exist_ok=True)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(line + '\n')


def random_sleep(min_sec=15, max_sec=18):
    return random.uniform(min_sec, max_sec)


async def login(page):
    log('登录中...')
    if COOKIE_FILE.exists():
        try:
            cookies = json.loads(COOKIE_FILE.read_text(encoding='utf-8'))
            if cookies:
                await page.context.add_cookies(cookies)
                log('已加载Cookie')
        except:
            pass

    await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
    await asyncio.sleep(random_sleep(2, 4))

    try:
        account_tab = await page.query_selector('.form-tab-account, a[data-tab="account"]')
        if account_tab:
            await account_tab.click()
            await asyncio.sleep(random_sleep(1, 2))
    except:
        pass

    await page.evaluate(f'''() => {{
        const inputs = document.querySelectorAll('input');
        for (const inp of inputs) {{
            const ph = inp.placeholder || '';
            if (ph.includes('用户名')) inp.value = '{USERNAME}';
            if (ph.includes('密码') && inp.type === 'password') inp.value = '{PASSWORD}';
        }}
    }}''')

    await asyncio.sleep(random_sleep(1, 2))

    try:
        login_btn = await page.query_selector('button:has-text("登录"), input[type="submit"]')
        if login_btn:
            await login_btn.click()
            await asyncio.sleep(random_sleep(3, 5))
            cookies = await page.context.cookies()
            COOKIE_FILE.write_text(json.dumps(cookies, ensure_ascii=False), encoding='utf-8')
            log('登录成功')
    except Exception as e:
        log(f'登录失败: {e}')


def get_existing_sheets():
    """获取Excel中已存在的所有sheet"""
    existing = set()
    if EXCEL_FILE.exists():
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
            for name in wb.sheetnames:
                if name.startswith('20'):
                    existing.add(name[:10])
            wb.close()
        except:
            pass
    return existing


def extract_date(text):
    """从文本提取日期"""
    import re
    patterns = [r'(\d{4})-(\d{1,2})-(\d{1,2})', r'(\d{4})/(\d{1,2})/(\d{1,2})']
    for pattern in patterns:
        m = re.search(pattern, text)
        if m:
            return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'
    return None


async def scroll_and_collect(page, target_year, target_month):
    """滚动到指定月份并收集所有链接"""
    collected = []

    # 先滚动到页面底部
    for _ in range(10):
        await page.evaluate('window.scrollBy(0, 500)')
        await asyncio.sleep(0.5)

    # 然后慢慢向上滚动
    current_scroll = await page.evaluate('document.body.scrollHeight')
    while current_scroll > 0:
        await page.evaluate('window.scrollBy(0, -500)')
        await asyncio.sleep(0.3)
        current_scroll -= 500

    await asyncio.sleep(random_sleep(2, 3))

    # 收集所有列表项
    items = await page.query_selector_all('.list-item, .data-item, [class*="item"], tr')

    for item in items:
        try:
            text = await item.inner_text()
            date_str = extract_date(text)
            if date_str:
                # 检查是否在目标月份
                try:
                    date = datetime.strptime(date_str, '%Y-%m-%d')
                    if date.year == target_year and date.month == target_month:
                        # 找到链接
                        link = await item.query_selector('a[href*="mysteel"], a[href*="jiancai"]')
                        if link:
                            href = await link.get_attribute('href')
                            if href:
                                period = 'AM' if '上午' in text or '早' in text else 'PM'
                                url = href if href.startswith('http') else 'https://jiancai.mysteel.com' + href
                                collected.append({
                                    'date': date_str,
                                    'period': period,
                                    'url': url
                                })
                except:
                    pass
        except:
            pass

    # 去重
    seen = {}
    unique = []
    for item in collected:
        key = (item['date'], item['period'])
        if key not in seen:
            seen[key] = True
            unique.append(item)

    return unique


async def fetch_and_save(page, date_str, period, url):
    """抓取单个日期的数据"""
    try:
        await page.goto(url, wait_until='domcontentloaded', timeout=60000)
        await asyncio.sleep(random_sleep(4, 6))

        # 提取数据
        data = await page.evaluate('''() => {
            const tables = document.querySelectorAll('table');
            const results = [];
            tables.forEach(table => {
                const rows = table.querySelectorAll('tr');
                rows.forEach(row => {
                    const cells = row.querySelectorAll('td');
                    if (cells.length >= 5) {
                        const material_name = cells[0]?.textContent?.trim();
                        const spec = cells[1]?.textContent?.trim();
                        const material_type = cells[2]?.textContent?.trim();
                        const brand = cells[3]?.textContent?.trim();
                        const price_text = cells[4]?.textContent?.trim();

                        if (['高线', '螺纹钢', '盘螺', '圆钢'].includes(material_name)) {
                            const price = parseInt(price_text.replace(/,/g, ''));
                            if (!isNaN(price)) {
                                results.push({
                                    material_name,
                                    spec,
                                    material_type,
                                    brand,
                                    price
                                });
                            }
                        }
                    }
                });
            });
            return results;
        }''')

        # 截图
        screenshot_b64 = None
        try:
            screenshot = await page.screenshot(full_page=True)
            screenshot_b64 = base64.b64encode(screenshot).decode()
        except:
            pass

        return list(data), screenshot_b64

    except Exception as e:
        log(f'  抓取失败: {e}')
        return [], None


def save_to_excel(prices, period, date_str, screenshot_b64):
    """保存到Excel"""
    try:
        if EXCEL_FILE.exists():
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE)
            except:
                log(f'  文件损坏，创建新文件')
                wb = openpyxl.Workbook()
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']
        else:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        sheet_name = date_str if period == 'AM' else f'{date_str}_PM'

        if sheet_name in wb.sheetnames:
            log(f'  Sheet已存在，跳过')
            wb.close()
            return False

        ws = wb.create_sheet(title=sheet_name)

        # 标题
        period_text = '下午(晚)' if period == 'PM' else '上午'
        ws.merge_cells('A1:K1')
        ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        # 表头
        header_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid') if period == 'PM' else \
            PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 数据
        fetch_time = datetime.now().strftime('%H:%M:%S')
        for i, price in enumerate(prices):
            row = 4 + i
            for col, val in enumerate([date_str, fetch_time, price['material_name'], price['spec'],
                                       price['material_type'], price['brand'], price['price'],
                                       '', '', '', '山东烟台'], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border

        # 截图
        if screenshot_b64:
            screenshot_path = DATA_DIR / f'screenshot_{date_str.replace("-", "")}_{period}.png'
            with open(screenshot_path, 'wb') as f:
                f.write(base64.b64decode(screenshot_b64))
            row = 4 + len(prices) + 2
            ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)
            img = Image(str(screenshot_path))
            img.width = 900
            img.height = 500
            img.anchor = f'A{row + 1}'
            ws.add_image(img)

        wb.save(EXCEL_FILE)
        wb.close()
        return True

    except Exception as e:
        log(f'  保存失败: {e}')
        return False


async def main():
    print()
    print('=' * 70)
    print('烟台钢筋价格 - 历史数据抓取')
    print('=' * 70)
    print()

    existing = get_existing_sheets()
    log(f'已有数据: {len(existing)} 个日期')

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            viewport={'width': 1920, 'height': 3000},
            locale='zh-CN',
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        )
        page = await context.new_page()

        await login(page)

        # 生成所有目标月份
        months = []
        year, month = START_YEAR, START_MONTH
        while (year < END_YEAR) or (year == END_YEAR and month <= END_MONTH):
            months.append((year, month))
            if month == 12:
                year += 1
                month = 1
            else:
                month += 1

        log(f'目标月份: {len(months)} 个')

        success_count = 0
        fail_count = 0
        skip_count = 0

        for year, month in months:
            month_str = f'{year}-{month:02d}'
            log(f'\n{"="*40}')
            log(f'处理月份: {month_str}')
            log(f'{"="*40}')

            # 打开市场页面
            log('打开市场页面...')
            await page.goto(MARKET_URL, wait_until='domcontentloaded', timeout=60000)
            await asyncio.sleep(random_sleep(3, 5))

            # 收集当月数据
            links = await scroll_and_collect(page, year, month)
            log(f'  找到 {len(links)} 条链接')

            # 过滤已存在
            tasks = []
            for link in links:
                if link['date'] not in existing:
                    tasks.append(link)
                else:
                    skip_count += 1

            log(f'  需要抓取: {len(tasks)} 条 (已有: {len(links) - len(tasks)} 条)')

            # 按日期排序
            tasks.sort(key=lambda x: x['date'])

            month_success = 0
            month_fail = 0

            for task in tasks:
                date_str = task['date']
                period = task['period']
                url = task['url']

                log(f'  抓取 {date_str} {period}...', end='', flush=True)

                prices, screenshot = await fetch_and_save(page, date_str, period, url)

                if prices:
                    if save_to_excel(prices, period, date_str, screenshot):
                        existing.add(date_str)
                        log(f' {len(prices)}条 [OK]')
                        success_count += 1
                        month_success += 1
                    else:
                        log(' 保存失败')
                        fail_count += 1
                        month_fail += 1
                else:
                    log(' 无数据')
                    fail_count += 1
                    month_fail += 1

                # 停顿15秒
                log('  停顿15秒...')
                await asyncio.sleep(random_sleep(14, 16))

            log(f'  月份完成: 成功 {month_success}, 失败 {month_fail}, 跳过 {len(links) - len(tasks)}')

            # 月份之间停顿30秒
            await asyncio.sleep(30)

    await browser.close()

    print()
    print('=' * 70)
    print('抓取完成')
    print('=' * 70)
    log(f'成功: {success_count} 条')
    log(f'失败: {fail_count} 条')
    log(f'跳过: {skip_count} 条')

    final_existing = get_existing_sheets()
    log(f'最终数据: {len(final_existing)} 个日期')


if __name__ == '__main__':
    asyncio.run(main())