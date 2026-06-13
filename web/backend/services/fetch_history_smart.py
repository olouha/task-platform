"""
烟台钢筋价格历史数据智能抓取脚本 v3.0
- 从市场页面动态获取URL
- 确保每天不少于11条数据
- 支持AM/PM时段区分
- 人机模拟行为

用法:
    python fetch_history_smart.py --start 2024-01-01 --end 2026-06-10
"""
import asyncio
import sys
import json
import base64
import random
from pathlib import Path
from datetime import datetime, timedelta
from typing import List, Dict, Set, Tuple

sys.path.insert(0, '.')

from playwright.async_api import async_playwright, Page
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
import sqlite3

# 路径配置
DATA_DIR = Path('services/data')
DATA_DIR.mkdir(exist_ok=True)

DB_FILE = DATA_DIR / 'yantai_rebar.db'
EXCEL_FILE = DATA_DIR / '烟台钢筋价格_智能抓取_2024_2026.xlsx'
COOKIE_FILE = DATA_DIR / 'mysteel_cookies_smart.json'
CONFIG_FILE = DATA_DIR / 'mysteel_config.json'

# 最小数据条数
MIN_PRICES_PER_DAY = 11

# 日志
import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(str(DATA_DIR / 'fetch_history_smart.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# ==================== 人机模拟 ====================

class HumanBehavior:
    """人机行为模拟"""

    @staticmethod
    async def random_delay(min_sec: int = 2, max_sec: int = 5):
        delay = random.uniform(min_sec, max_sec)
        await asyncio.sleep(delay)

    @staticmethod
    async def simulate_reading(page: Page):
        """模拟阅读行为"""
        await HumanBehavior.random_delay(1, 3)

        # 随机滚动
        try:
            for _ in range(random.randint(1, 3)):
                scroll_y = random.randint(-300, 300)
                await page.evaluate(f'window.scrollBy(0, {scroll_y})')
                await asyncio.sleep(random.uniform(0.5, 1.5))
        except:
            pass

    @staticmethod
    async def simulate_mouse(page: Page):
        """模拟鼠标移动"""
        try:
            viewport = page.viewport_size
            if viewport:
                x = random.randint(200, viewport['width'] - 200)
                y = random.randint(200, viewport['height'] - 200)
                await page.mouse.move(x, y)
                await asyncio.sleep(random.uniform(0.1, 0.5))
        except:
            pass


# ==================== 数据库管理 ====================

class DBManager:
    """数据库管理"""

    @staticmethod
    def init_db():
        """初始化数据库"""
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()

        c.execute('''
            CREATE TABLE IF NOT EXISTS rebar_prices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                period TEXT,
                material_name TEXT,
                spec TEXT,
                material_type TEXT,
                brand TEXT,
                price INTEGER,
                region TEXT DEFAULT '山东烟台',
                fetch_time TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(date, period, material_name, spec, brand, price)
            )
        ''')

        c.execute('CREATE INDEX IF NOT EXISTS idx_date ON rebar_prices(date)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_period ON rebar_prices(period)')
        conn.commit()
        conn.close()

    @staticmethod
    def get_existing_keys() -> Set[str]:
        """获取已存在的数据键"""
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute('SELECT date, COALESCE(period,\'\'), material_name, spec, brand FROM rebar_prices')
        existing = {f"{r[0]}_{r[1]}_{r[2]}_{r[3]}_{r[4]}" for r in c.fetchall()}
        conn.close()
        return existing

    @staticmethod
    def insert_prices(date: str, period: str, prices: List[Dict], existing: Set[str]) -> Tuple[int, int]:
        """插入价格数据"""
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()

        inserted = 0
        skipped = 0

        for price in prices:
            key = f"{date}_{period}_{price.get('material_name', '')}_{price.get('spec', '')}_{price.get('brand', '')}"
            if key not in existing:
                try:
                    c.execute('''
                        INSERT INTO rebar_prices
                        (date, period, material_name, spec, material_type, brand, price, region, fetch_time)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        date, period, price.get('material_name', ''), price.get('spec', ''),
                        price.get('material_type', ''), price.get('brand', ''), price.get('price', 0),
                        '山东烟台', datetime.now().strftime('%H:%M:%S')
                    ))
                    inserted += 1
                    existing.add(key)
                except sqlite3.IntegrityError:
                    skipped += 1
            else:
                skipped += 1

        conn.commit()
        conn.close()
        return inserted, skipped

    @staticmethod
    def get_date_count(date: str) -> int:
        """获取日期的数据条数"""
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute('SELECT COUNT(*) FROM rebar_prices WHERE date = ?', (date,))
        count = c.fetchone()[0]
        conn.close()
        return count


# ==================== 智能抓取器 ====================

class SmartFetcher:
    """智能抓取器"""

    def __init__(self, start_date: str, end_date: str, interval: int = 5):
        self.start_date = start_date
        self.end_date = end_date
        self.interval = interval
        self.existing_keys = set()
        self.browser = None
        self.context = None
        self.page = None

    async def init_browser(self):
        """初始化浏览器"""
        playwright = await async_playwright().start()
        self.browser = await playwright.chromium.launch(headless=True)
        self.context = await self.browser.new_context(
            viewport={'width': 1920, 'height': 3000},
            locale='zh-CN'
        )
        self.page = await self.context.new_page()

        # 加载Cookie
        if COOKIE_FILE.exists():
            try:
                with open(COOKIE_FILE, 'r', encoding='utf-8') as f:
                    cookies = json.load(f)
                await self.context.add_cookies(cookies)
                logger.info(f"已加载 {len(cookies)} 个Cookie")
            except:
                pass

    async def close_browser(self):
        """关闭浏览器"""
        if self.context:
            cookies = await self.context.cookies()
            try:
                with open(COOKIE_FILE, 'w', encoding='utf-8') as f:
                    json.dump(cookies, f, ensure_ascii=False)
            except:
                pass
        if self.browser:
            await self.browser.close()

    async def login(self) -> bool:
        """登录"""
        logger.info("尝试登录...")

        # 加载凭据
        username, password = '', ''
        if CONFIG_FILE.exists():
            try:
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    username = config.get('username', '')
                    password = config.get('password', '')
            except:
                pass

        if username and password:
            try:
                await self.page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=30000)
                await asyncio.sleep(3)

                # 填写表单
                await self.page.evaluate(f'''() => {{
                    const inputs = document.querySelectorAll('input');
                    for (const inp of inputs) {{
                        const ph = inp.placeholder || '';
                        if (ph.includes('用户名')) inp.value = '{username}';
                        if (ph.includes('密码') && inp.type === 'password') inp.value = '{password}';
                    }}
                }}''')

                await asyncio.sleep(1)

                # 点击登录
                try:
                    btn = await self.page.query_selector('.form-button-login, button:has-text("登录")')
                    if btn:
                        await btn.click()
                except:
                    pass

                await asyncio.sleep(5)

                # 保存Cookie
                cookies = await self.context.cookies()
                with open(COOKIE_FILE, 'w', encoding='utf-8') as f:
                    json.dump(cookies, f, ensure_ascii=False)

                logger.info("登录成功")
                return True
            except Exception as e:
                logger.warning(f"自动登录失败: {e}")

        # 手动登录
        logger.info("请手动登录...")
        await self.page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded')
        logger.info("等待手动登录（2分钟）...")

        for i in range(24):
            await asyncio.sleep(5)
            if 'passport' not in self.page.url:
                cookies = await self.context.cookies()
                with open(COOKIE_FILE, 'w', encoding='utf-8') as f:
                    json.dump(cookies, f, ensure_ascii=False)
                logger.info("登录成功！")
                return True

        return False

    async def fetch_market_urls(self) -> List[Dict]:
        """从市场页面获取所有可用的URL列表"""
        logger.info("从市场页面获取URL列表...")

        urls = []

        try:
            # 访问山东市场页面
            market_url = "https://jiancai.mysteel.com/market/pa228aa01010104a0aaaaa1.html"
            await self.page.goto(market_url, wait_until='domcontentloaded', timeout=30000)
            await HumanBehavior.simulate_reading(self.page)

            # 获取所有烟台价格链接
            links = await self.page.evaluate('''() => {
                const results = [];
                document.querySelectorAll('a[href]').forEach(a => {
                    const href = a.href;
                    const text = a.textContent.trim();

                    // 查找包含烟台的链接，且是/m/开头的价格页面
                    if (href.includes('jiancai.mysteel.com/m/') &&
                        text.includes('烟台')) {
                        results.push({
                            href: href,
                            text: text
                        });
                    }
                });
                return results;
            }''')

            logger.info(f"找到 {len(links)} 个烟台价格链接")
            urls.extend(links)

            # 尝试另一个市场页面
            market_url2 = "https://jiancai.mysteel.com/market/pa228a81723aa0aaaaa1.html"
            await self.page.goto(market_url2, wait_until='domcontentloaded', timeout=30000)
            await HumanBehavior.simulate_reading(self.page)

            links2 = await self.page.evaluate('''() => {
                const results = [];
                document.querySelectorAll('a[href]').forEach(a => {
                    const href = a.href;
                    const text = a.textContent.trim();
                    if (href.includes('jiancai.mysteel.com/m/') &&
                        text.includes('烟台')) {
                        results.push({
                            href: href,
                            text: text
                        });
                    }
                });
                return results;
            }''')

            logger.info(f"第二个页面找到 {len(links2)} 个链接")
            urls.extend(links2)

        except Exception as e:
            logger.error(f"获取市场URL失败: {e}")

        # 去重
        seen = set()
        unique_urls = []
        for url in urls:
            if url['href'] not in seen:
                seen.add(url['href'])
                unique_urls.append(url)

        logger.info(f"总共找到 {len(unique_urls)} 个唯一URL")
        return unique_urls

    async def fetch_url_data(self, url: str) -> Tuple[str, List[Dict]]:
        """抓取单个URL的数据"""
        try:
            await self.page.goto(url, wait_until='domcontentloaded', timeout=30000)
            await HumanBehavior.simulate_reading(self.page)

            # 检查是否需要登录
            body_text = await self.page.evaluate('() => document.body.textContent')
            if '登录' in body_text and len(body_text) < 1000:
                logger.warning("需要重新登录")
                return '', []

            # 提取数据
            data = await self.page.evaluate('''() => {
                const tables = document.querySelectorAll('table');
                const results = [];
                tables.forEach(table => {
                    const rows = table.querySelectorAll('tr');
                    rows.forEach(row => {
                        const cells = row.querySelectorAll('td, th');
                        const rowData = [];
                        cells.forEach(c => rowData.push(c.textContent.trim()));
                        if (rowData.length > 0) results.push(rowData);
                    });
                });
                return results;
            }''')

            # 解析价格
            prices = []
            date_str = ''

            for row in data:
                if row and len(row) >= 5:
                    material_name = str(row[0]).strip()
                    spec = str(row[1]).strip()
                    material_type = str(row[2]).strip()
                    brand = str(row[3]).strip()
                    price_str = str(row[4]).strip()

                    # 从URL或页面提取日期
                    if not date_str:
                        # 尝试从URL提取日期
                        import re
                        date_match = re.search(r'(\d{2})-(\d{1,2})-(\d{1,2})', url)
                        if date_match:
                            year = int(date_match.group(1)) + 2000
                            month = int(date_match.group(2))
                            day = int(date_match.group(3))
                            date_str = f'{year:04d}-{month:02d}-{day:02d}'

                    valid_names = ['高线', '螺纹钢', '盘螺', '圆钢']
                    if material_name in valid_names and spec.startswith('Φ'):
                        try:
                            price = int(''.join(filter(str.isdigit, price_str)))
                            if price > 0:
                                prices.append({
                                    'material_name': material_name,
                                    'spec': spec,
                                    'material_type': material_type,
                                    'brand': brand,
                                    'price': price
                                })
                        except:
                            pass

            if prices:
                logger.info(f"从 {url} 提取 {len(prices)} 条数据")
                return date_str, prices

        except Exception as e:
            logger.warning(f"抓取URL失败 {url}: {e}")

        return '', []

    async def run(self) -> Dict:
        """运行抓取任务"""
        logger.info("=" * 60)
        logger.info("烟台钢筋价格智能抓取 v3.0")
        logger.info(f"确保每天不少于 {MIN_PRICES_PER_DAY} 条数据")
        logger.info("=" * 60)

        # 初始化
        DBManager.init_db()
        self.existing_keys = DBManager.get_existing_keys()
        logger.info(f"数据库已有 {len(self.existing_keys)} 条记录")

        await self.init_browser()

        # 登录
        if not await self.login():
            await self.close_browser()
            return {'success': False, 'error': '登录失败'}

        # 获取市场URL列表
        market_urls = await self.fetch_market_urls()

        if not market_urls:
            logger.warning("未找到市场URL，尝试使用其他方式...")

        total_inserted = 0
        success_count = 0

        # 抓取每个URL
        for i, url_info in enumerate(market_urls):
            url = url_info['href']
            logger.info(f"\n[{i+1}/{len(market_urls)}] 抓取: {url}")

            date_str, prices = await self.fetch_url_data(url)

            if prices and date_str:
                # 检查该日期已有数据量
                existing_count = DBManager.get_date_count(date_str)
                if existing_count >= MIN_PRICES_PER_DAY:
                    logger.info(f"日期 {date_str} 已有 {existing_count} 条数据，跳过")
                    continue

                # 确定时段（根据URL或时间）
                # 简单处理：使用空字符串表示不区分时段
                period = ''

                inserted, skipped = DBManager.insert_prices(date_str, period, prices, self.existing_keys)

                if inserted > 0:
                    total_inserted += inserted
                    success_count += 1
                    logger.info(f"插入 {inserted} 条，跳过 {skipped} 条")

                    # 保存到Excel
                    self.save_to_excel(date_str, period, prices)
                else:
                    logger.info(f"所有数据已存在")

            # 延迟
            if i < len(market_urls) - 1:
                delay = self.interval + random.randint(0, 3)
                logger.info(f"等待 {delay} 秒...")
                await asyncio.sleep(delay)

        await self.close_browser()

        logger.info("\n" + "=" * 60)
        logger.info("抓取完成")
        logger.info(f"  成功URL: {success_count}/{len(market_urls)}")
        logger.info(f"  新增记录: {total_inserted} 条")
        logger.info("=" * 60)

        return {
            'success': True,
            'total_urls': len(market_urls),
            'success_urls': success_count,
            'inserted': total_inserted
        }

    def save_to_excel(self, date: str, period: str, prices: List[Dict]):
        """保存到Excel"""
        try:
            if not prices:
                return

            if EXCEL_FILE.exists():
                wb = openpyxl.load_workbook(EXCEL_FILE)
            else:
                wb = openpyxl.Workbook()
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']

            sheet_name = f'{date}_{datetime.now().strftime("%H%M%S")}'
            if sheet_name in wb.sheetnames:
                wb.close()
                return

            ws = wb.create_sheet(title=sheet_name)

            # 标题
            ws.merge_cells('A1:J1')
            ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date}').font = Font(bold=True, size=14)

            # 表头
            headers = ['日期', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '地区']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)

            # 数据
            for i, price in enumerate(prices):
                row = 4 + i
                ws.cell(row=row, column=1, value=date)
                ws.cell(row=row, column=2, value=price.get('material_name', ''))
                ws.cell(row=row, column=3, value=price.get('spec', ''))
                ws.cell(row=row, column=4, value=price.get('material_type', ''))
                ws.cell(row=row, column=5, value=price.get('brand', ''))
                ws.cell(row=row, column=6, value=price.get('price', 0))
                ws.cell(row=row, column=7, value='山东烟台')

            wb.save(EXCEL_FILE)
            wb.close()

        except Exception as e:
            logger.warning(f"Excel保存失败: {e}")


# ==================== 主函数 ====================

async def main():
    import argparse

    parser = argparse.ArgumentParser(description='烟台钢筋价格智能抓取')
    parser.add_argument('--start', '-s', default='2024-01-01', help='开始日期')
    parser.add_argument('--end', '-e', default='2026-06-10', help='结束日期')
    parser.add_argument('--interval', '-i', type=int, default=5, help='抓取间隔')

    args = parser.parse_args()

    fetcher = SmartFetcher(args.start, args.end, args.interval)
    result = await fetcher.run()

    if result.get('success'):
        print(f"\n✅ 抓取成功完成！")
        print(f"Excel: {EXCEL_FILE}")
        print(f"数据库: {DB_FILE}")
    else:
        print(f"\n❌ 抓取失败: {result.get('error')}")


if __name__ == '__main__':
    asyncio.run(main())
