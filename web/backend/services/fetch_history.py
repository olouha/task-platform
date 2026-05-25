"""
批量抓取历史价格数据
"""

import asyncio
from playwright.async_api import async_playwright
from pathlib import Path
import json
import re
import base64
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

DATA_DIR = Path('services/data')
COOKIE_FILE = DATA_DIR / 'mysteel_cookies.json'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
MARKET_URL = 'https://jiancai.mysteel.com/market/pa228aa010101a0a01010205aaaa1.html'

# 硬编码的可用URL（从市场页面获取）
AVAILABLE_URLS = [
    ('2026-05-15', 'AM', 'https://jiancai.mysteel.com/m/26051510/19B77109BDE6183C.html'),
    ('2026-05-15', 'PM', 'https://jiancai.mysteel.com/m/26051516/06AC8B0B0D2BB9BF.html'),
    ('2026-05-14', 'AM', 'https://jiancai.mysteel.com/m/26051410/25B3355C6617BD3C.html'),
    ('2026-05-14', 'PM', 'https://jiancai.mysteel.com/m/26051416/C9998D9EC5FE17F0.html'),
    ('2026-05-13', 'AM', 'https://jiancai.mysteel.com/m/26051310/C7E274318523C3AE.html'),
    ('2026-05-13', 'PM', 'https://jiancai.mysteel.com/m/26051316/B7EAA4BE8AB3DA35.html'),
    ('2026-05-12', 'AM', 'https://jiancai.mysteel.com/m/26051210/EE0F2403F1F3F713.html'),
    ('2026-05-12', 'PM', 'https://jiancai.mysteel.com/m26051216/C62C65295C20B15B.html'),
    ('2026-05-11', 'AM', 'https://jiancai.mysteel.com/m/26051110/99771C0982C1F177.html'),
    ('2026-05-09', 'AM', 'https://jiancai.mysteel.com/m/26050910/5E9741D6C569CA44.html'),
    ('2026-05-08', 'AM', 'https://jiancai.mysteel.com/m/26050810/C4A4549218C84CA8.html'),
    ('2026-05-07', 'AM', 'https://jiancai.mysteel.com/m/26050710/885B555DD260EF7B.html'),
    ('2026-05-06', 'AM', 'https://jiancai.mysteel.com/m/26050610/B37F4169FEA928E4.html'),
    ('2026-04-30', 'AM', 'https://jiancai.mysteel.com/m/26043010/13E96DB3D3E67D07.html'),
    ('2026-04-30', 'PM', 'https://jiancai.mysteel.com/m/26043015/7AE1A0219402C96C.html'),
    ('2026-04-29', 'AM', 'https://jiancai.mysteel.com/m/26042910/B642D57F54B3CC2F.html'),
    ('2026-04-28', 'AM', 'https://jiancai.mysteel.com/m/26042810/4D3396D3F9A2E4C8.html'),
    ('2026-04-28', 'PM', 'https://jiancai.mysteel.com/m/26042816/65A30D40DCD35CEF.html'),
    ('2026-04-27', 'AM', 'https://jiancai.mysteel.com/m/26042710/150F8A9D3B6A654B.html'),
    ('2026-04-27', 'PM', 'https://jiancai.mysteel.com/m/26042716/448B0C035FB7BE62.html'),
    ('2026-04-24', 'AM', 'https://jiancai.mysteel.com/m/26042410/E0D5A0D891C7BC12.html'),
    ('2026-04-24', 'PM', 'https://jiancai.mysteel.com/m/26042416/19B4ADC50620FD54.html'),
    ('2026-04-23', 'AM', 'https://jiancai.mysteel.com/m/26042310/3B83E3A4188ADD02.html'),
    ('2026-04-22', 'AM', 'https://jiancai.mysteel.com/m/26042210/0280430126617D73.html'),
    ('2026-04-21', 'AM', 'https://jiancai.mysteel.com/m/26042110/BD4C0673054A09E4.html'),
    ('2026-04-20', 'AM', 'https://jiancai.mysteel.com/m/26042010/984E6D78F31958CD.html'),
    ('2026-04-20', 'PM', 'https://jiancai.mysteel.com/m/26042015/562AB74C7B9C1957.html'),
    ('2026-04-17', 'AM', 'https://jiancai.mysteel.com/m/26041710/BA5BC57E9F2DEC78.html'),
    ('2026-04-17', 'PM', 'https://jiancai.mysteel.com/m/26041716/7B3CBAC69A611AD1.html'),
]


def get_sheet_name(date_str, period):
    fetch_time = datetime.now().strftime('%H%M%S')
    return f'{date_str}_{period}_{fetch_time}'


async def save_to_excel(prices, period, date_str, screenshot_b64):
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    sheet_name = get_sheet_name(date_str, period)
    ws = wb.create_sheet(title=sheet_name)

    # 标题
    period_text = '下午' if period == 'PM' else '上午'
    ws.merge_cells('A1:K1')
    ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    # 表头
    header_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid') if period == 'PM' else \
        PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 数据
    fetch_time = datetime.now().strftime('%H:%M:%S')
    for i, price in enumerate(prices):
        row = 4 + i
        for col, val in enumerate([date_str, fetch_time, price['material_name'], price['spec'],
                                   price['material_type'], price['brand'], price['price'],
                                   '', '', '', '山东烟台'], 1):
            ws.cell(row=row, column=col, value=val)

    # 截图
    if screenshot_b64:
        screenshot_path = DATA_DIR / f'screenshot_{date_str.replace("-", "")}_{period}.png'
        with open(screenshot_path, 'wb') as f:
            f.write(base64.b64decode(screenshot_b64))

        row = 4 + len(prices) + 2
        ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)
        img = Image(str(screenshot_path))
        img.width = 900
        img.height = 500
        img.anchor = f'A{row + 1}'
        ws.add_image(img)

    wb.save(EXCEL_FILE)
    wb.close()
    return sheet_name


async def fetch_prices(page, url):
    await page.goto(url, wait_until='domcontentloaded', timeout=60000)
    await page.wait_for_timeout(5000)

    screenshot = await page.screenshot(full_page=True)
    screenshot_b64 = base64.b64encode(screenshot).decode('utf-8')

    data = await page.evaluate('''() => {
        const tables = document.querySelectorAll('table');
        const results = [];
        tables.forEach(table => {
            const rows = table.querySelectorAll('tr');
            rows.forEach(row => {
                const cells = row.querySelectorAll('td');
                if (cells.length >= 5) {
                    const material_name = cells[0]?.textContent?.trim();
                    const spec = cells[1]?.textContent?.trim();
                    const material_type = cells[2]?.textContent?.trim();
                    const brand = cells[3]?.textContent?.trim();
                    const price = cells[4]?.textContent?.trim();

                    if (['高线', '螺纹钢', '盘螺', '圆钢'].includes(material_name) &&
                        spec && spec.startsWith('Phi') && price && /^\\d+$/.test(price)) {
                        results.push({
                            material_name,
                            spec,
                            material_type,
                            brand,
                            price: parseFloat(price),
                            region: '山东烟台'
                        });
                    }
                }
            });
        });
        return results;
    }''')

    return list(data), screenshot_b64


async def main():
    # 获取已存在的sheet
    existing_sheets = set()
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        existing_sheets = set(wb.sheetnames)
        wb.close()

    print(f'已有sheet: {len(existing_sheets)}个')

    # 过滤已抓取的
    tasks = []
    for date_str, period, url in AVAILABLE_URLS:
        # 检查是否已有该日期该时段的sheet
        has_existing = any(s.startswith(f'{date_str}_{period}_') for s in existing_sheets)
        if not has_existing:
            tasks.append((date_str, period, url))

    print(f'需要抓取: {len(tasks)} 条')
    if not tasks:
        print('所有数据已抓取')
        return

    # 登录
    cookies = []
    if COOKIE_FILE.exists():
        cookies = json.loads(COOKIE_FILE.read_text(encoding='utf-8'))

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(viewport={'width': 1920, 'height': 3000}, locale='zh-CN')
        if cookies:
            await context.add_cookies(cookies)
        page = await context.new_page()

        print('登录中...')
        await page.goto('https://passport.mysteel.com/', wait_until='domcontentloaded', timeout=60000)
        await page.wait_for_timeout(5000)

        username = 'M6616592358'
        password = 'mysteel573005'

        await page.evaluate(f'''() => {{
            const inputs = document.querySelectorAll('input');
            for (const inp of inputs) {{
                const ph = inp.placeholder || '';
                if (ph.includes('用户名')) inp.value = '{username}';
                if (ph.includes('密码') && inp.type === 'password') inp.value = '{password}';
            }}
        }}''')

        await page.wait_for_timeout(500)

        try:
            login_btn = await page.query_selector('.form-button-login, button:has-text("登录")')
            if login_btn:
                await login_btn.click()
        except:
            pass

        await page.wait_for_timeout(8000)
        print('登录完成\n')

        results = {}
        for i, (date_str, period, url) in enumerate(tasks):
            print(f'[{i+1}/{len(tasks)}] {date_str} {period}...', end='')

            try:
                prices, screenshot_b64 = await fetch_prices(page, url)
                if prices:
                    sheet_name = await save_to_excel(prices, period, date_str, screenshot_b64)
                    results[date_str] = results.get(date_str, {})
                    results[date_str][period] = {'sheet': sheet_name, 'count': len(prices)}
                    print(f' {len(prices)}条')
                else:
                    print(' 无数据')
            except Exception as e:
                print(f' 失败: {e}')

        await browser.close()

    print()
    print('=' * 60)
    print('抓取完成')
    print('=' * 60)
    for date_str in sorted(results.keys(), reverse=True):
        am = results[date_str].get('AM', {}).get('count', 0)
        pm = results[date_str].get('PM', {}).get('count', 0)
        print(f'{date_str}: AM={am}, PM={pm}')


if __name__ == '__main__':
    asyncio.run(main())