"""
合并两个Excel文件并更新SQLite数据库
"""
import sqlite3
import openpyxl
import re
from pathlib import Path
from datetime import datetime

DATA_DIR = Path('services/data')
DB_FILE = DATA_DIR / 'yantai_rebar.db'

def init_db():
    """初始化数据库"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('DROP TABLE IF EXISTS rebar_prices')
    c.execute('''
        CREATE TABLE rebar_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            fetch_time TEXT,
            material_name TEXT,
            spec TEXT,
            material_type TEXT,
            brand TEXT,
            price INTEGER,
            price_change TEXT,
            remark TEXT,
            region TEXT DEFAULT '山东烟台'
        )
    ''')
    c.execute('CREATE INDEX idx_date ON rebar_prices(date)')
    c.execute('CREATE INDEX idx_material ON rebar_prices(material_name)')
    c.execute('CREATE INDEX idx_spec ON rebar_prices(spec)')
    conn.commit()
    return conn

def parse_date(sheet_name):
    """从sheet名提取日期"""
    if '-' in sheet_name:
        parts = sheet_name.split('_')
        return parts[0]
    return None

def import_from_wb(conn, wb, source_name):
    """从workbook导入数据"""
    c = conn.cursor()
    data = []
    count = 0

    for sheet_name in wb.sheetnames:
        date = parse_date(sheet_name)
        if not date:
            continue
        try:
            datetime.strptime(date, '%Y-%m-%d')
        except:
            continue

        ws = wb[sheet_name]

        # 检测格式 - 检查Row1是否包含'品名'
        first_row_val = ws.cell(row=1, column=1).value
        if first_row_val and '品名' in str(first_row_val):
            # 最终版格式（无标题行）：Row1=品名, Row2=规格, ...
            start_row = 1
        elif first_row_val and '山东烟台' in str(first_row_val):
            # 完整版格式（有标题）：Row3开始是数据
            start_row = 3
        else:
            # 默认从Row1开始
            start_row = 1

        for row_idx in range(start_row, ws.max_row + 1):
            row_data = [ws.cell(row=row_idx, column=col).value for col in range(1, 8)]

            # 尝试不同的列布局
            material = None
            spec = None
            mat_type = None
            price_val = None

            # 最终版格式：品名,规格,材质,价格,...
            if len(row_data) >= 4:
                material = row_data[0]
                spec = row_data[1]
                mat_type = row_data[2]
                price_val = row_data[3]

            if not material or not isinstance(material, str):
                continue
            if '品名' in material or '截图' in material or '状态' in material:
                continue

            price = 0
            if price_val:
                if isinstance(price_val, (int, float)):
                    price = int(price_val)
                else:
                    match = re.search(r'(\d{3,5})', str(price_val))
                    if match:
                        price = int(match.group(1))

            if price > 0:
                data.append((date, '', str(material) if material else '', str(spec) if spec else '',
                           str(mat_type) if mat_type else '', '', price, '', '', '山东烟台'))
                count += 1

    return data

def main():
    print('=' * 60)
    print('合并数据并更新数据库')
    print('=' * 60)

    conn = init_db()
    all_data = []

    # 导入最终版
    print('\n处理: 最终版.xlsx')
    final_wb = openpyxl.load_workbook(DATA_DIR / '山东烟台钢筋价格_最终版.xlsx', read_only=True, data_only=True)
    data1 = import_from_wb(conn, final_wb, '最终版')
    print(f'  读取: {len(data1)} 条')
    all_data.extend(data1)
    final_wb.close()

    # 导入完整版
    print('\n处理: 完整版_数据+截图.xlsx')
    complete_wb = openpyxl.load_workbook(DATA_DIR / '山东烟台钢筋价格_完整版_数据+截图.xlsx', read_only=True, data_only=True)
    data2 = import_from_wb(conn, complete_wb, '完整版')
    print(f'  读取: {len(data2)} 条')
    all_data.extend(data2)
    complete_wb.close()

    # 去重
    print(f'\n总共: {len(all_data)} 条，开始去重...')
    unique_data = {}
    for row in all_data:
        date, _, material, spec, mat_type, brand, price, *rest = row
        key = (date, material, spec)
        if key not in unique_data or unique_data[key][6] < price:
            unique_data[key] = row

    final_data = list(unique_data.values())
    print(f'去重后: {len(final_data)} 条')

    # 插入数据库
    print('\n插入数据库...')
    c = conn.cursor()
    c.executemany('''INSERT INTO rebar_prices
        (date, fetch_time, material_name, spec, material_type, brand, price, price_change, remark, region)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', final_data)
    conn.commit()

    # 统计
    print('\n' + '=' * 60)
    print('数据库统计')
    print('=' * 60)
    print(f'总记录: {c.execute("SELECT COUNT(*) FROM rebar_prices").fetchone()[0]}')
    print(f'日期数: {c.execute("SELECT COUNT(DISTINCT date) FROM rebar_prices").fetchone()[0]}')
    r = c.execute('SELECT MIN(date), MAX(date) FROM rebar_prices').fetchone()
    print(f'范围: {r[0]} 到 {r[1]}')

    print('\n材料统计:')
    for row in c.execute('SELECT material_name, COUNT(*) FROM rebar_prices GROUP BY material_name ORDER BY COUNT(*) DESC'):
        print(f'  {row[0]}: {row[1]}条')

    print('\n规格统计:')
    for row in c.execute('SELECT spec, COUNT(*) FROM rebar_prices GROUP BY spec ORDER BY COUNT(*) DESC LIMIT 10'):
        print(f'  {row[0]}: {row[1]}条')

    conn.close()
    print(f'\n完成! 数据库: {DB_FILE}')

if __name__ == '__main__':
    main()