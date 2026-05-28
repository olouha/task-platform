# -*- coding: utf-8 -*-
"""
OCR 图片识别脚本 - 钢筋混凝土造价参考价
基于坐标列识别，精确提取价格数据
"""
import os
import re
from paddleocr import PaddleOCR
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from collections import defaultdict

BASE_DIR = r'C:\Users\admin\Desktop\近五年钢筋混凝土造价管理截图(1)\近五年钢筋混凝土造价管理截图'

ocr = PaddleOCR(use_angle_cls=True, lang='ch', show_log=False)

def extract_rebar_spec(text):
    """从钢筋文本中提取规格（HPB/HRB格式）

    格式:
    - HPB3006.5 -> 6.5
    - HPB30065 -> 6.5
    - HPB3008 -> 8
    - HPB30010 -> 10
    - HRB400≤12 -> 12
    - HRB40018 -> 18
    - HRB400中6 -> 6 (中=直径6)
    - HRB4008 -> 8 (2位数字)
    """
    # 清理文本中的干扰字符
    clean_text = text.replace('≤', '≤').replace('＞', '>').replace('＜', '<')

    # 1. HRB/HPB + 符号 + 数字 (如 HRB400≤12, HPB300>12)
    match = re.search(r'(?:HPB|HRB)\d*(?:≤|>|＜)\s*(\d+)', clean_text)
    if match:
        return match.group(1)

    # 2. HPB300 + 小数点数字（如HPB3006.5）
    match = re.search(r'HPB300(\d+(?:\.\d+)?)', clean_text)
    if match:
        num_str = match.group(1)
        if '.' in num_str:
            return num_str  # 直接返回 '6.5'

        # 处理无小数点的数字
        num_str = num_str.lstrip('0')
        if not num_str:
            return None

        if len(num_str) == 2:
            if num_str == '65':
                return '6.5'
            elif num_str.startswith('0'):
                return num_str[1]  # '08' -> '8'
            else:
                return num_str

        if len(num_str) == 4 and num_str.startswith('3'):
            return num_str[1:]  # '3010' -> '10'

        return num_str if int(num_str) <= 50 else None

    # 3. HRB400 + 数字（包括中/≤/>/等符号后的情况）
    # 匹配 HRB400 + 可选符号(中/≤/>/) + 数字
    match = re.search(r'HRB400[中中≤＞<]*(\d+)', clean_text)
    if match:
        num_str = match.group(1)
        return num_str  # 直接返回数字部分

    # 4. HRB400 + 2位数字 (如 HRB40018 -> 18)
    match = re.search(r'HRB400(\d{2})', clean_text)
    if match:
        return match.group(1)

    # 5. HRB400 + 3位数字以4开头 (如 HRB40018)
    match = re.search(r'HRB400(4\d{2})', clean_text)
    if match:
        return match.group(1)[1:]  # '418' -> '18'

    return None

def ocr_image(img_path):
    """OCR单张图片，返回带坐标的文本"""
    result = ocr.ocr(img_path, cls=True)
    if not result or not result[0]:
        return []

    items = []
    for line in result[0]:
        bbox, (text, conf) = line
        x1, y1 = bbox[0]
        x2, y2 = bbox[2]
        items.append({
            'x': (x1 + x2) / 2,
            'y': (y1 + y2) / 2,
            'text': text.strip(),
            'conf': conf
        })
    items.sort(key=lambda k: (round(k['y'] / 15) * 15, k['x']))
    return items

def parse_concrete_image(img_path):
    """解析混凝土截图 - 基于X坐标列"""
    items = ocr_image(img_path)

    data = []
    current_grade = None
    row_prices = {}

    for item in items:
        text = item['text']
        x, y = item['x'], item['y']

        grade_match = re.match(r'^C(\d+)$', text)
        if grade_match and 50 < x < 140:
            if current_grade and row_prices:
                data.append({'grade': current_grade, **row_prices})
            current_grade = text
            row_prices = {}

        if re.match(r'^\d{3,4}$', text):
            if 150 < x < 250:
                row_prices['yantai'] = text
            elif 280 < x < 360:
                row_prices['rushan'] = text

    if current_grade and row_prices:
        data.append({'grade': current_grade, **row_prices})

    return data

def parse_rebar_image(img_path):
    """解析钢筋截图 - 提取规格和价格"""
    items = ocr_image(img_path)

    data = []
    rows = defaultdict(list)

    # 按Y坐标分组到行 - 使用较小的行高以保持精度
    for item in items:
        y_group = round(item['y'] / 15) * 15
        rows[y_group].append(item)

    # 建立Y坐标到行的索引
    row_by_y = {}
    for y in sorted(rows.keys()):
        row_items = sorted(rows[y], key=lambda k: k['x'])
        row_text = ' '.join([item['text'] for item in row_items])
        row_by_y[y] = {'items': row_items, 'text': row_text}

    for y in sorted(rows.keys()):
        row_items = row_by_y[y]['items']
        row_text = row_by_y[y]['text']

        # 查找钢筋行：包含"钢筋"关键词
        if '钢筋' not in row_text:
            continue

        # 跳过"冷轧"、"螺纹"、"预应力"等非普通钢筋
        skip_patterns = ['冷轧', '预应力', '镀锌', '钢丝', '钢丝绳', '环氧', '钢航']
        if any(p in row_text for p in skip_patterns):
            continue

        # 提取等级：HPB 或 HRB
        grade = None
        if 'HPB' in row_text:
            grade = 'HPB'
        elif 'HRB' in row_text:
            grade = 'HRB'
        elif '螺纹' in row_text:
            grade = 'HRB'  # 螺纹钢筋通常是HRB
        else:
            # 尝试从上下文推断等级
            for offset in [-15, -30, -45, 15, 30, 45]:
                check_y = y + offset
                if check_y in row_by_y:
                    check_text = row_by_y[check_y]['text']
                    if 'HPB' in check_text:
                        grade = 'HPB'
                        break
                    elif 'HRB' in check_text:
                        grade = 'HRB'
                        break

        if not grade:
            grade = 'HPB'  # 默认HPB

        # 提取规格
        size = extract_rebar_spec(row_text)
        if not size:
            # 尝试纯数字格式（钢筋后面直接跟数字）
            simple_match = re.search(r'钢筋\s*(?:HRB|HPB)?[\s一-龥]*(\d+)', row_text)
            if simple_match:
                size = simple_match.group(1)
            else:
                num_match = re.search(r'钢筋(\d+(?:\.\d+)?)', row_text)
                if num_match:
                    size = num_match.group(1)

        if not size or not re.match(r'^\d+(?:\.\d+)?$', size):
            continue

        # 规格范围 6-50
        try:
            size_num = float(size)
            if not (6 <= size_num <= 50):
                continue
        except:
            continue

        # 找价格：
        # 1. 在当前行X > 250区域找数字（包含不同格式的截图）
        price = None
        price_y = None
        for item in row_items:
            text = item['text']
            x = item['x']
            if x > 250:
                # 匹配价格: 5280.00, 7450.00, 5,280.00 等（4位数+.00格式）
                price_match = re.match(r'^(\d{3,5}\. ?\d{2})$', text)
                if price_match:
                    # 处理带空格的情况，如 "4540. 00"
                    clean_price = text.replace(' ', '').replace(',', '')
                    price = clean_price.split('.')[0]
                    price_y = y
                    break
                # 匹配带逗号的价格: 5,280.00
                price_match2 = re.match(r'^(\d,\d{3}\. ?\d{2})$', text)
                if price_match2:
                    price = text.replace(',', '').replace(' ', '').split('.')[0]
                    price_y = y
                    break
                # 匹配价格+税率混合文本: 4031.3113.00% 或 4143.2913.00%
                price_match3 = re.search(r'(\d{3,5}\.\d{2})\d*13\. ?00%', text)
                if price_match3:
                    price = price_match3.group(1).split('.')[0]
                    price_y = y
                    break
                # 匹配纯价格数字（如4069.16）和税率分开的情况）
                price_match4 = re.match(r'^(\d{3,5}\.\d{2})$', text)
                if price_match4 and not text.startswith('17'):  # 跳过单位价格如17.67
                    price = price_match4.group(1).split('.')[0]
                    price_y = y
                    break
                # 三位数带小数点: 5.28 (单位kg)
                if re.match(r'^\d+\.\d{2}$', text):
                    continue  # 跳过单位价格

        # 2. 如果当前行没找到，检查上下行
        if not price or not price_y:
            for offset in [-15, -30, 15, 30, 45]:
                check_y = y + offset
                if check_y in row_by_y:
                    prev_items = row_by_y[check_y]['items']
                    for item in prev_items:
                        text = item['text']
                        x = item['x']
                        # 价格通常在X=250-550区域
                        if 200 < x < 600:
                            price_match = re.match(r'^(\d{3,5}\. ?\d{2})$', text)
                            if price_match:
                                clean_price = text.replace(' ', '').replace(',', '')
                                price = clean_price.split('.')[0]
                                price_y = check_y
                                break
                            price_match2 = re.match(r'^(\d,\d{3}\. ?\d{2})$', text)
                            if price_match2:
                                price = text.replace(',', '').replace(' ', '').split('.')[0]
                                price_y = check_y
                                break
                            # 匹配价格+税率混合文本
                            price_match3 = re.search(r'(\d{3,5}\.\d{2})\d*13\. ?00%', text)
                            if price_match3:
                                price = price_match3.group(1).split('.')[0]
                                price_y = check_y
                                break
                            # 匹配纯价格数字
                            price_match4 = re.match(r'^(\d{3,5}\.\d{2})$', text)
                            if price_match4 and not text.startswith('17'):
                                price = price_match4.group(1).split('.')[0]
                                price_y = check_y
                                break
                    if price:
                        break

        if price:
            data.append({
                'grade': grade if grade else 'HPB',  # 默认HPB
                'size': str(int(size_num) if size_num == int(size_num) else size_num),
                'spec': row_text[:60],
                'price': price
            })

    return data

def process_folder(year, quarter, folder_path):
    """处理一个季度文件夹"""
    print(f'  {year} {quarter}: ', end='')

    concrete_data = []
    rebar_data = []

    # 混凝土
    concrete_img = os.path.join(folder_path, '混凝土.png')
    if os.path.exists(concrete_img):
        concrete_data = parse_concrete_image(concrete_img)

    # 钢筋
    rebar_files = sorted([f for f in os.listdir(folder_path)
                         if f.startswith('钢筋') and f.endswith('.png')])
    for rebar_file in rebar_files:
        rebar_path = os.path.join(folder_path, rebar_file)
        data = parse_rebar_image(rebar_path)
        rebar_data.extend(data)

    print(f'混凝土 {len(concrete_data)}条, 钢筋 {len(rebar_data)}条')
    return concrete_data, rebar_data

def create_excel(all_concrete, all_rebar):
    """生成Excel"""
    wb = Workbook()

    ws_c = wb.active
    ws_c.title = '混凝土信息价'
    ws_c.append(['年份', '季度', '强度等级', '烟台含税(元/m³)', '蓬莱含税(元/m³)'])

    for year in sorted(all_concrete.keys()):
        for quarter in sorted(all_concrete[year].keys()):
            for item in all_concrete[year][quarter]:
                ws_c.append([
                    year, quarter, item['grade'],
                    item.get('yantai', ''), item.get('rushan', '')
                ])

    ws_r = wb.create_sheet('钢筋信息价')
    ws_r.append(['年份', '季度', '等级', '规格(mm)', '价格(含税元/吨)'])

    for year in sorted(all_rebar.keys()):
        for quarter in sorted(all_rebar[year].keys()):
            for item in all_rebar[year][quarter]:
                ws_r.append([year, quarter, item['grade'], item['size'], item['price']])

    for ws in [ws_c, ws_r]:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    from datetime import datetime
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output = os.path.join(BASE_DIR, f'造价参考价数据_{timestamp}.xlsx')
    wb.save(output)
    print(f'\n已生成: {output}')

def main():
    all_concrete = defaultdict(lambda: defaultdict(list))
    all_rebar = defaultdict(lambda: defaultdict(list))

    for year_folder in sorted(os.listdir(BASE_DIR)):
        year_path = os.path.join(BASE_DIR, year_folder)
        if not os.path.isdir(year_path):
            continue

        year_match = re.match(r'^(\d{4})年$', year_folder)
        if not year_match:
            continue
        year = year_match.group(1)

        for quarter_folder in sorted(os.listdir(year_path)):
            quarter_path = os.path.join(year_path, quarter_folder)
            if not os.path.isdir(quarter_path):
                continue

            q_match = re.search(r'第(.+?)季度', quarter_folder)
            quarter = f'第{q_match.group(1)}季度' if q_match else quarter_folder

            concrete, rebar = process_folder(year, quarter, quarter_path)

            if concrete:
                all_concrete[year][quarter] = concrete
            if rebar:
                all_rebar[year][quarter] = rebar

    create_excel(all_concrete, all_rebar)
    print('\n处理完成!')

if __name__ == '__main__':
    main()