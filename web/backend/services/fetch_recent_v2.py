"""抓取最近日期数据"""
import asyncio
import json
import base64
import random
from pathlib import Path
from datetime import datetime, timedelta
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

DATA_DIR = Path('services/data')
COOKIE_FILE = DATA_DIR / 'mysteel_cookies.json'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格_完整版.xlsx'

USERNAME = 'M6616592358'
PASSWORD = 'panhui199261'

LOG_FILE = DATA_DIR / 'logs' / f'fetch_recent_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'

# 正确的hash值（从已成功的URL中获取）
HASH_AM = '25B3355C6617BD3C'

def log(msg):
    print(f'[{datetime.now().strftime("%H:%M:%S")}] {msg}')
    LOG_FILE.parent.mkdir(exist_ok=True)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(f'[{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}] {msg}\n')

def random_sleep():
    return random.uniform(8, 12)

def get_url(date_str, period):
    """生成正确格式的URL"""
    date_part = date_str.replace('-', '')
    suffix = '10' if period == 'AM' else '16'
    return f'https://jiancai.mysteel.com/m/{date_part}{suffix}/{HASH_AM}.html'

async def login(page, context):
    log('登录中...')
    await page.goto('https://passport.mysteel.com/', wait_until='networkidle', timeout=60000)
    await page.wait_for_timeout(3000)
    try:
        tab = await page.query_selector('text=账号登录')
        if tab:
            await tab.click()
            await page.wait_for_timeout(2000)
    except:
        pass
    await page.evaluate(f'''
        () => {{
            const inputs = document.querySelectorAll('input');
            for (const inp of inputs) {{
                const ph = inp.placeholder || '';
                if (ph.includes('用户名') || ph.includes('手机')) inp.value = '{USERNAME}';
                if (ph.includes('密码') && inp.type === 'password') inp.value = '{PASSWORD}';
            }}
        }}
    ''')
    await page.wait_for_timeout(1000)
    try:
        btn = await page.query_selector('.form-button-login, button:has-text("登录")')
        if btn:
            await btn.click()
            for i in range(20):
                await page.wait_for_timeout(1000)
                if 'passport' not in page.url:
                    break
    except:
        pass
    cookies = await context.cookies()
    COOKIE_FILE.write_text(json.dumps(cookies, ensure_ascii=False), encoding='utf-8')
    log(f'登录成功，Cookie: {len(cookies)}条')

async def fetch_price_data(page, url, date_str, period):
    try:
        log(f'  访问: {url}')
        await page.goto(url, wait_until='domcontentloaded', timeout=60000)
        await page.wait_for_timeout(5000)

        for i in range(5):
            await page.evaluate('window.scrollBy(0, 500)')
            await page.wait_for_timeout(500)

        screenshot_b64 = None
        try:
            screenshot = await page.screenshot(full_page=True)
            screenshot_b64 = base64.b64encode(screenshot).decode()
        except:
            pass

        data = await page.evaluate('''
            () => {
                const results = [];
                const tables = document.querySelectorAll('table');
                tables.forEach(table => {
                    const rows = table.querySelectorAll('tr');
                    rows.forEach(row => {
                        const cells = row.querySelectorAll('td');
                        if (cells.length >= 5) {
                            const cells_text = [];
                            cells.forEach(c => cells_text.push(c.textContent.trim()));
                            results.push(cells_text);
                        }
                    });
                });
                return results;
            }
        ''')

        prices = []
        for row in data:
            if len(row) >= 5:
                material = row[0].strip()
                spec = row[1].strip() if len(row) > 1 else ''
                mat_type = row[2].strip() if len(row) > 2 else ''
                brand = row[3].strip() if len(row) > 3 else ''
                price_text = row[4].strip() if len(row) > 4 else ''
                valid_materials = ['高线', '螺纹钢', '盘螺', '圆钢', '拉丝材']
                if any(m in material for m in valid_materials):
                    import re
                    price_match = re.search(r'(\d{3,5})', price_text)
                    if price_match:
                        price = int(price_match.group(1))
                        prices.append({
                            'material_name': material,
                            'spec': spec,
                            'material_type': mat_type,
                            'brand': brand,
                            'price': price
                        })
        return prices, screenshot_b64
    except Exception as e:
        log(f'  抓取失败: {e}')
        return [], None

def save_to_excel(prices, period, date_str, screenshot_b64):
    try:
        if EXCEL_FILE.exists():
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE)
            except:
                log('  Excel文件损坏')
                return False
        else:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
        sheet_name = date_str if period == 'AM' else f'{date_str}_PM'
        if sheet_name in wb.sheetnames:
            log(f'  Sheet已存在，跳过')
            wb.close()
            return False
        ws = wb.create_sheet(title=sheet_name[:31])
        header_font = Font(bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid') if period == 'AM' else \
                      PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        period_text = '下午' if period == 'PM' else '上午'
        ws.merge_cells('A1:K1')
        ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {date_str} {period_text}').font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        fetch_time = datetime.now().strftime('%H:%M:%S')
        for i, price in enumerate(prices):
            row = 4 + i
            for col, val in enumerate([date_str, fetch_time, price['material_name'], price['spec'],
                                       price['material_type'], price['brand'], price['price'],
                                       '', '', '', '山东烟台'], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = border
        if screenshot_b64:
            screenshot_path = DATA_DIR / f'screenshot_{date_str.replace("-", "")}_{period}.png'
            with open(screenshot_path, 'wb') as f:
                f.write(base64.b64decode(screenshot_b64))
            row = 4 + len(prices) + 2
            ws.cell(row=row, column=1, value='当日截图').font = Font(bold=True, size=12)
            img_obj = Image(str(screenshot_path))
            img_obj.width = 900
            img_obj.height = 500
            img_obj.anchor = f'A{row + 1}'
            ws.add_image(img_obj)
        wb.save(EXCEL_FILE)
        wb.close()
        return True
    except Exception as e:
        log(f'  保存失败: {e}')
        return False

async def main():
    print()
    print('=' * 70)
    print('抓取最近日期数据')
    print('=' * 70)
    print()

    # 需要抓取的日期
    missing_dates = [
        ('2026-05-21', 'AM'),
        ('2026-05-22', 'AM'),
        ('2026-05-25', 'AM'),
        ('2026-05-26', 'AM'),
        ('2026-05-27', 'AM'),
    ]

    # 生成URL
    urls = [(date, period, get_url(date, period)) for date, period in missing_dates]
    log(f'需要抓取: {len(urls)} 个日期')

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(viewport={'width': 1920, 'height': 3000}, locale='zh-CN')
        page = await context.new_page()
        await login(page, context)

        success = 0
        fail = 0

        for date_str, period, url in urls:
            log(f'{date_str} {period}...')
            print(f'  -> 抓取中...', end='', flush=True)
            prices, screenshot = await fetch_price_data(page, url, date_str, period)
            if prices:
                if save_to_excel(prices, period, date_str, screenshot):
                    log(f' {len(prices)}条 [OK]')
                    success += 1
                else:
                    log(' 保存失败')
                    fail += 1
            else:
                log(' 无数据')
                fail += 1
            await asyncio.sleep(random_sleep())

        await browser.close()

    print()
    print('=' * 70)
    log(f'成功: {success} 条')
    log(f'失败: {fail} 条')

if __name__ == '__main__':
    asyncio.run(main())