"""
从Excel文件导入2025年钢筋价格数据到SQLite数据库
"""
import sqlite3
import openpyxl
from pathlib import Path
from datetime import datetime
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 配置
EXCEL_FILE = Path(__file__).parent / 'services' / 'data' / '山东烟台钢筋价格_完整版_数据+截图.xlsx'
DB_FILE = Path(__file__).parent / 'data' / 'yantai_rebar.db'

def import_2025_data():
    """从Excel导入2025年数据"""
    logger.info(f"[import] 开始导入2025年数据")
    logger.info(f"Excel: {EXCEL_FILE}")
    logger.info(f"DB: {DB_FILE}")

    # 打开Excel文件
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    # 筛选2025年的sheet
    sheets_2025 = [name for name in wb.sheetnames if name.startswith('2025')]
    logger.info(f"[import] 找到2025年数据: {len(sheets_2025)}个交易日")

    # 连接数据库
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    inserted = 0
    skipped = 0
    errors = []

    for sheet_name in sheets_2025:
        try:
            ws = wb[sheet_name]

            # 读取数据（从第4行开始）
            # Excel结构: 列1=品名, 列2=规格, 列3=材质, 列4=价格, 列7=时段
            for row_num in range(4, ws.max_row + 1):
                material_name = ws.cell(row=row_num, column=1).value
                spec = ws.cell(row=row_num, column=2).value
                material_type = ws.cell(row=row_num, column=3).value
                price = ws.cell(row=row_num, column=4).value
                fetch_time_raw = ws.cell(row=row_num, column=7).value

                # 跳过空行
                if not material_name:
                    break

                # 日期就是sheet名称
                date = sheet_name

                # 确定时段
                fetch_time = None
                if isinstance(fetch_time_raw, str):
                    if '09:00' in fetch_time_raw or '上午' in fetch_time_raw:
                        fetch_time = '09:00'
                    elif '15:00' in fetch_time_raw or '下午' in fetch_time_raw:
                        fetch_time = 'PM'

                # 插入数据
                try:
                    cursor.execute('''
                        INSERT INTO rebar_prices
                        (date, fetch_time, material_name, spec, material_type, brand, price, region)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        str(date),
                        fetch_time,
                        str(material_name),
                        str(spec),
                        str(material_type) if material_type else '',
                        '',  # 品牌列为空
                        int(price) if price and str(price).isdigit() else 0,
                        '山东烟台'
                    ))
                    inserted += 1
                except sqlite3.IntegrityError:
                    skipped += 1
                except Exception as e:
                    errors.append(f"{sheet_name} row {row_num}: {e}")

            # 每10个sheet提交一次
            if (sheets_2025.index(sheet_name) + 1) % 10 == 0:
                conn.commit()
                logger.info(f"[import] 进度: {sheets_2025.index(sheet_name) + 1}/{len(sheets_2025)}")

        except Exception as e:
            logger.error(f"[import] 处理sheet失败 {sheet_name}: {e}")
            errors.append(f"{sheet_name}: {e}")

    # 最终提交
    conn.commit()
    conn.close()
    wb.close()

    # 统计结果
    logger.info(f"[import] 导入完成")
    logger.info(f"  插入: {inserted}条")
    logger.info(f"  跳过: {skipped}条")
    if errors:
        logger.error(f"  错误: {len(errors)}条")
        for err in errors[:10]:
            logger.error(f"    {err}")

    # 验证
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(DISTINCT date) FROM rebar_prices WHERE date LIKE "2025%"')
    count_2025 = cursor.fetchone()[0]
    cursor.execute('SELECT MIN(date), MAX(date) FROM rebar_prices WHERE date LIKE "2025%"')
    date_range = cursor.fetchone()
    conn.close()

    logger.info(f"[import] 验证: 2025年数据 {count_2025}天, 范围 {date_range}")

if __name__ == '__main__':
    import_2025_data()
