import openpyxl
import os

data_dir = "services/data"
for f in os.listdir(data_dir):
    if f.endswith(".xlsx") and "钢筋" in f:
        path = os.path.join(data_dir, f)
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            total_sheets = len(wb.sheetnames)
            total_rows = 0
            for sn in wb.sheetnames[:10]:  # 只检查前10个sheet
                ws = wb[sn]
                for row in range(3, ws.max_row + 1):
                    if ws.cell(row=row, column=1).value and ws.cell(row=row, column=4).value:
                        total_rows += 1
            print(f"{f}")
            print(f"  总sheet数: {total_sheets}")
            print(f"  前10个sheet数据量: {total_rows}")
            wb.close()
        except Exception as e:
            print(f"{f}: Error - {e}")
