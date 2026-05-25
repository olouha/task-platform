"""
楼栋施工时间段解析 API
专门解析"楼号-开工至封顶时间表.xlsx"格式
"""

from datetime import datetime
from typing import Dict, List, Optional
from pathlib import Path
import re

from fastapi import APIRouter, UploadFile, File, HTTPException
import openpyxl

router = APIRouter(prefix="", tags=["楼栋施工时间"])


def parse_building_schedule(file_path: str) -> dict:
    """
    解析楼栋施工时间段文件

    格式：
    - 第1行：标题（如"一次结构钢筋调差（地下车库除外）"）
    - 奇数行（A列：楼号，B列：开工日期，C列：封顶日期）
    - 偶数行为空行
    """
    locations = []

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            cells = [str(cell.value).strip() if cell.value else '' for cell in row]

            # 获取楼号（第A列）
            building = cells[0] if len(cells) > 0 else ''
            if not building or building in ['', '楼号']:
                continue

            # 清理楼号格式（如 "1#" -> "1#楼"）
            building = building.strip()
            if building and not building.endswith('楼'):
                building = building + '楼'

            # 获取开工日期（第B列）
            start_date = _parse_date(cells[1] if len(cells) > 1 else '')

            # 获取封顶日期（第C列）
            end_date = _parse_date(cells[2] if len(cells) > 2 else '')

            if building:
                locations.append({
                    '楼号': building,
                    '开工日期': start_date,
                    '封顶日期': end_date,
                    '工期天数': _calc_days(start_date, end_date)
                })

        wb.close()

    except Exception as e:
        raise Exception(f"解析失败: {str(e)}")

    return {
        'success': True,
        'locations': locations,
        'total': len(locations),
        'title': _get_title(file_path)
    }


def _parse_date(date_str: str) -> str:
    """解析日期"""
    if not date_str:
        return ''

    # 处理 datetime 对象
    if isinstance(date_str, datetime):
        return date_str.strftime('%Y-%m-%d')

    # 处理字符串
    date_str = str(date_str).strip()

    # 匹配各种格式
    patterns = [
        r'(\d{4})-(\d{1,2})-(\d{1,2})',
        r'(\d{4})/(\d{1,2})/(\d{1,2})',
        r'(\d{4})年(\d{1,2})月(\d{1,2})日',
        r'(\d{4})(\d{2})(\d{2})',
    ]

    for pattern in patterns:
        match = re.search(pattern, date_str)
        if match:
            year, month, day = match.groups()
            return f"{year}-{int(month):02d}-{int(day):02d}"

    return date_str


def _calc_days(start: str, end: str) -> int:
    """计算工期天数"""
    if not start or not end:
        return 0

    try:
        s = datetime.strptime(start, '%Y-%m-%d')
        e = datetime.strptime(end, '%Y-%m-%d')
        return (e - s).days
    except:
        return 0


def _get_title(file_path: str) -> str:
    """获取文件标题"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active
        title = ws.cell(row=1, column=1).value or ''
        wb.close()
        return str(title).strip()
    except:
        return ''


# ============================================================
# API 端点
# ============================================================

@router.post("/upload", summary="上传楼栋施工时间表")
async def upload_building_schedule(file: UploadFile = File(...)):
    """上传并解析楼栋施工时间表"""
    allowed_extensions = ['.xlsx', '.xls']
    file_ext = Path(file.filename).suffix.lower()

    if file_ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail=f"不支持: {', '.join(allowed_extensions)}")

    upload_dir = Path('services/data/uploads')
    upload_dir.mkdir(parents=True, exist_ok=True)

    file_path = upload_dir / f"{datetime.now().strftime('%Y%m%d%H%M%S')}_schedule.xlsx"

    try:
        content = await file.read()
        with open(file_path, 'wb') as f:
            f.write(content)

        result = parse_building_schedule(str(file_path))

        # 清理文件
        try:
            file_path.unlink()
        except:
            pass

        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/list", summary="获取楼栋列表")
async def get_building_list():
    """获取已保存的楼栋施工时间段"""
    file_path = Path('services/data/楼栋施工时间段.json')

    if file_path.exists():
        import json
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data

    return {'locations': [], 'total': 0}


@router.post("/save", summary="保存楼栋施工时间")
async def save_building_schedule(locations: List[dict]):
    """保存楼栋施工时间段到文件"""
    import json

    file_path = Path('services/data/楼栋施工时间段.json')
    file_path.parent.mkdir(parents=True, exist_ok=True)

    data = {
        'updated_at': datetime.now().isoformat(),
        'locations': locations
    }

    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return {'success': True, 'total': len(locations)}