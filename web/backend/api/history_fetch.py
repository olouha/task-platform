"""
历史数据抓取 API
从现在开始往前一年的烟台造价信息（按月抓取）
"""

from fastapi import APIRouter, BackgroundTasks, HTTPException
from pydantic import BaseModel
from datetime import datetime, date, timedelta
import time
import logging

router = APIRouter(prefix="/api/history", tags=["历史数据抓取"])
logger = logging.getLogger(__name__)


class HistoryFetchResult(BaseModel):
    success: bool
    message: str
    timestamp: str
    months_fetched: int = 0
    months_skipped: int = 0
    total_prices: int = 0


def get_months_to_fetch(start_date: date, end_date: date) -> list:
    """获取需要抓取的月份列表"""
    months = []
    current = start_date.replace(day=1)
    while current <= end_date:
        # 检查是否已经是该月最后一天之后
        next_month = current.replace(day=28) + timedelta(days=4)
        if next_month.month != current.month:
            current = next_month.replace(day=1)
            continue
        months.append(current.strftime("%Y-%m"))
        # 移到下个月
        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)
    return months


def check_already_fetched(year_month: str) -> bool:
    """检查某月是否已抓取过"""
    from pathlib import Path
    import openpyxl

    excel_file = Path(__file__).parent.parent / "services" / "data" / "山东烟台造价信息历史.xlsx"

    if not excel_file.exists():
        return False

    try:
        wb = openpyxl.load_workbook(excel_file)
        # 检查是否存在该月份的sheet
        if year_month in wb.sheetnames:
            wb.close()
            return True
        wb.close()
        return False
    except Exception:
        return False


async def fetch_history_worker():
    """后台执行历史数据抓取"""
    try:
        from services.daily_fetch_ocr_v2 import YantaiPriceScraper
        from pathlib import Path
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        DATA_DIR = Path(__file__).parent.parent / "services" / "data"
        EXCEL_FILE = DATA_DIR / "山东烟台造价信息历史.xlsx"

        # 获取需要抓取的月份
        end_date = date.today()
        start_date = date(end_date.year - 1, end_date.month, 1)
        months = get_months_to_fetch(start_date, end_date)

        months_fetched = 0
        months_skipped = 0
        total_prices = 0

        print(f"[历史抓取] 需要抓取 {len(months)} 个月份")

        # 初始化抓取器
        scraper = YantaiPriceScraper()

        async with scraper:
            await scraper.login_with_captcha()

            for i, month in enumerate(months):
                print(f"[历史抓取] 正在抓取 {month} ({i+1}/{len(months)})")

                # 检查是否已抓取
                if check_already_fetched(month):
                    print(f"[历史抓取] {month} 已抓取，跳过")
                    months_skipped += 1
                    time.sleep(15)  # 仍然停顿以避免被封
                    continue

                try:
                    # 抓取该月数据
                    result = await scraper.fetch_month_prices(month)

                    if result and result.get('prices'):
                        # 保存到Excel
                        prices = result['prices']
                        total_prices += len(prices)

                        # 创建或更新Excel
                        if not EXCEL_FILE.exists():
                            wb = openpyxl.Workbook()
                            wb.remove(wb.active)
                        else:
                            wb = openpyxl.load_workbook(EXCEL_FILE)

                        # 创建sheet
                        ws = wb.create_sheet(month)

                        # 写入表头
                        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '地区']
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col, value=header)

                        # 写入数据
                        for row_idx, price in enumerate(prices, 2):
                            ws.cell(row=row_idx, column=1, value=price.get('date', ''))
                            ws.cell(row=row_idx, column=2, value=price.get('time', ''))
                            ws.cell(row=row_idx, column=3, value=price.get('material_name', ''))
                            ws.cell(row=row_idx, column=4, value=price.get('spec', ''))
                            ws.cell(row=row_idx, column=5, value=price.get('material_type', ''))
                            ws.cell(row=row_idx, column=6, value=price.get('brand', ''))
                            ws.cell(row=row_idx, column=7, value=price.get('price', 0))
                            ws.cell(row=row_idx, column=8, value=price.get('price_change', ''))
                            ws.cell(row=row_idx, column=9, value=price.get('remark', ''))
                            ws.cell(row=row_idx, column=10, value=price.get('region', '山东烟台'))

                        # 保存
                        wb.save(EXCEL_FILE)
                        wb.close()

                        months_fetched += 1
                        print(f"[历史抓取] {month} 抓取成功，{len(prices)} 条数据")

                except Exception as e:
                    print(f"[历史抓取] {month} 抓取失败: {e}")

                # 停顿15秒
                print(f"[历史抓取] 停顿15秒...")
                time.sleep(15)

        print(f"[历史抓取] 完成！抓取 {months_fetched} 个月，跳过 {months_skipped} 个月，共 {total_prices} 条数据")

        return HistoryFetchResult(
            success=True,
            message=f"历史数据抓取完成",
            timestamp=datetime.now().isoformat(),
            months_fetched=months_fetched,
            months_skipped=months_skipped,
            total_prices=total_prices
        )

    except Exception as e:
        logger.error(f"历史抓取异常: {e}")
        import traceback
        traceback.print_exc()
        return HistoryFetchResult(
            success=False,
            message=f"历史抓取异常: {str(e)}",
            timestamp=datetime.now().isoformat()
        )


@router.get("/fetch-history", response_model=HistoryFetchResult)
async def start_history_fetch(background_tasks: BackgroundTasks):
    """
    启动历史数据抓取任务
    从现在开始往前一年，每个月抓取一次，停顿15秒
    已抓取的不重复抓取
    """
    # 启动后台任务
    background_tasks.add_task(fetch_history_worker)

    return HistoryFetchResult(
        success=True,
        message="历史数据抓取任务已启动，请稍后刷新查看进度",
        timestamp=datetime.now().isoformat()
    )


@router.get("/fetch-status")
async def get_history_fetch_status():
    """获取历史数据抓取状态"""
    from pathlib import Path
    import openpyxl

    excel_file = Path(__file__).parent.parent / "services" / "data" / "山东烟台造价信息历史.xlsx"

    if not excel_file.exists():
        return {
            "exists": False,
            "total_sheets": 0,
            "sheets": []
        }

    try:
        wb = openpyxl.load_workbook(excel_file)
        sheets = wb.sheetnames
        wb.close()

        return {
            "exists": True,
            "total_sheets": len(sheets),
            "sheets": sheets
        }
    except Exception as e:
        return {
            "exists": False,
            "error": str(e)
        }


@router.get("/check-month")
async def check_month_status(year_month: str):
    """检查某月是否已抓取"""
    fetched = check_already_fetched(year_month)
    return {
        "year_month": year_month,
        "fetched": fetched
    }


@router.get("/months-to-fetch")
async def get_months_to_fetch_info():
    """获取需要抓取的月份列表"""
    end_date = date.today()
    start_date = date(end_date.year - 1, end_date.month, 1)
    months = get_months_to_fetch(start_date, end_date)

    result = []
    for month in months:
        result.append({
            "month": month,
            "fetched": check_already_fetched(month)
        })

    return {
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "total_months": len(months),
        "months": result
    }