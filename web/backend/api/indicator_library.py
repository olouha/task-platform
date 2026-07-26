"""
指标库 API
提供指标库项目管理、导入导出、验证等功能的 RESTful 接口
"""

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import StreamingResponse, JSONResponse
from typing import Optional, List, Dict, Any
from datetime import datetime
import logging
import io
import os

from models.indicator_library import (
    IndicatorLibrarySummary,
    IndicatorLibraryDetail,
    IndicatorLibraryCreate,
    ValidationResult,
    ImportResult,
    ImportPreviewResult,
)
from services.indicator_library_service import IndicatorLibraryService, get_indicator_library_service
from services.excel_parser_service import ExcelParserService
from api.deps import get_current_account, get_current_user_can_delete

router = APIRouter()
logger = logging.getLogger(__name__)


# ==================== 共享表头常量 ====================
# 明细表头：与 ExcelParserService.DETAIL_COLUMN_MAPPING 的 key 逐字对齐
# （导入按列名匹配），按系统归类排列：基本信息→造价→专项→材料→建筑指标。
# 模板下载与数据导出共用，保证"导出→改→导入"闭环。
DETAIL_HEADERS = [
    "序号", "项目名称", "业态", "项目所在地", "结构形式", "交付形式",
    "层数（地上/下）", "总面积（m2）", "檐高（m）",
    "地上建筑面积（m2）", "地下建筑面积（m2）",
    "平米造价（元/m2）", "总造价（元）",
    "地上土建造价", "地上安装造价", "地下土建造价", "地下安装造价",
    "措施费（元）", "室外造价（元）",
    "地上结构（元/㎡）", "地上安装（元/㎡）", "地下结构（元/㎡）", "地下安装（元/㎡）",
    "屋面（元/㎡）", "外墙（元/㎡）", "内墙（元/㎡）", "楼地面（元/㎡）",
    "电气（元/㎡）", "给排水（元/㎡）", "暖通（元/㎡）", "电梯（元/㎡）", "消防（元/㎡）", "措施（元/㎡）",
    "桩基造价（元）", "桩基平米造价（元/m2）",
    "基坑支护造价（元）", "基坑支护平米造价（元/m2）",
    "幕墙造价（元）", "幕墙平米造价（元/m2）",
    "精装修造价（元）", "精装修平米造价（元/m2）",
    "外墙保温造价（元）", "外墙保温平米造价（元/m2）",
    "外窗造价（元）", "外窗平米造价（元/m2）",
    "给排水造价（元）", "给排水平米造价（元/m2）",
    "采暖造价（元）", "采暖平米造价（元/m2）",
    "电气造价（元）", "电气平米造价（元/m2）",
    "暖通造价（元）", "暖通平米造价（元/m2）",
    "地上砼用量（m3）", "地上砼平米含量", "地上钢筋用量（t）", "地上钢筋平米含量",
    "地上模板用量（m2）", "地上模板平米含量",
    "地下砼用量（m3）", "地下砼平米含量", "地下钢筋用量（t）", "地下钢筋平米含量",
    "地下模板用量（m2）", "地下模板平米含量",
    "砌体含量（m³/㎡）",
    "电缆含量（m/㎡）", "管道含量（m/㎡）", "风管含量（㎡/㎡）",
    "开工时间", "竣工时间", "备注",
    "桩基形式",
    "室外平米造价（元/m2）",
    "墙地比\n（%）", "窗墙比\n（%）", "窗含量\n（㎡/㎡）", "门含量\n（㎡/㎡）", "内墙含量\n（㎡/㎡）",
    "阳台占比\n（%）", "装配率\n（%）", "装配构件含量\n（m3/m2）",
]


def get_service() -> IndicatorLibraryService:
    """获取指标库业务服务实例"""
    return get_indicator_library_service()


# ==================== 下载导入模板 ====================

@router.get("/template")
async def download_template() -> StreamingResponse:
    """
    下载指标库导入模板

    - 明细表：员工填写详细数据
    - 汇总表：系统从明细自动提取汇总
    - 支持数据验证下拉选择
    """
    logger.info("[download_template] 下载导入模板")

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = openpyxl.Workbook()

        # 定义样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # ==================== 明细表（员工填写）====================
        ws_detail = wb.active
        ws_detail.title = "明细"

        # 明细表表头 - 详细数据
        detail_headers = DETAIL_HEADERS

        for col, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 设置列宽
        for col in range(1, len(detail_headers) + 1):
            ws_detail.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

        # 设置行高
        ws_detail.row_dimensions[1].height = 30

        # 明细表数据验证下拉
        # 业态下拉 (C列)
        category_dv = DataValidation(type="list", formula1='"住宅,商业,办公,工业"', allow_blank=True)
        category_dv.prompt = "请选择业态"
        ws_detail.add_data_validation(category_dv)
        category_dv.add("C2:C1000")

        # 结构形式下拉 (E列)
        structure_dv = DataValidation(type="list", formula1='"框架结构,框架-剪力墙结构,剪力墙结构,框架-核心筒结构,框筒结构,钢结构"', allow_blank=True)
        structure_dv.prompt = "请选择结构形式"
        ws_detail.add_data_validation(structure_dv)
        structure_dv.add("E2:E1000")

        # 交付形式下拉 (F列)
        delivery_dv = DataValidation(type="list", formula1='"毛坯交付,精装修,带装修"', allow_blank=True)
        delivery_dv.prompt = "请选择交付形式"
        ws_detail.add_data_validation(delivery_dv)
        delivery_dv.add("F2:F1000")

        # 严格72列，与detail_headers(第69-96行)一一对应
        # 1-6: 序号,项目名称,业态,项目所在地,结构形式,交付形式
        # 7-9: 层数/总面积/檐高
        # 10-11: 地上面积/地下面积
        # 12-13: 平米造价/总造价
        # 14-17: 地上土建/安装,地下土建/安装造价
        # 18-19: 措施费/室外造价
        # 20-23: 地上结构/安装,地下结构/安装平米造价
        # 24-33: 屋面,外墙,内墙,楼地面,电气,给排水,暖通,电梯,消防,措施 平米造价
        # 34-35: 桩基造价/平米造价
        # 36-37: 基坑支护造价/平米造价
        # 38-39: 幕墙造价/平米造价
        # 40-41: 精装修造价/平米造价
        # 42-43: 外墙保温造价/平米造价
        # 44-45: 外窗造价/平米造价
        # 46-47: 给排水造价/平米造价
        # 48-49: 采暖造价/平米造价
        # 50-51: 电气造价/平米造价
        # 52-53: 暖通造价/平米造价
        # 54-55: 地上砼用量/地上砼平米含量
        # 56-57: 地上钢筋用量/地上钢筋平米含量
        # 58-59: 地上模板用量/地上模板平米含量
        # 60-61: 地下砼用量/地下砼平米含量
        # 62-63: 地下钢筋用量/地下钢筋平米含量
        # 64-65: 地下模板用量/地下模板平米含量
        # 66: 砌体含量
        # 67-69: 电缆含量/管道含量/风管含量
        # 70-72: 开工时间/竣工时间/备注
        sample_detail = [
            1, "示例住宅项目", "住宅", "山东烟台", "框架结构", "毛坯交付",      # 1-6 (6)
            "18/2", 25000, 54,                                                    # 7-9 (3) = 9
            22000, 3000,                                                          # 10-11 (2) = 11
            2350, 58750000,                                                       # 12-13 (2) = 13
            35000000, 10000000, 8000000, 3000000,                                 # 14-17 (4) = 17
            3000000, 2000000,                                                     # 18-19 (2) = 19
            1400, 50, 600, 20,                                                    # 20-23 (4) = 23
            200, 8, 100, 4, 300, 12, 280, 10, 150, 5,                             # 24-33 (10) = 33  屋面~措施
            500000, 20, 800000, 32, 2000000, 80, 5000000, 200,                     # 34-41 (8) = 41  桩基/基坑/幕墙/精装
            300000, 12, 200000, 8, 150000, 6, 180000, 7,                          # 42-49 (8) = 49  外墙保温/外窗/给排水/采暖
            160000, 6, 100000, 4,                                                  # 50-53 (4) = 53  电气/暖通
            8000, 0.32, 1000, 0.04, 1000, 0.04,                                   # 54-59 (6) = 59  地上砼/钢筋/模板
            2500, 0.1, 300, 0.012, 500, 0.02,                                      # 60-65 (6) = 65  地下砼/钢筋/模板
            0.5,                                                                   # 66 (1) = 66  砌体
            50, 80, 30,                                                            # 67-69 (3) = 69  电缆/管道/风管
            "2023-01", "2024-06", "",                                              # 70-72 (3) = 72  开工/竣工/备注
            "钢板桩", 80,                                                           # 73-74 (2) = 74  桩基形式/室外平米造价
            15, 30, 0.2, 0.05, 1.5, 10, 30, 0.1,                                    # 75-82 (8) = 82  建筑指标8项
        ]
        # 验证: 6+3+2+2+4+2+4+10+4+8+8+4+6+6+1+3+3 + 2 + 8 = 82
        for col, value in enumerate(sample_detail, 1):
            cell = ws_detail.cell(row=2, column=col, value=value)
            cell.border = thin_border

        # ==================== 填写指导（单独的工作表）====================
        ws_guide = wb.create_sheet("填写说明")

        guide_title_font = Font(bold=True, color="FFFFFF", size=14)
        guide_title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        guide_section_font = Font(bold=True, color="000000", size=12)
        guide_text_font = Font(color="000000", size=11)
        guide_green_font = Font(color="006400", size=11)
        guide_red_font = Font(color="DC143C", size=11)

        # 标题
        ws_guide.merge_cells("A1:H1")
        title_cell = ws_guide.cell(row=1, column=1, value="指标库导入模板填写说明")
        title_cell.font = guide_title_font
        title_cell.fill = guide_title_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_guide.row_dimensions[1].height = 35

        row = 3

        # 一、必填字段
        ws_guide.cell(row=row, column=1, value="一、必填字段（必须填写，否则系统无法识别）").font = guide_section_font
        row += 1
        required_fields = [
            ("序号", "必须填写正整数：1, 2, 3..."),
            ("项目名称", "填写项目全称，如：XX住宅小区一期"),
            ("业态", "从下拉列表选择：住宅/商业/办公/工业"),
            ("项目所在地", "填写地区，如：山东烟台、北京朝阳"),
            ("结构形式", "从下拉列表选择"),
        ]
        for field, desc in required_fields:
            ws_guide.cell(row=row, column=1, value=f"  • {field}").font = guide_text_font
            ws_guide.cell(row=row, column=2, value=desc).font = guide_text_font
            row += 1

        row += 1

        # 二、数值字段
        ws_guide.cell(row=row, column=1, value="二、数值字段（只填数字，不要带单位）").font = guide_section_font
        row += 1
        number_fields = [
            ("总面积（m2）", "只填数字，如：25000"),
            ("檐高（m）", "只填数字，如：54"),
            ("地上/地下建筑面积", "只填数字，如：22000、3000"),
            ("平米造价（元/m2）", "只填数字，如：2350"),
            ("总造价（元）", "只填数字，如：58750000"),
            ("各种造价字段", "只填数字，单位：元"),
            ("砼/钢筋用量", "只填数字，单位：m³或t"),
            ("平米含量", "只填数字，如：0.32、0.04"),
        ]
        for field, desc in number_fields:
            ws_guide.cell(row=row, column=1, value=f"  • {field}").font = guide_text_font
            ws_guide.cell(row=row, column=2, value=desc).font = guide_text_font
            row += 1

        row += 1

        # 三、时间格式
        ws_guide.cell(row=row, column=1, value="三、时间格式").font = guide_section_font
        row += 1
        ws_guide.cell(row=row, column=1, value="  • 开工/竣工时间").font = guide_text_font
        ws_guide.cell(row=row, column=2, value="格式：YYYY-MM，如：2023-01、2024-06").font = guide_text_font
        row += 2

        # 四、常见错误
        ws_guide.cell(row=row, column=1, value="四、常见错误（这些写法系统无法识别）").font = guide_section_font
        row += 1
        errors = [
            ("错误写法", "正确写法"),
            ("25000元", "25000"),
            ("54米", "54"),
            ("2023年1月", "2023-01"),
            ("示例项目", "XX住宅小区"),
        ]
        for wrong, correct in errors:
            ws_guide.cell(row=row, column=1, value=wrong).font = guide_red_font
            ws_guide.cell(row=row, column=2, value=correct).font = guide_green_font
            row += 1

        row += 1

        # 五、下拉选项
        ws_guide.cell(row=row, column=1, value="五、可下拉选择的字段").font = guide_section_font
        row += 1
        dropdowns = [
            ("业态", "住宅, 商业, 办公, 工业"),
            ("结构形式", "框架结构, 框架-剪力墙结构, 剪力墙结构, 框架-核心筒结构, 框筒结构, 钢结构"),
            ("交付形式", "毛坯交付, 精装修, 带装修"),
        ]
        for field, options in dropdowns:
            ws_guide.cell(row=row, column=1, value=f"  • {field}").font = guide_text_font
            ws_guide.cell(row=row, column=2, value=options).font = guide_text_font
            row += 1

        row += 1

        # 六、注意事项
        ws_guide.cell(row=row, column=1, value="六、注意事项").font = guide_section_font
        row += 1
        notes = [
            "1. 系统根据「序号」列识别有效数据行，序号为空的行会被忽略",
            "2. 明细表的序号列必须是正整数（如1, 2, 3...）",
            "3. 汇总表数据由系统自动从明细表提取，无需手动填写",
            "4. 请删除示例数据后再填写您的项目信息",
        ]
        for note in notes:
            ws_guide.cell(row=row, column=1, value=note).font = guide_text_font
            ws_guide.merge_cells(f"A{row}:H{row}")
            row += 1

        # 设置列宽
        ws_guide.column_dimensions["A"].width = 25
        ws_guide.column_dimensions["B"].width = 60

        # ==================== 汇总表（系统自动提取）====================
        ws_summary = wb.create_sheet("汇总")

        # 汇总表表头
        summary_headers = [
            "项目名称", "业态", "项目所在地", "结构形式", "交付形式",
            "层数（地上/下）", "总面积（m2）", "檐高（m）",
            "平米造价（元/m2）", "总造价（元）",
            "地上建筑面积（m2）", "地下建筑面积（m2）",
            "开工时间", "竣工时间",
        ]

        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")  # 绿色背景
            cell.alignment = header_alignment
            cell.border = thin_border

        # 设置列宽
        for col in range(1, len(summary_headers) + 1):
            ws_summary.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

        # 汇总表说明
        note_font = Font(italic=True, color="666666")
        ws_summary.cell(row=3, column=1, value="提示：汇总数据将从明细表自动提取，只需填写明细表即可").font = note_font
        ws_summary.merge_cells("A3:N3")

        # 从明细表提取汇总的公式（第一行示例）
        # 明细表第2行对应汇总第2行
        summary_formulas = [
            "=明细!B2",  # 项目名称
            "=明细!C2",  # 业态
            "=明细!D2",  # 项目所在地
            "=明细!E2",  # 结构形式
            "=明细!F2",  # 交付形式
            "=明细!G2",  # 层数
            "=明细!H2",  # 总面积
            "=明细!I2",  # 檐高
            "=明细!L2",  # 平米造价
            "=明细!M2",  # 总造价
            "=明细!J2",  # 地上面积
            "=明细!K2",  # 地下面积
            "=明细!BR2",  # 开工时间（明细第70列）
            "=明细!BS2",  # 竣工时间（明细第71列）
        ]

        for col, formula in enumerate(summary_formulas, 1):
            cell = ws_summary.cell(row=2, column=col, value=formula)
            cell.border = thin_border
            cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # 浅绿色背景

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "indicator_template.xlsx"

        logger.info("[download_template] 模板生成完成")

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
        )

    except ImportError:
        logger.error("[download_template] openpyxl 未安装")
        raise HTTPException(status_code=500, detail="服务器未安装 Excel 支持库")
    except Exception as e:
        logger.error(f"[download_template] 模板生成失败 | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"模板生成失败: {str(e)}")


# ==================== 汇总列表 ====================

@router.get("/summary", response_model=List[Dict[str, Any]])
async def get_summary_list(
    category: Optional[str] = Query(None, description="业态筛选"),
    location: Optional[str] = Query(None, description="所在地筛选"),
    limit: int = Query(100, ge=1, le=500, description="返回数量限制"),
    service: IndicatorLibraryService = Depends(get_service),
) -> List[Dict[str, Any]]:
    """
    获取指标库汇总列表

    - 支持按业态和所在地筛选
    - 返回简化的汇总信息
    """
    logger.info(f"[get_summary_list] 获取汇总列表 | category={category} | location={location} | limit={limit}")

    try:
        summaries = service.get_summary_list(
            category=category,
            location=location,
            limit=limit,
        )

        result = [s.model_dump() for s in summaries]
        logger.info(f"[get_summary_list] 返回 {len(result)} 条记录")
        return result

    except Exception as e:
        logger.error(f"[get_summary_list] 获取汇总列表失败 | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取汇总列表失败: {str(e)}")


# ==================== 自动导入（先校验后入库）====================

@router.post("/auto-import", response_model=Dict[str, Any])
async def auto_import(
    file: UploadFile = File(..., description="Excel文件"),
    service: IndicatorLibraryService = Depends(get_service),
    account: str = Depends(get_current_account),
) -> Dict[str, Any]:
    """
    自动导入 Excel 数据（先预览校验，有错误返回，无错误直接入库）

    - 解析 Excel 文件
    - 逐行校验数据
    - 有错误返回错误列表让用户处理
    - 无错误直接入库并返回成功结果
    """
    logger.info(f"[auto_import] 自动导入 | filename={file.filename}")

    try:
        # 验证文件类型
        if not (file.filename or '').endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 格式的 Excel 文件")

        # 读取文件内容
        content = await file.read()
        if len(content) == 0:
            raise HTTPException(status_code=400, detail="文件为空")

        # 执行自动导入（放线程池，避免 openpyxl 解析阻塞事件循环）
        result = await run_in_threadpool(service.auto_import, content, file.filename, account)
        logger.info(
            f"[auto_import] 导入完成 | success={result.success} | imported={result.imported} | total={result.total}"
        )
        return result.model_dump()

    except HTTPException:
        raise
    except ValueError as e:
        logger.warning(f"[auto_import] 导入失败 | error={e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"[auto_import] 自动导入失败 | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"自动导入失败: {str(e)}")


# ==================== 导入历史 ====================

@router.get("/import-history", response_model=List[Dict[str, Any]])
async def get_import_history(
    limit: int = Query(50, ge=1, le=200, description="返回数量限制"),
    service: IndicatorLibraryService = Depends(get_service),
) -> List[Dict[str, Any]]:
    """
    获取导入历史列表

    - 返回历史导入记录
    - 包含文件名、导入数量、成功/失败数等信息
    """
    logger.info(f"[get_import_history] 获取导入历史 | limit={limit}")

    try:
        history = service.get_import_history(limit)
        logger.info(f"[get_import_history] 返回 {len(history)} 条记录")
        return history

    except Exception as e:
        logger.error(f"[get_import_history] 获取导入历史失败 | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取导入历史失败: {str(e)}")


@router.get("/import-history/{import_id}", response_model=Dict[str, Any])
async def get_import_detail(
    import_id: int,
    service: IndicatorLibraryService = Depends(get_service),
) -> Dict[str, Any]:
    """
    获取导入详情

    - 返回指定导入记录的详细信息
    - 包含成功导入的详细数据
    """
    logger.info(f"[get_import_detail] 获取导入详情 | import_id={import_id}")

    try:
        detail = service.get_import_detail(import_id)

        if not detail:
            logger.warning(f"[get_import_detail] 导入记录不存在 | import_id={import_id}")
            raise HTTPException(status_code=404, detail=f"导入记录不存在: {import_id}")

        logger.info(f"[get_import_detail] 获取成功 | import_id={import_id}")
        return detail

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[get_import_detail] 获取导入详情失败 | import_id={import_id} | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取导入详情失败: {str(e)}")


# ==================== 版本历史 ====================

@router.get("/versions/{project_id}", response_model=List[Dict[str, Any]])
async def get_version_history(
    project_id: str,
    service: IndicatorLibraryService = Depends(get_service),
) -> List[Dict[str, Any]]:
    """
    获取项目版本历史

    - 返回项目的所有版本快照
    - 包含版本号、创建时间、来源文件等信息
    """
    logger.info(f"[get_version_history] 获取版本历史 | project_id={project_id}")

    try:
        versions = service.get_version_history(project_id)
        logger.info(f"[get_version_history] 返回 {len(versions)} 个版本")
        return versions

    except Exception as e:
        logger.error(f"[get_version_history] 获取版本历史失败 | project_id={project_id} | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取版本历史失败: {str(e)}")


@router.get("/versions/{project_id}/snapshot/{snapshot_id}", response_model=Dict[str, Any])
async def get_snapshot_detail(
    project_id: str,
    snapshot_id: str,
    service: IndicatorLibraryService = Depends(get_service),
) -> Dict[str, Any]:
    """
    获取快照详情

    - 返回指定快照的完整数据
    """
    logger.info(f"[get_snapshot_detail] 获取快照详情 | project_id={project_id} | snapshot_id={snapshot_id}")

    try:
        detail = service.get_snapshot_detail(snapshot_id)

        if not detail:
            logger.warning(f"[get_snapshot_detail] 快照不存在 | snapshot_id={snapshot_id}")
            raise HTTPException(status_code=404, detail=f"快照不存在: {snapshot_id}")

        logger.info(f"[get_snapshot_detail] 获取成功 | snapshot_id={snapshot_id}")
        return detail

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[get_snapshot_detail] 获取快照详情失败 | snapshot_id={snapshot_id} | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取快照详情失败: {str(e)}")


@router.post("/versions/{project_id}/rollback/{snapshot_id}")
async def rollback_version(
    project_id: str,
    snapshot_id: str,
    service: IndicatorLibraryService = Depends(get_service),
) -> Dict[str, bool]:
    """
    回滚到指定版本

    - 将项目恢复到指定快照的状态
    - 会自动保存当前版本为新快照
    """
    logger.info(f"[rollback_version] 回滚版本 | project_id={project_id} | snapshot_id={snapshot_id}")

    try:
        success = service.rollback_version(snapshot_id)

        if success:
            logger.info(f"[rollback_version] 回滚成功 | project_id={project_id} | snapshot_id={snapshot_id}")
        else:
            logger.warning(f"[rollback_version] 回滚失败 | snapshot_id={snapshot_id}")

        return {"success": success}

    except Exception as e:
        logger.error(f"[rollback_version] 回滚失败 | snapshot_id={snapshot_id} | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"回滚失败: {str(e)}")


# ==================== 数据一致性校验 ====================

@router.get("/data-sync", response_model=Dict[str, Any])
async def sync_check(
    service: IndicatorLibraryService = Depends(get_service),
) -> Dict[str, Any]:
    """
    前后端数据一致性校验

    - 返回数据库统计信息
    - 包含项目数、快照数、导入记录数等
    """
    logger.info("[sync_check] 执行数据一致性校验")

    try:
        result = service.sync_check()
        logger.info(f"[sync_check] 校验完成 | in_sync={result.get('in_sync')}")
        return result

    except Exception as e:
        logger.error(f"[sync_check] 校验失败 | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"数据一致性校验失败: {str(e)}")

# ==================== 创建项目 ====================

@router.post("/", response_model=Dict[str, Any], status_code=201)
async def create_project(
    data: IndicatorLibraryCreate,
    service: IndicatorLibraryService = Depends(get_service),
    account: str = Depends(get_current_account),
) -> Dict[str, Any]:
    """
    创建指标库项目

    - 验证必填字段和数据格式
    - 自动记录创建时间
    """
    logger.info(f"[create_project] 创建项目 | name={data.name} | category={data.category}")

    try:
        detail = service.create_project(data, account)
        logger.info(f"[create_project] 创建成功 | id={detail.id}")
        return detail.model_dump(exclude_none=True)

    except ValueError as e:
        logger.warning(f"[create_project] 创建失败 | error={e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"[create_project] 创建项目失败 | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"创建项目失败: {str(e)}")


# ==================== 更新项目 ====================

@router.put("/{project_id}", response_model=Dict[str, Any])
async def update_project(
    project_id: str,
    data: IndicatorLibraryCreate,
    admin_account: str = Depends(get_current_user_can_delete),
    service: IndicatorLibraryService = Depends(get_service),
) -> Dict[str, Any]:
    """
    更新指标库项目

    - 验证数据格式
    - 自动记录更新时间
    """
    logger.info(f"[update_project] 更新项目 | project_id={project_id}")

    try:
        detail = service.update_project(project_id, data)
        logger.info(f"[update_project] 更新成功 | project_id={project_id}")
        return detail.model_dump(exclude_none=True)

    except ValueError as e:
        error_msg = str(e)
        if "不存在" in error_msg:
            logger.warning(f"[update_project] 项目不存在 | project_id={project_id}")
            raise HTTPException(status_code=404, detail=error_msg)
        logger.warning(f"[update_project] 更新失败 | error={e}")
        raise HTTPException(status_code=400, detail=error_msg)
    except Exception as e:
        logger.error(f"[update_project] 更新项目失败 | project_id={project_id} | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"更新项目失败: {str(e)}")


# ==================== 删除项目 ====================

@router.delete("/{project_id}")
async def delete_project(
    project_id: str,
    admin_account: str = Depends(get_current_user_can_delete),
    service: IndicatorLibraryService = Depends(get_service),
) -> Dict[str, bool]:
    """
    删除指标库项目

    - 返回删除是否成功
    """
    logger.info(f"[delete_project] 删除项目 | project_id={project_id}")

    try:
        success = service.delete_project(project_id)

        if success:
            logger.info(f"[delete_project] 删除成功 | project_id={project_id}")
        else:
            logger.warning(f"[delete_project] 项目不存在 | project_id={project_id}")

        return {"success": success}

    except Exception as e:
        logger.error(f"[delete_project] 删除项目失败 | project_id={project_id} | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"删除项目失败: {str(e)}")


# ==================== 数据验证 ====================

@router.post("/validate", response_model=Dict[str, Any])
async def validate_data(
    data: IndicatorLibraryCreate,
    service: IndicatorLibraryService = Depends(get_service),
) -> Dict[str, Any]:
    """
    验证指标库数据

    - 执行三层验证：基础验证、逻辑验证、参考范围验证
    - 返回验证结果，包括错误和警告信息
    """
    logger.info(f"[validate_data] 验证数据 | name={data.name}")

    try:
        result = service.validate_data(data)
        logger.info(
            f"[validate_data] 验证完成 | passed={result.passed} | errors={len(result.errors)} | warnings={len(result.warnings)}"
        )
        return result.model_dump()

    except Exception as e:
        logger.error(f"[validate_data] 验证失败 | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"验证失败: {str(e)}")


# ==================== 导入预览 ====================

@router.post("/preview", response_model=Dict[str, Any])
async def preview_import(
    file: UploadFile = File(..., description="Excel文件"),
    service: IndicatorLibraryService = Depends(get_service),
    account: str = Depends(get_current_account),
) -> Dict[str, Any]:
    """
    预览 Excel 导入内容

    - 解析 Excel 文件但不实际导入数据库
    - 返回每条数据的验证结果
    """
    logger.info(f"[preview_import] 预览导入 | filename={file.filename}")

    try:
        # 验证文件类型
        if not (file.filename or '').endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 格式的 Excel 文件")

        # 读取文件内容
        content = await file.read()
        if len(content) == 0:
            raise HTTPException(status_code=400, detail="文件为空")

        # 执行预览（放线程池，避免 openpyxl 解析阻塞事件循环）
        result = await run_in_threadpool(service.preview_import, content, file.filename)
        logger.info(
            f"[preview_import] 预览完成 | total={result.total} | valid={result.valid_count} | error={result.error_count}"
        )
        return result.model_dump()

    except HTTPException:
        raise
    except ValueError as e:
        logger.warning(f"[preview_import] 预览失败 | error={e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"[preview_import] 预览导入失败 | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"预览导入失败: {str(e)}")


# ==================== 导入Excel ====================

@router.post("/import", response_model=Dict[str, Any])
async def import_from_excel(
    file: UploadFile = File(..., description="Excel文件"),
    service: IndicatorLibraryService = Depends(get_service),
    account: str = Depends(get_current_account),
) -> Dict[str, Any]:
    """
    从 Excel 文件导入指标库数据

    - 解析并验证 Excel 内容
    - 实际写入数据库
    - 返回导入结果统计
    """
    logger.info(f"[import_from_excel] 开始导入 | filename={file.filename}")

    try:
        # 验证文件类型
        if not (file.filename or '').endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 格式的 Excel 文件")

        # 读取文件内容
        content = await file.read()
        if len(content) == 0:
            raise HTTPException(status_code=400, detail="文件为空")

        # 执行导入（放线程池，避免 openpyxl 解析阻塞事件循环）
        result = await run_in_threadpool(service.import_from_excel, content, file.filename, account)
        logger.info(
            f"[import_from_excel] 导入完成 | imported={result.imported} | total={result.total} | errors={len(result.errors)}"
        )
        return result.model_dump()

    except HTTPException:
        raise
    except ValueError as e:
        logger.warning(f"[import_from_excel] 导入失败 | error={e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"[import_from_excel] 导入失败 | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


# ==================== 导出Excel ====================

@router.get("/export")
async def export_to_excel(
    category: Optional[str] = Query(None, description="业态筛选"),
    location: Optional[str] = Query(None, description="所在地筛选"),
    service: IndicatorLibraryService = Depends(get_service),
) -> StreamingResponse:
    """
    导出指标库数据到 Excel（总分 + 详细双 sheet）

    - 汇总 sheet：核心字段（总分模式），便于快速浏览
    - 明细 sheet：全字段，表头与导入模板逐字一致，可改后重新导入（闭环）
    - 支持 .xlsx（双 sheet）；若未装 openpyxl 则回退单表 CSV
    """
    logger.info(f"[export_to_excel] 开始导出 | category={category} | location={location}")

    try:
        # 全字段数据（不做 Summary 裁剪，保留所有 db 字段供明细回填）
        projects = service.get_full_projects(
            category=category,
            location=location,
            limit=10000,
        )
        logger.info(f"[export_to_excel] 取得项目 | count={len(projects)}")

        mapping = ExcelParserService.DETAIL_COLUMN_MAPPING

        def cell_value(header: str, project: Dict[str, Any], row_num: int) -> Any:
            """按表头从 db 行取值；序号与层数做特殊处理。"""
            if header == "序号":
                return row_num
            if header == "层数（地上/下）":
                fa, fb = project.get("floor_above"), project.get("floor_below")
                if fa is None and fb is None:
                    return ""
                return f"{fa if fa is not None else 0}/{fb if fb is not None else 0}"
            db_field = mapping.get(header)
            if db_field is None:
                return ""
            val = project.get(db_field)
            return val if val is not None else ""

        # 汇总表头（总分模式）
        summary_headers = [
            "项目名称", "业态", "项目所在地", "结构形式", "交付形式",
            "层数（地上/下）", "总面积（m2）", "檐高（m）",
            "地上建筑面积（m2）", "地下建筑面积（m2）",
            "平米造价（元/m2）", "总造价（元）",
            "开工时间", "竣工时间",
        ]

        use_xlsx = True
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            header_font = Font(bold=True, color="FFFFFF")
            detail_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            summary_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            def write_sheet(ws, headers, fill):
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                for idx, project in enumerate(projects, start=1):
                    for col, header in enumerate(headers, 1):
                        c = ws.cell(row=idx + 1, column=col, value=cell_value(header, project, idx))
                        c.border = thin_border
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14
                ws.row_dimensions[1].height = 28
                ws.freeze_panes = "A2"

            wb = openpyxl.Workbook()
            # 汇总 sheet（首 sheet，总分模式）
            ws_summary = wb.active
            ws_summary.title = "汇总"
            write_sheet(ws_summary, summary_headers, summary_fill)
            # 明细 sheet（详细模式，全字段，表头与导入模板一致，可重新导入）
            ws_detail = wb.create_sheet("明细")
            write_sheet(ws_detail, DETAIL_HEADERS, detail_fill)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

        except ImportError:
            logger.warning("[export_to_excel] openpyxl 未安装，回退 CSV（仅汇总）")
            use_xlsx = False
            import csv

            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(summary_headers)
            for idx, project in enumerate(projects, start=1):
                writer.writerow([cell_value(h, project, idx) for h in summary_headers])
            output.seek(0)
            output = io.BytesIO(output.getvalue().encode("utf-8-sig"))

        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ext = "xlsx" if use_xlsx else "csv"
        filename = f"指标库导出_{timestamp}.{ext}"
        media_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            if use_xlsx
            else "text/csv"
        )

        # 中文文件名百分号编码（RFC 5987），避免 Starlette latin-1 编码报错
        from urllib.parse import quote
        encoded_filename = quote(filename)

        logger.info(f"[export_to_excel] 导出完成 | records={len(projects)} | file={filename}")

        return StreamingResponse(
            output,
            media_type=media_type,
            headers={
                "Content-Disposition": f"attachment; filename=export.{ext}; filename*=UTF-8''{encoded_filename}",
            },
        )

    except Exception as e:
        logger.error(f"[export_to_excel] 导出失败 | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


# ==================== 项目详情 ====================
# 注意：动态路径参数路由必须放在所有静态 GET 路由（/export, /stats 等）之后，
# 否则会拦截同名静态路径（如 /export 被当成 project_id="export"）。

@router.get("/{project_id}", response_model=Dict[str, Any])
async def get_project_detail(
    project_id: str,
    service: IndicatorLibraryService = Depends(get_service),
) -> Dict[str, Any]:
    """
    获取指标库项目详情

    - 返回项目的完整信息
    - 包含所有造价、材料含量等详细字段
    """
    logger.info(f"[get_project_detail] 获取项目详情 | project_id={project_id}")

    try:
        detail = service.get_detail(project_id)

        if not detail:
            logger.warning(f"[get_project_detail] 项目不存在 | project_id={project_id}")
            raise HTTPException(status_code=404, detail=f"项目不存在: {project_id}")

        logger.info(f"[get_project_detail] 获取成功 | project_id={project_id}")
        return detail.model_dump(exclude_none=True)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[get_project_detail] 获取项目详情失败 | project_id={project_id} | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取项目详情失败: {str(e)}")


# ==================== 统计信息 ====================

@router.get("/stats/overview", response_model=Dict[str, Any])
async def get_stats_overview(
    service: IndicatorLibraryService = Depends(get_service),
) -> Dict[str, Any]:
    """
    获取指标库统计概览

    - 返回项目总数、各业态分布、各地区分布等信息
    """
    logger.info("[get_stats_overview] 获取统计概览")

    try:
        stats = service.get_stats()
        logger.info(f"[get_stats_overview] 统计完成 | total={stats.get('total', 0)}")
        return stats

    except Exception as e:
        logger.error(f"[get_stats_overview] 获取统计失败 | error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取统计失败: {str(e)}")


