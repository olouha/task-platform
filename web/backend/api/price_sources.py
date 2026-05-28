"""
价格来源 API
支持配置网站登录账号、API密钥等认证信息
"""

from fastapi import APIRouter, HTTPException
from typing import List, Optional
from pydantic import BaseModel
from datetime import datetime
import json
import os

router = APIRouter()


def get_excel_file():
    """获取可用的Excel文件路径"""
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    candidates = [
        os.path.join(base_dir, "services", "data", "山东烟台钢筋价格_current.xlsx"),
        os.path.join(base_dir, "services", "data", "山东烟台钢筋价格_完整版.xlsx"),
        os.path.join(base_dir, "services", "data", "山东烟台钢筋价格.xlsx"),
    ]
    for candidate in candidates:
        if os.path.exists(candidate):
            return candidate
    return None


# 修正Excel路径 - 使用动态查找
_excel_file = get_excel_file()


@router.get("/latest")
async def get_latest_price():
    """获取最新价格"""
    try:
        import openpyxl

        if not _excel_file or not os.path.exists(_excel_file):
            return {
                "success": False,
                "message": "暂无数据，请先运行爬虫抓取"
            }

        wb = openpyxl.load_workbook(_excel_file)
        sheet_names = wb.sheetnames

        if not sheet_names:
            return {
                "success": False,
                "message": "暂无数据"
            }

        # 获取最新sheet
        latest_sheet = sheet_names[-1]
        ws = wb[latest_sheet]

        # 读取数据行
        prices = []
        for row in range(4, ws.max_row + 1):
            date = ws.cell(row=row, column=1).value
            if date and str(date) != "当日截图":
                prices.append({
                    "date": str(date),
                    "time": ws.cell(row=row, column=2).value,
                    "material_name": ws.cell(row=row, column=3).value,
                    "spec": ws.cell(row=row, column=4).value,
                    "price_min": ws.cell(row=row, column=5).value,
                    "price_max": ws.cell(row=row, column=6).value,
                    "price_avg": ws.cell(row=row, column=7).value,
                    "region": ws.cell(row=row, column=8).value,
                    "source": ws.cell(row=row, column=9).value,
                })

        wb.close()

        return {
            "success": True,
            "sheet": latest_sheet,
            "prices": prices,
            "total_sheets": len(sheet_names),
            "all_sheets": sheet_names
        }

    except Exception as e:
        return {
            "success": False,
            "message": str(e)
        }


@router.get("/history")
async def get_price_history(days: int = 30):
    """获取历史价格"""
    try:
        import openpyxl
        from datetime import timedelta

        if not _excel_file or not os.path.exists(_excel_file):
            return {"success": False, "prices": []}

        wb = openpyxl.load_workbook(_excel_file)

        cutoff = datetime.now() - timedelta(days=days)
        all_prices = []

        for sheet_name in wb.sheetnames:
            sheet_date = datetime.strptime(sheet_name, "%Y-%m-%d")
            if sheet_date >= cutoff:
                ws = wb[sheet_name]
                for row in range(4, ws.max_row + 1):
                    date = ws.cell(row=row, column=1).value
                    if date and str(date) != "当日截图":
                        all_prices.append({
                            "date": str(date),
                            "material_name": ws.cell(row=row, column=3).value,
                            "spec": ws.cell(row=row, column=4).value,
                            "price_min": ws.cell(row=row, column=5).value,
                            "price_max": ws.cell(row=row, column=6).value,
                            "price_avg": ws.cell(row=row, column=7).value,
                            "region": ws.cell(row=row, column=8).value,
                        })

        wb.close()

        return {
            "success": True,
            "prices": all_prices
        }

    except Exception as e:
        return {"success": False, "message": str(e), "prices": []}


@router.get("/sheets")
async def get_all_sheets():
    """获取所有Sheet（日期）"""
    try:
        import openpyxl

        if not _excel_file or not os.path.exists(_excel_file):
            return {"success": False, "sheets": []}

        # 使用read_only模式快速读取
        wb = openpyxl.load_workbook(_excel_file, read_only=True)
        sheets = wb.sheetnames
        wb.close()

        return {
            "success": True,
            "sheets": sheets,
            "total": len(sheets)
        }

    except Exception as e:
        return {"success": False, "message": str(e), "sheets": []}


@router.post("/fetch")
async def fetch_price():
    """触发爬虫抓取"""
    try:
        import sys
        sys.path.insert(0, 'services')
        from yantai_rebar_scraper import YantaiRebarScraper, save_to_excel

        scraper = YantaiRebarScraper()

        # 检查是否今日已抓取
        can_fetch, reason = scraper._check_rate_limit()
        if not can_fetch:
            return {
                "success": False,
                "message": reason
            }

        result = scraper.fetch(force=True)

        if result.success:
            save_to_excel(result)
            return {
                "success": True,
                "message": "抓取成功",
                "prices": [
                    {
                        "material_name": p.material_name,
                        "spec": p.spec,
                        "price_min": p.price,
                        "price_max": p.price_max,
                        "price_avg": (p.price + p.price_max) / 2,
                    }
                    for p in result.prices
                ]
            }
        else:
            return {
                "success": False,
                "message": result.error_message
            }

    except Exception as e:
        return {
            "success": False,
            "message": str(e)
        }


@router.get("/status")
async def get_fetch_status():
    """获取抓取状态"""
    try:
        import sys
        sys.path.insert(0, 'services')
        from yantai_rebar_scraper import YantaiRebarScraper

        scraper = YantaiRebarScraper()
        can_fetch, reason = scraper._check_rate_limit()

        return {
            "can_fetch": can_fetch,
            "reason": reason,
            "excel_exists": Path(_excel_file).exists()
        }

    except Exception as e:
        return {
            "can_fetch": False,
            "reason": str(e),
            "excel_exists": False
        }