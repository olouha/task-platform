"""
调差计算模板生成API
生成Excel格式的调差输入模板
支持材料清单和价格的自动联动
"""
from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import FileResponse
from typing import Optional, List, Dict, Any
from datetime import datetime, timedelta
from pathlib import Path
import os
import logging

router = APIRouter()
logger = logging.getLogger(__name__)

TEMPLATE_DIR = 'services/data/templates'
os.makedirs(TEMPLATE_DIR, exist_ok=True)

# 默认品牌优先级（用于价格计算）
DEFAULT_BRANDS = ['莱钢', '莱钢永锋', '永锋', '石横', '日钢', '镔鑫', '西王', '三德']


@router.get("/template")
async def generate_template(
    project_name: str = Query("调差项目", description="项目名称"),
    rule_name: str = Query("标准调差规则", description="调差规则名称"),
    material_type: str = Query("钢筋", description="材料类型"),
    include_examples: bool = Query(True, description="是否包含示例数据")
):
    """
    生成调差计算Excel模板

    模板包含以下工作表：
    1. 【项目信息】- 项目基本信息
    2. 【材料清单】- 投标材料明细
    3. 【施工时间】- 各楼栋/阶段施工时间
    4. 【调差配置】- 调差规则配置
    5. 【填写说明】- 各表填写指南
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # 样式定义
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4A86C8', end_color='4A86C8', fill_type='solid')
        title_font = Font(bold=True, size=14, color='16325C')
        required_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        def set_header_row(ws, row, headers, widths=None):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            if widths:
                for i, w in enumerate(widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = w

        # ===== 1. 项目信息表 =====
        ws1 = wb.active
        ws1.title = '项目信息'

        ws1.merge_cells('A1:D1')
        ws1.cell(row=1, column=1, value='调差项目基本信息').font = title_font

        project_info = [
            ('项目名称', project_name, '必填'),
            ('合同编号', '', '必填'),
            ('甲方单位', '', '必填'),
            ('乙方单位', '', '必填'),
            ('合同签订日期', '', '选填'),
            ('项目地址', '', '选填'),
            ('建筑面积(㎡)', '', '选填'),
            ('结构类型', '', '选填'),
            ('层数(地上/地下)', '', '选填'),
            ('开工日期', '', '必填'),
            ('计划竣工日期', '', '必填'),
            ('调差规则', rule_name, '必填'),
            ('备注', '', '选填'),
        ]

        for i, (label, value, required) in enumerate(project_info, 3):
            ws1.cell(row=i, column=1, value=label).font = Font(bold=True)
            cell = ws1.cell(row=i, column=2, value=value)
            if required == '必填':
                cell.fill = required_fill
            ws1.cell(row=i, column=3, value=required)
            ws1.cell(row=i, column=2).border = thin_border
            ws1.cell(row=i, column=3).border = thin_border

        ws1.column_dimensions['A'].width = 18
        ws1.column_dimensions['B'].width = 30
        ws1.column_dimensions['C'].width = 10

        # ===== 2. 材料清单表 =====
        ws2 = wb.create_sheet('材料清单')

        ws2.merge_cells('A1:H1')
        ws2.cell(row=1, column=1, value='投标材料清单（请按以下格式填写）').font = title_font

        material_headers = ['序号', '材料名称', '规格型号', '单位', '投标单价(元)', '品牌/厂家', '供货方式', '备注']
        set_header_row(ws2, 3, material_headers, [8, 15, 15, 8, 15, 15, 12, 20])

        # 材料类型选项说明
        material_types = ['HRB400钢筋', 'HPB300钢筋', '盘螺', '高线', '圆钢', '其他钢材']
        ws2.cell(row=2, column=1, value='材料类型：' + '、'.join(material_types))

        # 示例数据
        if include_examples:
            examples = [
                (1, 'HRB400钢筋', 'Φ12', '吨', 4500, '莱钢', '乙供', ''),
                (2, 'HRB400钢筋', 'Φ14', '吨', 4450, '莱钢', '乙供', ''),
                (3, 'HRB400钢筋', 'Φ16', '吨', 4400, '莱钢', '乙供', ''),
                (4, 'HRB400钢筋', 'Φ18', '吨', 4380, '莱钢', '乙供', ''),
                (5, 'HRB400钢筋', 'Φ20', '吨', 4350, '莱钢', '乙供', ''),
                (6, 'HRB400钢筋', 'Φ22', '吨', 4320, '莱钢', '乙供', ''),
                (7, 'HRB400钢筋', 'Φ25', '吨', 4300, '莱钢', '乙供', ''),
                (8, '盘螺', 'Φ8', '吨', 4550, '中天', '乙供', ''),
            ]
            for row_idx, data in enumerate(examples, 4):
                for col_idx, value in enumerate(data, 1):
                    cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    if col_idx == 1:
                        cell.alignment = Alignment(horizontal='center')

        # 添加更多空行
        for row_idx in range(12, 32):
            for col_idx in range(1, 9):
                ws2.cell(row=row_idx, column=col_idx).border = thin_border

        # ===== 3. 施工时间表 =====
        ws3 = wb.create_sheet('施工时间')

        ws3.merge_cells('A1:G1')
        ws3.cell(row=1, column=1, value='楼栋/阶段施工时间（请按以下格式填写）').font = title_font

        ws3.cell(row=2, column=1, value='说明：每个楼栋或施工阶段填写一行，日期格式：YYYY-MM-DD')

        time_headers = ['序号', '楼栋/阶段名称', '结构部位', '开始日期', '结束日期', '施工天数', '备注']
        set_header_row(ws3, 4, time_headers, [8, 18, 15, 15, 15, 12, 20])

        if include_examples:
            time_examples = [
                (1, '地下室', '基础结构', '2024-01-01', '2024-04-30', '', '含钢筋绑扎、模板、混凝土'),
                (2, '1#楼', '主体结构', '2024-05-01', '2024-08-31', '', ''),
                (3, '2#楼', '主体结构', '2024-05-15', '2024-09-15', '', ''),
                (4, '3#楼', '主体结构', '2024-06-01', '2024-10-31', '', ''),
            ]
            for row_idx, data in enumerate(time_examples, 5):
                for col_idx, value in enumerate(data, 1):
                    cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border

        for row_idx in range(9, 29):
            for col_idx in range(1, 8):
                ws3.cell(row=row_idx, column=col_idx).border = thin_border

        # ===== 4. 调差配置表 =====
        ws4 = wb.create_sheet('调差配置')

        ws4.merge_cells('A1:D1')
        ws4.cell(row=1, column=1, value='调差规则配置（请选择或填写）').font = title_font

        config_items = [
            ('配置项', '选项/值', '说明'),
            ('风险幅度(%)', '±3', '价格波动容忍范围，超出部分才调差'),
            ('基准价来源', '造价信息', '可选项：造价信息/钢铁网/投标价/合同价'),
            ('是否分阶段', '是', '可选项：是/否'),
            ('阶段划分', '地下室/主体/装饰', '按施工阶段分开计算'),
            ('增值税率(%)', '9', '可选项：3/9/13'),
            ('跌价处理', '按实计算', '可选项：扣回/不调整/按实计算'),
            ('节假日处理', '顺延取价', '可选项：顺延/取前后均价/取上月价'),
            ('价格取整', '到元', '可选项：到元/保留2位小数'),
            ('是否含税', '含税价', '可选项：含税价/除税价'),
        ]

        for row_idx, row_data in enumerate(config_items, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws4.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx == 3:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                elif col_idx == 1:
                    cell.font = Font(bold=True)

        ws4.column_dimensions['A'].width = 20
        ws4.column_dimensions['B'].width = 20
        ws4.column_dimensions['C'].width = 40

        # ===== 5. 填写说明 =====
        ws5 = wb.create_sheet('填写说明')

        ws5.merge_cells('A1:B1')
        ws5.cell(row=1, column=1, value='调差计算模板填写说明').font = title_font

        instructions = [
            '',
            '【使用流程】',
            '1. 下载本模板',
            '2. 在「项目信息」表中填写项目基本信息',
            '3. 在「材料清单」表中填写投标材料明细（含投标单价）',
            '4. 在「施工时间」表中填写各楼栋/阶段的施工时间',
            '5. 在「调差配置」表中确认或修改调差规则',
            '6. 将填写好的模板导入系统执行调差计算',
            '',
            '【注意事项】',
            '• 红色标记的单元格为必填项',
            '• 日期格式：YYYY-MM-DD（如：2024-01-15）',
            '• 价格单位：元/吨或元/m³',
            '• 材料名称和规格需与价格数据库中的记录一致',
            '',
            '【调差计算公式】',
            '调整金额 = 工程量 × (施工期均价 - 基准价 ± 风险幅度) × (1 + 增值税率)',
            '',
            '【示例数据说明】',
            '• 材料清单中的数据为示例，请替换为实际数据',
            '• 施工时间中的数据为示例，请根据实际情况填写',
        ]

        for row_idx, text in enumerate(instructions, 3):
            cell = ws5.cell(row=row_idx, column=1, value=text)
            if text.startswith('【'):
                cell.font = Font(bold=True, color='4A86C8')

        ws5.column_dimensions['A'].width = 80

        # 保存文件
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_name = f'调差计算模板_{project_name}_{timestamp}.xlsx'
        file_path = os.path.join(TEMPLATE_DIR, file_name)
        wb.save(file_path)

        logger.info(f"[generate_template] 模板生成成功 | {file_name}")

        return {
            'success': True,
            'file_name': file_name,
            'file_path': file_path,
            'message': '模板生成成功'
        }

    except ImportError:
        logger.error("[generate_template] openpyxl未安装")
        raise HTTPException(status_code=500, detail='Excel库未安装，请联系管理员')
    except Exception as e:
        logger.error(f"[generate_template] 模板生成失败 | {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/download/{file_name}")
async def download_template(file_name: str):
    """下载调差计算模板"""
    file_path = os.path.join(TEMPLATE_DIR, file_name)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail='模板文件不存在')

    return FileResponse(
        path=file_path,
        filename=file_name,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@router.get("/list")
async def list_templates():
    """列出所有生成的模板文件"""
    try:
        files = []
        if os.path.exists(TEMPLATE_DIR):
            for f in os.listdir(TEMPLATE_DIR):
                if f.endswith('.xlsx'):
                    full_path = os.path.join(TEMPLATE_DIR, f)
                    stat = os.stat(full_path)
                    files.append({
                        'name': f,
                        'size': stat.st_size,
                        'created': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S')
                    })

        files.sort(key=lambda x: x['created'], reverse=True)
        return {'success': True, 'files': files}
    except Exception as e:
        logger.error(f"[list_templates] 获取模板列表失败 | {e}")
        return {'success': True, 'files': []}


# ============================================================
# 材料清单自动联动 API
# ============================================================

def load_db_prices() -> List[Dict]:
    """从SQLite数据库加载价格数据"""
    import sqlite3

    db_file = 'services/data/yantai_rebar.db'
    if not os.path.exists(db_file):
        logger.warning(f"[load_db_prices] 数据库不存在: {db_file}")
        return []

    try:
        conn = sqlite3.connect(db_file)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute('''
            SELECT date, material_name, spec, brand, price, material_type
            FROM rebar_prices
            WHERE price > 0
            ORDER BY date DESC
        ''')

        prices = []
        for row in cursor.fetchall():
            prices.append({
                'date': row['date'],
                'material_name': row['material_name'],
                'spec': row['spec'] or '',
                'brand': row['brand'] or '',
                'price': row['price'],
                'material_type': row['material_type'] or ''
            })

        conn.close()
        logger.info(f"[load_db_prices] 加载价格数据 {len(prices)} 条")
        return prices

    except Exception as e:
        logger.error(f"[load_db_prices] 加载失败 | {e}", exc_info=True)
        return []


def get_material_list_from_prices(prices: List[Dict]) -> List[Dict]:
    """
    从价格数据中提取材料清单（规格和品牌组合）
    返回标准化的材料列表
    """
    # 按品名+规格+品牌分组，取最新价格
    material_map = {}

    for p in prices:
        key = (p['material_name'], p['spec'], p['brand'])
        if key not in material_map or p['date'] > material_map[key]['latest_date']:
            material_map[key] = {
                'material_name': p['material_name'],
                'spec': p['spec'],
                'brand': p['brand'],
                'latest_date': p['date'],
                'latest_price': p['price']
            }

    # 转换为列表并排序
    materials = list(material_map.values())
    materials.sort(key=lambda x: (x['material_name'], x['spec'], x['brand']))

    return materials


def get_latest_price_for_material(
    prices: List[Dict],
    material_name: str,
    spec: str = None,
    brands: List[str] = None
) -> Optional[float]:
    """获取指定材料的最新价格"""
    for p in prices:
        if p['material_name'] != material_name:
            continue
        if spec and p['spec'] != spec:
            continue
        if brands and p['brand'] not in brands:
            continue
        return p['price']
    return None


@router.get("/materials", summary="获取材料清单（自动从价格库提取）")
async def get_auto_materials(
    material_type: str = Query("钢筋", description="材料类型：钢筋/混凝土/全部"),
    include_all_specs: bool = Query(True, description="是否包含所有规格")
):
    """
    自动获取材料清单

    从价格数据库中提取所有品名、规格、品牌组合，
    用于自动填充调差模板的材料清单。

    返回格式：
    - material_name: 材料名称（如"螺纹钢"、"高线"）
    - spec: 规格型号（如"Φ12"、"Φ14"）
    - brand: 品牌（如"莱钢"、"石横"）
    - latest_price: 最新价格
    - latest_date: 价格日期
    """
    logger.info(f"[get_auto_materials] 获取材料清单 | type={material_type}")

    try:
        prices = load_db_prices()

        if not prices:
            # 数据库为空，返回默认材料列表
            default_materials = [
                {'material_name': '螺纹钢', 'spec': 'Φ12', 'brand': '莱钢', 'latest_price': 0, 'latest_date': None},
                {'material_name': '螺纹钢', 'spec': 'Φ14', 'brand': '莱钢', 'latest_price': 0, 'latest_date': None},
                {'material_name': '螺纹钢', 'spec': 'Φ16', 'brand': '莱钢', 'latest_price': 0, 'latest_date': None},
                {'material_name': '螺纹钢', 'spec': 'Φ18', 'brand': '莱钢', 'latest_price': 0, 'latest_date': None},
                {'material_name': '螺纹钢', 'spec': 'Φ20', 'brand': '莱钢', 'latest_price': 0, 'latest_date': None},
                {'material_name': '螺纹钢', 'spec': 'Φ22', 'brand': '莱钢', 'latest_price': 0, 'latest_date': None},
                {'material_name': '螺纹钢', 'spec': 'Φ25', 'brand': '莱钢', 'latest_price': 0, 'latest_date': None},
                {'material_name': '螺纹钢', 'spec': 'Φ28', 'brand': '莱钢', 'latest_price': 0, 'latest_date': None},
                {'material_name': '高线', 'spec': 'Φ6', 'brand': '莱钢', 'latest_price': 0, 'latest_date': None},
                {'material_name': '高线', 'spec': 'Φ8', 'brand': '莱钢', 'latest_price': 0, 'latest_date': None},
                {'material_name': '盘螺', 'spec': 'Φ8', 'brand': '莱钢', 'latest_price': 0, 'latest_date': None},
                {'material_name': '圆钢', 'spec': 'Φ12-22', 'brand': '莱钢', 'latest_price': 0, 'latest_date': None},
            ]
            logger.warning("[get_auto_materials] 价格数据库为空，返回默认材料列表")
            return {
                'success': True,
                'source': 'default',
                'materials': default_materials,
                'total_count': len(default_materials)
            }

        # 筛选材料类型
        filtered_prices = prices
        if material_type == '钢筋':
            filtered_prices = [p for p in prices if p['material_name'] in ['螺纹钢', '高线', '盘螺', '圆钢']]
        elif material_type == '混凝土':
            filtered_prices = [p for p in prices if '混凝土' in p['material_name']]

        # 提取材料清单
        materials = get_material_list_from_prices(filtered_prices)

        logger.info(f"[get_auto_materials] 返回 {len(materials)} 种材料")

        return {
            'success': True,
            'source': 'database',
            'materials': materials,
            'total_count': len(materials),
            'price_date_range': {
                'latest': max(p['latest_date'] for p in materials) if materials else None,
                'earliest': min(p['latest_date'] for p in materials) if materials else None
            }
        }

    except Exception as e:
        logger.error(f"[get_auto_materials] 获取材料清单失败 | {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/prices/period-average", summary="获取施工期均价")
async def get_period_average_price(
    material_name: str = Query(..., description="材料名称"),
    start_date: str = Query(..., description="开始日期 YYYY-MM-DD"),
    end_date: str = Query(..., description="结束日期 YYYY-MM-DD"),
    spec: str = Query(None, description="规格型号"),
    brands: str = Query(None, description="品牌偏好，逗号分隔")
):
    """
    获取指定材料在施工期间的平均价格

    用于调差计算中的施工期均价获取
    """
    logger.info(f"[get_period_average_price] 获取均价 | material={material_name}, start={start_date}, end={end_date}")

    try:
        prices = load_db_prices()

        if not prices:
            logger.warning("[get_period_average_price] 价格数据库为空")
            return {
                'success': True,
                'material_name': material_name,
                'start_date': start_date,
                'end_date': end_date,
                'avg_price': 0,
                'data_count': 0,
                'source': 'default'
            }

        # 解析品牌偏好
        brand_list = [b.strip() for b in brands.split(',')] if brands else DEFAULT_BRANDS

        # 筛选符合条件的记录
        filtered = []
        for p in prices:
            if not (start_date <= p['date'] <= end_date):
                continue
            if p['material_name'] != material_name:
                continue
            if spec and p['spec'] != spec:
                continue
            if brand_list and p['brand'] not in brand_list:
                continue
            filtered.append(p['price'])

        # 计算均价
        if filtered:
            avg_price = sum(filtered) / len(filtered)
        else:
            # 如果没有匹配的数据，尝试不使用品牌过滤
            for p in prices:
                if start_date <= p['date'] <= end_date and p['material_name'] == material_name:
                    if spec is None or p['spec'] == spec:
                        filtered.append(p['price'])

            avg_price = sum(filtered) / len(filtered) if filtered else 0

        result = {
            'success': True,
            'material_name': material_name,
            'spec': spec,
            'start_date': start_date,
            'end_date': end_date,
            'avg_price': round(avg_price, 2),
            'data_count': len(filtered),
            'source': 'database'
        }

        logger.info(f"[get_period_average_price] 均价: {avg_price:.2f}, 数据点: {len(filtered)}")
        return result

    except Exception as e:
        logger.error(f"[get_period_average_price] 获取均价失败 | {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/prices/batch-period-average", summary="批量获取施工期均价")
async def batch_get_period_average(
    materials: str = Query(..., description="材料列表，逗号分隔，格式：品名@规格"),
    start_date: str = Query(..., description="开始日期 YYYY-MM-DD"),
    end_date: str = Query(..., description="结束日期 YYYY-MM-DD"),
    brands: str = Query(None, description="品牌偏好，逗号分隔")
):
    """
    批量获取多个材料的施工期均价

    materials 格式：螺纹钢@Φ12,螺纹钢@Φ14,高线@Φ8
    """
    logger.info(f"[batch_get_period_average] 批量获取均价 | materials={materials}")

    try:
        prices = load_db_prices()
        brand_list = [b.strip() for b in brands.split(',')] if brands else DEFAULT_BRANDS

        # 解析材料列表
        material_list = []
        for m in materials.split(','):
            if '@' in m:
                name, spec = m.split('@', 1)
                material_list.append({'name': name.strip(), 'spec': spec.strip()})
            else:
                material_list.append({'name': m.strip(), 'spec': None})

        results = []
        for mat in material_list:
            mat_name = mat['name']
            mat_spec = mat['spec']

            # 筛选数据
            filtered = []
            for p in prices:
                if not (start_date <= p['date'] <= end_date):
                    continue
                if p['material_name'] != mat_name:
                    continue
                if mat_spec and p['spec'] != mat_spec:
                    continue
                if brand_list and p['brand'] not in brand_list:
                    continue
                filtered.append(p['price'])

            # 计算均价
            if filtered:
                avg_price = sum(filtered) / len(filtered)
            else:
                avg_price = 0

            results.append({
                'material_name': mat_name,
                'spec': mat_spec,
                'avg_price': round(avg_price, 2),
                'data_count': len(filtered)
            })

        return {
            'success': True,
            'start_date': start_date,
            'end_date': end_date,
            'results': results,
            'total_count': len(results)
        }

    except Exception as e:
        logger.error(f"[batch_get_period_average] 批量获取均价失败 | {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/template/with-prices", summary="生成带价格的模板")
async def generate_template_with_prices(
    project_name: str = Query("调差项目", description="项目名称"),
    rule_name: str = Query("标准调差规则", description="调差规则名称"),
    material_type: str = Query("钢筋", description="材料类型"),
    price_date: str = Query(None, description="价格日期 YYYY-MM-DD，默认最新")
):
    """
    生成带自动填充价格的调差模板

    自动从价格数据库获取：
    1. 材料清单（品名+规格+品牌）
    2. 最新价格
    """
    logger.info(f"[generate_template_with_prices] 生成模板 | project={project_name}, type={material_type}")

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        # 获取材料清单和价格
        prices = load_db_prices()

        if prices:
            # 筛选材料类型
            if material_type == '钢筋':
                filtered = [p for p in prices if p['material_name'] in ['螺纹钢', '高线', '盘螺', '圆钢']]
            else:
                filtered = prices

            materials = get_material_list_from_prices(filtered)
        else:
            # 默认材料列表
            materials = [
                {'material_name': '螺纹钢', 'spec': 'Φ12', 'brand': '莱钢', 'latest_price': 0},
                {'material_name': '螺纹钢', 'spec': 'Φ14', 'brand': '莱钢', 'latest_price': 0},
                {'material_name': '螺纹钢', 'spec': 'Φ16', 'brand': '莱钢', 'latest_price': 0},
                {'material_name': '螺纹钢', 'spec': 'Φ18', 'brand': '莱钢', 'latest_price': 0},
                {'material_name': '螺纹钢', 'spec': 'Φ20', 'brand': '莱钢', 'latest_price': 0},
                {'material_name': '螺纹钢', 'spec': 'Φ22', 'brand': '莱钢', 'latest_price': 0},
                {'material_name': '螺纹钢', 'spec': 'Φ25', 'brand': '莱钢', 'latest_price': 0},
                {'material_name': '高线', 'spec': 'Φ6', 'brand': '莱钢', 'latest_price': 0},
                {'material_name': '高线', 'spec': 'Φ8', 'brand': '莱钢', 'latest_price': 0},
                {'material_name': '盘螺', 'spec': 'Φ8', 'brand': '莱钢', 'latest_price': 0},
                {'material_name': '圆钢', 'spec': 'Φ12-22', 'brand': '莱钢', 'latest_price': 0},
            ]

        # 创建工作簿
        wb = openpyxl.Workbook()

        # 样式定义
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4A86C8', end_color='4A86C8', fill_type='solid')
        title_font = Font(bold=True, size=14, color='16325C')
        required_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        price_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        def set_header_row(ws, row, headers, widths=None):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            if widths:
                for i, w in enumerate(widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = w

        # ===== 1. 项目信息表 =====
        ws1 = wb.active
        ws1.title = '项目信息'

        ws1.merge_cells('A1:D1')
        ws1.cell(row=1, column=1, value='调差项目基本信息').font = title_font

        project_info = [
            ('项目名称', project_name, '必填'),
            ('合同编号', '', '必填'),
            ('甲方单位', '', '必填'),
            ('乙方单位', '', '必填'),
            ('合同签订日期', '', '选填'),
            ('项目地址', '', '选填'),
            ('建筑面积(㎡)', '', '选填'),
            ('结构类型', '', '选填'),
            ('层数(地上/地下)', '', '选填'),
            ('开工日期', '', '必填'),
            ('计划竣工日期', '', '必填'),
            ('调差规则', rule_name, '必填'),
            ('备注', '', '选填'),
        ]

        for i, (label, value, required) in enumerate(project_info, 3):
            ws1.cell(row=i, column=1, value=label).font = Font(bold=True)
            cell = ws1.cell(row=i, column=2, value=value)
            if required == '必填':
                cell.fill = required_fill
            ws1.cell(row=i, column=3, value=required)
            ws1.cell(row=i, column=2).border = thin_border
            ws1.cell(row=i, column=3).border = thin_border

        ws1.column_dimensions['A'].width = 18
        ws1.column_dimensions['B'].width = 30
        ws1.column_dimensions['C'].width = 10

        # ===== 2. 材料清单表（自动填充） =====
        ws2 = wb.create_sheet('材料清单')

        ws2.merge_cells('A1:I1')
        ws2.cell(row=1, column=1, value=f'投标材料清单（自动填充，价格来源：系统价格库）').font = title_font

        # 添加价格日期信息
        price_date_str = price_date or (materials[0]['latest_date'] if materials and 'latest_date' in materials else '最新')
        ws2.cell(row=2, column=1, value=f'价格参考日期：{price_date_str}').font = Font(italic=True, color='666666')

        material_headers = ['序号', '材料名称', '规格型号', '单位', '投标单价(元)', '当前参考价(元)', '品牌/厂家', '供货方式', '备注']
        set_header_row(ws2, 3, material_headers, [8, 15, 12, 8, 15, 18, 15, 10, 15])

        # 自动填充材料清单
        for idx, mat in enumerate(materials, 1):
            row_num = idx + 3
            ws2.cell(row=row_num, column=1, value=idx).border = thin_border
            ws2.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
            ws2.cell(row=row_num, column=2, value=mat['material_name']).border = thin_border
            ws2.cell(row=row_num, column=3, value=mat['spec']).border = thin_border
            ws2.cell(row=row_num, column=4, value='吨').border = thin_border
            ws2.cell(row=row_num, column=5).border = thin_border  # 投标单价留空，用户填写
            ws2.cell(row=row_num, column=5).fill = required_fill

            # 当前参考价（自动填充）
            price_cell = ws2.cell(row=row_num, column=6, value=mat.get('latest_price', 0) or '')
            price_cell.border = thin_border
            price_cell.fill = price_fill
            price_cell.number_format = '#,##0'

            ws2.cell(row=row_num, column=7, value=mat['brand']).border = thin_border
            ws2.cell(row=row_num, column=8, value='乙供').border = thin_border
            ws2.cell(row=row_num, column=9).border = thin_border

        # 额外添加10行空行供用户补充
        start_row = len(materials) + 4
        for row_idx in range(start_row, start_row + 10):
            ws2.cell(row=row_idx, column=1, value=row_idx - 3).border = thin_border
            ws2.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')
            for col_idx in range(2, 10):
                ws2.cell(row=row_idx, column=col_idx).border = thin_border

        # ===== 3. 施工时间表 =====
        ws3 = wb.create_sheet('施工时间')

        ws3.merge_cells('A1:G1')
        ws3.cell(row=1, column=1, value='楼栋/阶段施工时间').font = title_font

        ws3.cell(row=2, column=1, value='说明：每个楼栋或施工阶段填写一行，日期格式：YYYY-MM-DD')

        time_headers = ['序号', '楼栋/阶段名称', '结构部位', '开始日期', '结束日期', '施工天数', '备注']
        set_header_row(ws3, 4, time_headers, [8, 18, 15, 15, 15, 12, 20])

        for row_idx in range(5, 15):
            for col_idx in range(1, 8):
                ws3.cell(row=row_idx, column=col_idx).border = thin_border

        # ===== 4. 调差配置表 =====
        ws4 = wb.create_sheet('调差配置')

        ws4.merge_cells('A1:D1')
        ws4.cell(row=1, column=1, value='调差规则配置').font = title_font

        config_items = [
            ('配置项', '选项/值', '说明'),
            ('风险幅度(%)', '±3', '价格波动容忍范围，超出部分才调差'),
            ('基准价来源', '造价信息', '可选项：造价信息/钢铁网/投标价/合同价'),
            ('是否分阶段', '否', '可选项：是/否'),
            ('增值税率(%)', '9', '可选项：3/9/13'),
            ('跌价处理', '按实计算', '可选项：扣回/不调整/按实计算'),
        ]

        for row_idx, row_data in enumerate(config_items, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws4.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx == 3:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                elif col_idx == 1:
                    cell.font = Font(bold=True)

        ws4.column_dimensions['A'].width = 20
        ws4.column_dimensions['B'].width = 20
        ws4.column_dimensions['C'].width = 40

        # 保存文件
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_name = f'调差计算模板_{project_name}_{timestamp}.xlsx'
        file_path = os.path.join(TEMPLATE_DIR, file_name)
        wb.save(file_path)

        logger.info(f"[generate_template_with_prices] 模板生成成功 | {file_name}, 材料数: {len(materials)}")

        return {
            'success': True,
            'file_name': file_name,
            'file_path': file_path,
            'materials_count': len(materials),
            'message': f'模板生成成功，自动填充了 {len(materials)} 种材料'
        }

    except ImportError:
        logger.error("[generate_template_with_prices] openpyxl未安装")
        raise HTTPException(status_code=500, detail='Excel库未安装')
    except Exception as e:
        logger.error(f"[generate_template_with_prices] 模板生成失败 | {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
