"""
定时抓取任务 API
可用于 Cloudflare Workers cron 触发
或被其他服务调用
"""

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from datetime import datetime
import logging

router = APIRouter()
logger = logging.getLogger(__name__)


class FetchResult(BaseModel):
    success: bool
    message: str
    timestamp: str
    prices_count: int = 0


@router.get("/fetch-today")
async def fetch_today_prices():
    """
    抓取今日价格
    可被定时任务调用
    """
    try:
        import asyncio
        import sys
        from pathlib import Path

        # 切换到 services 目录以确保模块导入正确
        services_dir = Path(__file__).parent.parent / "services"
        sys.path.insert(0, str(services_dir))

        from yantai_rebar_scraper import YantaiRebarScraper, save_to_excel

        scraper = YantaiRebarScraper()

        # 检查是否今日已抓取
        can_fetch, reason = scraper._check_rate_limit()

        if not can_fetch:
            # 今日已抓取，直接返回成功
            return FetchResult(
                success=True,
                message=f"今日已抓取: {reason}",
                timestamp=datetime.now().isoformat(),
                prices_count=0
            )

        # 执行抓取
        result = await scraper.fetch_async(force=True)

        if result.success and result.prices:
            # 保存到Excel
            save_to_excel(result)

            return FetchResult(
                success=True,
                message=f"抓取成功",
                timestamp=result.fetched_at,
                prices_count=len(result.prices)
            )
        else:
            return FetchResult(
                success=False,
                message=result.error_message or "抓取失败",
                timestamp=datetime.now().isoformat()
            )

    except Exception as e:
        logger.error(f"抓取异常: {e}")
        import traceback
        traceback.print_exc()
        return FetchResult(
            success=False,
            message=f"抓取异常: {str(e)}",
            timestamp=datetime.now().isoformat()
        )


@router.get("/status")
async def get_status():
    """获取抓取状态"""
    from pathlib import Path

    last_fetch_file = Path(__file__).parent.parent / "services" / "logs" / "yantai_last_fetch.json"

    if last_fetch_file.exists():
        import json
        with open(last_fetch_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return {
                "last_fetch": data.get('last_fetch'),
                "success": data.get('success'),
                "prices_count": data.get('prices_count'),
                "region": data.get('region', '山东烟台'),
                "today_fetched": data.get('last_fetch', '')[:10] == datetime.now().date().isoformat()
            }

    return {
        "last_fetch": None,
        "success": None,
        "prices_count": 0,
        "today_fetched": False
    }


@router.post("/force-fetch")
async def force_fetch():
    """强制重新抓取（忽略每日限制）"""
    try:
        import asyncio
        import sys
        from pathlib import Path

        services_dir = Path(__file__).parent.parent / "services"
        sys.path.insert(0, str(services_dir))

        from yantai_rebar_scraper import YantaiRebarScraper, save_to_excel

        scraper = YantaiRebarScraper()
        result = await scraper.fetch_async(force=True)

        if result.success and result.prices:
            save_to_excel(result)
            return FetchResult(
                success=True,
                message=f"强制抓取成功",
                timestamp=result.fetched_at,
                prices_count=len(result.prices)
            )
        else:
            return FetchResult(
                success=False,
                message=result.error_message or "强制抓取失败",
                timestamp=datetime.now().isoformat()
            )

    except Exception as e:
        return FetchResult(
            success=False,
            message=f"强制抓取异常: {str(e)}",
            timestamp=datetime.now().isoformat()
        )


@router.get("/latest")
async def get_latest():
    """获取最新抓取的数据"""
    try:
        import openpyxl
        from pathlib import Path

        excel_file = Path(__file__).parent.parent / "services" / "data" / "山东烟台钢筋价格.xlsx"

        if not excel_file.exists():
            return {"success": False, "message": "暂无数据"}

        wb = openpyxl.load_workbook(excel_file)
        latest_sheet = wb.sheetnames[-1] if wb.sheetnames else None

        if not latest_sheet:
            return {"success": False, "message": "暂无Sheet"}

        ws = wb[latest_sheet]
        prices = []

        for row in range(4, ws.max_row + 1):
            date = ws.cell(row=row, column=1).value
            if date and str(date) != "当日截图" and str(date).startswith("2026"):
                prices.append({
                    "date": str(date),
                    "time": ws.cell(row=row, column=2).value,
                    "material_name": ws.cell(row=row, column=3).value,
                    "spec": ws.cell(row=row, column=4).value,
                    "brand": ws.cell(row=row, column=6).value,
                    "price": ws.cell(row=row, column=7).value
                })

        wb.close()

        return {
            "success": True,
            "sheet": latest_sheet,
            "prices_count": len(prices),
            "prices": prices[:10]  # 返回前10条
        }

    except Exception as e:
        return {"success": False, "message": str(e)}


# 本地定时任务脚本路径
LOCAL_SCRIPT = Path(__file__).parent.parent / "services" / "auto_fetch.bat"


@router.get("/setup-local-cron")
async def setup_local_cron():
    """生成Windows定时任务脚本"""
    script_content = '''@echo off
REM 山东烟台钢筋价格自动抓取脚本
REM 每天早上8点自动执行

echo [%date% %time%] 开始抓取钢筋价格...

cd /d "%~dp0"
cd services

REM 设置Python路径
set PYTHONPATH=%CD%

REM 执行抓取（忽略每日限制，用于定时任务）
python -c "
import asyncio
import sys
sys.path.insert(0, '.')
from yantai_rebar_scraper import YantaiRebarScraper, save_to_excel

scraper = YantaiRebarScraper()
result = asyncio.run(scraper.fetch_async(force=True))

if result.success and result.prices:
    save_to_excel(result)
    print(f'抓取成功: {len(result.prices)} 条数据')
else:
    print(f'抓取失败: {result.error_message}')
"

echo [%date% %time%] 抓取完成

REM 保留日志
echo [%date% %time%] >> logs/cron.log
'''

    try:
        with open(LOCAL_SCRIPT, 'w', encoding='utf-8') as f:
            f.write(script_content)
        return {
            "success": True,
            "script_path": str(LOCAL_SCRIPT),
            "usage": "使用 Windows 任务计划程序 创建每天早上8点执行此脚本的任务"
        }
    except Exception as e:
        return {"success": False, "message": str(e)}