"""
人工抓取API - 使用员工手动登录的Cookie
员工在浏览器登录后，导出Cookie，然后通过此接口触发抓取
"""

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import Optional
import json
import base64
from pathlib import Path
from datetime import datetime
import asyncio
from playwright.async_api import async_playwright
import openpyxl
import hashlib
import logging

from services.fetch_status_manager import get_status_manager, FetchStatus, FetchRecord

router = APIRouter(prefix="/api/fetch", tags=["价格抓取"])
logger = logging.getLogger(__name__)

DATA_DIR = Path(__file__).parent / 'data'
EXCEL_FILE = DATA_DIR / '山东烟台钢筋价格.xlsx'
TODAY = datetime.now().date().isoformat()
TIME_NOW = datetime.now().strftime('%H%M%S')


class ManualCookieInput(BaseModel):
    """手动Cookie输入"""
    cookies: list  # Cookie列表
    data_source: str = "manual"  # 数据来源标识


class FetchTriggerRequest(BaseModel):
    """抓取触发请求"""
    force: bool = False  # 强制重新抓取（跳过去重检查）


class FetchResult(BaseModel):
    """抓取结果"""
    success: bool
    period: str
    count: int
    message: str
    timestamp: str


def calculate_data_hash(data: list) -> str:
    """计算数据哈希"""
    sorted_data = sorted([
        (d['material_name'], d['spec'], d['brand'], d['price'])
        for d in data
    ])
    return hashlib.md5(json.dumps(sorted_data, ensure_ascii=False).encode()).hexdigest()


def save_to_excel(data: list, period: str, record_id: str) -> bool:
    """保存数据到Excel"""
    logger.info(f"[save_to_excel] 保存数据 | period={period}, count={len(data)}")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE) if EXCEL_FILE.exists() else openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        sheet_name = f'{TODAY}_{period}_{TIME_NOW}'
        ws = wb.create_sheet(title=sheet_name)

        period_label = '下午(较晚)' if period == 'PM' else '上午'
        ws.merge_cells('A1:K1')
        ws.cell(row=1, column=1, value=f'山东烟台钢筋价格 - {TODAY} {period_label}').font = openpyxl.styles.Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')

        from openpyxl.styles import PatternFill
        header_fill = PatternFill(start_color='FF6B6B' if period == 'PM' else '4472C4', end_color='FF6B6B' if period == 'PM' else '4472C4', fill_type='solid')
        headers = ['日期', '时间', '品名', '规格', '材质', '品牌/钢厂', '单价(元/吨)', '涨跌', '备注', '钢号', '地区']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True, size=12, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        fetch_time = datetime.now().strftime('%H:%M:%S')
        for i, item in enumerate(data):
            row = 4 + i
            for col, val in enumerate([TODAY, fetch_time, item['material_name'], item['spec'],
                item.get('material_type', ''), item['brand'], item['price'], '', '', '', '山东烟台'], 1):
                ws.cell(row=row, column=col, value=val)

        wb.save(EXCEL_FILE)
        wb.close()

        # 保存哈希
        hash_file = DATA_DIR / 'data_hashes.json'
        hashes = json.load(open(hash_file)) if hash_file.exists() else {}
        hashes[f'{TODAY}_{period}'] = calculate_data_hash(data)
        with open(hash_file, 'w') as f:
            json.dump(hashes, f, ensure_ascii=False)

        logger.info(f"[save_to_excel] 保存成功 | sheet={sheet_name}, count={len(data)}")
        return True
    except Exception as e:
        logger.error(f"[save_to_excel] 保存失败 | {e}", exc_info=True)
        return False


async def manual_fetch_with_cookies(cookies: list, period: str = None, record_id: str = None) -> FetchResult:
    """使用手动Cookie进行抓取"""
    logger.info(f"[manual_fetch_with_cookies] 开始抓取 | cookies_count={len(cookies)}, period={period}")
    if period is None:
        current_hour = datetime.now().hour
        period = 'AM' if current_hour < 14 else 'PM'

    timestamp = datetime.now().isoformat()

    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True, args=['--no-sandbox'])
            context = await browser.new_context(viewport={'width': 1920, 'height': 1080}, locale='zh-CN')
            page = await context.new_page()

            await context.add_cookies(cookies)
            logger.info(f"[manual_fetch_with_cookies] 已加载 {len(cookies)} 个Cookie")

            url = 'https://jiancai.mysteel.com/mysteel/market/pa228aa010101a0a01010205aaaa1.html'
            logger.info(f"[manual_fetch_with_cookies] 访问 | url={url}")
            await page.goto(url, wait_until='domcontentloaded', timeout=60000)
            await asyncio.sleep(5)

            current_url = page.url
            if 'captcha' in current_url or 'passport' in current_url:
                logger.warning(f"[manual_fetch_with_cookies] Cookie失效")
                return FetchResult(
                    success=False,
                    period=period,
                    count=0,
                    message='Cookie已失效，需要重新登录',
                    timestamp=timestamp
                )

            data = await page.evaluate('''() => {
                const results = [];
                const tables = document.querySelectorAll('table');
                tables.forEach(table => {
                    const rows = table.querySelectorAll('tr');
                    rows.forEach(row => {
                        const cells = row.querySelectorAll('td');
                        if (cells.length >= 5) {
                            const material_name = cells[0]?.textContent?.trim();
                            const spec = cells[1]?.textContent?.trim();
                            const brand = cells[3]?.textContent?.trim();
                            const price_str = cells[4]?.textContent?.trim();
                            if (['高线', '螺纹钢', '盘螺', '圆钢'].includes(material_name) &&
                                spec && spec.startsWith('Φ') && price_str && /^\\d+$/.test(price_str)) {
                                results.push({
                                    material_name,
                                    spec,
                                    material_type: cells[2]?.textContent?.trim() || '',
                                    brand,
                                    price: parseInt(price_str, 10)
                                });
                            }
                        }
                    });
                });
                return results;
            }''')

            logger.info(f"[manual_fetch_with_cookies] 提取数据 | count={len(data)}")

            if not data:
                logger.warning("[manual_fetch_with_cookies] 未获取到数据")
                return FetchResult(
                    success=False,
                    period=period,
                    count=0,
                    message='未获取到价格数据',
                    timestamp=timestamp
                )

            success = save_to_excel(data, period, record_id)

            if success:
                period_label = '上午' if period == 'AM' else '下午(较晚)'
                logger.info(f"[manual_fetch_with_cookies] 抓取成功 | count={len(data)}")
                return FetchResult(
                    success=True,
                    period=period,
                    count=len(data),
                    message=f'{period_label}抓取成功，{len(data)}条数据',
                    timestamp=timestamp
                )
            else:
                logger.error("[manual_fetch_with_cookies] 数据保存失败")
                return FetchResult(
                    success=False,
                    period=period,
                    count=0,
                    message='数据保存失败',
                    timestamp=timestamp
                )

    except Exception as e:
        logger.error(f"[manual_fetch_with_cookies] 抓取异常 | {e}", exc_info=True)
        return FetchResult(
            success=False,
            period=period,
            count=0,
            message=f'抓取异常: {str(e)}',
            timestamp=timestamp
        )


# ============================================================
# API 端点
# ============================================================

@router.get("/status")
async def get_fetch_status():
    """获取抓取状态"""
    logger.info("[get_fetch_status] 获取抓取状态")
    manager = get_status_manager()
    result = manager.get_summary()
    logger.info(f"[get_fetch_status] 返回状态 | today_fetched={result.get('today_fetched')}")
    return result


@router.post("/manual")
async def trigger_manual_fetch(request: ManualCookieInput):
    """
    手动触发抓取

    请求参数:
    ```json
    {
        "cookies": [
            {"name": "JSESSIONID", "value": "...", "domain": ".mysteel.com"},
            {"name": "wm_ni", "value": "...", "domain": ".mysteel.com"}
        ],
        "data_source": "manual"
    }
    ```
    """
    logger.info(f"[trigger_manual_fetch] 手动触发 | cookies_count={len(request.cookies)}")
    manager = get_status_manager()
    current_hour = datetime.now().hour
    period = 'AM' if current_hour < 14 else 'PM'

    if not manager.is_fetched_today(TODAY, period):
        record_id = f"{TODAY}_{period}_{TIME_NOW}"
    else:
        existing = manager.get_period_record(TODAY, period)
        record_id = existing.id if existing else f"{TODAY}_{period}_{TIME_NOW}"

    record = FetchRecord(
        id=record_id,
        date=TODAY,
        period=period,
        status=FetchStatus.RUNNING,
        count=0,
        timestamp=datetime.now().isoformat(),
        requires_manual=False
    )
    manager.add_record(record)

    try:
        result = await manual_fetch_with_cookies(request.cookies, period, record_id)

        if result.success:
            manager.update_record(
                record_id,
                status=FetchStatus.SUCCESS,
                count=result.count,
                error_message=""
            )
            logger.info(f"[trigger_manual_fetch] 成功 | count={result.count}")
        else:
            manager.update_record(
                record_id,
                status=FetchStatus.FAILED if 'Cookie失效' in result.message else FetchStatus.MANUAL_REQUIRED,
                error_message=result.message
            )
            logger.warning(f"[trigger_manual_fetch] 失败 | message={result.message}")

        return result

    except Exception as e:
        logger.error(f"[trigger_manual_fetch] 异常 | {e}", exc_info=True)
        manager.update_record(
            record_id,
            status=FetchStatus.FAILED,
            error_message=str(e)
        )
        return FetchResult(
            success=False,
            period=period,
            count=0,
            message=f'抓取失败: {str(e)}',
            timestamp=datetime.now().isoformat()
        )


@router.post("/auto")
async def trigger_auto_fetch():
    """
    自动抓取（尝试自动登录）
    注意：由于验证码问题，大概率会失败，失败后会标记为需要手动操作
    """
    logger.info("[trigger_auto_fetch] 自动抓取")
    manager = get_status_manager()
    current_hour = datetime.now().hour
    period = 'AM' if current_hour < 14 else 'PM'

    if not manager.should_auto_fetch():
        logger.info("[trigger_auto_fetch] 不在自动抓取时段")
        return {
            "success": False,
            "message": "当前时间不在自动抓取时段内",
            "period": period
        }

    if manager.is_fetched_today(TODAY, period):
        logger.info(f"[trigger_auto_fetch] 今日已抓取 | period={period}")
        return {
            "success": False,
            "message": f"今日{period}时段已成功抓取",
            "period": period
        }

    record_id = f"{TODAY}_{period}_{TIME_NOW}"
    record = FetchRecord(
        id=record_id,
        date=TODAY,
        period=period,
        status=FetchStatus.RUNNING,
        count=0,
        timestamp=datetime.now().isoformat(),
        requires_manual=False
    )
    manager.add_record(record)

    try:
        from services.daily_fetch_ocr_v2 import YantaiPriceScraper
        from config.mysteel import MYSTEEL_USERNAME, MYSTEEL_PASSWORD

        async with YantaiPriceScraper() as scraper:
            is_logged_in = await scraper.check_login_status()

            if not is_logged_in:
                logger.info("[trigger_auto_fetch] 尝试自动登录")
                await scraper.login_with_captcha()
                await asyncio.sleep(2, 3)

                if not await scraper.check_login_status():
                    manager.update_record(
                        record_id,
                        status=FetchStatus.MANUAL_REQUIRED,
                        error_message="自动登录失败，需要手动操作"
                    )
                    logger.warning("[trigger_auto_fetch] 自动登录失败")
                    return {
                        "success": False,
                        "message": "自动登录失败（验证码），需要手动操作",
                        "period": period,
                        "requires_manual": True
                    }

            data = await scraper.fetch_prices()

            if not data:
                manager.update_record(
                    record_id,
                    status=FetchStatus.FAILED,
                    error_message="未获取到数据"
                )
                logger.warning("[trigger_auto_fetch] 未获取到数据")
                return {
                    "success": False,
                    "message": "未获取到数据",
                    "period": period
                }

            success = await asyncio.to_thread(scraper.save_to_excel, data, calculate_data_hash(data))
            if success:
                period_label = '上午' if period == 'AM' else '下午(较晚)'
                manager.update_record(
                    record_id,
                    status=FetchStatus.SUCCESS,
                    count=len(data),
                    error_message=""
                )
                logger.info(f"[trigger_auto_fetch] 成功 | count={len(data)}")
                return {
                    "success": True,
                    "message": f'{period_label}自动抓取成功，{len(data)}条数据',
                    "period": period,
                    "count": len(data)
                }
            else:
                manager.update_record(
                    record_id,
                    status=FetchStatus.FAILED,
                    error_message="数据保存失败"
                )
                logger.error("[trigger_auto_fetch] 保存失败")
                return {
                    "success": False,
                    "message": "数据保存失败",
                    "period": period
                }

    except Exception as e:
        logger.error(f"[trigger_auto_fetch] 异常 | {e}", exc_info=True)
        manager.update_record(
            record_id,
            status=FetchStatus.FAILED,
            error_message=str(e)
        )
        return {
            "success": False,
            "message": f"自动抓取失败: {str(e)}",
            "period": period
        }


@router.get("/manual-required")
async def get_manual_required():
    """获取需要手动操作的日期列表"""
    logger.info("[get_manual_required] 查询需要手动操作的日期")
    manager = get_status_manager()
    dates = manager.get_manual_required_dates(days=7)
    logger.info(f"[get_manual_required] 返回 {len(dates)} 个日期")
    return {
        "dates": dates,
        "total": len(dates)
    }


@router.post("/clear-old")
async def clear_old_records(days: int = 30):
    """清理旧记录"""
    logger.info(f"[clear_old_records] 清理 {days} 天前的记录")
    manager = get_status_manager()
    manager.clear_old_records(days)
    logger.info("[clear_old_records] 清理完成")
    return {"success": True, "message": f"已清理{days}天前的记录"}


@router.get("/export-cookie-guide")
async def get_cookie_export_guide():
    """获取Cookie导出指南"""
    logger.info("[get_cookie_export_guide] 获取导出指南")
    return {
        "title": "如何导出浏览器Cookie",
        "chrome_steps": [
            "1. 在已登录的网站页面按F12打开开发者工具",
            "2. 切换到 Application 标签页",
            "3. 左侧选择 Cookies",
            "4. 选中 .mysteel.com 的Cookie",
            "5. 右键 → Copy as cURL (bash)",
            "将复制的JSON数据发送到本接口"
        ],
        "edge_steps": [
            "1. 在已登录的网站页面按F12打开开发者工具",
            "2. 切换到 Application 标签页",
            "3. 左侧选择 Cookies",
            '4. 点击"导出"按钮',
            "选择JSON格式并下载",
            "将文件内容发送到本接口"
        ],
        "api_endpoint": "POST /api/fetch/manual"
    }


@router.get("/excel-data")
async def get_excel_data():
    """获取Excel文件（下载）"""
    logger.info("[get_excel_data] 获取Excel数据")
    if not EXCEL_FILE.exists():
        logger.warning("[get_excel_data] Excel文件不存在")
        raise HTTPException(status_code=404, detail="Excel文件不存在")

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet_count = len(wb.sheetnames)
    wb.close()

    manager = get_status_manager()
    today_records = manager.get_today_records()

    logger.info(f"[get_excel_data] 返回 | sheets={sheet_count}, today_records={len(today_records)}")
    return {
        "file_exists": True,
        "total_sheets": sheet_count,
        "today_records": [
            {
                "period": r.period,
                "status": r.status.value,
                "count": r.count,
                "time": r.timestamp
            }
            for r in today_records
        ],
        "download_url": "/api/fetch/download"
    }


@router.get("/download")
async def download_excel():
    """下载Excel文件"""
    logger.info("[download_excel] 下载Excel")
    if not EXCEL_FILE.exists():
        logger.warning("[download_excel] Excel文件不存在")
        raise HTTPException(status_code=404, detail="Excel文件不存在")

    from fastapi.responses import FileResponse
    logger.info("[download_excel] 开始下载")
    return FileResponse(
        path=str(EXCEL_FILE),
        filename=f'山东烟台钢筋价格_{TODAY}.xlsx',
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )