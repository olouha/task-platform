"""
调差价格获取 API
从已存储的价格数据中获取施工期均价
"""

from datetime import datetime, timedelta
from typing import List, Dict, Optional, Any
from pathlib import Path

from fastapi import APIRouter, HTTPException, Query
from pydantic import BaseModel

from services.adjustment_engine import PriceData

router = APIRouter(prefix="/adjustment-prices", tags=["调差价格获取"])


# ============================================================
# 数据模型
# ============================================================

class PriceQuery(BaseModel):
    """价格查询请求"""
    材料名称: str
    规格: Optional[str] = None
    开始日期: str  # YYYY-MM-DD
    结束日期: str   # YYYY-MM-DD
    品牌偏好: Optional[List[str]] = None  # 优先使用的品牌


class PriceResponse(BaseModel):
    """价格响应"""
    材料名称: str
    均价: float
    数据点数: int
    价格列表: List[Dict]
    查询参数: Dict


class PeriodPriceResult(BaseModel):
    """施工期价格结果"""
    材料名称: str
    施工期均价: float
    基准价: float
    涨跌幅度: float
    数据来源: str
    时间范围: Dict[str, str]


# ============================================================
# 价格数据获取
# ============================================================

def load_yantai_prices() -> List[Dict]:
    """从Excel加载烟台钢筋价格数据"""
    prices = []

    excel_file = Path('services/data/山东烟台钢筋价格.xlsx')
    if not excel_file.exists():
        return prices

    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_file, data_only=True)

        # 遍历所有sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
                try:
                    date_val = row[0].value  # 日期
                    time_val = row[1].value  # 时间
                    material_name = row[2].value  # 品名
                    spec = row[3].value  # 规格
                    material_type = row[4].value  # 材质
                    brand = row[5].value  # 品牌
                    price = row[6].value  # 单价

                    if material_name and price and isinstance(price, (int, float)):
                        # 解析日期
                        if isinstance(date_val, datetime):
                            date_str = date_val.strftime('%Y-%m-%d')
                        elif isinstance(date_val, str) and len(date_val) >= 10:
                            date_str = date_val[:10]
                        else:
                            continue

                        prices.append({
                            'date': date_str,
                            'time': str(time_val) if time_val else '',
                            'material_name': str(material_name),
                            'spec': str(spec) if spec else '',
                            'material_type': str(material_type) if material_type else '',
                            'brand': str(brand) if brand else '',
                            'price': float(price),
                            'region': '山东烟台'
                        })
                except:
                    continue

        wb.close()
    except Exception as e:
        print(f"加载Excel失败: {e}")

    return prices


def get_material_prices(
    material_name: str,
    start_date: str,
    end_date: str,
    all_prices: List[Dict] = None,
    spec: Optional[str] = None,
    brands: Optional[List[str]] = None
) -> Dict[str, Any]:
    """
    获取指定材料的施工期价格

    参数:
    - material_name: 材料名称（如"钢筋"、"螺纹钢"）
    - start_date: 开始日期
    - end_date: 结束日期
    - all_prices: 所有价格数据（用于避免重复加载）
    - spec: 规格过滤
    - brands: 品牌偏好
    """
    if all_prices is None:
        all_prices = load_yantai_prices()

    # 筛选符合条件的记录
    filtered = []
    for p in all_prices:
        if start_date <= p.get('date', '') <= end_date:
            # 匹配材料名称（支持模糊匹配）
            pname = p.get('material_name', '')
            if (material_name in pname or pname in material_name or
                '钢筋' in material_name and '钢筋' in pname or
                '高线' in pname or '盘螺' in pname or '螺纹' in pname):

                # 规格过滤
                if spec and spec not in p.get('spec', ''):
                    continue

                filtered.append(p)

    # 按品牌偏好排序
    if brands:
        def brand_score(p):
            brand = p.get('brand', '')
            for i, b in enumerate(brands):
                if b in brand or brand in b:
                    return i
            return len(brands)
        filtered.sort(key=brand_score)

    # 计算算术平均值
    if filtered:
        avg_price = sum(p['price'] for p in filtered) / len(filtered)
    else:
        avg_price = 0

    return {
        'material_name': material_name,
        'start_date': start_date,
        'end_date': end_date,
        'avg_price': round(avg_price, 2),
        'count': len(filtered),
        'prices': filtered
    }


def calculate_period_average(
    material_name: str,
    start_date: str,
    end_date: str,
    all_prices: List[Dict] = None,
    brands: Optional[List[str]] = None
) -> float:
    """计算施工期平均价格（简化方法）"""
    result = get_material_prices(material_name, start_date, end_date, all_prices, brands=brands)
    return result.get('avg_price', 0)


# ============================================================
# 默认品牌优先级
# ============================================================

STEEL_BRANDS_PRIORITY = [
    '莱钢', '莱钢永锋', '永锋', '石横', '日钢', '镔鑫',
    '西王', '三德', '南钢', '沙钢', '马钢', '武钢'
]


# ============================================================
# API 端点
# ============================================================

@router.get("/query", summary="查询施工期价格")
async def query_price(
    material: str = Query(..., description="材料名称"),
    start_date: str = Query(..., description="开始日期 YYYY-MM-DD"),
    end_date: str = Query(..., description="结束日期 YYYY-MM-DD"),
    spec: Optional[str] = Query(None, description="规格过滤"),
    brands: Optional[str] = Query(None, description="品牌偏好，逗号分隔")
):
    """
    查询施工期价格数据

    返回指定时间段内的材料均价及明细
    """
    # 解析品牌偏好
    brand_list = None
    if brands:
        brand_list = [b.strip() for b in brands.split(',')]

    all_prices = load_yantai_prices()

    result = get_material_prices(
        material_name=material,
        start_date=start_date,
        end_date=end_date,
        all_prices=all_prices,
        spec=spec,
        brands=brand_list
    )

    return {
        'success': True,
        'data': {
            'material_name': result['material_name'],
            'start_date': result['start_date'],
            'end_date': result['end_date'],
            'avg_price': result['avg_price'],
            'data_points': result['count'],
            'prices': [
                {
                    'date': p['date'],
                    'price': p['price'],
                    'brand': p.get('brand', ''),
                    'spec': p.get('spec', '')
                }
                for p in result['prices'][:50]  # 最多返回50条
            ]
        }
    }


@router.get("/period-average", summary="获取施工期均价")
async def get_period_average(
    material: str = Query(..., description="材料名称"),
    start_date: str = Query(..., description="开始日期 YYYY-MM-DD"),
    end_date: str = Query(..., description="结束日期 YYYY-MM-DD"),
    brands: Optional[str] = Query(None, description="品牌偏好，逗号分隔")
):
    """获取施工期平均价格（简化接口）"""
    brand_list = [b.strip() for b in brands.split(',')] if brands else STEEL_BRANDS_PRIORITY

    avg_price = calculate_period_average(material, start_date, end_date, brands=brand_list)

    return {
        'success': True,
        'material': material,
        'start_date': start_date,
        'end_date': end_date,
        'avg_price': avg_price,
        'brands_preferred': brand_list[:5] if brand_list else []
    }


@router.get("/base-price", summary="获取基准价")
async def get_base_price(
    material: str = Query(..., description="材料名称"),
    base_date: str = Query(..., description="基准日期 YYYY-MM-DD"),
    spec: Optional[str] = Query(None, description="规格"),
    brands: Optional[str] = Query(None, description="品牌偏好，逗号分隔")
):
    """
    获取材料基准价

    基准价通常取自招标/签约时期的价格
    """
    brand_list = [b.strip() for b in brands.split(',')] if brands else STEEL_BRANDS_PRIORITY

    # 基准价取基准日期的价格，如果没有则取最近的价格
    all_prices = load_yantai_prices()

    # 查找基准日期附近的数据
    base_prices = []
    for p in all_prices:
        if p['date'] == base_date or abs((datetime.strptime(p['date'], '%Y-%m-%d') - datetime.strptime(base_date, '%Y-%m-%d')).days) <= 7:
            if material in p.get('material_name', '') or '钢筋' in material:
                base_prices.append(p)

    if base_prices:
        avg = sum(p['price'] for p in base_prices) / len(base_prices)
    else:
        # 没有找到，使用最近价格
        result = get_material_prices(material, base_date, base_date, all_prices)
        avg = result.get('avg_price', 0)

        # 如果还是没有，使用默认值
        if avg == 0:
            if '钢筋' in material or '螺纹' in material:
                avg = 4500  # 默认钢筋价格
            else:
                avg = 0

    return {
        'success': True,
        'material': material,
        'base_date': base_date,
        'base_price': round(avg, 2)
    }


@router.get("/adjustment-prices", summary="获取调差价格（包含基准价和施工期价）")
async def get_adjustment_prices(
    material: str = Query(..., description="材料名称"),
    base_date: str = Query(..., description="基准日期 YYYY-MM-DD"),
    period_start: str = Query(..., description="施工期开始日期"),
    period_end: str = Query(..., description="施工期结束日期"),
    brands: Optional[str] = Query(None, description="品牌偏好，逗号分隔")
):
    """
    获取调差所需的完整价格数据

    包含：基准价、施工期均价、涨跌幅度
    """
    brand_list = [b.strip() for b in brands.split(',')] if brands else STEEL_BRANDS_PRIORITY

    all_prices = load_yantai_prices()

    # 获取基准价
    base_result = get_material_prices(material, base_date, base_date, all_prices)
    base_price = base_result.get('avg_price', 0)

    # 如果基准日没有数据，使用前后7天的数据
    if base_price == 0:
        base_dt = datetime.strptime(base_date, '%Y-%m-%d')
        for offset in range(1, 8):
            prev_date = (base_dt - timedelta(days=offset)).strftime('%Y-%m-%d')
            result = get_material_prices(material, prev_date, prev_date, all_prices)
            if result.get('avg_price', 0) > 0:
                base_price = result['avg_price']
                break

    # 获取施工期均价
    period_result = get_material_prices(material, period_start, period_end, all_prices, brands=brand_list)
    period_price = period_result.get('avg_price', 0)

    # 计算涨跌
    if base_price > 0:
        change_pct = ((period_price - base_price) / base_price) * 100
    else:
        change_pct = 0

    return {
        'success': True,
        'material': material,
        'base_date': base_date,
        'period': {
            'start': period_start,
            'end': period_end
        },
        'prices': {
            'base': base_price,
            'period_avg': period_price,
            'change_pct': round(change_pct, 2)
        },
        'data_points': {
            'base': base_result.get('count', 0),
            'period': period_result.get('count', 0)
        }
    }


@router.get("/batch", summary="批量获取材料价格")
async def batch_get_prices(
    materials: str = Query(..., description="材料列表，逗号分隔"),
    base_date: str = Query(..., description="基准日期 YYYY-MM-DD"),
    period_start: str = Query(..., description="施工期开始日期"),
    period_end: str = Query(..., description="施工期结束日期")
):
    """批量获取多个材料的调差价格"""
    material_list = [m.strip() for m in materials.split(',')]

    all_prices = load_yantai_prices()
    results = []

    for material in material_list:
        # 基准价
        base_result = get_material_prices(material, base_date, base_date, all_prices)
        base_price = base_result.get('avg_price', 0)

        # 施工期价
        period_result = get_material_prices(material, period_start, period_end, all_prices)
        period_price = period_result.get('avg_price', 0)

        # 默认值
        if base_price == 0 and ('钢筋' in material or '螺纹' in material):
            base_price = 4500

        results.append({
            'material': material,
            'base_price': round(base_price, 2),
            'period_avg': round(period_price, 2),
            'change_pct': round(((period_price - base_price) / base_price * 100) if base_price > 0 else 0, 2),
            'data_points': period_result.get('count', 0)
        })

    return {
        'success': True,
        'materials': results,
        'query': {
            'base_date': base_date,
            'period_start': period_start,
            'period_end': period_end
        }
    }


@router.get("/sources", summary="获取价格来源列表")
async def get_price_sources():
    """获取支持的价格来源"""
    return {
        'sources': [
            {'id': 'mysteel', 'name': '我的钢铁网', 'region': '全国'},
            {'id': 'yantai', 'name': '烟台造价信息', 'region': '山东烟台'},
            {'id': 'qingdao', 'name': '青岛材价信息', 'region': '青岛'},
            {'id': 'bid', 'name': '投标价', 'region': '项目合同'},
        ]
    }


@router.get("/latest", summary="获取最新价格")
async def get_latest_price(
    material: str = Query(..., description="材料名称"),
    spec: Optional[str] = Query(None, description="规格")
):
    """获取材料的最新价格"""
    all_prices = load_yantai_prices()

    # 筛选最近的数据
    latest_prices = []
    for p in all_prices:
        if material in p.get('material_name', '') or '钢筋' in material:
            if spec is None or spec in p.get('spec', ''):
                latest_prices.append(p)

    if latest_prices:
        latest_prices.sort(key=lambda x: x['date'], reverse=True)
        return {
            'success': True,
            'material': material,
            'latest': {
                'date': latest_prices[0]['date'],
                'price': latest_prices[0]['price'],
                'brand': latest_prices[0].get('brand', ''),
                'spec': latest_prices[0].get('spec', '')
            },
            'recent_avg': round(sum(p['price'] for p in latest_prices[:5]) / min(5, len(latest_prices)), 2)
        }

    return {
        'success': True,
        'material': material,
        'latest': None,
        'note': '暂无数据'
    }