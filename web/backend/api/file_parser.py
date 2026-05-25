"""
文件解析 API - 快速解析 Excel/TXT/CSV
用于解析工程量底稿，提取材料清单和施工时间段
"""

import os
import re
import threading
from datetime import datetime
from typing import Dict, Any
from pathlib import Path

from fastapi import APIRouter, UploadFile, File, HTTPException
import openpyxl

router = APIRouter(prefix="/file-parser", tags=["文件解析"])

# 异步任务存储
_parse_tasks: Dict[str, dict] = {}

# ============================================================
# 快速解析函数
# ============================================================

def _parse_excel_fast(file_path: str) -> dict:
    """快速解析Excel - 支持多部位分时段"""
    materials = []
    locations = {}

    try:
        if not os.path.exists(file_path):
            raise Exception(f"文件不存在: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active

        # 尝试识别格式：调差格式 vs 标准格式
        first_cell = str(ws.cell(row=1, column=1).value or '')
        second_cell = str(ws.cell(row=2, column=1).value or '')

        # 检测是否为调差工程量格式（标题含"调差"，第二行有"单位"列）
        if '调差' in first_cell or ('单位' in second_cell and '钢筋' in second_cell):
            print(f"[解析] 检测为调差工程量格式，第一行: {first_cell}")
            result = _parse_quantity_file(ws)
            wb.close()
            return result

        # 否则按标准格式解析
        header_cols = {}
        header_row_idx = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(10, ws.max_row)), start=1):
            cells = [str(cell.value).strip() if cell.value else '' for cell in row]
            if not cells or len(cells) < 2:
                continue

            for col_idx, cell in enumerate(cells):
                if not cell or col_idx >= 20:
                    continue

                cell_upper = cell.upper()
                if header_cols.get('name') is None and ('名称' in cell or '材料' in cell or 'NAME' in cell_upper):
                    header_cols['name'] = col_idx
                if header_cols.get('quantity') is None and ('工程量' in cell or '数量' in cell or 'QTY' in cell_upper):
                    header_cols['quantity'] = col_idx
                if header_cols.get('price') is None and ('单价' in cell or '投标' in cell or 'PRICE' in cell_upper):
                    header_cols['price'] = col_idx
                if header_cols.get('spec') is None and ('规格' in cell or '型号' in cell or 'SPEC' in cell_upper):
                    header_cols['spec'] = col_idx
                if header_cols.get('location') is None and ('楼栋' in cell or '部位' in cell or '楼号' in cell or 'LOCATION' in cell_upper):
                    header_cols['location'] = col_idx
                if header_cols.get('start_date') is None and ('开始' in cell or '开工' in cell or 'START' in cell_upper):
                    header_cols['start_date'] = col_idx
                if header_cols.get('end_date') is None and ('结束' in cell or '竣工' in cell or 'END' in cell_upper):
                    header_cols['end_date'] = col_idx

            if header_cols.get('name') is not None:
                header_row_idx = row_idx
                break

        if header_cols.get('name') is None:
            header_cols['name'] = 0
        if header_cols.get('quantity') is None:
            header_cols['quantity'] = min(1, ws.max_column - 1)

        row_count = 0
        start_row = header_row_idx + 1 if header_row_idx > 0 else 2

        for row in ws.iter_rows(min_row=start_row, max_row=min(ws.max_row, start_row + 149)):
            cells = [str(cell.value).strip() if cell.value else '' for cell in row]

            if len(cells) <= header_cols.get('name', 0):
                continue

            name = cells[header_cols.get('name', 0)] if header_cols.get('name', 0) < len(cells) else ''
            if not name or len(name) < 2:
                continue

            skip_keywords = ['合计', '小计', '总计', '项目', '编号', '备注', '说明', '序号']
            if any(kw in name for kw in skip_keywords):
                continue

            location = ''
            if header_cols.get('location') is not None and header_cols['location'] < len(cells):
                location = cells[header_cols['location']]

            start_date = ''
            end_date = ''
            if header_cols.get('start_date') is not None and header_cols['start_date'] < len(cells):
                start_date = _parse_date(cells[header_cols['start_date']])
            if header_cols.get('end_date') is not None and header_cols['end_date'] < len(cells):
                end_date = _parse_date(cells[header_cols['end_date']])

            quantity = 0
            if header_cols.get('quantity') is not None and header_cols['quantity'] < len(cells):
                q_str = re.sub(r'[^\d.]', '', cells[header_cols['quantity']])
                if q_str:
                    try:
                        quantity = float(q_str)
                    except:
                        pass

            price = 0
            if header_cols.get('price') is not None and header_cols['price'] < len(cells):
                p_str = re.sub(r'[^\d.]', '', cells[header_cols['price']])
                if p_str:
                    try:
                        price = float(p_str)
                    except:
                        pass

            material = {
                '名称': name,
                '规格': '',
                '单位': 't',
                '工程量': quantity,
                '投标单价': price,
                '基准价': 0,
                '部位': location,
                '开始日期': start_date,
                '结束日期': end_date,
                '行号': row_count + start_row
            }
            materials.append(material)

            if location and (start_date or end_date):
                if location not in locations:
                    locations[location] = {}
                if start_date:
                    locations[location]['开始日期'] = start_date
                if end_date:
                    locations[location]['结束日期'] = end_date

            row_count += 1
            if row_count >= 150:
                break

        wb.close()

    except Exception as e:
        raise Exception(f"Excel解析失败: {str(e)}")

    return {
        '材料清单': materials,
        '总数': len(materials),
        '解析行数': len(materials),
        '部位时段': locations
    }


def _normalize_building(building: str) -> str:
    """标准化楼号格式（统一为 X#楼 格式）"""
    if not building:
        return ''
    building = building.strip()
    # 移除"楼"后缀以便统一处理
    if building.endswith('楼'):
        building = building[:-1]
    # 如果不是以#结尾，加上#
    if not building.endswith('#'):
        building = building + '#'
    # 加上楼
    building = building + '楼'
    return building


def _match_building(building: str, locations: dict) -> dict:
    """匹配楼号，返回对应的时间段信息"""
    if not building:
        return {}

    # 尝试直接匹配
    if building in locations:
        return locations[building]

    # 尝试标准化后匹配
    norm = _normalize_building(building)
    if norm in locations:
        return locations[norm]

    # 尝试去掉"楼"后缀后匹配
    if building.endswith('楼'):
        without_lou = building[:-1]
        if without_lou in locations:
            return locations[without_lou]
        norm2 = _normalize_building(without_lou)
        if norm2 in locations:
            return locations[norm2]

    return {}


def _parse_quantity_file(ws) -> dict:
    """
    解析调差工程量文件格式

    格式：
    - 行1：标题（含¥符号）
    - 行2：列标题（单位、材料钢筋等）
    - 行3+：楼号 | 单位 | 普通钢筋量 | 高强钢筋量
    """
    materials = []
    locations = {}

    # 读取楼栋施工时间段
    schedule_file = Path('services/data/楼栋施工时间段.json')
    schedule_data = {}
    if schedule_file.exists():
        import json
        with open(schedule_file, 'r', encoding='utf-8') as f:
            schedule_data = json.load(f)
            for loc in schedule_data.get('locations', []):
                # 标准化楼号
                norm_building = _normalize_building(loc['楼号'])
                locations[norm_building] = {
                    '开始日期': loc['开工日期'],
                    '结束日期': loc['封顶日期']
                }

    print(f"[解析调差表] 已加载 {len(locations)} 个楼栋时间段")
    print(f"[解析调差表] 楼栋列表: {list(locations.keys())}")

    # 解析数据行（从第3行开始）
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
        cells = [str(cell.value).strip() if cell.value else '' for cell in row]

        if len(cells) < 2:
            continue

        # A列：楼号
        building = cells[0].strip()
        if not building or building in ['楼号', '合计']:
            continue

        # 标准化楼号
        norm_building = _normalize_building(building)
        print(f"[解析] 处理楼号: '{building}' -> '{norm_building}'")

        # B列：单位
        unit = cells[1] if len(cells) > 1 else 't'

        # C列：普通钢筋量
        normal_qty = _parse_number(cells[2] if len(cells) > 2 else '0')

        # D列：高强钢筋量
        high_qty = _parse_number(cells[3] if len(cells) > 3 else '0')

        # 获取该楼栋的时间段
        period = locations.get(norm_building, {})
        start_date = period.get('开始日期', '')
        end_date = period.get('结束日期', '')

        if normal_qty > 0:
            material = {
                '名称': '普通钢筋',
                '规格': '',
                '单位': unit,
                '工程量': normal_qty,
                '投标单价': 0,
                '基准价': 0,
                '部位': norm_building,
                '开始日期': start_date,
                '结束日期': end_date,
                '行号': row_idx
            }
            materials.append(material)

        if high_qty > 0:
            material = {
                '名称': '高强钢筋',
                '规格': '',
                '单位': unit,
                '工程量': high_qty,
                '投标单价': 0,
                '基准价': 0,
                '部位': norm_building,
                '开始日期': start_date,
                '结束日期': end_date,
                '行号': row_idx
            }
            materials.append(material)

    print(f"[解析调差表] 解析结果: {len(materials)} 条材料")

    return {
        '材料清单': materials,
        '总数': len(materials),
        '解析行数': len(materials),
        '部位时段': locations
    }


def _parse_number(s: str) -> float:
    """解析数字"""
    if not s:
        return 0
    try:
        # 移除千分位逗号
        s = s.replace(',', '')
        return float(s)
    except:
        return 0


def _parse_csv_fast(file_path: str) -> dict:
    """快速解析CSV（限制200行）"""
    materials = []

    try:
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            lines = f.readlines()

        header_cols = {'name': 0, 'quantity': 1}
        if lines:
            header = lines[0].split(',')
            for col_idx, col in enumerate(header[:10]):
                col = col.strip().strip('"')
                if '名称' in col or '材料' in col:
                    header_cols['name'] = col_idx
                elif '工程量' in col or '数量' in col:
                    header_cols['quantity'] = col_idx

        for row_idx, line in enumerate(lines[1:min(len(lines), 151)], start=2):
            parts = line.strip().split(',')
            if len(parts) <= header_cols['name']:
                continue

            name = parts[header_cols['name']].strip().strip('"')
            if not name or len(name) < 2:
                continue

            quantity = 0
            if len(parts) > header_cols['quantity']:
                try:
                    quantity = float(re.sub(r'[^\d.]', '', parts[header_cols['quantity']]))
                except:
                    pass

            materials.append({
                '名称': name,
                '规格': '',
                '单位': 't',
                '工程量': quantity,
                '投标单价': 0,
                '基准价': 0,
                '行号': row_idx
            })

    except Exception as e:
        raise Exception(f"CSV解析失败: {str(e)}")

    return {
        '材料清单': materials,
        '总数': len(materials),
        '解析行数': len(materials)
    }


def _parse_text_for_period(file_path: str) -> dict:
    """从文本/Excel中解析施工时间段"""
    period_info = {'开始日期': '', '结束日期': '', '阶段列表': []}

    try:
        content = ''

        if file_path.endswith('.txt'):
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                content = f.read()
        elif file_path.endswith(('.xlsx', '.xls', '.csv')):
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(max_row=50):
                for cell in row:
                    if cell.value:
                        content += str(cell.value) + ' '
            wb.close()

        # 提取日期
        date_pattern = r'(\d{4}[-/年]\d{1,2}[-/月]\d{1,2})'
        dates = re.findall(date_pattern, content)

        if len(dates) >= 2:
            period_info['开始日期'] = dates[0].replace('年', '-').replace('月', '-').replace('/', '-')
            period_info['结束日期'] = dates[-1].replace('年', '-').replace('月', '-').replace('/', '-')

        # 提取阶段
        phase_keywords = ['地下室', '地库', '基础', '主体', '结构', '封顶', '砌体', '装修', '竣工']
        period_info['阶段列表'] = [kw for kw in phase_keywords if kw in content]

    except Exception as e:
        period_info['error'] = str(e)

    return period_info


def parse_async(task_id: str, file_path: str):
    """后台线程解析"""
    try:
        result = _parse_excel_fast(file_path)
        _parse_tasks[task_id] = {
            'status': 'completed',
            'result': result,
            'completed_at': datetime.now().isoformat()
        }
    except Exception as e:
        _parse_tasks[task_id] = {
            'status': 'failed',
            'error': str(e),
            'completed_at': datetime.now().isoformat()
        }


# ============================================================
# API 端点
# ============================================================

@router.post("/upload", summary="上传并解析工程量底稿")
async def upload_and_parse(file: UploadFile = File(...)):
    """上传并解析工程量底稿 - 快速模式（最多150条）"""
    allowed_extensions = ['.xlsx', '.xls', '.csv']
    file_ext = Path(file.filename).suffix.lower()

    if file_ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail=f"不支持: {', '.join(allowed_extensions)}")

    upload_dir = Path('services/data/uploads')
    upload_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    # 使用安全的文件名
    safe_filename = f"{timestamp}.xlsx"
    file_path = upload_dir / safe_filename

    try:
        content = await file.read()

        # 保存文件
        with open(file_path, 'wb') as f:
            f.write(content)

        # 同步快速解析
        if file_ext == '.csv':
            result = _parse_csv_fast(str(file_path))
        else:
            result = _parse_excel_fast(str(file_path))

        # 清理文件
        try:
            file_path.unlink()
        except:
            pass

        return {
            'success': result['总数'] > 0,
            'file': {'name': file.filename, 'size': len(content)},
            'result': result
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/upload-async", summary="异步上传解析（大文件）")
async def upload_async(file: UploadFile = File(...)):
    """异步上传 - 返回任务ID"""
    import uuid

    allowed_extensions = ['.xlsx', '.xls', '.csv']
    file_ext = Path(file.filename).suffix.lower()

    if file_ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail="不支持的文件格式")

    upload_dir = Path('services/data/uploads')
    upload_dir.mkdir(parents=True, exist_ok=True)

    task_id = str(uuid.uuid4())[:8]
    file_path = upload_dir / f"{task_id}_{file.filename}"

    try:
        content = await file.read()
        with open(file_path, 'wb') as f:
            f.write(content)

        _parse_tasks[task_id] = {'status': 'processing', 'file': str(file_path)}

        thread = threading.Thread(target=parse_async, args=(task_id, str(file_path)))
        thread.daemon = True
        thread.start()

        return {
            'success': True,
            'task_id': task_id,
            'message': '文件已上传，解析中...',
            'check_url': f'/api/file-parser/status/{task_id}'
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/status/{task_id}", summary="查询异步任务状态")
async def get_task_status(task_id: str):
    """查询异步解析任务状态"""
    task = _parse_tasks.get(task_id)

    if not task:
        return {'status': 'not_found', 'message': '任务不存在'}

    return {
        'status': task.get('status'),
        'completed_at': task.get('completed_at'),
        'result': task.get('result'),
        'error': task.get('error')
    }


@router.post("/parse-period", summary="解析施工时间段")
async def parse_period(file: UploadFile = File(...)):
    """解析施工时间段文件"""
    allowed_extensions = ['.xlsx', '.xls', '.csv', '.txt']
    file_ext = Path(file.filename).suffix.lower()

    if file_ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail=f"不支持: {file_ext}")

    upload_dir = Path('services/data/uploads')
    upload_dir.mkdir(parents=True, exist_ok=True)

    file_path = upload_dir / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"

    try:
        content = await file.read()
        with open(file_path, 'wb') as f:
            f.write(content)

        result = _parse_text_for_period(str(file_path))
        file_path.unlink()

        return {
            'success': True,
            'file': file.filename,
            'period_info': result
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/preview", summary="快速预览文件（前20行）")
async def preview_file(file: UploadFile = File(...)):
    """快速预览文件内容"""
    allowed_extensions = ['.xlsx', '.xls', '.csv']
    file_ext = Path(file.filename).suffix.lower()

    if file_ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail="不支持的文件格式")

    content = await file.read()

    try:
        import tempfile

        with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
        ws = wb.active

        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20), start=1):
            cells = [str(cell.value) if cell.value else '' for cell in row[:10]]
            rows.append({'row': row_idx, 'cells': cells})

        wb.close()
        Path(tmp_path).unlink()

        return {
            'success': True,
            'preview': rows,
            'total_rows': ws.max_row,
            'total_cols': min(ws.max_column, 10)
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/templates", summary="获取解析模板")
async def get_templates():
    """获取预定义的解析模板"""
    return {
        'templates': [
            {'id': 'standard', 'name': '标准工程量清单', 'columns': ['名称', '规格', '单位', '工程量', '单价', '基准价']},
            {'id': 'simple', 'name': '简化清单', 'columns': ['名称', '工程量']},
        ]
    }