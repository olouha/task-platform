"""
指标库 - 分析报告 API
基于历史指标数据生成项目分析报告
按照《指标库编写流程》规范实现
数据存储在本地SQLite indicator_projects 表
"""

import logging
import io
from datetime import datetime
from typing import Optional, List, Dict
from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Depends
from pydantic import BaseModel, Field
from starlette.responses import StreamingResponse

from services.local_indicator_service import LocalIndicatorService
from services.indicator_service import IndicatorService, CORRECTION_FACTORS
from services.indicator_library_service import IndicatorLibraryService, get_indicator_library_service
from api.deps import get_current_account, get_current_user_can_delete

logger = logging.getLogger(__name__)
router = APIRouter(tags=["指标库分析报告"])


def get_indicator_service():
    return LocalIndicatorService()


def get_library_service() -> IndicatorLibraryService:
    """获取指标库业务服务（与「指标库管理」页同源）"""
    return get_indicator_library_service()


# ============================================================
# 格式转换工具：IndicatorLibraryService 项目 -> 算法 legacy format
# ============================================================

def _enrich_raw_project(project: Dict) -> Dict:
    """
    对存储层原始记录做 legacy 字段兜底。

    新录入数据（经「指标库管理」页 / Excel 导入）通常只填 above_rebar_unit /
    above_concrete_unit 平米含量，而未直接填 steel / concrete。此处补齐，保证
    后续 IndicatorService._to_legacy_format 不丢钢筋 / 混凝土关键字段。

    单位换算：above_rebar_unit 为 t/㎡，legacy steel 为 kg/㎡，故 ×1000。
    缺失或非法值保持原样（None），由算法侧兜底处理。
    """
    p = dict(project)
    if p.get("steel") is None and p.get("above_rebar_unit") is not None:
        try:
            p["steel"] = round(float(p["above_rebar_unit"]) * 1000, 2)
        except (TypeError, ValueError):
            logger.warning(f"[_enrich_raw_project] above_rebar_unit 转换失败 | value={p.get('above_rebar_unit')}")
    if p.get("concrete") is None and p.get("above_concrete_unit") is not None:
        try:
            p["concrete"] = float(p["above_concrete_unit"])
        except (TypeError, ValueError):
            logger.warning(f"[_enrich_raw_project] above_concrete_unit 转换失败 | value={p.get('above_concrete_unit')}")
    return p


def _detail_to_legacy_format(detail: Dict) -> Dict:
    """
    将 IndicatorLibraryService.get_detail() 返回的项目（IndicatorLibraryDetail 序列化后的字典）
    转换为匹配算法所需的 legacy 嵌套格式。

    IndicatorLibraryDetail 不直接含 steel / concrete 字段（新库改用 above_rebar_unit /
    above_concrete_unit 平米含量），这里做合理兜底：
      - steel(kg/㎡)       ← above_rebar_unit(t/㎡) × 1000
      - concrete(m³/㎡)    ← above_concrete_unit(m³/㎡)
    其余可能缺失的字段（height / floor_above / unit_cost 等）给 0 或 None，保证算法不崩。
    """
    steel: Optional[float] = detail.get("steel")
    if steel is None and detail.get("above_rebar_unit") is not None:
        try:
            steel = round(float(detail["above_rebar_unit"]) * 1000, 2)
        except (TypeError, ValueError):
            steel = None

    concrete: Optional[float] = detail.get("concrete")
    if concrete is None and detail.get("above_concrete_unit") is not None:
        try:
            concrete = float(detail["above_concrete_unit"])
        except (TypeError, ValueError):
            concrete = None

    return {
        "id": detail.get("id"),
        "name": detail.get("name"),
        "category": detail.get("category"),
        "location": detail.get("location"),
        "structure": detail.get("structure"),
        "floor_above": detail.get("floor_above") or 0,
        "floor_below": detail.get("floor_below") or 0,
        "area_total": detail.get("area_total") or 0,
        "area_above": detail.get("area_above"),
        "area_below": detail.get("area_below"),
        "height": detail.get("height") or 0,
        "indicators": {
            "total_cost": detail.get("total_cost"),
            "unit_cost": detail.get("unit_cost") or 0,
            "unit_structure": detail.get("unit_structure"),
            "unit_installation": detail.get("unit_installation"),
            "unit_decoration": detail.get("unit_decoration"),
            "unit_measure": detail.get("unit_measure"),
        },
        "material_content": {
            "steel": steel,
            "concrete": concrete,
        },
    }


# ============================================================
# API 请求/响应模型
# ============================================================

class GenerateReportRequest(BaseModel):
    project: Dict
    indicators: Dict
    material_content: Optional[Dict] = None


class AnalyzeProjectRequest(BaseModel):
    project_id: str = Field(..., description="指标库中要分析的项目 ID")


# ============================================================
# 分析报告生成
# ============================================================

@router.post("/generate")
async def generate_report(
    request: GenerateReportRequest,
    indicator_service: LocalIndicatorService = Depends(get_indicator_service)
):
    """
    生成分析报告

    根据项目基本信息和指标数据，匹配最相似的历史项目，
    应用修正系数，生成详细的对比分析报告。
    """
    logger.info(f"[generate_report] 生成分析报告 | 项目: {request.project.get('name')}, 业态: {request.project.get('category')}")

    try:
        # 从本地数据库获取指标库
        database_flat = indicator_service.get_indicator_projects(limit=500)
        database = [IndicatorService._to_legacy_format(p) for p in database_flat]

        # 查找匹配的指标
        matched = IndicatorService.find_matched_indicators(request.project, database)

        # 指标对比分析
        target_material = request.material_content
        comparison = IndicatorService.analyze_comparison(
            request.indicators.get("unit_cost", 0),
            target_material.get("steel") if target_material else None,
            target_material.get("concrete") if target_material else None,
            matched,
            database
        )

        # 造价分解
        breakdown = IndicatorService.generate_cost_breakdown(request.indicators)

        # 修正系数
        corrections = IndicatorService.generate_corrections(request.project, matched)

        # 生成建议
        suggestions = IndicatorService.generate_suggestions(request.indicators, comparison, matched)

        # 风险提示
        warnings = IndicatorService.generate_risk_warnings(request.project, comparison)

        # 生成报告ID
        report_id = f"RPT-{datetime.now().strftime('%Y%m%d%H%M%S')}"

        logger.info(f"[generate_report] 报告生成完成 | report_id={report_id}, 匹配数={len(matched)}")

        return {
            "report_id": report_id,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "project_name": request.project.get("name"),
            "matched_indicators": matched,
            "comparison": comparison,
            "cost_breakdown": breakdown,
            "corrections": corrections,
            "suggestions": suggestions,
            "risk_warnings": warnings
        }

    except Exception as e:
        logger.error(f"[generate_report] 报告生成失败 | 错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"报告生成失败: {str(e)}")


@router.post("/analyze")
async def analyze_project(
    request: AnalyzeProjectRequest,
    library_service: IndicatorLibraryService = Depends(get_library_service),
):
    """
    对指标库中已存在的某个项目进行分析。

    数据源与「指标库管理」页同源（IndicatorLibraryService）：选中的项目作为 target，
    其余项目作为 database 做匹配 / 修正 / 对比。
    """
    logger.info(f"[analyze_project] 分析指标库项目 | project_id={request.project_id}")

    try:
        # 1. 取 target 项目（IndicatorLibraryService.get_detail，与指标库管理页一致）
        target_detail = library_service.get_detail(request.project_id)
        if not target_detail:
            logger.warning(f"[analyze_project] 项目不存在 | project_id={request.project_id}")
            raise HTTPException(status_code=404, detail=f"指标库中不存在项目 {request.project_id}")

        target_project = _detail_to_legacy_format(target_detail.model_dump())
        project_for_match = {
            "name": target_project.get("name"),
            "category": target_project.get("category"),
            "location": target_project.get("location"),
            "structure": target_project.get("structure"),
            "height": target_project.get("height"),
            "floor_above": target_project.get("floor_above"),
            "floor_below": target_project.get("floor_below"),
            "area_total": target_project.get("area_total"),
        }
        target_indicators = target_project.get("indicators", {}) or {}
        target_material = target_project.get("material_content", {}) or {}

        # 2. 取 database（全部项目，排除 target 自身）—— 同样来自 IndicatorLibraryService 存储层
        database_flat = library_service._storage_service.get_indicator_projects(limit=500)
        database = [
            IndicatorService._to_legacy_format(_enrich_raw_project(p))
            for p in database_flat
            if p.get("id") != request.project_id
        ]
        logger.info(f"[analyze_project] 数据库候选数 | count={len(database)} | target={request.project_id}")

        # 3. 复用现有算法：匹配 / 对比 / 修正 / 建议 / 风险
        matched = IndicatorService.find_matched_indicators(project_for_match, database)

        comparison = IndicatorService.analyze_comparison(
            target_indicators.get("unit_cost", 0) or 0,
            target_material.get("steel"),
            target_material.get("concrete"),
            matched,
            database,
        )

        breakdown = IndicatorService.generate_cost_breakdown(target_indicators)

        corrections = IndicatorService.generate_corrections(project_for_match, matched)

        suggestions = IndicatorService.generate_suggestions(target_indicators, comparison, matched)

        warnings = IndicatorService.generate_risk_warnings(project_for_match, comparison)

        report_id = f"RPT-{datetime.now().strftime('%Y%m%d%H%M%S')}"

        logger.info(f"[analyze_project] 报告生成完成 | report_id={report_id}, project_id={request.project_id}, 匹配数={len(matched)}")

        return {
            "report_id": report_id,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "project_id": request.project_id,
            "project_name": target_project.get("name"),
            "project_snapshot": target_detail.model_dump(exclude_none=True),  # 该项目完整数据，前端可展示
            "matched_indicators": matched,
            "comparison": comparison,
            "cost_breakdown": breakdown,
            "corrections": corrections,
            "suggestions": suggestions,
            "risk_warnings": warnings,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[analyze_project] 失败 | project_id={request.project_id} | {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"分析失败: {str(e)}")


@router.get("/match")
async def match_indicators(
    category: str = Query(..., description="业态类型"),
    location: str = Query(..., description="项目地区"),
    structure: str = Query(..., description="结构形式"),
    height: float = Query(..., description="檐高(m)"),
    indicator_service: LocalIndicatorService = Depends(get_indicator_service)
):
    """快速匹配指标"""
    logger.info(f"[match_indicators] 匹配指标 | 业态={category}, 地区={location}, 层高={height}")

    target = {
        "category": category,
        "location": location,
        "structure": structure,
        "height": height
    }

    database_flat = indicator_service.get_indicator_projects(limit=500, category=category)
    database = [IndicatorService._to_legacy_format(p) for p in database_flat]
    matched = IndicatorService.find_matched_indicators(target, database)

    logger.info(f"[match_indicators] 匹配完成 | 找到 {len(matched)} 个匹配项目")

    return {
        "category": category,
        "location": location,
        "height": height,
        "matched_count": len(matched),
        "matches": matched
    }


@router.get("/correction-factors")
async def get_correction_factors():
    """获取修正系数表"""
    logger.info("[get_correction_factors] 获取修正系数")

    return {
        "factors": CORRECTION_FACTORS,
        "description": {
            "height": "高度修正：檐高越高，垂直运输成本越高",
            "structure": "结构形式修正：剪力墙结构钢筋含量高于框架结构",
            "region": "地区修正：一线城市造价高于其他城市"
        }
    }


# ============================================================
# 指标库项目管理
# ============================================================

@router.get("/database/summary")
async def get_database_summary(indicator_service: LocalIndicatorService = Depends(get_indicator_service)):
    """获取指标库汇总信息"""
    logger.info("[get_database_summary] 获取指标库汇总")

    database = indicator_service.get_indicator_projects(limit=1000)

    # 按业态分组统计
    by_category = {}
    for item in database:
        cat = item.get("category", "未知")
        if cat not in by_category:
            by_category[cat] = []
        by_category[cat].append(item)

    # 按地区分组
    by_location = {}
    for item in database:
        loc = item.get("location", "未知")
        if loc not in by_location:
            by_location[loc] = []
        by_location[loc].append(item)

    # 数据来源统计
    by_source = {}
    for item in database:
        src = item.get("source", "未知")
        by_source[src] = by_source.get(src, 0) + 1

    def avg(lst):
        return round(sum(lst) / len(lst), 2) if lst else 0

    summary = {
        "total_count": len(database),
        "by_category": {
            cat: {
                "count": len(items),
                "avg_unit_cost": avg([i.get("unit_cost", 0) for i in items if i.get("unit_cost")]),
                "avg_steel": avg([i.get("steel", 0) for i in items if i.get("steel")])
            }
            for cat, items in by_category.items()
        },
        "by_location": {loc: len(items) for loc, items in by_location.items()},
        "by_source": by_source,
        "price_range": {
            "min": min((i.get("unit_cost", 0) for i in database if i.get("unit_cost")), default=0),
            "max": max((i.get("unit_cost", 0) for i in database if i.get("unit_cost")), default=0)
        }
    }

    logger.info(f"[get_database_summary] 汇总完成 | 总项目数={len(database)}")

    return summary


@router.get("/database/list")
async def list_database_projects(
    category: Optional[str] = Query(None, description="按业态筛选"),
    location: Optional[str] = Query(None, description="按地区筛选"),
    limit: int = Query(20, ge=1, le=100, description="返回数量限制"),
    indicator_service: LocalIndicatorService = Depends(get_indicator_service)
):
    """获取指标库项目列表"""
    logger.info(f"[list_database_projects] 查询项目 | category={category}, location={location}")

    projects = indicator_service.get_indicator_projects(category=category, location=location, limit=limit)

    return {
        "total": len(projects),
        "projects": projects
    }


@router.get("/database/init-sample")
async def init_sample_data(indicator_service: LocalIndicatorService = Depends(get_indicator_service)):
    """初始化示例数据（仅当数据库为空时）"""
    logger.info("[init_sample_data] 初始化示例数据")

    # 检查是否已有数据
    existing = indicator_service.get_indicator_projects(limit=1)
    if existing:
        logger.info("[init_sample_data] 数据库已有数据，跳过初始化")
        return {"success": True, "message": "数据库已有数据", "count": len(existing)}

    # 示例数据
    sample_projects = [
        {
            "name": "烟台市住宅项目A区",
            "category": "住宅",
            "location": "山东",
            "structure": "剪力墙结构",
            "floor_above": 18,
            "floor_below": 1,
            "area_total": 25000,
            "height": 54,
            "unit_cost": 2350,
            "unit_structure": 1200,
            "unit_installation": 450,
            "unit_decoration": 400,
            "unit_measure": 300,
            "steel": 48,
            "concrete": 0.42,
            "source": "sample"
        },
        {
            "name": "烟台商业综合体项目",
            "category": "商业",
            "location": "山东",
            "structure": "框架结构",
            "floor_above": 12,
            "floor_below": 2,
            "area_total": 45000,
            "height": 48,
            "unit_cost": 3200,
            "unit_structure": 1800,
            "unit_installation": 600,
            "unit_decoration": 500,
            "unit_measure": 300,
            "steel": 55,
            "concrete": 0.48,
            "source": "sample"
        },
        {
            "name": "烟台市办公写字楼",
            "category": "办公",
            "location": "山东",
            "structure": "框架核心筒结构",
            "floor_above": 25,
            "floor_below": 3,
            "area_total": 60000,
            "height": 100,
            "unit_cost": 3800,
            "unit_structure": 2100,
            "unit_installation": 700,
            "unit_decoration": 600,
            "unit_measure": 400,
            "steel": 65,
            "concrete": 0.52,
            "source": "sample"
        },
        {
            "name": "烟台住宅项目B区",
            "category": "住宅",
            "location": "山东",
            "structure": "框架结构",
            "floor_above": 11,
            "floor_below": 0,
            "area_total": 18000,
            "height": 33,
            "unit_cost": 2100,
            "unit_structure": 1050,
            "unit_installation": 420,
            "unit_decoration": 380,
            "unit_measure": 250,
            "steel": 42,
            "concrete": 0.38,
            "source": "sample"
        },
        {
            "name": "烟台工业区厂房项目",
            "category": "工业",
            "location": "山东",
            "structure": "钢结构",
            "floor_above": 3,
            "floor_below": 0,
            "area_total": 12000,
            "height": 12,
            "unit_cost": 1800,
            "unit_structure": 900,
            "unit_installation": 350,
            "unit_decoration": 200,
            "unit_measure": 350,
            "steel": 35,
            "concrete": 0.25,
            "source": "sample"
        }
    ]

    result = indicator_service.import_indicator_projects(sample_projects)

    logger.info(f"[init_sample_data] 初始化完成 | 成功={result['imported']}, 总数={result['total']}")

    return {
        "success": True,
        "imported": result["imported"],
        "total": result["total"],
        "message": "示例数据初始化完成"
    }


@router.get("/database/{project_id}")
async def get_database_project(
    project_id: str,
    indicator_service: LocalIndicatorService = Depends(get_indicator_service)
):
    """获取指标库单个项目"""
    logger.info(f"[get_database_project] 获取项目 | id={project_id}")

    project = indicator_service.get_indicator_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")

    return project


@router.post("/database/")
async def create_database_project(
    project: Dict,
    indicator_service: LocalIndicatorService = Depends(get_indicator_service)
):
    """创建指标库项目"""
    logger.info(f"[create_database_project] 创建项目 | name={project.get('name')}")

    result = indicator_service.create_indicator_project(project)
    if result:
        logger.info(f"[create_database_project] 创建成功 | id={result.get('id')}")
        return result
    else:
        raise HTTPException(status_code=500, detail="创建失败")


@router.put("/database/{project_id}")
async def update_database_project(
    project_id: str,
    project: Dict,
    admin_account: str = Depends(get_current_user_can_delete),
    indicator_service: LocalIndicatorService = Depends(get_indicator_service)
):
    """更新指标库项目"""
    logger.info(f"[update_database_project] 更新项目 | id={project_id}")

    success = indicator_service.update_indicator_project(project_id, project)
    if success:
        updated = indicator_service.get_indicator_project(project_id)
        logger.info(f"[update_database_project] 更新成功 | id={project_id}")
        return updated
    else:
        raise HTTPException(status_code=500, detail="更新失败")


@router.delete("/database/{project_id}")
async def delete_database_project(
    project_id: str,
    admin_account: str = Depends(get_current_user_can_delete),
    indicator_service: LocalIndicatorService = Depends(get_indicator_service)
):
    """删除指标库项目"""
    logger.info(f"[delete_database_project] 删除项目 | id={project_id}")

    success = indicator_service.delete_indicator_project(project_id)
    if success:
        logger.info(f"[delete_database_project] 删除成功 | id={project_id}")
    else:
        logger.warning(f"[delete_database_project] 项目不存在 | id={project_id}")

    return {"success": success}


@router.post("/quality-check")
async def check_quality(
    project: Dict,
    indicators: Dict,
    indicator_service: LocalIndicatorService = Depends(get_indicator_service)
):
    """质量审核"""
    logger.info(f"[check_quality] 质量审核 | 项目: {project.get('name')}")

    result = IndicatorService.quality_check(project, indicators)

    logger.info(f"[check_quality] 审核完成 | passed={result.get('passed')}, warnings={len(result.get('warnings', []))}")

    return result


@router.get("/reference-ranges")
async def get_reference_ranges():
    """获取参考指标范围"""
    logger.info("[get_reference_ranges] 获取参考范围")

    return {
        "ranges": {
            "住宅": {
                "框架结构": {
                    "unit_cost": {"min": 1800, "max": 2500, "unit": "元/㎡"},
                    "steel": {"min": 35, "max": 50, "unit": "kg/㎡"},
                    "concrete": {"min": 0.35, "max": 0.45, "unit": "m³/㎡"}
                },
                "剪力墙结构": {
                    "unit_cost": {"min": 2200, "max": 3000, "unit": "元/㎡"},
                    "steel": {"min": 45, "max": 65, "unit": "kg/㎡"},
                    "concrete": {"min": 0.40, "max": 0.50, "unit": "m³/㎡"}
                }
            },
            "商业": {
                "框架结构": {
                    "unit_cost": {"min": 2500, "max": 4000, "unit": "元/㎡"},
                    "steel": {"min": 50, "max": 70, "unit": "kg/㎡"},
                    "concrete": {"min": 0.40, "max": 0.55, "unit": "m³/㎡"}
                }
            },
            "办公": {
                "框架结构": {
                    "unit_cost": {"min": 2800, "max": 4500, "unit": "元/㎡"},
                    "steel": {"min": 55, "max": 75, "unit": "kg/㎡"},
                    "concrete": {"min": 0.45, "max": 0.60, "unit": "m³/㎡"}
                }
            }
        },
        "description": "数据来源：指标库编写流程规范"
    }


# ============================================================
# 导入导出
# ============================================================

@router.post("/import")
async def import_indicator(
    file: UploadFile = File(...),
    indicator_service: LocalIndicatorService = Depends(get_indicator_service),
    account: str = Depends(get_current_account)
):
    """
    导入指标数据

    从Excel文件导入指标库数据。
    """
    logger.info(f"[import_indicator] 导入指标 | 文件: {file.filename}")

    # 验证文件类型
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持Excel文件格式(.xlsx, .xls)")

    try:
        import openpyxl

        # 读取文件内容
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active

        # 解析Excel
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        projects = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            # 过滤 None 值并构建字典
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = value
            if row_dict.get("name"):
                row_dict["uploaded_by"] = account
                projects.append(row_dict)

        if not projects:
            return {"success": False, "message": "Excel文件中没有找到有效数据"}

        # 批量导入
        result = indicator_service.import_indicator_projects(projects)

        logger.info(f"[import_indicator] 导入完成 | 成功={result['imported']}, 总数={result['total']}")

        return {
            "success": True,
            "imported": result["imported"],
            "total": result["total"],
            "errors": result["errors"]
        }

    except ImportError:
        logger.warning("[import_indicator] openpyxl未安装，使用简化解析")
        return {"success": False, "message": "openpyxl未安装，无法解析Excel"}
    except Exception as e:
        logger.error(f"[import_indicator] 导入失败 | 错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


@router.get("/export")
async def export_database(
    format: str = Query("json", description="导出格式: json/excel"),
    category: Optional[str] = Query(None, description="按业态筛选"),
    indicator_service: LocalIndicatorService = Depends(get_indicator_service)
):
    """
    导出指标库数据

    导出指标库为JSON或Excel格式。
    """
    logger.info(f"[export_database] 导出指标库 | format={format}, category={category}")

    projects = indicator_service.get_indicator_projects(category=category, limit=1000)

    if format == "json":
        return {
            "version": "1.0",
            "update_date": datetime.now().strftime("%Y-%m-%d"),
            "total_count": len(projects),
            "projects": projects
        }
    else:
        try:
            import openpyxl
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "指标库"

            if projects:
                # 写入表头
                headers = list(projects[0].keys())
                ws.append(headers)
                # 写入数据
                for project in projects:
                    ws.append([project.get(h) for h in headers])

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            logger.info(f"[export_database] Excel导出完成 | 记录数={len(projects)}")

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=indicator_database_{datetime.now().strftime('%Y%m%d')}.xlsx"}
            )
        except ImportError:
            logger.error("[export_database] openpyxl未安装")
            raise HTTPException(status_code=501, detail="openpyxl未安装，无法导出Excel")
        except Exception as e:
            logger.error(f"[export_database] 导出失败 | {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")