"""
楼栋调差计算 API
结合楼栋施工时间表、价格数据和调差规则进行计算
"""

from datetime import datetime
from typing import Dict, List, Optional
from pathlib import Path

from fastapi import APIRouter, HTTPException, Query
from pydantic import BaseModel

from services.adjustment_engine import AdjustmentEngine, CalculationInput, PriceData, QuantityData
from models.adjustment_rules import AdjustmentRuleConfig, PRESET_RULES
from api.adjustment_prices import get_material_prices, load_yantai_prices

router = APIRouter(prefix="", tags=["楼栋调差计算"])


# ============================================================
# 数据模型
# ============================================================

class BuildingSchedule(BaseModel):
    """楼栋施工时间"""
    楼号: str
    开工日期: str
    封顶日期: str
    工期天数: int


class BuildingMaterial(BaseModel):
    """楼栋材料配置"""
    楼号: str
    材料名称: str
    工程量: float
    单位: str = "t"
    基准价: Optional[float] = None


class BuildingCalculationRequest(BaseModel):
    """楼栋调差计算请求"""
    buildings: List[BuildingSchedule]  # 楼栋施工时间表
    materials: List[BuildingMaterial]  # 每个楼栋的材料配置
    rule_name: str = "朱家庄"  # 调差规则名称
    base_date: Optional[str] = None  # 基准日期，如果不指定则取最早开工日期


class BuildingAdjustmentResult(BaseModel):
    """单个楼栋的调差结果"""
    楼号: str
    施工期: Dict[str, str]  # {"开始": "2024-06-29", "结束": "2024-12-06"}
    工期天数: int
    材料: List[Dict]  # 材料明细
    调差总金额: float
    含税调差总金额: float


class CalculationSummary(BaseModel):
    """计算汇总"""
    总楼栋数: int
    总调差金额: float
    总含税调差金额: float
    楼栋明细: List[BuildingAdjustmentResult]


# ============================================================
# 核心计算函数
# ============================================================

def calculate_building_adjustment(
    building: BuildingSchedule,
    building_materials: List[BuildingMaterial],
    rule_config: Dict,
    base_date: str,
    all_prices: List[Dict] = None
) -> Dict:
    """
    计算单个楼栋的调差金额

    参数:
    - building: 楼栋施工时间
    - building_materials: 该楼栋的材料配置
    - rule_config: 调差规则配置
    - base_date: 基准日期
    - all_prices: 所有价格数据（用于避免重复加载）
    """
    if all_prices is None:
        all_prices = load_yantai_prices()

    # 构建配置对象
    config = AdjustmentRuleConfig(**rule_config)
    engine = AdjustmentEngine(config)

    # 获取该楼栋的材料
    materials_for_building = [m for m in building_materials if m.楼号 == building.楼号]

    if not materials_for_building:
        return {
            "楼号": building.楼号,
            "施工期": {"开始": building.开工日期, "结束": building.封顶日期},
            "工期天数": building.工期天数,
            "材料": [],
            "调差总金额": 0,
            "含税调差总金额": 0,
            "备注": "无材料配置"
        }

    # 构建计算输入
    base_prices = {}
    period_prices = {}
    quantities = []

    for mat in materials_for_building:
        material_name = mat.材料名称

        # 获取基准价
        if mat.基准价 and mat.基准价 > 0:
            base_price = mat.基准价
        else:
            # 从价格数据中获取基准价
            base_result = get_material_prices(material_name, base_date, base_date, all_prices)
            base_price = base_result.get('avg_price', 0)
            # 如果没有数据，使用默认值
            if base_price == 0:
                if '钢筋' in material_name:
                    base_price = 4500
                elif '混凝土' in material_name:
                    base_price = 400

        base_prices[material_name] = base_price

        # 获取施工期价格
        period_result = get_material_prices(
            material_name,
            building.开工日期,
            building.封顶日期,
            all_prices
        )
        period_avg_price = period_result.get('avg_price', 0)

        # 构建施工期价格列表（用于计算引擎）
        period_prices[material_name] = [
            PriceData(date=building.开工日期, price=period_avg_price, source="施工期均价")
        ]

        # 添加工程量数据
        quantities.append(QuantityData(
            material_name=material_name,
            quantity=mat.工程量,
            unit=mat.单位,
            phase=building.楼号  # 用楼号作为阶段标识
        ))

    # 执行计算
    input_data = CalculationInput(
        base_prices=base_prices,
        period_prices=period_prices,
        quantities=quantities
    )

    try:
        result = engine.calculate(input_data)

        # 转换为楼栋格式
        building_result = {
            "楼号": building.楼号,
            "施工期": {"开始": building.开工日期, "结束": building.封顶日期},
            "工期天数": building.工期天数,
            "材料": [
                {
                    "材料名称": detail.材料名称,
                    "工程量": detail.工程量,
                    "单位": detail.工程量单位,
                    "基准价": detail.基准价,
                    "施工均价": detail.施工均价,
                    "风险幅度": detail.风险幅度,
                    "是否超幅": detail.是否超幅,
                    "调整单价": detail.调整单价,
                    "调整金额": detail.调整金额,
                    "含税调整金额": detail.含税调整金额,
                    "计算公式": detail.计算公式
                }
                for detail in result.明细
            ],
            "调差总金额": result.调差总金额,
            "含税调差总金额": result.调差总金额  # 计算引擎已含税
        }

        return building_result

    except Exception as e:
        return {
            "楼号": building.楼号,
            "施工期": {"开始": building.开工日期, "结束": building.封顶日期},
            "工期天数": building.工期天数,
            "材料": [],
            "调差总金额": 0,
            "含税调差总金额": 0,
            "错误": str(e)
        }


def calculate_all_buildings(
    buildings: List[BuildingSchedule],
    materials: List[BuildingMaterial],
    rule_name: str = "朱家庄",
    base_date: str = None
) -> Dict:
    """
    计算所有楼栋的调差金额

    参数:
    - buildings: 楼栋施工时间表
    - materials: 所有楼栋的材料配置
    - rule_name: 调差规则名称
    - base_date: 基准日期
    """
    # 获取规则配置
    preset_config = PRESET_RULES.get(rule_name)
    if not preset_config:
        raise HTTPException(status_code=400, detail=f"规则 '{rule_name}' 不存在")

    # 转换为计算引擎配置格式
    from api.adjustments import _convert_preset_to_config
    rule_config = _convert_preset_to_config(preset_config, rule_name)

    # 确定基准日期
    if not base_date:
        # 使用最早的开工日期
        valid_dates = [b.开工日期 for b in buildings if b.开工日期]
        base_date = min(valid_dates) if valid_dates else "2024-06-01"

    # 加载价格数据（一次加载，多次使用）
    all_prices = load_yantai_prices()

    # 计算每个楼栋
    results = []
    total_adjustment = 0
    total_with_tax = 0

    for building in buildings:
        result = calculate_building_adjustment(
            building, materials, rule_config, base_date, all_prices
        )
        results.append(result)
        total_adjustment += result.get("调差总金额", 0)
        total_with_tax += result.get("含税调差总金额", 0)

    return {
        "规则名称": rule_name,
        "基准日期": base_date,
        "总楼栋数": len(buildings),
        "总调差金额": round(total_adjustment, 2),
        "总含税调差金额": round(total_with_tax, 2),
        "楼栋明细": results
    }


# ============================================================
# API 端点
# ============================================================

@router.post("/calculate", summary="楼栋调差计算")
async def calculate_building_adjustments(request: BuildingCalculationRequest):
    """
    根据楼栋施工时间表和材料配置进行调差计算

    请求参数:
    - buildings: 楼栋施工时间表列表
    - materials: 每个楼栋的材料配置列表
    - rule_name: 调差规则名称（默认"朱家庄"）
    - base_date: 基准日期（可选，默认取最早开工日期）
    """
    try:
        result = calculate_all_buildings(
            buildings=request.buildings,
            materials=request.materials,
            rule_name=request.rule_name,
            base_date=request.base_date
        )

        return {
            "success": True,
            "data": result
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/calculate-from-file", summary="从Excel文件计算")
async def calculate_from_excel(
    schedule_file: str = Query(..., description="楼栋施工时间Excel文件路径"),
    material_file: str = Query(..., description="材料配置Excel文件路径"),
    rule_name: str = Query("朱家庄", description="调差规则名称"),
    base_date: str = Query(None, description="基准日期")
):
    """
    从Excel文件直接计算调差

    - schedule_file: 楼栋施工时间Excel文件（楼号、开工日期、封顶日期）
    - material_file: 材料配置Excel文件（楼号、材料名称、工程量、基准价）
    - rule_name: 调差规则名称
    - base_date: 基准日期
    """
    try:
        import openpyxl

        # 解析施工时间表
        buildings = []
        wb_schedule = openpyxl.load_workbook(schedule_file, data_only=True, read_only=True)
        ws_schedule = wb_schedule.active

        for row in ws_schedule.iter_rows(min_row=2, max_row=ws_schedule.max_row):
            building = str(row[0].value or '').strip()
            if not building or building == '楼号':
                continue

            start_date = row[1].value
            end_date = row[2].value

            if isinstance(start_date, datetime):
                start_date = start_date.strftime('%Y-%m-%d')
            else:
                start_date = str(start_date)[:10] if start_date else ''

            if isinstance(end_date, datetime):
                end_date = end_date.strftime('%Y-%m-%d')
            else:
                end_date = str(end_date)[:10] if end_date else ''

            # 计算工期天数
            days = 0
            if start_date and end_date:
                try:
                    from datetime import datetime as dt
                    days = (dt.strptime(end_date, '%Y-%m-%d') - dt.strptime(start_date, '%Y-%m-%d')).days
                except:
                    pass

            buildings.append(BuildingSchedule(
                楼号=building,
                开工日期=start_date,
                封顶日期=end_date,
                工期天数=days
            ))

        wb_schedule.close()

        # 解析材料配置
        materials = []
        wb_material = openpyxl.load_workbook(material_file, data_only=True, read_only=True)
        ws_material = wb_material.active

        # 假设格式：楼号、材料名称、工程量、单位、基准价
        for row in ws_material.iter_rows(min_row=2, max_row=ws_material.max_row):
            building = str(row[0].value or '').strip()
            material_name = str(row[1].value or '').strip()
            quantity = float(row[2].value or 0)
            unit = str(row[3].value or 't').strip()
            base_price = float(row[4].value or 0) if len(row) > 4 else None

            if building and material_name and quantity > 0:
                materials.append(BuildingMaterial(
                    楼号=building,
                    材料名称=material_name,
                    工程量=quantity,
                    单位=unit,
                    基准价=base_price
                ))

        wb_material.close()

        # 计算
        result = calculate_all_buildings(
            buildings=buildings,
            materials=materials,
            rule_name=rule_name,
            base_date=base_date
        )

        return {
            "success": True,
            "data": result
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/parse-schedule", summary="解析楼栋施工时间表")
async def parse_schedule_file(file_path: str = Query(..., description="Excel文件路径")):
    """
    解析楼栋施工时间Excel文件，返回结构化数据

    支持格式：
    - A列：楼号
    - B列：开工日期
    - C列：封顶日期
    """
    try:
        from api.building_schedule import parse_building_schedule

        result = parse_building_schedule(file_path)
        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/rules", summary="获取可用调差规则")
async def get_available_rules():
    """获取所有可用的调差规则列表"""
    rules = []
    for name, config in PRESET_RULES.items():
        basic_info = config.get('基础信息', {})
        price_rule = config.get('价格规则', {})

        rules.append({
            "name": name,
            "description": config.get('项目名称', ''),
            "formula": config.get('计算公式', {}).get('调差公式模板', ''),
            "materials": [m.get('名称', '') for m in basic_info.get('调差项目', [])],
            "price_source": price_rule.get('基准价来源', ''),
            "risk_percent": price_rule.get('风险幅度', {})
        })

    return {
        "success": True,
        "rules": rules
    }


@router.post("/quick-calculate", summary="快速计算单楼栋调差")
async def quick_calculate(
    building_name: str = Query(..., description="楼号，如 1#"),
    start_date: str = Query(..., description="开工日期 YYYY-MM-DD"),
    end_date: str = Query(..., description="封顶日期 YYYY-MM-DD"),
    material: str = Query("钢筋", description="材料名称"),
    quantity: float = Query(..., description="工程量"),
    base_price: float = Query(None, description="基准价，不指定则自动获取"),
    rule_name: str = Query("朱家庄", description="调差规则名称")
):
    """
    快速计算单个楼栋、单个材料的调差金额

    适用于快速估算，无需完整配置
    """
    try:
        # 计算工期天数
        from datetime import datetime as dt
        days = (dt.strptime(end_date, '%Y-%m-%d') - dt.strptime(start_date, '%Y-%m-%d')).days

        # 构建楼栋和材料
        building = BuildingSchedule(
            楼号=building_name,
            开工日期=start_date,
            封顶日期=end_date,
            工期天数=days
        )

        building_material = BuildingMaterial(
            楼号=building_name,
            材料名称=material,
            工程量=quantity,
            基准价=base_price
        )

        # 获取规则
        preset_config = PRESET_RULES.get(rule_name)
        if not preset_config:
            raise HTTPException(status_code=400, detail=f"规则 '{rule_name}' 不存在")

        # 转换配置
        from api.adjustments import _convert_preset_to_config
        rule_config = _convert_preset_to_config(preset_config, rule_name)

        # 确定基准日期
        base_date = start_date  # 使用开工日期作为基准日期

        # 计算
        all_prices = load_yantai_prices()
        result = calculate_building_adjustment(
            building, [building_material], rule_config, base_date, all_prices
        )

        return {
            "success": True,
            "data": result
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))