"""
数据管理 API
提供数据导出、备份、导入、清洗等功能
"""
from fastapi import APIRouter, HTTPException, Query, Depends, Header
from api.deps import get_current_account
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime
import sqlite3
import json
import shutil
import os
from pathlib import Path
import logging

router = APIRouter()
logger = logging.getLogger(__name__)

# 数据库路径
DB_FILE = 'data/yantai_rebar.db'
BACKUP_DIR = 'services/data/backups'
EXPORT_DIR = 'services/data/exports'

# 确保目录存在
os.makedirs(BACKUP_DIR, exist_ok=True)
os.makedirs(EXPORT_DIR, exist_ok=True)


def get_db_connection():
    """获取数据库连接"""
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


class DataStats(BaseModel):
    """数据统计"""
    total_records: int
    total_dates: int
    date_range: dict
    materials: dict
    duplicates: int
    last_updated: Optional[str]


class ExportResult(BaseModel):
    """导出结果"""
    success: bool
    file_path: Optional[str]
    file_name: Optional[str]
    records_count: int
    message: str


class ImportResult(BaseModel):
    """导入结果"""
    success: bool
    imported: int
    skipped: int
    errors: List[str]


class CleanResult(BaseModel):
    """清洗结果"""
    success: bool
    removed_duplicates: int
    fixed_records: int
    remaining_records: int


@router.get("/stats", response_model=DataStats)
async def get_data_stats():
    """
    获取数据统计信息
    """
    try:
        conn = get_db_connection()
        c = conn.cursor()

        # 总记录数
        c.execute('SELECT COUNT(*) FROM rebar_prices')
        total = c.fetchone()[0]

        # 日期数
        c.execute('SELECT COUNT(DISTINCT date) FROM rebar_prices')
        dates = c.fetchone()[0]

        # 日期范围
        c.execute('SELECT MIN(date), MAX(date) FROM rebar_prices')
        range_row = c.fetchone()

        # 重复记录数
        c.execute('''
            SELECT COUNT(*) FROM rebar_prices
            WHERE id NOT IN (
                SELECT MIN(id) FROM rebar_prices
                GROUP BY date, material_name, spec, brand, price
            )
        ''')
        duplicates = c.fetchone()[0]

        # 材料统计
        c.execute('''
            SELECT material_name, COUNT(*) as cnt
            FROM rebar_prices
            GROUP BY material_name
            ORDER BY cnt DESC
        ''')
        materials = {row[0]: row[1] for row in c.fetchall()}

        # 最后更新时间
        c.execute('SELECT MAX(fetch_time) FROM rebar_prices')
        last_updated = c.fetchone()[0]

        conn.close()

        return DataStats(
            total_records=total,
            total_dates=dates,
            date_range={'start': range_row[0], 'end': range_row[1]},
            materials=materials,
            duplicates=duplicates,
            last_updated=last_updated
        )
    except Exception as e:
        logger.error(f"获取数据统计失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export", response_model=ExportResult)
async def export_data(
    format: str = Query("xlsx", description="导出格式: xlsx 或 csv"),
    start_date: Optional[str] = Query(None, description="开始日期"),
    end_date: Optional[str] = Query(None, description="结束日期"),
    material: Optional[str] = Query(None, description="品名筛选")
):
    """
    导出数据到 Excel 或 CSV 文件
    """
    try:
        conn = get_db_connection()
        c = conn.cursor()

        # 构建查询
        sql = '''
            SELECT date, material_name, spec, material_type, brand, price, region, fetch_time, fetch_time AS created_at
            FROM rebar_prices
            WHERE 1=1
        '''
        params = []

        if start_date:
            sql += ' AND date >= ?'
            params.append(start_date)
        if end_date:
            sql += ' AND date <= ?'
            params.append(end_date)
        if material:
            sql += ' AND material_name LIKE ?'
            params.append(f'%{material}%')

        sql += ' ORDER BY date DESC, material_name, spec'

        c.execute(sql, params)
        rows = c.fetchall()
        conn.close()

        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if format == 'csv':
            file_name = f'yantai_rebar_export_{timestamp}.csv'
            file_path = os.path.join(EXPORT_DIR, file_name)

            # 导出CSV
            with open(file_path, 'w', encoding='utf-8-sig') as f:
                # 写入表头
                f.write('日期,品名,规格,材质,品牌,价格,地区,抓取时间,创建时间\n')
                for row in rows:
                    f.write(f'{row["date"]},{row["material_name"]},{row["spec"]},{row["material_type"]},{row["brand"]},{row["price"]},{row["region"]},{row["fetch_time"]},{row["created_at"]}\n')
        else:
            # 导出Excel
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment

                file_name = f'yantai_rebar_export_{timestamp}.xlsx'
                file_path = os.path.join(EXPORT_DIR, file_name)

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = '钢筋价格数据'

                # 表头
                headers = ['日期', '品名', '规格', '材质', '品牌', '价格(元/吨)', '地区', '抓取时间', '创建时间']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

                # 数据
                for row_idx, row in enumerate(rows, 2):
                    ws.cell(row=row_idx, column=1, value=row['date'])
                    ws.cell(row=row_idx, column=2, value=row['material_name'])
                    ws.cell(row=row_idx, column=3, value=row['spec'])
                    ws.cell(row=row_idx, column=4, value=row['material_type'])
                    ws.cell(row=row_idx, column=5, value=row['brand'])
                    ws.cell(row=row_idx, column=6, value=row['price'])
                    ws.cell(row=row_idx, column=7, value=row['region'])
                    ws.cell(row=row_idx, column=8, value=row['fetch_time'])
                    ws.cell(row=row_idx, column=9, value=row['created_at'])

                wb.save(file_path)
            except ImportError:
                # 如果没有openpyxl，回退到CSV
                file_name = f'yantai_rebar_export_{timestamp}.csv'
                file_path = os.path.join(EXPORT_DIR, file_name)
                with open(file_path, 'w', encoding='utf-8-sig') as f:
                    f.write('日期,品名,规格,材质,品牌,价格,地区,抓取时间,创建时间\n')
                    for row in rows:
                        f.write(f'{row["date"]},{row["material_name"]},{row["spec"]},{row["material_type"]},{row["brand"]},{row["price"]},{row["region"]},{row["fetch_time"]},{row["created_at"]}\n')

        logger.info(f"[export_data] 导出完成 | file={file_name} | records={len(rows)}")

        return ExportResult(
            success=True,
            file_path=file_path,
            file_name=file_name,
            records_count=len(rows),
            message=f"成功导出 {len(rows)} 条记录"
        )
    except Exception as e:
        logger.error(f"导出数据失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/download/{file_name}")
async def download_file(file_name: str):
    """
    下载导出文件
    """
    file_path = os.path.join(EXPORT_DIR, file_name)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="文件不存在")

    return FileResponse(
        path=file_path,
        filename=file_name,
        media_type='application/octet-stream'
    )


@router.post("/backup", response_model=dict)
async def backup_database():
    """
    备份数据库
    """
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f'yantai_rebar_backup_{timestamp}.db'
        backup_path = os.path.join(BACKUP_DIR, backup_name)

        # 复制数据库
        shutil.copy2(DB_FILE, backup_path)

        # 同时备份为SQL
        sql_name = f'yantai_rebar_backup_{timestamp}.sql'
        sql_path = os.path.join(BACKUP_DIR, sql_name)

        conn = get_db_connection()
        c = conn.cursor()
        with open(sql_path, 'w', encoding='utf-8') as f:
            # 写入表结构
            c.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='rebar_prices'")
            create_sql = c.fetchone()[0]
            f.write(f'{create_sql};\n\n')

            # 写入数据
            c.execute('SELECT * FROM rebar_prices')
            for row in c:
                f.write(f"INSERT INTO rebar_prices VALUES ({', '.join(repr(v) for v in row)});\n")

        conn.close()

        logger.info(f"[backup_database] 备份完成 | {backup_name}")

        return {
            'success': True,
            'backup_files': [backup_name, sql_name],
            'message': f'备份完成，共2个文件'
        }
    except Exception as e:
        logger.error(f"备份数据库失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/backups")
async def list_backups():
    """
    列出所有备份文件
    """
    try:
        backups = []
        for f in os.listdir(BACKUP_DIR):
            if f.endswith('.db') or f.endswith('.sql'):
                full_path = os.path.join(BACKUP_DIR, f)
                stat = os.stat(full_path)
                backups.append({
                    'name': f,
                    'size': stat.st_size,
                    'created': datetime.fromtimestamp(stat.st_ctime).isoformat()
                })

        # 按时间倒序
        backups.sort(key=lambda x: x['created'], reverse=True)

        return {
            'success': True,
            'backups': backups
        }
    except Exception as e:
        logger.error(f"获取备份列表失败: {e}", exc_info=True)
        return {'success': True, 'backups': []}


@router.post("/clean", response_model=CleanResult)
async def clean_data():
    """
    清洗数据：删除重复记录
    """
    try:
        conn = get_db_connection()
        c = conn.cursor()

        # 获取清洗前的记录数
        c.execute('SELECT COUNT(*) FROM rebar_prices')
        before_count = c.fetchone()[0]

        # 找出重复记录
        c.execute('''
            SELECT COUNT(*) FROM rebar_prices
            WHERE id NOT IN (
                SELECT MIN(id) FROM rebar_prices
                GROUP BY date, material_name, spec, brand, price
            )
        ''')
        duplicate_count = c.fetchone()[0]

        # 删除重复记录（保留每组的第一条）
        c.execute('''
            DELETE FROM rebar_prices
            WHERE id NOT IN (
                SELECT MIN(id) FROM rebar_prices
                GROUP BY date, material_name, spec, brand, price
            )
        ''')
        conn.commit()

        # 获取清洗后的记录数
        c.execute('SELECT COUNT(*) FROM rebar_prices')
        after_count = c.fetchone()[0]

        # 清理空闲空间
        c.execute('VACUUM')
        conn.close()

        logger.info(f"[clean_data] 清洗完成 | 删除重复: {duplicate_count} | 剩余: {after_count}")

        return CleanResult(
            success=True,
            removed_duplicates=duplicate_count,
            fixed_records=0,
            remaining_records=after_count
        )
    except Exception as e:
        logger.error(f"清洗数据失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/import")
async def import_data(file_path: str = Query(..., description="要导入的文件路径"), account: str = Depends(get_current_account)):
    """
    导入数据（从已存在的文件）
    注意：实际的文件上传功能需要通过其他方式实现
    """
    logger.info(f"[import_data] 导入 | file={file_path} | by={account}")
    try:
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="文件不存在")

        conn = get_db_connection()
        c = conn.cursor()

        imported = 0
        skipped = 0
        errors = []

        if file_path.endswith('.csv'):
            import csv
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    try:
                        c.execute('''
                            INSERT INTO rebar_prices (date, material_name, spec, material_type, brand, price, region, fetch_time)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            row.get('日期', row.get('date', '')),
                            row.get('品名', row.get('material_name', '')),
                            row.get('规格', row.get('spec', '')),
                            row.get('材质', row.get('material_type', '')),
                            row.get('品牌', row.get('brand', '')),
                            int(row.get('价格', row.get('price', 0))),
                            row.get('地区', row.get('region', '山东烟台')),
                            row.get('抓取时间', row.get('fetch_time', ''))
                        ))
                        imported += 1
                    except sqlite3.IntegrityError:
                        skipped += 1
                    except Exception as e:
                        errors.append(str(e))

        conn.commit()
        conn.close()

        return ImportResult(
            success=True,
            imported=imported,
            skipped=skipped,
            errors=errors[:10]  # 只返回前10个错误
        )
    except Exception as e:
        logger.error(f"导入数据失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
