import openpyxl

# 检查 完整版
wb1 = openpyxl.load_workbook("services/data/山东烟台钢筋价格_完整版.xlsx", read_only=True, data_only=True)
print("=== 完整版 ===")
print(f"Sheets: {len(wb1.sheetnames)}")
print(f"前5个: {wb1.sheetnames[:5]}")
wb1.close()

# 检查 current
wb2 = openpyxl.load_workbook("services/data/山东烟台钢筋价格_current.xlsx", read_only=True, data_only=True)
print()
print("=== current ===")
print(f"Sheets: {len(wb2.sheetnames)}")
print(f"前5个: {wb2.sheetnames[:5]}")
wb2.close()
