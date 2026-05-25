# -*- coding: utf-8 -*-
import openpyxl
import json

file_path = r'C:\Users\admin\Desktop\西安神州数码科技园22F办公室装修项目-四家投标单位报价对比.xlsx'

wb = openpyxl.load_workbook(file_path)
result = {}
result['sheets'] = wb.sheetnames

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    result[sheet_name] = {
        'rows': ws.max_row,
        'cols': ws.max_column,
        'data': []
    }

    for row in range(1, min(150, ws.max_row + 1)):
        row_data = {}
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val is not None:
                row_data[str(col)] = str(val)
        if row_data:
            result[sheet_name]['data'].append({'row': row, 'cells': row_data})

with open('excel_data.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print('Data saved to excel_data.json')